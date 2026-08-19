from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import Decimal
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.api.afg import apply_afg_route_requests
from app.database import Base
from app.models.enums import (
    MetalTypeEnum,
    PosDocumentTypeEnum,
    PosRateSourceEnum,
    PosTradeSideEnum,
    ProductStatusEnum,
    ProductTypeEnum,
    RoleEnum,
)
from app.models.pos_document import PosDocument
from app.models.pos_session import PosSession
from app.models.product import Product
from app.models.transaction import Transaction
from app.models.transaction_line import TransactionLine
from app.models.user import User
from app.schemas.afg import (
    AfgLogBucketOut,
    AfgLogBucketSummaryOut,
    AfgLogWorkspaceOut,
    AfgMeltLotOut,
    AfgMeltQueueOut,
    AfgRouteRequest,
    AfgWorkspaceDocumentOut,
    AfgWorkspaceLineOut,
    AfgWorkspaceSummaryOut,
)
from app.services.document_artifact_service import (
    ArtifactSyncContext,
    LOG_CONTROL_SHEET,
    LOG_SHEET,
    _build_log_workbook_bytes,
    parse_log_workbook_inputs_from_workbook,
)
from app.utils.helpers import utc_now


def _workspace_line(*, metal_type: str, line_no: int, destination: str | None = "undecided") -> AfgWorkspaceLineOut:
    now = utc_now()
    return AfgWorkspaceLineOut(
        id=uuid4(),
        transaction_id=uuid4(),
        document_sequence_no=2,
        document_number="SG-2026-000002",
        session_id=uuid4(),
        session_code="LOGARK1",
        line_no=line_no,
        customer_name="Denis Thor Chrisistensen",
        customer_phone="24917296",
        customer_email=None,
        issued_at=now,
        product_id=None,
        product_number=None,
        reference_number=f"REF-{line_no:03d}",
        product_type="jewelry",
        metal_type=metal_type,
        weight_grams=Decimal("15.00"),
        purity_karat="22",
        purity_percentage=Decimal("91.70"),
        pure_gold_grams=Decimal("13.76"),
        rate_dkk=Decimal("859.48"),
        margin_percent=Decimal("0.00"),
        line_total_dkk=Decimal("8787.14"),
        product_status=None,
        operation_destination=destination,
        operation_classification="standard",
        is_gdpr_locked=False,
        product_notes="Mevcut not",
        created_at=now,
    )


def _workspace_document(line: AfgWorkspaceLineOut) -> AfgWorkspaceDocumentOut:
    now = utc_now()
    return AfgWorkspaceDocumentOut(
        sequence_no=2,
        document_number="SG-2026-000002",
        session_id=line.session_id,
        document_kind="afregningsbilag",
        document_title="Afregningsbilag",
        status="confirmed",
        trade_side="buy_from_customer",
        customer_name=line.customer_name,
        customer_phone=line.customer_phone,
        customer_email=line.customer_email,
        customer_address=None,
        issued_at=now,
        confirmed_at=now,
        gross_amount_dkk=Decimal("8787.14"),
        net_amount_dkk=Decimal("8787.14"),
        total_weight_grams=Decimal("15.00"),
        total_pure_gold_grams=Decimal("13.76"),
        line_count=1,
        operation_state="awaiting_decision",
        has_locked_products=False,
        lines=[line],
    )


def _workspace_bucket(*, lines: list[AfgWorkspaceLineOut], metal_bucket: str) -> AfgLogBucketOut:
    lot = AfgMeltLotOut(
        id=uuid4(),
        metal_bucket=metal_bucket,
        sent_date=date(2026, 3, 30),
        purchased_from_date=date(2026, 3, 30),
        before_weight_grams=Decimal("15.00"),
        before_amount_dkk=Decimal("8787.14"),
        before_pure_gold_grams=Decimal("13.76"),
        after_pure_gold_grams=Decimal("13.00"),
        insurance_dkk=Decimal("0.00"),
        shipping_dkk=Decimal("0.00"),
        refining_dkk=Decimal("0.00"),
        sale_date=None,
        quote_eur=None,
        exchange_rate_dkk=Decimal("7.45"),
        payout_total_dkk=None,
        notes=None,
        cost_total_dkk=Decimal("0.00"),
        estimated_sale_value_dkk=None,
        net_after_costs_dkk=None,
        bridge_difference_dkk=None,
        advance_per_gram_dkk=None,
        created_at=utc_now(),
        updated_at=utc_now(),
    )
    documents = [_workspace_document(line) for line in lines]
    return AfgLogBucketOut(
        metal_bucket=metal_bucket,
        summary=AfgLogBucketSummaryOut(
            total_documents=len(documents),
            total_lines=len(lines),
            awaiting_lines=len(lines),
            routed_lines=0,
            split_line_count=0,
            melt_line_count=0,
            melt_lot_count=1,
            total_weight_grams=Decimal("15.00") if lines else Decimal("0.00"),
            total_pure_gold_grams=Decimal("13.76") if lines else Decimal("0.00"),
            total_amount_dkk=Decimal("8787.14") if lines else Decimal("0.00"),
        ),
        documents=documents,
        split_groups=[],
        melt_queue=AfgMeltQueueOut(
            line_count=1 if lines else 0,
            total_weight_grams=Decimal("15.00") if lines else Decimal("0.00"),
            total_pure_gold_grams=Decimal("13.76") if lines else Decimal("0.00"),
            total_amount_dkk=Decimal("8787.14") if lines else Decimal("0.00"),
            earliest_purchase_date=date(2026, 3, 30) if lines else None,
            latest_purchase_date=date(2026, 3, 30) if lines else None,
            document_numbers=["SG-2026-000002"] if lines else [],
        ),
        melt_lots=[lot] if lines else [],
    )


def test_log_ark1_workbook_round_trip_uses_ark1_sheet_inputs():
    gold_line = _workspace_line(metal_type="yellow_gold", line_no=1)
    workspace = AfgLogWorkspaceOut(
        summary=AfgWorkspaceSummaryOut(total_documents=1, total_amount_dkk=Decimal("8787.14"), total_pure_gold_grams=Decimal("13.76")),
        gold=_workspace_bucket(lines=[gold_line], metal_bucket="gold"),
        silver=_workspace_bucket(lines=[], metal_bucket="silver"),
    )

    workbook_bytes = _build_log_workbook_bytes(
        workspace,
        year=2026,
        sync_context=ArtifactSyncContext(
            kind="log",
            key="2026",
            artifact_key="log.live.2026",
            base_version="1",
            contract_version="log-v2",
        ),
    )

    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    assert LOG_SHEET in workbook.sheetnames
    assert LOG_CONTROL_SHEET not in workbook.sheetnames
    assert workbook.active.title == LOG_SHEET

    sheet = workbook[LOG_SHEET]
    sheet["G10"] = "S"
    sheet["B41"] = 200

    buffer = __import__("io").BytesIO()
    workbook.save(buffer)
    parsed = parse_log_workbook_inputs_from_workbook(buffer.getvalue(), year=2026)

    assert len(parsed.route_updates) == 1
    assert parsed.route_updates[0].payload.destination == "inventory"
    assert parsed.route_updates[0].payload.classification == "standard"
    assert parsed.route_updates[0].payload.note == "Mevcut not"
    assert len(parsed.lot_updates) == 1
    assert parsed.lot_updates[0].payload.insurance_dkk == Decimal("200")


@pytest.mark.asyncio
async def test_log_route_batch_apply_is_atomic_on_failure():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
    async with Session() as session:
        admin = User(
            email="log-admin@test.local",
            password_hash="x",
            name="Admin",
            role=RoleEnum.ADMIN,
        )
        customer = User(
            email="log-customer@test.local",
            password_hash="x",
            name="Customer",
            role=RoleEnum.CUSTOMER,
            phone="24917296",
        )
        session.add_all([admin, customer])
        await session.flush()

        pos_session = PosSession(
            session_code="LOGATMC1",
            display_token="display-log-atomic",
            clerk_user_id=admin.id,
            customer_id=customer.id,
            trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
            rate_source=PosRateSourceEnum.LIVE,
        )
        session.add(pos_session)
        await session.flush()

        document = PosDocument(
            pos_session_id=pos_session.id,
            document_type=PosDocumentTypeEnum.PURCHASE_RECEIPT,
            gross_amount_dkk=Decimal("20000.00"),
            net_amount_dkk=Decimal("20000.00"),
            customer_name=customer.name,
            customer_phone=customer.phone,
        )
        session.add(document)
        await session.flush()

        transaction = Transaction(
            pos_session_id=pos_session.id,
            pos_document_sequence_no=document.sequence_no,
            trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER.value,
            status="confirmed",
            customer_id=customer.id,
            clerk_user_id=admin.id,
            gross_amount_dkk=Decimal("20000.00"),
            net_amount_dkk=Decimal("20000.00"),
            confirmed_at=utc_now(),
        )
        session.add(transaction)
        await session.flush()

        line_one = TransactionLine(
            transaction_id=transaction.id,
            line_no=1,
            product_type="jewelry",
            metal_type="yellow_gold",
            weight_grams=Decimal("15.00"),
            purity_karat="22",
            purity_percentage=Decimal("91.70"),
            pure_gold_grams=Decimal("13.76"),
            rate_dkk=Decimal("859.48"),
            margin_percent=Decimal("0.00"),
            line_total_dkk=Decimal("8787.14"),
        )
        line_two = TransactionLine(
            transaction_id=transaction.id,
            line_no=2,
            product_type="jewelry",
            metal_type="yellow_gold",
            weight_grams=Decimal("10.00"),
            purity_karat="21.6",
            purity_percentage=Decimal("90.00"),
            pure_gold_grams=Decimal("9.00"),
            rate_dkk=Decimal("842.56"),
            margin_percent=Decimal("0.00"),
            line_total_dkk=Decimal("8425.60"),
        )
        session.add_all([line_one, line_two])
        await session.commit()
        line_one_id = line_one.id
        line_two_id = line_two.id

        # Atomiklik tetikleyicisi: ikinci istekte ürün oluşturma kasıtlı
        # patlatılır. (Eski tetikleyici GDPR eritme engeliydi; 0.3.8'de o engel
        # bilgiye dönüştü — testin amacı olan atomiklik aynı kalır.)
        from app.api import afg as afg_module

        real_create = afg_module.create_product_service
        call_count = {"n": 0}

        async def failing_create(*args, **kwargs):
            call_count["n"] += 1
            if call_count["n"] >= 2:
                raise RuntimeError("test: ikinci ürün oluşturma kasıtlı patladı")
            return await real_create(*args, **kwargs)

        afg_module.create_product_service = failing_create
        try:
            with pytest.raises(Exception) as exc_info:
                await apply_afg_route_requests(
                    db=session,
                    route_requests=[
                        AfgRouteRequest(line_ids=[line_one_id], destination="inventory", classification="standard"),
                        AfgRouteRequest(line_ids=[line_two_id], destination="inventory", classification="standard"),
                    ],
                    actor_id=admin.id,
                )
        finally:
            afg_module.create_product_service = real_create

        assert "kasıtlı" in str(exc_info.value).lower()

        product_count = await session.scalar(select(func.count()).select_from(Product))
        assert product_count == 0

        reloaded_lines = (
            await session.execute(
                select(TransactionLine).where(TransactionLine.id.in_([line_one_id, line_two_id])).order_by(TransactionLine.line_no.asc())
            )
        ).scalars().all()
        assert all(item.product_id is None for item in reloaded_lines)

    await engine.dispose()


@pytest.mark.asyncio
async def test_afg_route_apply_clears_existing_product_note_when_workbook_note_is_blank():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
    async with Session() as session:
        admin = User(
            email="route-clear-admin@test.local",
            password_hash="x",
            name="Admin",
            role=RoleEnum.ADMIN,
        )
        customer = User(
            email="route-clear-customer@test.local",
            password_hash="x",
            name="Customer",
            role=RoleEnum.CUSTOMER,
            phone="24917296",
        )
        session.add_all([admin, customer])
        await session.flush()

        pos_session = PosSession(
            session_code="ROUTCLR1",
            display_token="display-route-clear",
            clerk_user_id=admin.id,
            customer_id=customer.id,
            trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
            rate_source=PosRateSourceEnum.LIVE,
        )
        session.add(pos_session)
        await session.flush()

        issued_at = utc_now() - timedelta(days=30)
        document = PosDocument(
            pos_session_id=pos_session.id,
            document_type=PosDocumentTypeEnum.PURCHASE_RECEIPT,
            issued_at=issued_at,
            gross_amount_dkk=Decimal("20000.00"),
            net_amount_dkk=Decimal("20000.00"),
            customer_name=customer.name,
            customer_phone=customer.phone,
        )
        session.add(document)
        await session.flush()

        transaction = Transaction(
            pos_session_id=pos_session.id,
            pos_document_sequence_no=document.sequence_no,
            trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER.value,
            status="confirmed",
            customer_id=customer.id,
            clerk_user_id=admin.id,
            gross_amount_dkk=Decimal("20000.00"),
            net_amount_dkk=Decimal("20000.00"),
            confirmed_at=issued_at,
        )
        session.add(transaction)
        await session.flush()

        product = Product(
            product_number="CL01",
            display_name="Route clear gold",
            product_type=ProductTypeEnum.JEWELRY,
            metal_type=MetalTypeEnum.YELLOW_GOLD,
            weight_grams=Decimal("15.00"),
            purity_karat="22K",
            purity_percentage=Decimal("91.70"),
            pure_gold_grams=Decimal("13.76"),
            unit_count=1,
            total_weight_grams=Decimal("15.00"),
            purchase_date=issued_at,
            purchase_price_dkk=Decimal("20000.00"),
            gold_rate_at_purchase=Decimal("859.48"),
            commission=Decimal("0.00"),
            seller_customer_id=customer.id,
            gdpr_release_date=issued_at + timedelta(days=14),
            is_gdpr_locked=False,
            status=ProductStatusEnum.IN_INVENTORY,
            notes="Eski workbook notu",
        )
        session.add(product)
        await session.flush()

        line = TransactionLine(
            transaction_id=transaction.id,
            line_no=1,
            product_id=product.id,
            product_type="jewelry",
            metal_type="yellow_gold",
            weight_grams=Decimal("15.00"),
            purity_karat="22K",
            purity_percentage=Decimal("91.70"),
            pure_gold_grams=Decimal("13.76"),
            rate_dkk=Decimal("859.48"),
            margin_percent=Decimal("0.00"),
            line_total_dkk=Decimal("20000.00"),
        )
        session.add(line)
        await session.commit()

        await apply_afg_route_requests(
            db=session,
            route_requests=[AfgRouteRequest(line_ids=[line.id], destination="inventory", classification="standard")],
            actor_id=admin.id,
        )

        await session.refresh(product)
        assert product.notes is None

    await engine.dispose()
