"""M3 — inventory analiz özetindeki skipped_zero_stock ARTIK GERÇEK sayaç.

Eski özet max(0, 420 - len(rows)) uydurmasını içeriyordu (ilk test
dosyasının 420 satırlık sabitinden kalmış); operatör taşıma kararını bu
özete göre verdiği için atlanan sıfır stok satırlarının gerçek sayısı
raporlanmalı.
"""

from io import BytesIO

from openpyxl import Workbook

from app.services.legacy_migration_service import parse_legacy_inventory


def _build_lager_workbook(zero_stock_rows: int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lager"
    # Market snapshot hücreleri (K4-K7).
    sheet["K4"] = 450.0
    sheet["K5"] = 8.5
    sheet["K6"] = 950.0
    sheet["K7"] = 900.0
    # Bölüm başlığı (ilk 4 sütun taranır).
    sheet["A2"] = "Guldbarrer"
    # İki geçerli külçe satırı: A=ref, C=ad, D=birim gram, E=adet, F=toplam gram, H=alış, K=shop.
    sheet["A3"] = "1234"
    sheet["C3"] = "Barre 50g"
    sheet["D3"] = 50.0
    sheet["E3"] = 1
    sheet["F3"] = 50.0
    sheet["H3"] = 25000
    sheet["K3"] = 32000
    sheet["A4"] = "1235"
    sheet["C4"] = "Barre 1g"
    sheet["D4"] = 1.0
    sheet["E4"] = 5
    sheet["F4"] = 5.0
    sheet["H4"] = 2000
    row = 5
    for _ in range(zero_stock_rows):
        # Sıfır adet/gram satır — parse tarafından atlanır.
        sheet.cell(row=row, column=1, value="")
        sheet.cell(row=row, column=4, value=0)
        sheet.cell(row=row, column=5, value=0)
        sheet.cell(row=row, column=6, value=0)
        row += 1

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_skipped_zero_stock_counts_real_rows() -> None:
    records, _snapshot, skipped = parse_legacy_inventory(_build_lager_workbook(3), "hash1")

    assert len(records) == 2  # yalnız geçerli satırlar
    assert skipped == 3  # atlanan sıfır stok satırlarının GERÇEK sayısı


def test_skipped_zero_stock_is_zero_when_no_rows_skipped() -> None:
    records, _snapshot, skipped = parse_legacy_inventory(_build_lager_workbook(0), "hash2")

    assert len(records) == 2
    assert skipped == 0
