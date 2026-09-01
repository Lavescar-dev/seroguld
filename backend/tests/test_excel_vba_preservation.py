from __future__ import annotations

import io
import zipfile
from pathlib import Path

from openpyxl import load_workbook

REFERENCE_ROOT = Path(__file__).resolve().parents[2] / "referans"
AFG_TEMPLATE = REFERENCE_ROOT / "Afregningsbilag ( alis frontumuz).xlsm"


def _zip_names(content: bytes) -> set[str]:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        return set(archive.namelist())


def test_xlsm_vba_project_survives_openpyxl_roundtrip() -> None:
    original = AFG_TEMPLATE.read_bytes()
    assert "xl/vbaProject.bin" in _zip_names(original)

    # AFG üretici/senkron yolu .xlsm'i keep_vba=True ile açıp kaydeder;
    # bu turda VBA projesi kaybolmamalı.
    workbook = load_workbook(io.BytesIO(original), keep_vba=True, data_only=False)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    assert "xl/vbaProject.bin" in _zip_names(buffer.getvalue())


def test_hidden_row_flag_survives_save_and_keeps_vba() -> None:
    """Dinamik satır görünürlüğü: hidden bayrağıyla kayıt VBA'yı bozmaz ve
    bayrak tura (round-trip) sağ kalır — Excel'de satır gizli açılır."""
    workbook = load_workbook(io.BytesIO(AFG_TEMPLATE.read_bytes()), keep_vba=True, data_only=False)
    sheet = workbook["Afregningsbilag"]
    sheet.row_dimensions[22].hidden = True
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()

    assert "xl/vbaProject.bin" in _zip_names(buffer.getvalue())
    reloaded = load_workbook(io.BytesIO(buffer.getvalue()), keep_vba=True, data_only=False)
    assert reloaded["Afregningsbilag"].row_dimensions[22].hidden is True
    reloaded.close()
