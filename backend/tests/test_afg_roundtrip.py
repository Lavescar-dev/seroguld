from __future__ import annotations

import asyncio
import io
from decimal import Decimal

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.api.v2 import _apply_afg_workspace_artifact_inputs
from app.api.pos import get_pos_document_detail
from app.database import Base
from app.models.enums import (
    MetalTypeEnum,
    PosDocumentTypeEnum,
    PosRateSourceEnum,
    PosSessionStatusEnum,
    PosTradeSideEnum,
    ProductTypeEnum,
    RoleEnum,
)
from app.models.pos_document import PosDocument
from app.models.pos_session import PosSession
from app.models.pos_session_line import PosSessionLine
from app.models.transaction import Transaction
from app.models.transaction_line import TransactionLine
from app.models.user import User
from app.schemas.pos import (
    PosWorkspaceBankInfo,
    PosWorkspaceCustomerUpdate,
    PosWorkspaceFinalizeRequest,
    PosWorkspaceGoldRowInput,
    PosWorkspaceInvoiceGoldRowInput,
    PosWorkspaceInvoiceGoldSheetUpdate,
    PosWorkspaceInvoiceMiscRowInput,
    PosWorkspaceInvoiceMiscSheetUpdate,
    PosWorkspaceMarketRates,
    PosWorkspaceNumberingUpdate,
    PosWorkspaceSectionsUpdate,
)
from app.services.document_artifact_service import (
    AFG_FACTURA_GOLD_FOOTER_START,
    AFG_FACTURA_GOLD_ROW_END,
    AFG_FACTURA_GOLD_ROW_START,
    AFG_FACTURA_GOLD_SHEET,
    AFG_FACTURA_MISC_ROW_END,
    AFG_FACTURA_MISC_ROW_START,
    AFG_FACTURA_MISC_SHEET,
    AFG_VARIABLES_SHEET,
    ArtifactSyncContext,
    _build_afg_workbook_bytes_from_detail,
    _build_afg_workbook_bytes_from_workspace,
    parse_afg_workspace_inputs_from_workbook,
)
from app.services.pos_service import (
    finalize_purchase_workspace,
    open_purchase_document_for_edit,
    replace_purchase_workspace_sections,
    update_purchase_workspace_customer,
    update_purchase_workspace_draft_customer,
)
from app.utils.helpers import quantize_2, to_decimal


def _save_workbook_bytes(workbook) -> bytes:
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _workspace_sections() -> PosWorkspaceSectionsUpdate:
    return PosWorkspaceSectionsUpdate(
        gold_rows=[
            PosWorkspaceGoldRowInput(
                karat=Decimal("24.0"),
                gram=Decimal("22.00"),
                avance_percent=Decimal("0.00"),
            )
        ],
        silver_rows=[],
        bank_info=PosWorkspaceBankInfo(reg_number="5512", account_number="0725397984"),
        market_rates=PosWorkspaceMarketRates(
            gold_24k_dkk=Decimal("937.99"),
            silver_dkk=Decimal("14.56"),
        ),
        payment_method="bank",
        afg_note="KDV ve Uniconta notu",
        purchase_vat_enabled=True,
        purchase_vat_rate_percent=Decimal("25.00"),
        numbering=PosWorkspaceNumberingUpdate(
            afregnings_number_next="1003",
            invoice_number_next="1001",
        ),
        invoice_gold=PosWorkspaceInvoiceGoldSheetUpdate(
            rows=[
                PosWorkspaceInvoiceGoldRowInput(
                    row_key="invoice_gold:1",
                    code="3",
                    fineness="925",
                    gram=Decimal("2.50"),
                )
            ],
            footer_lines=["Line A", "Line B", "Line C"],
        ),
        invoice_misc=PosWorkspaceInvoiceMiscSheetUpdate(
            rows=[
                PosWorkspaceInvoiceMiscRowInput(
                    row_key="invoice_misc:1",
                    text="Polering",
                    quantity=Decimal("2.00"),
                    unit_price_dkk=Decimal("125.50"),
                )
            ]
        ),
    )


def _create_users() -> tuple[User, User]:
    clerk = User(
        email="afg-roundtrip-clerk@test.local",
        password_hash="x",
        name="Clerk",
        role=RoleEnum.ADMIN,
    )
    customer = User(
        email="afg-roundtrip-customer@test.local",
        password_hash="x",
        name="Customer",
        role=RoleEnum.CUSTOMER,
        phone="24917296",
        postal_code="7160",
    )
    return clerk, customer


def test_afg_workspace_workbook_round_trip_preserves_companion_sheet_inputs():
    async def run() -> None:
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
        async with Session() as session:
            clerk, customer = _create_users()
            session.add_all([clerk, customer])
            await session.flush()

            pos_session = PosSession(
                session_code="AFGRT100",
                display_token="display-afg-roundtrip-workspace",
                clerk_user_id=clerk.id,
                customer_id=customer.id,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
                live_rate_dkk=Decimal("937.99"),
                rate_source=PosRateSourceEnum.LIVE,
                margin_percent_internal=Decimal("0.00"),
                status=PosSessionStatusEnum.DRAFT,
                visible_snapshot={},
            )
            session.add(pos_session)
            await session.commit()
            await session.refresh(pos_session)
            session_id = pos_session.id
            session_id = pos_session.id

            workspace = await replace_purchase_workspace_sections(
                session,
                pos_session=pos_session,
                payload=_workspace_sections(),
            )

            workbook_bytes = _build_afg_workbook_bytes_from_workspace(
                workspace,
                    sync_context=ArtifactSyncContext(
                        kind="alis-workspace",
                        key=str(session_id),
                        artifact_key=f"alis.workspace.{session_id}",
                        base_version="",
                    ),
                )
            parsed = parse_afg_workspace_inputs_from_workbook(workbook_bytes)

            assert parsed.sections.numbering is not None
            assert parsed.sections.numbering.afregnings_number_next == "1003"
            assert parsed.sections.numbering.invoice_number_next == "1001"
            assert parsed.sections.market_rates is not None
            assert parsed.sections.market_rates.gold_24k_dkk == Decimal("937.99")
            assert parsed.sections.market_rates.silver_dkk == Decimal("14.56")
            assert parsed.sections.purchase_vat_enabled is True
            assert parsed.sections.purchase_vat_rate_percent == Decimal("25.00")
            assert parsed.sections.afg_note == "KDV ve Uniconta notu"

            invoice_gold_row = next(
                row for row in parsed.sections.invoice_gold.rows if row.row_key == "invoice_gold:1"
            )
            invoice_misc_row = next(
                row for row in parsed.sections.invoice_misc.rows if row.row_key == "invoice_misc:1"
            )

            assert invoice_gold_row.code == "3"
            assert invoice_gold_row.fineness == "925"
            assert invoice_gold_row.gram == Decimal("2.50")
            assert parsed.sections.invoice_gold.footer_lines == ["Line A", "Line B", "Line C"]

            assert invoice_misc_row.text == "Polering"
            assert invoice_misc_row.quantity == Decimal("2.00")
            assert invoice_misc_row.unit_price_dkk == Decimal("125.50")

        await engine.dispose()

    asyncio.run(run())


def test_afg_workspace_callback_applies_lines_without_selected_customer_when_customer_fields_blank():
    async def run() -> None:
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
        async with Session() as session:
            clerk, _ = _create_users()
            session.add(clerk)
            await session.flush()

            pos_session = PosSession(
                session_code="AFGRT150",
                display_token="display-afg-roundtrip-no-customer",
                clerk_user_id=clerk.id,
                customer_id=None,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
                live_rate_dkk=Decimal("937.99"),
                rate_source=PosRateSourceEnum.LIVE,
                margin_percent_internal=Decimal("0.00"),
                status=PosSessionStatusEnum.DRAFT,
                visible_snapshot={},
            )
            session.add(pos_session)
            await session.commit()
            await session.refresh(pos_session)
            session_id = pos_session.id

            workspace = await replace_purchase_workspace_sections(
                session,
                pos_session=pos_session,
                payload=_workspace_sections(),
            )

            workbook_bytes = _build_afg_workbook_bytes_from_workspace(
                workspace,
                sync_context=ArtifactSyncContext(
                    kind="alis-workspace",
                    key=str(session_id),
                    artifact_key=f"alis.workspace.{session_id}",
                    base_version="",
                ),
            )

            updated_workspace = await _apply_afg_workspace_artifact_inputs(
                session,
                pos_session=pos_session,
                workbook_bytes=workbook_bytes,
            )

            assert updated_workspace.customer.customer_id is None
            assert updated_workspace.customer.name == ""
            assert updated_workspace.market_rates.gold_24k_dkk == Decimal("937.99")
            assert any(row.gram == Decimal("22.00") for row in updated_workspace.gold_rows)

        await engine.dispose()

    asyncio.run(run())


def test_afg_workspace_shadow_customer_round_trip_without_linked_customer():
    async def run() -> None:
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
        async with Session() as session:
            clerk, _ = _create_users()
            session.add(clerk)
            await session.flush()

            pos_session = PosSession(
                session_code="AFGRT160",
                display_token="display-afg-roundtrip-shadow",
                clerk_user_id=clerk.id,
                customer_id=None,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
                live_rate_dkk=Decimal("937.99"),
                rate_source=PosRateSourceEnum.LIVE,
                margin_percent_internal=Decimal("0.00"),
                status=PosSessionStatusEnum.DRAFT,
                visible_snapshot={},
            )
            session.add(pos_session)
            await session.commit()
            await session.refresh(pos_session)
            session_id = pos_session.id

            workspace = await update_purchase_workspace_draft_customer(
                session,
                pos_session=pos_session,
                payload=PosWorkspaceCustomerUpdate(
                    name="Denis Thor Chrisistensen",
                    phone="+45 22 25 55 04",
                    address="Valby Langgade 84",
                    postal_code="2500",
                    cpr_number="0102031234",
                    identity_doc_number="P1234567",
                ),
            )

            assert workspace.customer.customer_id is None
            assert workspace.customer.name == "Denis Thor Chrisistensen"
            assert workspace.customer.identity_doc_number == "P1234567"

            workspace = await replace_purchase_workspace_sections(
                session,
                pos_session=pos_session,
                payload=_workspace_sections(),
            )

            workbook_bytes = _build_afg_workbook_bytes_from_workspace(
                workspace,
                sync_context=ArtifactSyncContext(
                    kind="alis-workspace",
                    key=str(session_id),
                    artifact_key=f"alis.workspace.{session_id}",
                    base_version="",
                ),
            )
            workbook = load_workbook(io.BytesIO(workbook_bytes), keep_vba=True)
            sheet = workbook["Afregningsbilag"]
            sheet["D16"] = "Ada Lovelace"
            sheet["G17"] = "ZXCV-9876"
            sheet["G18"] = "+45 33 44 55 66"
            updated_workspace = await _apply_afg_workspace_artifact_inputs(
                session,
                pos_session=pos_session,
                workbook_bytes=_save_workbook_bytes(workbook),
            )

            assert updated_workspace.customer.customer_id is None
            assert updated_workspace.customer.name == "Ada Lovelace"
            assert updated_workspace.customer.identity_doc_number == "ZXCV-9876"
            assert updated_workspace.customer.phone == "+45 33 44 55 66"

        await engine.dispose()

    asyncio.run(run())


def test_linked_customer_clear_stays_session_local_and_reaches_final_document():
    async def run() -> None:
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
        async with Session() as session:
            clerk, customer = _create_users()
            session.add_all([clerk, customer])
            await session.flush()

            pos_session = PosSession(
                session_code="AFGRT170",
                display_token="display-afg-roundtrip-linked-clear",
                clerk_user_id=clerk.id,
                customer_id=customer.id,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
                live_rate_dkk=Decimal("937.99"),
                rate_source=PosRateSourceEnum.LIVE,
                margin_percent_internal=Decimal("0.00"),
                status=PosSessionStatusEnum.DRAFT,
                visible_snapshot={},
            )
            session.add(pos_session)
            await session.commit()
            await session.refresh(pos_session)

            await update_purchase_workspace_customer(
                session,
                pos_session=pos_session,
                payload=PosWorkspaceCustomerUpdate(
                    name=None,
                    email=None,
                    phone=None,
                    address=None,
                    postal_code=None,
                    city=None,
                    cpr_number=None,
                    identity_doc_number=None,
                ),
            )
            assert customer.name == "Customer"
            assert customer.phone == "24917296"

            workspace = await replace_purchase_workspace_sections(
                session,
                pos_session=pos_session,
                payload=_workspace_sections(),
            )
            assert workspace.customer.customer_id == customer.id
            assert workspace.customer.name == ""
            assert workspace.customer.phone is None

            finalized = await finalize_purchase_workspace(
                session,
                pos_session=pos_session,
                payload=PosWorkspaceFinalizeRequest(notes="linked clear"),
            )
            document = await session.get(PosDocument, finalized.document_sequence_no)
            assert document is not None
            assert document.customer_name is None
            assert document.customer_phone is None
            assert document.customer_address is None
            assert document.vat_rate_percent == Decimal("25.00")
            assert document.vat_amount_dkk == quantize_2(document.net_amount_dkk * Decimal("0.25"))
            assert document.gross_amount_dkk == document.net_amount_dkk + document.vat_amount_dkk

        await engine.dispose()

    asyncio.run(run())


def test_afg_document_detail_and_final_workbook_keep_structured_market_rates():
    async def run() -> None:
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
        async with Session() as session:
            clerk, customer = _create_users()
            session.add_all([clerk, customer])
            await session.flush()

            pos_session = PosSession(
                session_code="AFGRT200",
                display_token="display-afg-roundtrip-detail",
                clerk_user_id=clerk.id,
                customer_id=customer.id,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
                live_rate_dkk=Decimal("937.99"),
                rate_source=PosRateSourceEnum.LIVE,
                margin_percent_internal=Decimal("0.00"),
                status=PosSessionStatusEnum.DRAFT,
                visible_snapshot={},
            )
            session.add(pos_session)
            await session.commit()
            await session.refresh(pos_session)
            session_id = pos_session.id

            workspace = await replace_purchase_workspace_sections(
                session,
                pos_session=pos_session,
                payload=_workspace_sections(),
            )
            pos_session = await session.get(PosSession, pos_session.id)
            assert pos_session is not None
            pos_session.status = PosSessionStatusEnum.CONFIRMED
            await session.flush()

            pos_document = PosDocument(
                pos_session_id=pos_session.id,
                document_type=PosDocumentTypeEnum.PURCHASE_RECEIPT,
                gross_amount_dkk=workspace.summary.total_amount_dkk,
                net_amount_dkk=workspace.summary.total_amount_dkk,
                vat_rate_percent=Decimal("0.00"),
                vat_amount_dkk=Decimal("0.00"),
                customer_name=customer.name,
                customer_phone=customer.phone,
                notes=pos_session.notes,
            )
            session.add(pos_document)
            await session.flush()

            transaction = Transaction(
                pos_session_id=pos_session.id,
                pos_document_sequence_no=pos_document.sequence_no,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER.value,
                status="confirmed",
                customer_id=customer.id,
                clerk_user_id=clerk.id,
                gross_amount_dkk=workspace.summary.total_amount_dkk,
                net_amount_dkk=workspace.summary.total_amount_dkk,
                vat_rate_percent=Decimal("0.00"),
                vat_amount_dkk=Decimal("0.00"),
            )
            session.add(transaction)
            await session.flush()
            session.add(
                TransactionLine(
                    transaction_id=transaction.id,
                    line_no=1,
                    product_type=ProductTypeEnum.JEWELRY.value,
                    metal_type=MetalTypeEnum.YELLOW_GOLD.value,
                    weight_grams=Decimal("22.00"),
                    purity_karat="24K",
                    purity_percentage=Decimal("99.90"),
                    pure_gold_grams=Decimal("21.98"),
                    rate_dkk=Decimal("937.99"),
                    margin_percent=Decimal("0.00"),
                    line_total_dkk=workspace.summary.total_amount_dkk,
                )
            )
            await session.commit()

            detail = await get_pos_document_detail(pos_document.sequence_no, db=session, _=clerk)
            silver_unit_price = quantize_2(Decimal("14.56") * (Decimal("925") / Decimal("999")))
            invoice_gold_row = next(row for row in detail.invoice_gold.rows if row.row_key == "invoice_gold:1")

            assert detail.market_rates.gold_24k_dkk == Decimal("937.99")
            assert detail.market_rates.silver_dkk == Decimal("14.56")
            assert invoice_gold_row.unit_price_dkk == silver_unit_price

            workbook_bytes = _build_afg_workbook_bytes_from_detail(
                detail,
                sync_context=ArtifactSyncContext(
                    kind="alis-document",
                    key=str(pos_document.sequence_no),
                    artifact_key=f"alis.document.{pos_document.sequence_no}",
                    base_version="",
                ),
            )
            workbook = load_workbook(io.BytesIO(workbook_bytes), keep_vba=True, data_only=False)
            vars_sheet = workbook[AFG_VARIABLES_SHEET]

            assert quantize_2(to_decimal(vars_sheet["C4"].value)) == Decimal("937.99")
            assert quantize_2(to_decimal(vars_sheet["C5"].value)) == Decimal("14.56")

        await engine.dispose()

    asyncio.run(run())


def test_open_purchase_document_for_edit_preserves_companion_sheet_state_after_finalize():
    async def run() -> None:
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
        async with Session() as session:
            clerk, customer = _create_users()
            session.add_all([clerk, customer])
            await session.flush()

            pos_session = PosSession(
                session_code="AFGRT300",
                display_token="display-afg-roundtrip-reopen",
                clerk_user_id=clerk.id,
                customer_id=customer.id,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
                live_rate_dkk=Decimal("937.99"),
                rate_source=PosRateSourceEnum.LIVE,
                margin_percent_internal=Decimal("0.00"),
                status=PosSessionStatusEnum.DRAFT,
                visible_snapshot={},
            )
            session.add(pos_session)
            await session.commit()
            await session.refresh(pos_session)

            await replace_purchase_workspace_sections(
                session,
                pos_session=pos_session,
                payload=_workspace_sections(),
            )
            pos_session = await session.get(PosSession, pos_session.id)
            assert pos_session is not None

            finalized = await finalize_purchase_workspace(
                session,
                pos_session=pos_session,
                payload=PosWorkspaceFinalizeRequest(notes="Roundtrip finalize"),
            )
            reopened = await open_purchase_document_for_edit(
                session,
                sequence_no=finalized.document_sequence_no,
                clerk_user=clerk,
            )

            assert reopened.numbering_preview.afregnings_number_next == "1003"
            assert reopened.numbering_preview.invoice_number_next == "1001"
            assert reopened.invoice_gold.rows[0].code == "3"
            assert reopened.invoice_gold.rows[0].fineness == "925"
            assert reopened.invoice_gold.rows[0].gram == Decimal("2.50")
            assert reopened.invoice_gold.footer_lines == ["Line A", "Line B", "Line C"]
            assert reopened.invoice_misc.rows[0].text == "Polering"
            assert reopened.invoice_misc.rows[0].quantity == Decimal("2.00")
            assert reopened.invoice_misc.rows[0].unit_price_dkk == Decimal("125.50")

        await engine.dispose()

    asyncio.run(run())


def test_blank_afg_callback_workbook_does_not_clear_existing_workspace():
    async def run() -> None:
        engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

        Session = async_sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)
        async with Session() as session:
            clerk, customer = _create_users()
            session.add_all([clerk, customer])
            await session.flush()

            pos_session = PosSession(
                session_code="AFGRT400",
                display_token="display-afg-roundtrip-blank-guard",
                clerk_user_id=clerk.id,
                customer_id=customer.id,
                trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
                live_rate_dkk=Decimal("937.99"),
                rate_source=PosRateSourceEnum.LIVE,
                margin_percent_internal=Decimal("0.00"),
                status=PosSessionStatusEnum.DRAFT,
                visible_snapshot={},
            )
            session.add(pos_session)
            await session.commit()
            await session.refresh(pos_session)
            session_id = pos_session.id

            workspace = await replace_purchase_workspace_sections(
                session,
                pos_session=pos_session,
                payload=PosWorkspaceSectionsUpdate(
                    gold_rows=[
                        PosWorkspaceGoldRowInput(
                            karat=Decimal("24.0"),
                            gram=Decimal("22.00"),
                            avance_percent=Decimal("0.00"),
                        )
                    ],
                    silver_rows=[],
                    market_rates=PosWorkspaceMarketRates(
                        gold_24k_dkk=Decimal("937.99"),
                        silver_dkk=Decimal("14.56"),
                    ),
                ),
            )
            workbook_bytes = _build_afg_workbook_bytes_from_workspace(
                workspace,
                sync_context=ArtifactSyncContext(
                    kind="alis-workspace",
                    key=str(session_id),
                    artifact_key=f"alis.workspace.{session_id}",
                    base_version="",
                ),
            )
            workbook = load_workbook(io.BytesIO(workbook_bytes), keep_vba=True, data_only=False)
            sheet = workbook["Afregningsbilag"]
            for row_idx in range(22, 35):
                sheet[f"F{row_idx}"] = 0
            for cell_ref in ["D16", "G16", "D17", "G17", "D18", "G18", "G19", "D41", "D42"]:
                sheet[cell_ref] = None
            factura_gold_sheet = workbook[AFG_FACTURA_GOLD_SHEET]
            for row_idx in range(AFG_FACTURA_GOLD_ROW_START, AFG_FACTURA_GOLD_ROW_END + 1):
                for col in ("A", "C", "E"):
                    factura_gold_sheet[f"{col}{row_idx}"] = None
            for row_idx in range(AFG_FACTURA_GOLD_FOOTER_START, AFG_FACTURA_GOLD_FOOTER_START + 3):
                factura_gold_sheet[f"B{row_idx}"] = None
            factura_misc_sheet = workbook[AFG_FACTURA_MISC_SHEET]
            for row_idx in range(AFG_FACTURA_MISC_ROW_START, AFG_FACTURA_MISC_ROW_END + 1):
                for col in ("C", "E", "F"):
                    factura_misc_sheet[f"{col}{row_idx}"] = None

            buffer = io.BytesIO()
            workbook.save(buffer)

            try:
                await _apply_afg_workspace_artifact_inputs(
                    session,
                    pos_session=pos_session,
                    workbook_bytes=buffer.getvalue(),
                )
            except HTTPException as exc:
                assert exc.status_code == 409
            else:
                raise AssertionError("Blank workbook callback should have been rejected")

            await session.rollback()
            persisted_lines = (
                await session.scalars(select(PosSessionLine).where(PosSessionLine.pos_session_id == session_id))
            ).all()
            assert len(persisted_lines) == 1
            assert persisted_lines[0].weight_grams == Decimal("22.00")

        await engine.dispose()

    asyncio.run(run())
