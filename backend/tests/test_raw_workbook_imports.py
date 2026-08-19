from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.schemas.afg import (
    AfgLogBucketOut,
    AfgLogBucketSummaryOut,
    AfgLogWorkspaceOut,
    AfgMeltLotOut,
    AfgMeltQueueOut,
    AfgWorkspaceDocumentOut,
    AfgWorkspaceLineOut,
    AfgWorkspaceSummaryOut,
)
from app.schemas.inventory import (
    InventoryGridRowOut,
    InventoryMarketPricesOut,
    InventoryWorkspaceOut,
    InventoryWorkspaceSummaryOut,
)
from app.services.document_artifact_service import (
    build_inventory_reconcile_preview,
    parse_afg_workspace_inputs_from_workbook,
    parse_inventory_workbook_inputs_from_workbook,
    parse_log_workbook_inputs_from_workbook,
)
from app.utils.helpers import utc_now


REFERENCE_ROOT = Path(__file__).resolve().parents[2] / "referans"


def _save_workbook_bytes(workbook) -> bytes:
    payload = io.BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def _raw_log_line(*, metal_type: str, line_no: int) -> AfgWorkspaceLineOut:
    now = utc_now()
    return AfgWorkspaceLineOut(
        id=uuid4(),
        transaction_id=uuid4(),
        document_sequence_no=2,
        document_number="SG-2026-000002",
        session_id=uuid4(),
        session_code="RAWLOG01",
        line_no=line_no,
        customer_name="Denis Thor Chrisistensen",
        customer_phone="24917296",
        customer_email=None,
        issued_at=now,
        product_id=None,
        product_number=None,
        reference_number=f"REF-{line_no:03d}",
        product_type="jewelry",
        metal_type=metal_type,
        weight_grams=Decimal("15.00"),
        purity_karat="22",
        purity_percentage=Decimal("91.70"),
        pure_gold_grams=Decimal("13.76"),
        rate_dkk=Decimal("859.48"),
        margin_percent=Decimal("0.00"),
        line_total_dkk=Decimal("8787.14"),
        product_status=None,
        operation_destination="undecided",
        operation_classification="standard",
        is_gdpr_locked=False,
        product_notes="Not korunmali",
        created_at=now,
    )


def _raw_log_document(line: AfgWorkspaceLineOut) -> AfgWorkspaceDocumentOut:
    now = utc_now()
    return AfgWorkspaceDocumentOut(
        sequence_no=2,
        document_number="SG-2026-000002",
        session_id=line.session_id,
        document_kind="afregningsbilag",
        document_title="Afregningsbilag",
        status="confirmed",
        trade_side="buy_from_customer",
        customer_name=line.customer_name,
        customer_phone=line.customer_phone,
        customer_email=line.customer_email,
        customer_address=None,
        issued_at=now,
        confirmed_at=now,
        gross_amount_dkk=Decimal("8787.14"),
        net_amount_dkk=Decimal("8787.14"),
        total_weight_grams=Decimal("15.00"),
        total_pure_gold_grams=Decimal("13.76"),
        line_count=1,
        operation_state="awaiting_decision",
        has_locked_products=False,
        lines=[line],
    )


def _raw_log_bucket(*, lines: list[AfgWorkspaceLineOut], metal_bucket: str) -> AfgLogBucketOut:
    lot = AfgMeltLotOut(
        id=uuid4(),
        metal_bucket=metal_bucket,
        sent_date=date(2026, 3, 30),
        purchased_from_date=date(2026, 3, 30),
        before_weight_grams=Decimal("15.00"),
        before_amount_dkk=Decimal("8787.14"),
        before_pure_gold_grams=Decimal("13.76"),
        after_pure_gold_grams=Decimal("13.00"),
        insurance_dkk=Decimal("0.00"),
        shipping_dkk=Decimal("0.00"),
        refining_dkk=Decimal("0.00"),
        sale_date=None,
        quote_eur=None,
        exchange_rate_dkk=Decimal("7.45"),
        payout_total_dkk=None,
        notes=None,
        cost_total_dkk=Decimal("0.00"),
        estimated_sale_value_dkk=None,
        net_after_costs_dkk=None,
        bridge_difference_dkk=None,
        advance_per_gram_dkk=None,
        created_at=utc_now(),
        updated_at=utc_now(),
    )
    documents = [_raw_log_document(line) for line in lines]
    return AfgLogBucketOut(
        metal_bucket=metal_bucket,
        summary=AfgLogBucketSummaryOut(
            total_documents=len(documents),
            total_lines=len(lines),
            awaiting_lines=len(lines),
            routed_lines=0,
            split_line_count=0,
            melt_line_count=0,
            melt_lot_count=1 if lines else 0,
            total_weight_grams=Decimal("15.00") if lines else Decimal("0.00"),
            total_pure_gold_grams=Decimal("13.76") if lines else Decimal("0.00"),
            total_amount_dkk=Decimal("8787.14") if lines else Decimal("0.00"),
        ),
        documents=documents,
        split_groups=[],
        melt_queue=AfgMeltQueueOut(
            line_count=1 if lines else 0,
            total_weight_grams=Decimal("15.00") if lines else Decimal("0.00"),
            total_pure_gold_grams=Decimal("13.76") if lines else Decimal("0.00"),
            total_amount_dkk=Decimal("8787.14") if lines else Decimal("0.00"),
            earliest_purchase_date=date(2026, 3, 30) if lines else None,
            latest_purchase_date=date(2026, 3, 30) if lines else None,
            document_numbers=["SG-2026-000002"] if lines else [],
        ),
        melt_lots=[lot] if lines else [],
    )


def test_afg_reference_workbook_parses_without_sync_metadata():
    workbook = load_workbook(
        REFERENCE_ROOT / "Afregningsbilag ( alis frontumuz).xlsm",
        keep_vba=True,
        data_only=False,
    )
    sheet = workbook["Afregningsbilag"]
    vars_sheet = workbook["Variable værdier"]
    sheet["D16"] = "Denis Thor Chrisistensen"
    sheet["G18"] = "+45 22 25 55 04"
    sheet["C40"] = "Overførsel"
    sheet["D41"] = "5512"
    sheet["D42"] = "0725397984"
    sheet["F22"] = Decimal("22.00")
    sheet["F30"] = Decimal("3.50")
    vars_sheet["C4"] = Decimal("936.18")
    vars_sheet["C5"] = Decimal("14.61")
    vars_sheet["C14"] = "1003"
    vars_sheet["D14"] = "1001"

    parsed = parse_afg_workspace_inputs_from_workbook(_save_workbook_bytes(workbook))

    assert parsed.base_version is None
    assert parsed.customer.name == "Denis Thor Chrisistensen"
    assert parsed.customer.phone == "+45 22 25 55 04"
    assert parsed.sections.payment_method == "bank"
    assert parsed.sections.bank_info.reg_number == "5512"
    assert parsed.sections.bank_info.account_number == "0725397984"
    assert parsed.sections.gold_rows[0].gram == Decimal("22.00")
    assert parsed.sections.silver_rows[0].gram == Decimal("3.50")
    assert parsed.sections.market_rates is not None
    assert parsed.sections.market_rates.gold_24k_dkk == Decimal("936.18")
    assert parsed.sections.numbering is not None
    assert parsed.sections.numbering.afregnings_number_next == "1003"


def test_inventory_reference_workbook_parses_by_product_number_without_sync_metadata():
    row_id = uuid4()
    workspace = InventoryWorkspaceOut(
        market_prices=InventoryMarketPricesOut(
            gold=Decimal("882.08"),
            silver=Decimal("14.60"),
            platinum=Decimal("423.90"),
            palladium=Decimal("335.69"),
        ),
        summary=InventoryWorkspaceSummaryOut(),
        rows=[
            InventoryGridRowOut(
                id=row_id,
                product_number="S2512",
                reference_number=None,
                main_category="kulce",
                subcategory=None,
                product_type="bar",
                metal_type="yellow_gold",
                status="in_inventory",
                operation_destination="inventory",
                operation_classification="standard",
                lager_dato="2026-03-01",
                urun="Eski Urun",
                saflik_label="24K / Barren",
                purity_percentage=Decimal("99.90"),
                birim_gram=Decimal("10.00"),
                adet=1,
                toplam_gram=Decimal("10.00"),
                has_metal_grams=Decimal("9.99"),
                alis_fiyati_dkk=Decimal("5000.00"),
                spot_degeri_dkk=Decimal("7000.00"),
                shop_fiyati_dkk=None,
                shop_sync_status=None,
                length_cm=None,
                width_mm=None,
                thickness_mm=None,
                producer=None,
                storage_location=None,
                needs_cleaning=False,
                is_gdpr_locked=False,
                primary_photo=None,
                photo_count=0,
                has_ai_description=False,
                ai_description_approved=False,
                notes=None,
            )
        ],
    )

    workbook = load_workbook(REFERENCE_ROOT / "Depolama.xlsx", data_only=False)
    sheet = workbook["Lager"]
    sheet["K4"] = Decimal("2850.00")
    sheet["K5"] = Decimal("8.50")
    sheet["K6"] = Decimal("280.00")
    sheet["K7"] = Decimal("190.00")
    for row_idx in range(11, 420):
        sheet[f"K{row_idx}"] = None
    sheet["B11"] = datetime(2026, 3, 30)
    sheet["C11"] = "Bar · Sarı Altın"
    sheet["D11"] = Decimal("25.00")
    sheet["E11"] = 2
    sheet["H11"] = Decimal("24589.00")
    sheet["K11"] = "S2512"
    sheet["N11"] = "40cm"
    sheet["O11"] = Decimal("1.20")
    sheet["P11"] = Decimal("0.50")
    sheet["Q11"] = "Umicore"

    parsed = parse_inventory_workbook_inputs_from_workbook(
        _save_workbook_bytes(workbook),
        current_workspace=workspace,
    )

    assert parsed.base_version is None
    assert parsed.market_prices.gold == Decimal("2850.00")
    assert parsed.market_prices.silver == Decimal("8.50")
    assert row_id in parsed.product_updates
    update = parsed.product_updates[row_id]
    assert update.display_name == "Bar · Sarı Altın"
    assert update.weight_grams == Decimal("25.00")
    assert update.unit_count == 2
    assert update.purchase_price_dkk == Decimal("24589.00")
    assert update.length_cm == "40cm"
    assert update.width_mm == Decimal("1.20")
    assert update.thickness_mm == Decimal("0.50")
    assert update.producer == "Umicore"


def test_inventory_reference_workbook_reconcile_preview_lists_only_controlled_changes():
    row_id = uuid4()
    workspace = InventoryWorkspaceOut(
        market_prices=InventoryMarketPricesOut(
            gold=Decimal("882.08"),
            silver=Decimal("14.60"),
            platinum=Decimal("423.90"),
            palladium=Decimal("335.69"),
        ),
        summary=InventoryWorkspaceSummaryOut(),
        rows=[
            InventoryGridRowOut(
                id=row_id,
                product_number="S2512",
                reference_number=None,
                main_category="kulce",
                subcategory=None,
                product_type="bar",
                metal_type="yellow_gold",
                status="in_inventory",
                operation_destination="inventory",
                operation_classification="standard",
                lager_dato="2026-03-01",
                urun="Eski Urun",
                saflik_label="24K / Barren",
                purity_percentage=Decimal("99.90"),
                birim_gram=Decimal("10.00"),
                adet=1,
                toplam_gram=Decimal("10.00"),
                has_metal_grams=Decimal("9.99"),
                alis_fiyati_dkk=Decimal("5000.00"),
                spot_degeri_dkk=Decimal("7000.00"),
                shop_fiyati_dkk=None,
                shop_sync_status=None,
                length_cm=None,
                width_mm=None,
                thickness_mm=None,
                producer=None,
                storage_location=None,
                needs_cleaning=False,
                is_gdpr_locked=False,
                primary_photo=None,
                photo_count=0,
                has_ai_description=False,
                ai_description_approved=False,
                notes=None,
            )
        ],
    )

    workbook = load_workbook(REFERENCE_ROOT / "Depolama.xlsx", data_only=False)
    sheet = workbook["Lager"]
    for row_idx in range(11, 420):
        sheet[f"K{row_idx}"] = None
    sheet["K4"] = Decimal("2850.00")
    sheet["B11"] = datetime(2026, 3, 30)
    sheet["C11"] = "Bar · Sarı Altın"
    sheet["D11"] = Decimal("25.00")
    sheet["E11"] = 2
    sheet["H11"] = Decimal("24589.00")
    sheet["K11"] = "S2512"
    sheet["N11"] = "40cm"
    sheet["O11"] = Decimal("1.20")
    sheet["P11"] = Decimal("0.50")
    sheet["Q11"] = "Umicore"

    parsed = parse_inventory_workbook_inputs_from_workbook(
        _save_workbook_bytes(workbook),
        current_workspace=workspace,
    )
    preview = build_inventory_reconcile_preview(workspace, parsed)

    assert preview.editable is True
    assert any(change.cell_ref == "K4" for change in preview.changes)
    assert any(change.cell_ref == "C11" and "S2512" in change.label for change in preview.changes)
    assert not preview.blocking_errors


def test_log_reference_ark1_parses_without_sync_metadata():
    gold_line = _raw_log_line(metal_type="yellow_gold", line_no=1)
    workspace = AfgLogWorkspaceOut(
        summary=AfgWorkspaceSummaryOut(
            total_documents=1,
            total_amount_dkk=Decimal("8787.14"),
            total_pure_gold_grams=Decimal("13.76"),
        ),
        gold=_raw_log_bucket(lines=[gold_line], metal_bucket="gold"),
        silver=_raw_log_bucket(lines=[], metal_bucket="silver"),
    )

    workbook = load_workbook(REFERENCE_ROOT / "Log sistemi- afg verileri buraya yazdiriyorum..xlsx", data_only=False)
    sheet = workbook["Ark1"]
    for row_idx in range(10, 34):
        sheet[f"A{row_idx}"] = None
        sheet[f"G{row_idx}"] = None
    for row_idx in range(58, 88):
        sheet[f"A{row_idx}"] = None
        sheet[f"G{row_idx}"] = None
    for row_idx in range(6, 18):
        sheet[f"R{row_idx}"] = None
    sheet["A10"] = 2
    sheet["G10"] = "x"
    sheet["R6"] = 2
    sheet["B41"] = 200

    parsed = parse_log_workbook_inputs_from_workbook(
        _save_workbook_bytes(workbook),
        year=2026,
        current_workspace=workspace,
    )

    assert parsed.base_version is None
    assert len(parsed.route_updates) == 1
    assert parsed.route_updates[0].payload.line_ids == [gold_line.id]
    assert parsed.route_updates[0].payload.destination == "inventory"
    assert parsed.route_updates[0].payload.classification == "separate_storage"
    assert parsed.route_updates[0].payload.note is None
    assert not parsed.lot_creates
    assert len(parsed.lot_updates) == 1
    assert parsed.lot_updates[0].payload.insurance_dkk == Decimal("200")
