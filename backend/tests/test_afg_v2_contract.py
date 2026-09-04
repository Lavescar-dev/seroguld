from __future__ import annotations

import io
import zipfile
from decimal import Decimal

from openpyxl import load_workbook

from app.services import document_artifact_service as core
from app.services.document_artifact_service import (
    AFG_TEMPLATE_NAME,
    ArtifactSyncContext,
    _apply_afg_customer_cells,
    _apply_afg_detail_rows,
    _apply_afg_footer_cells,
    _open_template,
    _save_workbook_bytes,
)
from app.services.document_artifact_afg import parse_afg_workspace_inputs_from_workbook
from app.services.pos_workspace_state import _workspace_calculators_from_note


def _template_sheet():
    workbook = _open_template(AFG_TEMPLATE_NAME, keep_vba=True)
    return workbook, workbook["Afregningsbilag"]


def test_customer_values_go_right_of_labels_and_cpr_is_minimized() -> None:
    workbook, sheet = _template_sheet()
    _apply_afg_customer_cells(
        sheet,
        name="Mads Jensen",
        cpr_number="150385-1001",
        address="Prøvevej 12, 2. th.",
        city="København N",
        postal_code="2200",
        phone="+45 20 00 10 01",
        email="test@example.com",
        identity_doc="Kørekort KK-1",
    )
    # Etiketler dokunulmadı
    assert str(sheet["C16"].value).startswith("Navn")
    assert str(sheet["C17"].value).startswith("Adresse")
    assert str(sheet["C18"].value).startswith("Postnr")
    # Değerler etiketlerin sağında
    assert sheet["D16"].value == "Mads Jensen"
    assert sheet["D17"].value == "Prøvevej 12, 2. th."
    assert sheet["D18"].value == "2200 København N"
    # CPR yalnız doğum tarihi bölümü; tire ve son 4 hane belgeye girmez
    assert sheet["G16"].value == "150385"
    assert sheet["G17"].value == "Kørekort KK-1"
    assert sheet["G18"].value == "+45 20 00 10 01"
    assert sheet["G19"].value == "test@example.com"
    workbook.close()


def test_v2_parser_reads_dg_cells_and_never_returns_cpr() -> None:
    workbook, sheet = _template_sheet()
    vars_sheet = workbook["Variable værdier"]
    vars_sheet["C4"] = Decimal("615.50")
    vars_sheet["C5"] = Decimal("7.80")
    _apply_afg_customer_cells(
        sheet,
        name="Ada Lovelace",
        cpr_number="1503851001",
        address="Gammelvej 3",
        city="Hvidovre",
        postal_code="2650",
        phone="+45 11 22 33 44",
        email="ada@example.com",
        identity_doc=None,
    )
    parsed = parse_afg_workspace_inputs_from_workbook(_save_workbook_bytes(workbook))
    assert parsed.customer.name == "Ada Lovelace"
    assert parsed.customer.address == "Gammelvej 3"
    assert parsed.customer.postal_code == "2650"
    assert parsed.customer.city == "Hvidovre"
    # G16 (doğum tarihi görüntüsü) kayıtlı tam CPR'nin üzerine yazılmamalı:
    # alan hiç set edilmez → mutation'da mevcut değer korunur.
    assert "cpr_number" not in parsed.customer.model_fields_set


def test_v1_marked_draft_still_parses_with_old_cell_map() -> None:
    workbook, sheet = _template_sheet()
    # 0.3.5 yazıcısı: değerler etiket hücrelerinin üzerine (C16/F16...).
    workbook["Variable værdier"]["C4"] = Decimal("615.50")
    workbook["Variable værdier"]["C5"] = Decimal("7.80")
    sheet["C16"] = "Eski Taslak"
    sheet["F18"] = "+45 99 88 77 66"
    core._write_sync_sheet(
        workbook,
        context=ArtifactSyncContext(
            kind="alis-workspace",
            key="test",
            artifact_key="alis.workspace.test",
            base_version="v1",
            contract_version="afg-v1",
        ),
    )
    parsed = parse_afg_workspace_inputs_from_workbook(_save_workbook_bytes(workbook))
    assert parsed.customer.name == "Eski Taslak"
    assert parsed.customer.phone == "+45 99 88 77 66"


def test_compact_final_document_renders_only_filled_rows() -> None:
    workbook, sheet = _template_sheet()
    gold_rows = [
        {
            "row_key": f"gold:{key}",
            "label": f"Guld {key}k",
            "karat": Decimal(key),
            "lodighed": lodighed,
            "gram": Decimal("4.80") if key == "14" else Decimal("0.00"),
            "avance_percent": Decimal("8.50"),
            "unit_price_dkk": Decimal("501.03") if key == "14" else Decimal("0.00"),
            "line_total_dkk": Decimal("2404.94") if key == "14" else Decimal("0.00"),
        }
        for key, lodighed in (("8", "333"), ("14", "585"), ("18", "750"), ("21", "875"), ("21.6", "900"), ("22", "916"), ("24", "999"))
    ]
    silver_rows = [
        {"row_key": "silver:2", "type_code": "2", "label": "Finsølv999‰", "lodighed": "999", "gram": Decimal("0"), "avance_percent": Decimal("0"), "unit_price_dkk": Decimal("0"), "line_total_dkk": Decimal("0")},
        {"row_key": "silver:5", "type_code": "5", "label": "Plet", "lodighed": "", "gram": Decimal("120.00"), "avance_percent": Decimal("0"), "unit_price_dkk": Decimal("0.02"), "line_total_dkk": Decimal("2.40")},
    ]
    bar_rows = [
        {"bar_type": "gold", "gram": Decimal("10.00"), "avance_percent": Decimal("0"), "unit_price_dkk": Decimal("620.00"), "line_total_dkk": Decimal("6200.00")},
        {"bar_type": "silver", "gram": Decimal("0.00"), "avance_percent": Decimal("0"), "unit_price_dkk": Decimal("0"), "line_total_dkk": Decimal("0")},
    ]
    _apply_afg_detail_rows(sheet, gold_rows, silver_rows, bar_rows)
    content = _save_workbook_bytes(workbook)

    reloaded = load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    # Sıra: 14K altın (22), Guldbarre (23), Plet (24); kalan satırlar boş.
    assert out["C22"].value == "Guld"
    assert float(out["D22"].value) == 14
    assert float(out["F22"].value) == 4.8
    assert float(out["H22"].value) == 2404.94
    assert out["C23"].value == "Guldbarre"
    assert out["E23"].value == "999.9"
    assert float(out["H23"].value) == 6200.0
    assert out["C24"].value == "Plet"
    assert float(out["F24"].value) == 120.0
    for idx in range(25, 38):
        for column in "ABCDEFGH":
            assert out[f"{column}{idx}"].value is None, f"{column}{idx} boş olmalı"
    # Gizli blok metal kimliğiyle doğru: 14K gramı E80'de (79=8K, 82=14K? sıra 8,9,10,14 → 14K = 82)
    assert float(out["E82"].value) == 4.8
    reloaded.close()


def test_footer_keeps_bold_sero_guld_run(monkeypatch) -> None:
    from app.config import Settings

    monkeypatch.setattr(
        core,
        "get_settings",
        lambda: Settings(_env_file=None, database_url="sqlite+aiosqlite:///test.db"),
    )
    workbook, sheet = _template_sheet()
    _apply_afg_footer_cells(sheet)
    content = _save_workbook_bytes(workbook)
    reloaded = load_workbook(io.BytesIO(content), keep_vba=True, rich_text=True)
    footer = reloaded["Afregningsbilag"]["C53"].value
    text = str(footer)
    assert text.startswith("Sero Guld")
    assert "Valby Langgade 84" in text
    assert "CVR-nr: DK34093083" in text
    # Rich text: ilk run kalın firma adı (' ApS' soneki çıkarılır).
    blocks = list(footer)
    assert getattr(blocks[0].font, "b", False) is True
    assert blocks[0].text == "Sero Guld og Sølv"
    line2 = str(reloaded["Afregningsbilag"]["C54"].value)
    assert "Tlf.: 22255504" in line2
    assert "info@seroguld.dk" in line2
    assert "www.seroguld.dk" in line2
    reloaded.close()


def test_calculator_defaults_are_knife_weights_and_old_notes_survive() -> None:
    defaults = _workspace_calculators_from_note({})
    assert [str(row.unit_weight) for row in defaults.gold_rows] == ["4.00", "6.00", "8.00", "10.00", "0.00"]
    assert [str(row.unit_weight) for row in defaults.silver_rows][:4] == ["1.75", "3.51", "7.02", "17.54"]
    # Eski notta operatörün girdiği 13 korunur (merge-by-key).
    old_note = {"calculators": {"gold_rows": [{"row_key": "calc_gold:2", "unit_weight": "13", "count": "2"}]}}
    merged = _workspace_calculators_from_note(old_note)
    row = next(r for r in merged.gold_rows if r.row_key == "calc_gold:2")
    assert str(row.unit_weight) == "13.00"
    assert str(row.total_weight) == "26.00"


def test_compact_final_document_appends_ptpd_after_silver() -> None:
    workbook, sheet = _template_sheet()
    gold_rows = [
        {"row_key": "gold:14", "label": "Guld 14k", "karat": Decimal("14"), "lodighed": "585",
         "gram": Decimal("4.80"), "avance_percent": Decimal("0"), "unit_price_dkk": Decimal("501.03"),
         "line_total_dkk": Decimal("2404.94")},
    ]
    silver_rows = []
    bar_rows = []
    ptpd_rows = [
        {"metal": "platinum", "gram": Decimal("12.50"), "avance_percent": Decimal("5"),
         "unit_price_dkk": Decimal("300.00"), "line_total_dkk": Decimal("3750.00")},
        {"metal": "palladium", "gram": Decimal("0"), "avance_percent": Decimal("0"),
         "unit_price_dkk": Decimal("0"), "line_total_dkk": Decimal("0")},
    ]
    _apply_afg_detail_rows(sheet, gold_rows, silver_rows, bar_rows, ptpd_rows)
    content = _save_workbook_bytes(workbook)

    reloaded = load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    # Sıra: 14K altın (22), Platin (23); Palladium gramsız → yazılmaz.
    assert out["C22"].value == "Guld"
    assert out["C23"].value == "Platin"
    assert int(out["A23"].value) == 8
    assert out["E23"].value == "950"
    assert float(out["F23"].value) == 12.5
    assert float(out["H23"].value) == 3750.0
    for idx in range(24, 38):
        for column in "ABCDEFGH":
            assert out[f"{column}{idx}"].value is None, f"{column}{idx} boş olmalı"
    reloaded.close()


def test_draft_writer_puts_ptpd_rows_on_35_36_and_backup_vlookup_rows() -> None:
    from app.schemas.pos import PosWorkspaceMarketRates, PosWorkspacePtPdRowOut
    from app.services.document_artifact_service import (
        _apply_afg_market_rate_cells,
        _apply_afg_workspace_rows,
    )

    workbook = _open_template(AFG_TEMPLATE_NAME, keep_vba=True)
    sheet = workbook["Afregningsbilag"]
    vars_sheet = workbook["Variable værdier"]

    ptpd_rows = [
        PosWorkspacePtPdRowOut(
            row_key="ptpd:platinum", metal="platinum", label="Platin", lodighed="950",
            purity_percentage=Decimal("95.00"), gram=Decimal("3.20"), avance_percent=Decimal("0"),
            rate_dkk=Decimal("310.00"), unit_price_dkk=Decimal("310.00"), line_total_dkk=Decimal("992.00"),
        ),
        PosWorkspacePtPdRowOut(
            row_key="ptpd:palladium", metal="palladium", label="Palladium", lodighed="500",
            purity_percentage=Decimal("50.00"), gram=Decimal("7.00"), avance_percent=Decimal("10"),
            rate_dkk=Decimal("250.00"), unit_price_dkk=Decimal("225.00"), line_total_dkk=Decimal("1575.00"),
        ),
    ]
    _apply_afg_workspace_rows(sheet, [], [], [], ptpd_rows)
    market_rates = PosWorkspaceMarketRates(
        eur_dkk_fx=Decimal("7.45"),
        gold_rates_dkk={"24": Decimal("615.50")},
        silver_rates_dkk={"999": Decimal("7.80")},
        gold_24k_dkk=Decimal("615.50"),
        silver_dkk=Decimal("7.80"),
        platinum_dkk=Decimal("310.00"),
        palladium_dkk=Decimal("250.00"),
    )
    _apply_afg_market_rate_cells(vars_sheet, market_rates)
    content = _save_workbook_bytes(workbook)

    reloaded = load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    variables = reloaded["Variable værdier"]
    # Taslak satırları 35 (Platin) ve 36 (Palladium).
    assert int(out["A35"].value) == 8 and out["C35"].value == "Platin"
    assert float(out["F35"].value) == 3.2
    assert int(out["A36"].value) == 9 and out["C36"].value == "Palladium"
    assert float(out["H36"].value) == 1575.0
    # VLOOKUP yedek satırları ($A$4:$B$12 aralığında) + J17/J18 oranları.
    assert int(variables["A11"].value) == 8 and variables["B11"].value == "Platin"
    assert int(variables["A12"].value) == 9 and variables["B12"].value == "Palladium"
    assert float(variables["J17"].value) == 310.0
    assert float(variables["J18"].value) == 250.0
    reloaded.close()

    # Parser taslaktan ptpd girişlerini geri okur.
    parsed = parse_afg_workspace_inputs_from_workbook(content)
    ptpd_inputs = {row.metal: row for row in parsed.sections.ptpd_rows}
    assert float(ptpd_inputs["platinum"].gram) == 3.2
    assert float(ptpd_inputs["palladium"].gram) == 7.0
    assert float(ptpd_inputs["palladium"].avance_percent) == 10.0


# --- Satır görünürlüğü: Excel orijinali — satırlar dursun, içleri boş ---
# Karat → satır: 22'den itibaren sırayla 8,14,18,21,21.6,22,24 (22K = 27).
# Gümüş → satır: 30'dan itibaren 2,3,4,5 (Finsølv = 30); Guldbarre = 29.
# Default (hide_blank_rows=False) hiçbir satırı gizlemez; 18705a9 gizlemesi
# hide_blank_rows=True bayrağıyla geri alınabilir.


def _gold_rows(filled: dict[str, str] | None = None) -> list["PosWorkspaceGoldRowOut"]:
    from app.schemas.pos import PosWorkspaceGoldRowOut

    lodighed = {"8": "333", "14": "585", "18": "750", "21": "875", "21.6": "900", "22": "916", "24": "999"}
    filled = filled or {}
    return [
        PosWorkspaceGoldRowOut(
            row_key=f"gold:{karat}",
            karat=Decimal(karat),
            label=f"Guld {karat}k",
            lodighed=lodighed[karat],
            purity_percentage=Decimal("91.60"),
            gram=Decimal(filled.get(karat, "0")),
            avance_percent=Decimal("8.50"),
            rate_dkk=Decimal("615.50"),
            unit_price_dkk=Decimal("501.03") if filled.get(karat) else Decimal("0"),
            line_total_dkk=Decimal("2404.94") if filled.get(karat) else Decimal("0"),
        )
        for karat in ("8", "14", "18", "21", "21.6", "22", "24")
    ]


def _silver_rows(filled: dict[str, str] | None = None) -> list["PosWorkspaceSilverRowOut"]:
    from app.schemas.pos import PosWorkspaceSilverRowOut

    meta = {"2": ("Finsølv999‰", "999"), "3": ("Sterlingsølv", "925"), "4": ("3 tårnet sølv", "830"), "5": ("Plet", "")}
    filled = filled or {}
    return [
        PosWorkspaceSilverRowOut(
            row_key=f"silver:{key}",
            type_code=key,
            label=meta[key][0],
            lodighed=meta[key][1],
            purity_percentage=Decimal("92.50"),
            gram=Decimal(filled.get(key, "0")),
            avance_percent=Decimal("0"),
            rate_dkk=Decimal("7.80"),
            unit_price_dkk=Decimal("7.80") if filled.get(key) else Decimal("0"),
            line_total_dkk=Decimal("23.40") if filled.get(key) else Decimal("0"),
        )
        for key in ("2", "3", "4", "5")
    ]


def _gold_bar_row(gram: str) -> "PosWorkspaceBarRowOut":
    from app.schemas.pos import PosWorkspaceBarRowOut

    return PosWorkspaceBarRowOut(
        row_key="bar:gold",
        bar_type="gold",
        label="Guldbarre",
        lodighed="999.9",
        purity_percentage=Decimal("99.99"),
        gram=Decimal(gram),
        avance_percent=Decimal("0"),
        rate_dkk=Decimal("620.00"),
        unit_price_dkk=Decimal("620.00"),
        line_total_dkk=Decimal("7750.00") if Decimal(gram) > 0 else Decimal("0"),
    )


def _visible_rows(out, start: int = 21, end: int = 37) -> set[int]:
    return {idx for idx in range(start, end + 1) if not out.row_dimensions[idx].hidden}


def test_draft_shows_all_rows_by_default() -> None:
    workbook, sheet = _template_sheet()
    core._apply_afg_workspace_rows(sheet, _gold_rows({"22": "5.00"}), _silver_rows(), [], [])
    core._apply_afg_row_visibility(sheet)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    # Excel orijinali: müşteri yalnız 22K satmış olsa bile 21-37 TÜMÜ görünür
    # — boş satırlar gizlenmez, içleri boş kalır.
    assert _visible_rows(out) == set(range(21, 38))
    assert float(out["F27"].value) == 5.0
    # I alt satırı (38) dokunulmaz.
    assert not out.row_dimensions[38].hidden
    reloaded.close()


def test_draft_blank_splitter_visible_and_clean() -> None:
    workbook, sheet = _template_sheet()
    # Yalnız altın tarafı dolu — 18705a9 şeridi gizlerdi; Excel orijinalinde
    # şerit daima görünür ama yazı artığı taşımaz.
    core._apply_afg_workspace_rows(sheet, _gold_rows({"22": "5.00"}), _silver_rows(), [], [])
    core._apply_afg_row_visibility(sheet)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    assert not out.row_dimensions[29].hidden
    # Ayraç temiz: boş Guldbarre yazı artıkları ("Guldbarre 24 0,00") taşımaz.
    assert out["A29"].value is None
    assert out["C29"].value is None
    assert out["D29"].value is None
    assert out["F29"].value is None
    # Sarıya boyanmadı — şablonun açık mavisi korunur.
    fill = out["C29"].fill
    assert not (fill.patternType == "solid" and str(fill.fgColor.rgb or "").endswith("FFC000"))
    reloaded.close()


def test_draft_fills_guldbarre_row_yellow_when_present() -> None:
    workbook, sheet = _template_sheet()
    core._apply_afg_workspace_rows(sheet, _gold_rows({"22": "5.00"}), _silver_rows(), [_gold_bar_row("12.50")], [])
    core._apply_afg_row_visibility(sheet)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    assert _visible_rows(out) == set(range(21, 38))
    assert int(out["A29"].value) == 6
    assert float(out["D29"].value) == 24
    assert float(out["H29"].value) == 7750.0
    # C:H kolonları şablon karat sarısına (FFC000) boyanır.
    for column in "CDEFGH":
        fill = out[f"{column}29"].fill
        assert fill.patternType == "solid" and str(fill.fgColor.rgb).endswith("FFC000"), column
    # İnce ayraç ölçüsü (ht=7.5) normal grid ölçüsüne döner.
    height = out.row_dimensions[29].height
    assert height is None or height > 10
    reloaded.close()


def test_calculator_defaults_keep_all_rows_visible_and_data_alive() -> None:
    workbook, sheet = _template_sheet()
    core._apply_afg_workspace_rows(sheet, _gold_rows(), _silver_rows(), [], [])
    core._apply_afg_calculator_cells(sheet, _workspace_calculators_from_note({}))
    core._apply_afg_row_visibility(sheet)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    # Excel orijinali: kniv varsayılanları (J22=4, J32=1.75) satır açmaz da
    # gizlemez de — grid daima tam görünür, hesaplayıcı verisi yaşamaya devam eder.
    assert _visible_rows(out) == set(range(21, 38))
    assert float(out["J22"].value) == 4.0
    assert float(out["J32"].value) == 1.75
    reloaded.close()


def test_hide_blank_rows_flag_restores_legacy_hiding() -> None:
    # 18705a9 gizleme davranışı tek bayrakla geri gelir.
    scenarios: list[tuple[list, list, list, set[int]]] = [
        (_gold_rows({"22": "5.00"}), _silver_rows(), [], {27}),
        (_gold_rows({"22": "5.00"}), _silver_rows({"2": "3.00"}), [], {27, 29, 30}),
        (_gold_rows(), _silver_rows({"2": "3.00"}), [], {30}),
    ]
    for gold, silver, bars, expected in scenarios:
        workbook, sheet = _template_sheet()
        core._apply_afg_workspace_rows(sheet, gold, silver, bars, [])
        core._apply_afg_row_visibility(sheet, hide_blank_rows=True)
        reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
        out = reloaded["Afregningsbilag"]
        assert _visible_rows(out) == expected
        reloaded.close()
    # Hesaplayıcı varsayılanları gizlemeyi AÇMAZ (eski sözleşme korunur);
    # veri gizli satırda yaşamaya devam eder.
    workbook, sheet = _template_sheet()
    core._apply_afg_workspace_rows(sheet, _gold_rows(), _silver_rows(), [], [])
    core._apply_afg_calculator_cells(sheet, _workspace_calculators_from_note({}))
    core._apply_afg_row_visibility(sheet, hide_blank_rows=True)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    assert _visible_rows(out) == set()
    assert float(out["J22"].value) == 4.0
    reloaded.close()


def test_final_keeps_all_rows_visible_and_clears_calculator_ghosts() -> None:
    workbook, sheet = _template_sheet()
    gold_rows = [
        {"row_key": "gold:14", "label": "Guld 14k", "karat": Decimal("14"), "lodighed": "585",
         "gram": Decimal("4.80"), "avance_percent": Decimal("8.50"), "unit_price_dkk": Decimal("501.03"),
         "line_total_dkk": Decimal("2404.94")},
    ]
    silver_rows = [
        {"row_key": "silver:5", "type_code": "5", "label": "Plet", "lodighed": "",
         "gram": Decimal("120.00"), "avance_percent": Decimal("0"), "unit_price_dkk": Decimal("0.02"),
         "line_total_dkk": Decimal("2.40")},
    ]
    _apply_afg_detail_rows(sheet, gold_rows, silver_rows, [])
    core._apply_afg_row_visibility(sheet)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    # Excel orijinali final: 14K (22) ve Plet (23) dolu; kalan satırlar
    # görünür ama içleri boş.
    assert out["C22"].value == "Guld"
    assert out["C23"].value == "Plet"
    assert _visible_rows(out) == set(range(21, 38))
    # Şablon hesaplayıcı hayaletleri (J22=10, J32=1.754) finalden kalkar.
    for idx in range(22, 38):
        for column in "JKL":
            assert out[f"{column}{idx}"].value is None, f"{column}{idx}"
    reloaded.close()

    # 18705a9 gizlemeli final yolu bayrakla geri gelir.
    workbook, sheet = _template_sheet()
    _apply_afg_detail_rows(sheet, gold_rows, silver_rows, [])
    core._apply_afg_row_visibility(sheet, blank_splitter_visible=False, hide_blank_rows=True)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    assert _visible_rows(out) == {22, 23}
    reloaded.close()


def test_hidden_rows_keep_vba_and_values_in_xlsm() -> None:
    workbook, sheet = _template_sheet()
    core._apply_afg_workspace_rows(sheet, _gold_rows({"22": "5.00"}), _silver_rows(), [], [])
    core._apply_afg_row_visibility(sheet, hide_blank_rows=True)
    content = _save_workbook_bytes(workbook)
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        assert "xl/vbaProject.bin" in archive.namelist()
    # Değerler gizli satırda durur: unhide edildiğinde veri kaybolmaz.
    reloaded = load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    assert out.row_dimensions[22].hidden is True
    reloaded.close()


def test_draft_splitter_visible_when_only_silver_side_filled() -> None:
    workbook, sheet = _template_sheet()
    core._apply_afg_workspace_rows(sheet, _gold_rows(), _silver_rows({"2": "3.00"}), [], [])
    core._apply_afg_row_visibility(sheet)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    # Görünürlük veriye bağlı değil: şerit daima görünür (şablon gibi).
    assert _visible_rows(out) == set(range(21, 38))
    reloaded.close()


def test_rebuild_clears_hidden_flags_from_older_files() -> None:
    workbook, sheet = _template_sheet()
    core._apply_afg_workspace_rows(sheet, _gold_rows({"22": "5.00"}), _silver_rows(), [], [])
    # 18705a9 üretiminden kalan gizli satır bayrağını simüle et: yeni build
    # şablon paritesi için bayrakları temizler (migrasyon).
    sheet.row_dimensions[27].hidden = True
    sheet.row_dimensions[30].hidden = True
    core._apply_afg_row_visibility(sheet)
    reloaded = load_workbook(io.BytesIO(_save_workbook_bytes(workbook)), keep_vba=True, data_only=False)
    out = reloaded["Afregningsbilag"]
    assert _visible_rows(out) == set(range(21, 38))
    reloaded.close()


def test_summary_cells_blank_bank_fields_and_never_print_none_text() -> None:
    workbook, sheet = _template_sheet()
    core._apply_afg_summary_cells(
        sheet,
        net_amount_dkk=Decimal("6161.10"),
        vat_amount_dkk=Decimal("0"),
        gross_amount_dkk=Decimal("6161.10"),
        payment_method="cash",
        reg_number="5512",
        account_number="0725397984",
        note="None",
    )
    assert sheet["C40"].value == "Kontant"
    # Şablon kuralı (0.3.32): nakit/eksik banka bilgisinde hücre BOŞ kalır.
    assert sheet["D41"].value is None
    assert sheet["D42"].value is None
    # Geçmiş üretimden sızan 'None' metni belgeye yazılmaz.
    assert sheet["D44"].value is None
    core._apply_afg_summary_cells(
        sheet,
        net_amount_dkk=Decimal("6161.10"),
        vat_amount_dkk=Decimal("0"),
        gross_amount_dkk=Decimal("6161.10"),
        payment_method="bank",
        reg_number="5512",
        account_number="0725397984",
        note="Testnot",
    )
    assert sheet["C40"].value == "Overførsel"
    assert sheet["D41"].value == "5512"
    assert sheet["D42"].value == "0725397984"
    assert sheet["D44"].value == "Testnot"
    workbook.close()
