"""Re-runnable ETL: build the depolama (inventory) seed from Recai Bey's two
real depolama workbooks.

The two source files are NOT committed to the repo (they are large and are the
shop's live inventory). Obtain the current pair from Recai Bey and re-run this
script whenever the catalogue is refreshed. Expected inputs (in --src dir):
  - Oversigt.xlsx                     (old/complete inventory: bars, coins, jewelry)
  - Ny lagerstyring af smykker.xlsx   (new 2026 jewelry catalog S2500+, + 266 photos)

Outputs (into backend/seed_data/depolama/, the bundled seed the app loads on a
fresh install):
  - inventory_seed.json   normalized products from BOTH files (status-classified)
  - photos/               AVIF photo pool (no per-product link) + manifest.json

Status is read from row TEXT (Danish) reinforced by fill colour:
  smelt/smeltet -> melted (+ melt_reason/log ref)   solgt -> sold (+ invoice)
  skal smeltes  -> undecided (pending melt)          else -> in_inventory

Usage (backend venv python — has openpyxl + pillow_avif):
  python scripts/seed/extract_inventory.py --src "C:/path/to/source-dir"
  python scripts/seed/extract_inventory.py --src ... --no-photos   # products only
"""
from __future__ import annotations

import argparse
import json
import re
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

# backend/scripts/seed/extract_inventory.py -> backend/seed_data/depolama
DEFAULT_OUT = Path(__file__).resolve().parents[2] / 'seed_data' / 'depolama'
_args_src = None  # set in __main__
OUT = DEFAULT_OUT
OVERSIGT: Path = Path()
NY: Path = Path()

# ---- status keyword classification -------------------------------------------
MELT_RE = re.compile(r'\bsmelt', re.I)
SOLD_RE = re.compile(r'\bsolgt', re.I)
PENDING_MELT_RE = re.compile(r'skal\s+smeltes', re.I)
LOG_REF_RE = re.compile(r'log[\s\-]*?(\d+)', re.I)
FAKTURA_RE = re.compile(r'faktura[\s\-]*?(\d+)', re.I)


def classify_status(row_texts: list[str]) -> tuple[str, str | None]:
    """Return (status, reference_note). row_texts = all string cells in the row."""
    joined = ' | '.join(t for t in row_texts if t)
    if PENDING_MELT_RE.search(joined):
        return 'undecided', joined_ref(joined)  # pending melt
    if MELT_RE.search(joined):
        m = LOG_REF_RE.search(joined)
        return 'melted', f'Smelt (Log {m.group(1)})' if m else 'Smeltet'
    if SOLD_RE.search(joined):
        m = FAKTURA_RE.search(joined)
        return 'sold', f'Solgt (Faktura {m.group(1)})' if m else 'Solgt'
    return 'in_inventory', None


def joined_ref(joined: str) -> str:
    m = LOG_REF_RE.search(joined)
    return f'Skal smeltes (Log {m.group(1)})' if m else 'Skal smeltes'


# ---- metal / type inference --------------------------------------------------
def infer_metal(name: str, category: str) -> str:
    s = f'{name} {category}'.lower()
    if 'hvidguld' in s or 'hvid guld' in s:
        return 'white_gold'
    if 'sølv' in s or 'solv' in s or 'silver' in s:
        return 'silver'
    if 'platin' in s:
        return 'platinum'
    if 'palladium' in s:
        return 'palladium'
    return 'yellow_gold'


def infer_type(name: str, category: str) -> str:
    s = f'{name} {category}'.lower()
    cat = category.lower()
    if 'barre' in cat or 'barrer' in cat or 'bar' in cat.split():
        return 'bar'
    if 'mønt' in cat or 'monter' in cat or 'mønter' in cat:
        return 'bar'  # investment coin -> closest coarse enum (fine detail in inventory_category)
    if 'ring' in s and 'øre' not in s and 'ørering' not in s:
        return 'ring'
    if 'armlænke' in s or 'armring' in s or 'armbånd' in s or 'armb' in s:
        return 'bracelet'
    if 'ørering' in s or 'ørestik' in s or 'ørehæng' in s or 'øre' in s:
        return 'earring'
    if 'halskæde' in s or 'halsk' in s or 'vedhæng' in s or 'kors' in s:
        return 'necklace'
    if 'kæde' in s:
        return 'chain'
    return 'jewelry'


PURITY_RE = re.compile(r'(\d{1,2})\s*(?:karat|kt|k)\b', re.I)


def infer_purity(name: str) -> tuple[str | None, float | None]:
    m = PURITY_RE.search(name or '')
    if not m:
        return None, None
    k = int(m.group(1))
    if 1 <= k <= 24:
        return f'{k}K', round(k / 24 * 100, 2)
    return None, None


@dataclass
class SeedProduct:
    source_file: str
    source_row: int
    legacy_code: str | None
    category: str | None
    display_name: str
    product_type: str
    metal_type: str
    weight_grams: float
    unit_count: int
    total_weight_grams: float | None
    pure_gold_grams: float | None
    purchase_price_dkk: float
    market_price_dkk: float | None
    purchase_date: str | None
    purity_karat: str | None
    purity_percentage: float | None
    length_cm: str | None
    width_mm: float | None
    thickness_mm: float | None
    producer: str | None
    afg_number: str | None
    status: str
    status_note: str | None
    inventory_category: str | None
    source_fill_color: str | None = None


MELT_FILL_COLORS = {'theme7/tint0.6', 'FFFF0000'}


def fill_rgb(cell):
    f = cell.fill
    if f is None or f.patternType is None:
        return None
    fg = f.fgColor
    if fg is None:
        return None
    if fg.type == 'rgb' and fg.rgb and fg.rgb != '00000000':
        return str(fg.rgb)
    if fg.type == 'theme':
        return f'theme{fg.theme}/tint{round(fg.tint or 0, 2)}'
    if fg.type == 'indexed':
        return f'idx{fg.indexed}'
    return None


def _num(v):
    if v in (None, ''):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _clean(v):
    if v in (None, ''):
        return None
    return str(v).strip()


def parse_file(path: Path, source_label: str, attr_cols: dict) -> tuple[list[SeedProduct], Counter]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Lager']
    header = next((r for r in range(1, 20) if str(ws.cell(r, 3).value).strip().lower() == 'vare'), 8)
    products: list[SeedProduct] = []
    skipped = Counter()
    current_cat = None
    for r in range(header + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        weight = _num(ws.cell(r, 4).value)
        is_date_b = isinstance(b, datetime)
        # A section header is NON-numeric text in col A (e.g. "Guldsmykker");
        # bare numeric A cells (stray 1608/1612 markers) are not categories.
        if a not in (None, '') and not is_date_b and weight is None and c in (None, ''):
            a_text = str(a).strip()
            if not re.fullmatch(r'\d+(?:[.,]\d+)?', a_text):
                current_cat = a_text
            continue
        if weight is None or weight <= 0:
            continue
        name = _clean(c)
        if not name:
            skipped['no_name'] += 1
            continue
        price = _num(ws.cell(r, 8).value) or _num(ws.cell(r, 9).value)  # H købspris else I market
        if not price or price <= 0:
            skipped['no_price'] += 1
            continue
        # gather all string cells in the row for status classification
        row_texts = [str(ws.cell(r, col).value) for col in range(1, ws.max_column + 1)
                     if isinstance(ws.cell(r, col).value, str)]
        status, note = classify_status(row_texts)
        color = fill_rgb(ws.cell(r, 3)) or fill_rgb(ws.cell(r, 1))
        # colour fallback: melt-coloured rows with no explicit text are still melts
        if status == 'in_inventory' and color in MELT_FILL_COLORS:
            status, note = 'melted', note or 'Smeltet (farve)'
        pk, pp = infer_purity(name)
        code = _clean(a)
        products.append(SeedProduct(
            source_file=source_label,
            source_row=r,
            legacy_code=code if code and code.lower() != 'foto' else None,
            category=current_cat,
            display_name=name,
            product_type=infer_type(name, current_cat or ''),
            metal_type=infer_metal(name, current_cat or ''),
            weight_grams=round(weight, 2),
            unit_count=int(_num(ws.cell(r, 5).value) or 1),
            total_weight_grams=round(_num(ws.cell(r, 6).value), 2) if _num(ws.cell(r, 6).value) else None,
            pure_gold_grams=round(_num(ws.cell(r, 7).value), 2) if _num(ws.cell(r, 7).value) else None,
            purchase_price_dkk=round(price, 2),
            market_price_dkk=round(_num(ws.cell(r, 9).value), 2) if _num(ws.cell(r, 9).value) else None,
            purchase_date=b.date().isoformat() if is_date_b else None,
            purity_karat=pk,
            purity_percentage=pp,
            length_cm=_clean(ws.cell(r, attr_cols['length'] + 1).value) if 'length' in attr_cols else None,
            width_mm=_num(ws.cell(r, attr_cols['width'] + 1).value) if 'width' in attr_cols else None,
            thickness_mm=_num(ws.cell(r, attr_cols['thickness'] + 1).value) if 'thickness' in attr_cols else None,
            producer=_clean(ws.cell(r, attr_cols['producer'] + 1).value) if 'producer' in attr_cols else None,
            afg_number=_clean(ws.cell(r, attr_cols['afg'] + 1).value) if 'afg' in attr_cols else None,
            status=status,
            status_note=note,
            inventory_category=current_cat,
            source_fill_color=color,
        ))
    wb.close()
    return products, skipped


def extract_products() -> list[SeedProduct]:
    ny, ny_skip = parse_file(NY, 'Ny lagerstyring af smykker.xlsx',
                             {'length': 12, 'width': 13, 'thickness': 14, 'producer': 15, 'afg': 21})
    ov, ov_skip = parse_file(OVERSIGT, 'Oversigt.xlsx',
                             {'length': 13, 'width': 14, 'thickness': 15, 'producer': 16})
    print(f'Ny lagerstyring: {len(ny)} products (skipped {dict(ny_skip)})')
    print(f'Oversigt:        {len(ov)} products (skipped {dict(ov_skip)})')
    all_products = ny + ov
    print(f'TOTAL: {len(all_products)} products')
    print('  status:', dict(Counter(p.status for p in all_products)))
    print('  type:  ', dict(Counter(p.product_type for p in all_products)))
    print('  metal: ', dict(Counter(p.metal_type for p in all_products)))
    print('  category:', dict(Counter(p.category for p in all_products)))
    return all_products


NS = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
}


def _billeder_number_labels(path: Path) -> dict[tuple[int, int], str]:
    """Map (row0, col0) -> gallery number label from the Billeder sheet text cells."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Billeder']
    labels: dict[tuple[int, int], str] = {}
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        for c, v in enumerate(row):
            if v in (None, ''):
                continue
            s = str(v).strip()
            if re.fullmatch(r'\d+(?:[.,]\d+)?', s):  # gallery numbers incl 2.5 / 9.5
                labels[(r, c)] = s
    wb.close()
    return labels


def extract_photos() -> list[dict]:
    """Extract 266 embedded PNGs from the Billeder sheet, convert to AVIF, tag each
    with its nearest gallery number. Photos are a POOL (no per-product link)."""
    from PIL import Image
    import pillow_avif  # noqa: F401  registers AVIF plugin
    import io

    photo_dir = OUT / 'photos'
    photo_dir.mkdir(parents=True, exist_ok=True)
    z = zipfile.ZipFile(NY)

    # resolve Billeder -> sheet xml -> drawing -> image rels (see inspect_billeder)
    wbx = ET.fromstring(z.read('xl/workbook.xml'))
    name_to_rid = {s.get('name'): s.get(f'{{{NS["r"]}}}id') for s in wbx.iter(f'{{{NS["main"]}}}sheet')}
    rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    rid_to_target = {rl.get('Id'): rl.get('Target') for rl in rels}
    sheet_base = ('xl/' + rid_to_target[name_to_rid['Billeder']].lstrip('/')).split('/')[-1]
    srels = ET.fromstring(z.read(f'xl/worksheets/_rels/{sheet_base}.rels'))
    drawing_target = next(rl.get('Target') for rl in srels if 'drawing' in (rl.get('Type') or ''))
    drawing_base = drawing_target.split('/')[-1]
    drels = ET.fromstring(z.read(f'xl/drawings/_rels/{drawing_base}.rels'))
    embed_to_img = {rl.get('Id'): rl.get('Target') for rl in drels}
    draw = ET.fromstring(z.read('xl/drawings/' + drawing_base))

    number_labels = _billeder_number_labels(NY)

    def nearest_number(row0: int, col0: int) -> str | None:
        best = None
        for (lr, lc), lbl in number_labels.items():
            if abs(lc - col0) <= 1 and lr <= row0 and (row0 - lr) <= 12:
                if best is None or lr > best[0]:
                    best = (lr, lbl)
        return best[1] if best else None

    manifest: list[dict] = []
    seen_files: dict[str, str] = {}  # image target -> output filename (dedup reused anchors)
    idx = 0
    for anchor in list(draw):
        frm = anchor.find(f'{{{NS["xdr"]}}}from')
        if frm is None:
            continue
        col0 = int(frm.find(f'{{{NS["xdr"]}}}col').text)
        row0 = int(frm.find(f'{{{NS["xdr"]}}}row').text)
        blip = anchor.find(f'.//{{{NS["a"]}}}blip')
        embed = blip.get(f'{{{NS["r"]}}}embed') if blip is not None else None
        img_target = embed_to_img.get(embed)
        if not img_target:
            continue
        img_path = 'xl/' + img_target.replace('../', '')
        gallery = nearest_number(row0, col0)
        if img_path in seen_files:
            manifest.append({'file': seen_files[img_path], 'gallery_number': gallery,
                             'anchor': [row0, col0], 'reused': True})
            continue
        raw = z.read(img_path)
        idx += 1
        gtag = (gallery or 'x').replace('.', '_').replace(',', '_')
        out_name = f'depolama_{idx:03d}_g{gtag}.avif'
        try:
            im = Image.open(io.BytesIO(raw))
            if im.mode not in ('RGB', 'RGBA'):
                im = im.convert('RGB')
            im.save(photo_dir / out_name, format='AVIF', quality=55)
        except Exception as exc:  # noqa: BLE001
            out_name = f'depolama_{idx:03d}_g{gtag}.png'
            (photo_dir / out_name).write_bytes(raw)
            print(f'  ! AVIF failed for {img_path} ({exc}); kept PNG')
        seen_files[img_path] = out_name
        manifest.append({'file': out_name, 'gallery_number': gallery,
                         'anchor': [row0, col0], 'source_image': img_path.split('/')[-1]})
    z.close()
    (photo_dir / 'manifest.json').write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
    total_bytes = sum(p.stat().st_size for p in photo_dir.glob('*.avif')) + \
        sum(p.stat().st_size for p in photo_dir.glob('*.png'))
    print(f'Photos: {len(seen_files)} unique files, {len(manifest)} manifest entries, '
          f'{total_bytes/1_000_000:.1f} MB total')
    return manifest


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Extract depolama inventory seed from the two source workbooks.')
    parser.add_argument('--src', required=True, help='Directory holding Oversigt.xlsx + "Ny lagerstyring af smykker.xlsx"')
    parser.add_argument('--out', default=str(DEFAULT_OUT), help='Output dir (default: backend/seed_data/depolama)')
    parser.add_argument('--no-photos', action='store_true', help='Products only; skip photo extraction')
    ns = parser.parse_args()

    src = Path(ns.src).expanduser()
    OVERSIGT = src / 'Oversigt.xlsx'
    NY = src / 'Ny lagerstyring af smykker.xlsx'
    for f in (OVERSIGT, NY):
        if not f.exists():
            parser.error(f'source file not found: {f}')
    OUT = Path(ns.out).expanduser()
    OUT.mkdir(parents=True, exist_ok=True)

    products = extract_products()
    (OUT / 'inventory_seed.json').write_text(
        json.dumps([asdict(p) for p in products], ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'\nWrote {OUT / "inventory_seed.json"}')
    seen = set()
    for p in products:
        if p.status not in seen:
            seen.add(p.status)
            print(f'  [{p.status}] {p.display_name[:40]} note={p.status_note} type={p.product_type} metal={p.metal_type} purity={p.purity_karat} price={p.purchase_price_dkk}')
    if not ns.no_photos:
        print()
        extract_photos()
