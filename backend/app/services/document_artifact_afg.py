from __future__ import annotations

import io
from decimal import Decimal
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from app.schemas.document_artifact import DocumentArtifactCellChangeOut, DocumentArtifactReconcilePreviewOut
from app.schemas.pos import (
    PosWorkspaceCustomerUpdate,
    PosWorkspaceInvoiceGoldSheetUpdate,
    PosWorkspaceInvoiceMiscSheetUpdate,
    PosWorkspaceMarketRates,
    PosWorkspaceNumberingUpdate,
    PosWorkspaceSectionsUpdate,
    PosWorkspaceSilverRowInput,
    PosWorkspaceGoldRowInput,
    PosWorkspaceBarRowInput,
    PosWorkspacePtPdRowInput,
)

if TYPE_CHECKING:
    from app.schemas.pos import PosDocumentDetailOut, PosWorkspaceOut
    from app.services.document_artifact_service import AfgWorkspaceArtifactInputs, ArtifactSyncContext


def _core():
    from app.services import document_artifact_service as core

    return core


def build_afg_workbook_bytes_from_workspace(workspace: "PosWorkspaceOut", *, sync_context: "ArtifactSyncContext") -> bytes:
    core = _core()
    workbook = core._open_template(core.AFG_TEMPLATE_NAME, keep_vba=True)
    core._touch_calc_flags(workbook)
    core._apply_afg_sheet_modes(workbook, can_write=True)
    sheet = workbook["Afregningsbilag"]
    factura_gold_sheet = workbook[core.AFG_FACTURA_GOLD_SHEET]
    factura_misc_sheet = workbook[core.AFG_FACTURA_MISC_SHEET]
    vars_sheet = workbook[core.AFG_VARIABLES_SHEET]

    document_number = workspace.numbering_preview.afregnings_number_next or workspace.session.session_code
    invoice_number = workspace.numbering_preview.invoice_number_next or "—"
    issued_at = workspace.session.updated_at
    sheet["H6"] = document_number
    sheet["H7"] = core._excel_datetime(issued_at)
    sheet["D2"] = core._excel_datetime(issued_at)
    core._apply_afg_market_rate_cells(vars_sheet, workspace.market_rates)
    vars_sheet["C14"] = document_number
    vars_sheet["D14"] = invoice_number
    core._apply_afg_customer_cells(
        sheet,
        name=workspace.customer.name,
        cpr_number=workspace.customer.cpr_number,
        address=workspace.customer.address,
        city=workspace.customer.city,
        postal_code=workspace.customer.postal_code,
        phone=workspace.customer.phone,
        email=workspace.customer.email,
        identity_doc=workspace.customer.identity_doc_number,
    )
    core._apply_afg_workspace_rows(
        sheet,
        core._afg_gold_rows_from_workspace(workspace),
        core._afg_silver_rows_from_workspace(workspace),
        workspace.bar_rows,
        workspace.ptpd_rows,
        extra_rows=getattr(workspace, "extra_rows", None) or [],
    )
    core._apply_afg_footer_cells(sheet)
    core._apply_afg_declaration_cells(sheet)
    core._apply_afg_calculator_cells(sheet, workspace.calculators)
    core._apply_afg_summary_cells(
        sheet,
        net_amount_dkk=core.quantize_2(core.to_decimal(workspace.summary.net_amount_dkk)),
        vat_amount_dkk=core.quantize_2(core.to_decimal(workspace.summary.vat_amount_dkk)),
        gross_amount_dkk=core.quantize_2(core.to_decimal(workspace.summary.gross_amount_dkk)),
        payment_method=workspace.payment_method,
        reg_number=workspace.bank_info.reg_number,
        account_number=workspace.bank_info.account_number,
        note=workspace.afg_note,
    )
    core._apply_afg_factura_gold_sheet(
        factura_gold_sheet,
        customer_name=workspace.customer.name,
        issued_at=issued_at,
        invoice_number=invoice_number,
        rows=workspace.invoice_gold.rows,
        footer_lines=workspace.invoice_gold.footer_lines,
        vat_enabled=workspace.purchase_vat_enabled,
        note=workspace.afg_note,
    )
    core._apply_afg_factura_misc_sheet(
        factura_misc_sheet,
        issued_at=issued_at,
        invoice_number=invoice_number,
        rows=workspace.invoice_misc.rows,
        vat_enabled=workspace.purchase_vat_enabled,
        note=workspace.afg_note,
    )
    core._write_sync_sheet(workbook, context=sync_context)
    return core._save_workbook_bytes(workbook)


def build_afg_workbook_bytes_from_detail(detail: "PosDocumentDetailOut", *, sync_context: "ArtifactSyncContext") -> bytes:
    core = _core()
    workbook = core._open_template(core.AFG_TEMPLATE_NAME, keep_vba=True)
    core._touch_calc_flags(workbook)
    core._apply_afg_sheet_modes(workbook, can_write=False)
    sheet = workbook["Afregningsbilag"]
    factura_gold_sheet = workbook[core.AFG_FACTURA_GOLD_SHEET]
    factura_misc_sheet = workbook[core.AFG_FACTURA_MISC_SHEET]
    vars_sheet = workbook[core.AFG_VARIABLES_SHEET]
    issued_at = detail.issued_at
    sheet["H6"] = detail.document_number
    sheet["H7"] = core._excel_datetime(issued_at)
    sheet["D2"] = core._excel_datetime(issued_at)
    core._apply_afg_market_rate_cells(vars_sheet, detail.market_rates)
    vars_sheet["C14"] = detail.numbering_preview.afregnings_number_next or detail.document_number
    vars_sheet["D14"] = detail.numbering_preview.invoice_number_next or "—"
    core._apply_afg_customer_cells(
        sheet,
        name=detail.customer_name,
        cpr_number=detail.customer_cpr,
        address=detail.customer_address,
        city=detail.customer_city,
        postal_code=detail.customer_postal_code,
        phone=detail.customer_phone,
        email=detail.customer_email,
        identity_doc=detail.customer_identity_doc_number,
    )
    core._apply_afg_detail_rows(sheet, core._aggregate_detail_gold_rows(detail), core._aggregate_detail_silver_rows(detail), core._aggregate_detail_bar_rows(detail), core._aggregate_detail_ptpd_rows(detail))
    core._apply_afg_footer_cells(sheet)
    core._apply_afg_declaration_cells(sheet)
    core._apply_afg_summary_cells(
        sheet,
        net_amount_dkk=core.quantize_2(core.to_decimal(detail.net_amount_dkk)),
        vat_amount_dkk=core.quantize_2(core.to_decimal(detail.vat_amount_dkk)),
        gross_amount_dkk=core.quantize_2(core.to_decimal(detail.gross_amount_dkk)),
        payment_method=(detail.payment_method or "bank") if hasattr(detail, "payment_method") else "bank",
        reg_number=detail.bank_reg_number,
        account_number=detail.bank_account_number,
        note=detail.notes,
    )
    core._apply_afg_factura_gold_sheet(
        factura_gold_sheet,
        customer_name=detail.customer_name,
        issued_at=issued_at,
        invoice_number=detail.numbering_preview.invoice_number_next or "—",
        rows=detail.invoice_gold.rows,
        footer_lines=detail.invoice_gold.footer_lines,
        vat_enabled=core.to_decimal(detail.vat_amount_dkk) > 0,
        note=detail.notes,
    )
    core._apply_afg_factura_misc_sheet(
        factura_misc_sheet,
        issued_at=issued_at,
        invoice_number=detail.numbering_preview.invoice_number_next or "—",
        rows=detail.invoice_misc.rows,
        vat_enabled=core.to_decimal(detail.vat_amount_dkk) > 0,
        note=detail.notes,
    )
    core._write_sync_sheet(workbook, context=sync_context)
    return core._save_workbook_bytes(workbook)


def workspace_normalized_afg_values(workspace: "PosWorkspaceOut") -> dict[str, str]:
    core = _core()
    values = {
        f"{core.AFG_PRIMARY_SHEET}!D16": workspace.customer.name or "—",
        f"{core.AFG_PRIMARY_SHEET}!G16": core.cpr_birth_part(workspace.customer.cpr_number) or "—",
        f"{core.AFG_PRIMARY_SHEET}!D17": (workspace.customer.address or "").strip() or "—",
        f"{core.AFG_PRIMARY_SHEET}!G17": workspace.customer.identity_doc_number or "—",
        f"{core.AFG_PRIMARY_SHEET}!D18": core._compose_afg_postal_line(workspace.customer.postal_code, workspace.customer.city) or "—",
        f"{core.AFG_PRIMARY_SHEET}!G18": workspace.customer.phone or "—",
        f"{core.AFG_PRIMARY_SHEET}!G19": workspace.customer.email or "—",
        f"{core.AFG_PRIMARY_SHEET}!C40": "Kontant" if workspace.payment_method == "cash" else "Overførsel",
        f"{core.AFG_PRIMARY_SHEET}!D41": "—" if workspace.payment_method == "cash" else (workspace.bank_info.reg_number or "—"),
        f"{core.AFG_PRIMARY_SHEET}!D42": "—" if workspace.payment_method == "cash" else (workspace.bank_info.account_number or "—"),
        f"{core.AFG_PRIMARY_SHEET}!A42": "1" if workspace.purchase_vat_enabled else "0",
        f"{core.AFG_PRIMARY_SHEET}!D44": workspace.afg_note or "—",
        f"{core.AFG_VARIABLES_SHEET}!C10": core._fmt_decimal(workspace.market_rates.eur_dkk_fx),
        f"{core.AFG_VARIABLES_SHEET}!C4": core._fmt_decimal(workspace.market_rates.gold_24k_dkk),
        f"{core.AFG_VARIABLES_SHEET}!C5": core._fmt_decimal(workspace.market_rates.silver_dkk),
        f"{core.AFG_VARIABLES_SHEET}!C14": workspace.numbering_preview.afregnings_number_next or "—",
        f"{core.AFG_VARIABLES_SHEET}!D14": workspace.numbering_preview.invoice_number_next or "—",
    }
    matrix_keys = (
        ("J4", "8", "gold"),
        ("J5", "14", "gold"),
        ("J6", "18", "gold"),
        ("J7", "21", "gold"),
        ("J8", "21.6", "gold"),
        ("J9", "22", "gold"),
        ("J10", "24", "gold"),
        ("J11", "999", "silver"),
        ("J12", "925", "silver"),
        ("J13", "830", "silver"),
    )
    for dkk_cell, rate_key, kind in matrix_keys:
        dkk_value = (
            workspace.market_rates.gold_rates_dkk.get(rate_key)
            if kind == "gold"
            else workspace.market_rates.silver_rates_dkk.get(rate_key)
        )
        values[f"{core.AFG_VARIABLES_SHEET}!{dkk_cell}"] = core._fmt_decimal(dkk_value)
    values[f"{core.AFG_VARIABLES_SHEET}!J14"] = core._fmt_decimal(workspace.market_rates.plet_dkk)
    values[f"{core.AFG_VARIABLES_SHEET}!J15"] = core._fmt_decimal(workspace.market_rates.gold_bar_dkk)
    values[f"{core.AFG_VARIABLES_SHEET}!J16"] = core._fmt_decimal(workspace.market_rates.silver_bar_dkk)
    values[f"{core.AFG_VARIABLES_SHEET}!J17"] = core._fmt_decimal(workspace.market_rates.platinum_dkk)
    values[f"{core.AFG_VARIABLES_SHEET}!J18"] = core._fmt_decimal(workspace.market_rates.palladium_dkk)
    for row in workspace.ptpd_rows:
        ptpd_idx = core.AFG_PTPD_PLATINUM_ROW if row.metal == "platinum" else core.AFG_PTPD_PALLADIUM_ROW
        values[f"{core.AFG_PRIMARY_SHEET}!B{ptpd_idx}"] = core._fmt_decimal(row.avance_percent)
        values[f"{core.AFG_PRIMARY_SHEET}!F{ptpd_idx}"] = core._fmt_decimal(row.gram)
    for row in workspace.bar_rows:
        bar_idx = core.AFG_BAR_GOLD_ROW if row.bar_type == "gold" else core.AFG_BAR_SILVER_ROW
        values[f"{core.AFG_PRIMARY_SHEET}!B{bar_idx}"] = core._fmt_decimal(row.avance_percent)
        values[f"{core.AFG_PRIMARY_SHEET}!F{bar_idx}"] = core._fmt_decimal(row.gram)
    for idx, row in enumerate(core._afg_gold_rows_from_workspace(workspace), start=core.AFG_GOLD_ROW_START):
        values[f"{core.AFG_PRIMARY_SHEET}!B{idx}"] = core._fmt_decimal(row.avance_percent)
        values[f"{core.AFG_PRIMARY_SHEET}!F{idx}"] = core._fmt_decimal(row.gram)
    for idx, row in enumerate(core._afg_silver_rows_from_workspace(workspace), start=core.AFG_SILVER_ROW_START):
        values[f"{core.AFG_PRIMARY_SHEET}!B{idx}"] = core._fmt_decimal(row.avance_percent)
        values[f"{core.AFG_PRIMARY_SHEET}!F{idx}"] = core._fmt_decimal(row.gram)
    calculator_payload = core._afg_calculator_rows_payload(workspace.calculators)
    calculator_by_key = {
        str(row.get("row_key") or "").strip(): row
        for row in calculator_payload["gold_rows"] + calculator_payload["silver_rows"]
        if isinstance(row, dict)
    }
    for row_key, unit_cell, count_cell, _total_cell in (*core.AFG_GOLD_CALCULATOR_ROWS, *core.AFG_SILVER_CALCULATOR_ROWS):
        row = calculator_by_key.get(row_key, {})
        values[f"{core.AFG_PRIMARY_SHEET}!{unit_cell}"] = core._fmt_decimal(row.get("unit_weight"))
        values[f"{core.AFG_PRIMARY_SHEET}!{count_cell}"] = core._fmt_decimal(row.get("count"))
    for row in workspace.invoice_gold.rows:
        row_index = int(str(row.row_key).split(":", 1)[1]) + core.AFG_FACTURA_GOLD_ROW_START - 1
        values[f"{core.AFG_FACTURA_GOLD_SHEET}!A{row_index}"] = row.code or "—"
        values[f"{core.AFG_FACTURA_GOLD_SHEET}!C{row_index}"] = row.fineness or "—"
        values[f"{core.AFG_FACTURA_GOLD_SHEET}!E{row_index}"] = core._fmt_decimal(row.gram)
    for index, line in enumerate(workspace.invoice_gold.footer_lines, start=core.AFG_FACTURA_GOLD_FOOTER_START):
        values[f"{core.AFG_FACTURA_GOLD_SHEET}!B{index}"] = line or "—"
    for row in workspace.invoice_misc.rows:
        row_index = int(str(row.row_key).split(":", 1)[1]) + core.AFG_FACTURA_MISC_ROW_START - 1
        values[f"{core.AFG_FACTURA_MISC_SHEET}!C{row_index}"] = row.text or "—"
        values[f"{core.AFG_FACTURA_MISC_SHEET}!E{row_index}"] = core._fmt_decimal(row.quantity) if row.quantity is not None else "—"
        values[f"{core.AFG_FACTURA_MISC_SHEET}!F{row_index}"] = core._fmt_decimal(row.unit_price_dkk)
    return values


def parse_afg_workspace_inputs_from_workbook(content: bytes, *, legacy_percent_avance: bool = False):
    core = _core()

    # R2-07: B kolonu YENI belgelerde HAM kr/g ("Mer pris") tasir; yalniz
    # TARIHSEL (eski) calisma kitaplarinda yuzde-fraksiyondur. legacy cagiran
    # (historical_afg_import) True gecer; canli round-trip ham okur — boylece
    # negatif (−15) ve 0<v<1 kr/g degerleri 100x bozulmaz.
    def _avance_from_cell(raw):
        if legacy_percent_avance:
            return core._percent_from_excel(raw)
        return core.quantize_2(core._decimal_from_excel(raw))
    workbook = load_workbook(io.BytesIO(content), keep_vba=True, data_only=False)
    calculated_workbook = load_workbook(io.BytesIO(content), keep_vba=True, data_only=True)
    if core.AFG_PRIMARY_SHEET not in workbook.sheetnames:
        raise ValueError("Afregningsbilag sheet bulunamadı.")
    sheet = workbook[core.AFG_PRIMARY_SHEET]
    calculated_sheet = calculated_workbook[core.AFG_PRIMARY_SHEET]
    factura_gold_sheet = workbook[core.AFG_FACTURA_GOLD_SHEET] if core.AFG_FACTURA_GOLD_SHEET in workbook.sheetnames else None
    calculated_factura_gold_sheet = (
        calculated_workbook[core.AFG_FACTURA_GOLD_SHEET]
        if core.AFG_FACTURA_GOLD_SHEET in calculated_workbook.sheetnames
        else None
    )
    factura_misc_sheet = workbook[core.AFG_FACTURA_MISC_SHEET] if core.AFG_FACTURA_MISC_SHEET in workbook.sheetnames else None
    calculated_factura_misc_sheet = (
        calculated_workbook[core.AFG_FACTURA_MISC_SHEET]
        if core.AFG_FACTURA_MISC_SHEET in calculated_workbook.sheetnames
        else None
    )
    vars_sheet = workbook[core.AFG_VARIABLES_SHEET] if core.AFG_VARIABLES_SHEET in workbook.sheetnames else None
    calculated_vars_sheet = (
        calculated_workbook[core.AFG_VARIABLES_SHEET]
        if core.AFG_VARIABLES_SHEET in calculated_workbook.sheetnames
        else None
    )

    def numeric_cell_value(source_sheet, calculated_sheet, cell_ref: str):
        source_cell = source_sheet[cell_ref]
        if source_cell.data_type != "f":
            return source_cell.value
        calculated_value = calculated_sheet[cell_ref].value if calculated_sheet is not None else None
        if calculated_value is None or calculated_value == "":
            raise ValueError(
                f"{source_sheet.title}!{cell_ref} formülünün kaydedilmiş sonucu yok. "
                "Dosyayı Excel'de açıp yeniden hesaplayarak kaydedin."
            )
        return calculated_value

    # Hücre düzeni sözleşme sürümüne bağlı: afg-v1 (0.3.5 ve öncesi) değerleri
    # etiket hücrelerine (C16/F16...) yazar; afg-v2 etiketlerin sağına (D/G).
    contract_version = core.AFG_CONTRACT_VERSION
    if core.SYNC_SHEET_NAME in workbook.sheetnames:
        try:
            contract_version = core._read_sync_sheet(workbook).contract_version or contract_version
        except ValueError:
            pass

    if contract_version == "afg-v1":
        address, city = core._split_afg_address_line(sheet["C17"].value)
        customer = PosWorkspaceCustomerUpdate(
            name=core._clean_text(sheet["C16"].value),
            cpr_number=core._clean_text(sheet["F16"].value),
            address=address,
            city=city,
            identity_doc_number=core._clean_text(sheet["F17"].value),
            postal_code=core._clean_text(sheet["C18"].value),
            phone=core._clean_text(sheet["F18"].value),
            email=core._clean_text(sheet["F19"].value),
        )
    else:
        postal_code, city = core._split_afg_postal_line(sheet["D18"].value)
        # CPR bilinçli olarak geri OKUNMAZ: G16 yalnız doğum tarihi bölümünü
        # gösterir; kayıtlı tam CPR'nin üzerine yazılmamalıdır (alan hiç set
        # edilmez → model_fields_set'e girmez → mevcut değer korunur).
        customer = PosWorkspaceCustomerUpdate(
            name=core._clean_text(sheet["D16"].value),
            address=core._clean_text(sheet["D17"].value),
            city=city,
            identity_doc_number=core._clean_text(sheet["G17"].value),
            postal_code=postal_code,
            phone=core._clean_text(sheet["G18"].value),
            email=core._clean_text(sheet["G19"].value),
        )

    payment_method = core._payment_method_from_excel(sheet["C40"].value)
    reg_number = core._clean_text(sheet["D41"].value)
    account_number = core._clean_text(sheet["D42"].value)
    purchase_vat_enabled = core._boolean_from_excel(sheet["A42"].value)
    afg_note = core._clean_text(sheet["D44"].value)

    gold_rows = [
        PosWorkspaceGoldRowInput(
            karat=Decimal(row_key.split(":", 1)[1]),
            gram=core._decimal_from_excel(numeric_cell_value(sheet, calculated_sheet, f"F{idx}")),
            avance_percent=_avance_from_cell(numeric_cell_value(sheet, calculated_sheet, f"B{idx}")),
        )
        for idx, row_key in enumerate(core.AFG_GOLD_ROW_KEYS, start=core.AFG_GOLD_ROW_START)
    ]
    silver_rows = [
        PosWorkspaceSilverRowInput(
            type_code=row_key.split(":", 1)[1],
            gram=core._decimal_from_excel(numeric_cell_value(sheet, calculated_sheet, f"F{idx}")),
            avance_percent=_avance_from_cell(numeric_cell_value(sheet, calculated_sheet, f"B{idx}")),
        )
        for idx, row_key in enumerate(core.AFG_SILVER_ROW_KEYS, start=core.AFG_SILVER_ROW_START)
    ]
    bar_rows = [
        PosWorkspaceBarRowInput(
            bar_type=bar_type,
            gram=core._decimal_from_excel(numeric_cell_value(sheet, calculated_sheet, f"F{idx}")),
            avance_percent=_avance_from_cell(numeric_cell_value(sheet, calculated_sheet, f"B{idx}")),
        )
        for bar_type, idx in (("gold", core.AFG_BAR_GOLD_ROW), ("silver", core.AFG_BAR_SILVER_ROW))
    ]
    ptpd_rows = [
        PosWorkspacePtPdRowInput(
            metal=metal,
            gram=core._decimal_from_excel(numeric_cell_value(sheet, calculated_sheet, f"F{idx}")),
            avance_percent=_avance_from_cell(numeric_cell_value(sheet, calculated_sheet, f"B{idx}")),
        )
        for metal, idx in (("platinum", core.AFG_PTPD_PLATINUM_ROW), ("palladium", core.AFG_PTPD_PALLADIUM_ROW))
    ]

    market_rates = None
    numbering = None
    if vars_sheet is not None:
        fx = (
            core._decimal_from_excel(numeric_cell_value(vars_sheet, calculated_vars_sheet, "C10"))
            if core._clean_text(vars_sheet["C10"].value) is not None
            else Decimal("7.45")
        )
        def _matrix_cell(column: str, row_idx: int) -> Decimal:
            ref = f"{column}{row_idx}"
            if core._clean_text(vars_sheet[ref].value) is None:
                return Decimal("0")
            return core._decimal_from_excel(numeric_cell_value(vars_sheet, calculated_vars_sheet, ref))

        gold_matrix_rows = {"8": 4, "14": 5, "18": 6, "21": 7, "21.6": 8, "22": 9, "24": 10}
        silver_matrix_rows = {"999": 11, "925": 12, "830": 13}
        # Kanonik birim J sütununda DKK/g'dir. 0.3.5 ve öncesi taslaklarda J,
        # I (EUR) × C10 (fx) ile yazılmış aynaydı; J boş kalan çok eski
        # dosyalarda I×fx çevrimine düşülür.
        gold_rate_cells = {
            key: core.quantize_2(_matrix_cell("J", row_idx) or (_matrix_cell("I", row_idx) * fx))
            for key, row_idx in gold_matrix_rows.items()
        }
        silver_rate_cells = {
            key: core.quantize_2(_matrix_cell("J", row_idx) or (_matrix_cell("I", row_idx) * fx))
            for key, row_idx in silver_matrix_rows.items()
        }
        plet_cell = core.quantize_2(_matrix_cell("J", 14) or (_matrix_cell("I", 14) * fx))
        gold_bar_cell = core.quantize_2(_matrix_cell("J", 15))
        silver_bar_cell = core.quantize_2(_matrix_cell("J", 16))
        platinum_cell = core.quantize_2(_matrix_cell("J", 17))
        palladium_cell = core.quantize_2(_matrix_cell("J", 18))
        has_matrix_values = any(value > 0 for value in [*gold_rate_cells.values(), *silver_rate_cells.values()])
        if has_matrix_values:
            gold_24k_dkk = (
                core._decimal_from_excel(numeric_cell_value(vars_sheet, calculated_vars_sheet, "C4"))
                if core._clean_text(vars_sheet["C4"].value) is not None
                else gold_rate_cells["24"]
            )
            silver_dkk = (
                core._decimal_from_excel(numeric_cell_value(vars_sheet, calculated_vars_sheet, "C5"))
                if core._clean_text(vars_sheet["C5"].value) is not None
                else silver_rate_cells["999"]
            )
            market_rates = core._build_afg_market_rates_from_workspace(
                PosWorkspaceMarketRates(
                    eur_dkk_fx=fx,
                    gold_rates_dkk=gold_rate_cells,
                    silver_rates_dkk=silver_rate_cells,
                    gold_24k_dkk=gold_24k_dkk,
                    silver_dkk=silver_dkk,
                    plet_dkk=plet_cell,
                    gold_bar_dkk=gold_bar_cell,
                    silver_bar_dkk=silver_bar_cell,
                    platinum_dkk=platinum_cell,
                    palladium_dkk=palladium_cell,
                )
            )
        else:
            gold_24k = core._clean_text(vars_sheet["C4"].value)
            silver = core._clean_text(vars_sheet["C5"].value)
            if gold_24k is not None or silver is not None:
                market_rates = core._build_afg_market_rates_from_workspace(
                    PosWorkspaceMarketRates(
                        eur_dkk_fx=fx,
                        gold_rates_dkk={},
                        silver_rates_dkk={},
                        gold_24k_dkk=core._decimal_from_excel(numeric_cell_value(vars_sheet, calculated_vars_sheet, "C4")),
                        silver_dkk=core._decimal_from_excel(numeric_cell_value(vars_sheet, calculated_vars_sheet, "C5")),
                    )
                )
        numbering = PosWorkspaceNumberingUpdate(
            afregnings_number_next=core._clean_text(vars_sheet["C14"].value),
            invoice_number_next=core._clean_text(vars_sheet["D14"].value),
        )
    calculators = core._parse_afg_calculators_from_sheet(sheet)

    invoice_gold_rows = []
    invoice_gold_footer_lines: list[str] = []
    if factura_gold_sheet is not None:
        for index, row_idx in enumerate(range(core.AFG_FACTURA_GOLD_ROW_START, core.AFG_FACTURA_GOLD_ROW_END + 1), start=1):
            invoice_gold_rows.append(
                {
                    "row_key": f"invoice_gold:{index}",
                    "code": core._clean_text(factura_gold_sheet[f"A{row_idx}"].value),
                    "fineness": core._clean_text(factura_gold_sheet[f"C{row_idx}"].value),
                    "gram": core._decimal_from_excel(
                        numeric_cell_value(factura_gold_sheet, calculated_factura_gold_sheet, f"E{row_idx}")
                    ),
                }
            )
        invoice_gold_footer_lines = [
            core._clean_text(factura_gold_sheet[f"B{row_idx}"].value) or ""
            for row_idx in range(core.AFG_FACTURA_GOLD_FOOTER_START, core.AFG_FACTURA_GOLD_FOOTER_START + 3)
        ]

    invoice_misc_rows = []
    if factura_misc_sheet is not None:
        for index, row_idx in enumerate(range(core.AFG_FACTURA_MISC_ROW_START, core.AFG_FACTURA_MISC_ROW_END + 1), start=1):
            invoice_misc_rows.append(
                {
                    "row_key": f"invoice_misc:{index}",
                    "text": core._clean_text(factura_misc_sheet[f"C{row_idx}"].value),
                    "quantity": (
                        core._decimal_from_excel(
                            numeric_cell_value(factura_misc_sheet, calculated_factura_misc_sheet, f"E{row_idx}")
                        )
                        if core._clean_text(factura_misc_sheet[f"E{row_idx}"].value) is not None
                        else None
                    ),
                    "unit_price_dkk": core._decimal_from_excel(
                        numeric_cell_value(factura_misc_sheet, calculated_factura_misc_sheet, f"F{row_idx}")
                    ),
                }
            )

    normalized_values: dict[str, str] = {}
    sheet_lookup = {
        core.AFG_PRIMARY_SHEET: sheet,
        core.AFG_VARIABLES_SHEET: vars_sheet,
        core.AFG_FACTURA_GOLD_SHEET: factura_gold_sheet,
        core.AFG_FACTURA_MISC_SHEET: factura_misc_sheet,
    }
    calculated_sheet_lookup = {
        core.AFG_PRIMARY_SHEET: calculated_sheet,
        core.AFG_VARIABLES_SHEET: calculated_vars_sheet,
        core.AFG_FACTURA_GOLD_SHEET: calculated_factura_gold_sheet,
        core.AFG_FACTURA_MISC_SHEET: calculated_factura_misc_sheet,
    }
    for cell in core.AFG_EDITABLE_CELLS:
        sheet_name = cell.get("sheet", core.AFG_PRIMARY_SHEET)
        cell_sheet = sheet_lookup.get(sheet_name)
        if cell_sheet is None:
            continue
        value = cell_sheet[cell["cell_ref"]].value
        if cell["input_kind"] in {"decimal", "percent"}:
            value = numeric_cell_value(cell_sheet, calculated_sheet_lookup.get(sheet_name), cell["cell_ref"])
        normalized_values[core._afg_editable_cell_key(cell)] = core._normalized_cell_value(
            value,
            input_kind=cell["input_kind"],
        )

    base_version = None
    if core.SYNC_SHEET_NAME in workbook.sheetnames:
        try:
            metadata = core._read_sync_sheet(workbook)
            if metadata.kind == "alis-workspace":
                base_version = metadata.base_version
        except ValueError:
            base_version = None

    return core.AfgWorkspaceArtifactInputs(
        customer=customer,
        sections=PosWorkspaceSectionsUpdate(
            gold_rows=gold_rows,
            silver_rows=silver_rows,
            bar_rows=bar_rows,
            ptpd_rows=ptpd_rows,
            bank_info={
                "reg_number": reg_number,
                "account_number": account_number,
            },
            payment_method=payment_method,
            afg_note=afg_note,
            purchase_vat_enabled=purchase_vat_enabled,
            purchase_vat_rate_percent=Decimal("25.00") if purchase_vat_enabled else Decimal("0.00"),
            market_rates=market_rates,
            calculators=calculators,
            numbering=numbering,
            invoice_gold=PosWorkspaceInvoiceGoldSheetUpdate(
                rows=invoice_gold_rows,
                footer_lines=invoice_gold_footer_lines,
            ),
            invoice_misc=PosWorkspaceInvoiceMiscSheetUpdate(
                rows=invoice_misc_rows,
            ),
        ),
        normalized_values=normalized_values,
        base_version=base_version,
    )


def build_afg_workspace_reconcile_preview(workspace: "PosWorkspaceOut", parsed) -> DocumentArtifactReconcilePreviewOut:
    core = _core()
    current_values = workspace_normalized_afg_values(workspace)
    changes: list[DocumentArtifactCellChangeOut] = []
    for cell in core.AFG_EDITABLE_CELLS:
        cell_ref = cell["cell_ref"]
        sheet_name = cell.get("sheet", core.AFG_PRIMARY_SHEET)
        cell_key = core._afg_editable_cell_key(cell)
        old_value = current_values.get(cell_key, "—")
        new_value = parsed.normalized_values.get(cell_key, "—")
        if old_value == new_value:
            continue
        changes.append(
            DocumentArtifactCellChangeOut(
                sheet=sheet_name,
                cell_ref=cell_ref,
                label=cell["label"],
                old_value=old_value,
                new_value=new_value,
            )
        )
    warnings: list[str] = []
    warnings.append("Yalnız kontrollü alanlar sisteme uygulanır; FAKTURA ve BRUGSANVISNING sheet değişiklikleri domain’e yazılmaz.")
    if not changes:
        warnings.append("Workbook mevcut taslakla aynı; uygulanacak değişiklik yok.")
    return DocumentArtifactReconcilePreviewOut(
        editable=True,
        changes=changes,
        warnings=warnings,
    )
