from __future__ import annotations

import hashlib
import json
import re
import unicodedata
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.afg import build_log_workspace
from app.config import get_settings
from app.database import AsyncSessionLocal
from app.models.legacy_migration import (
    LegacyMigrationFile,
    LegacyMigrationLink,
    LegacyMigrationRecord,
    LegacyMigrationRun,
)
from app.models.product import Product
from app.models.user import User
from app.schemas.product import ProductCreate
from app.services.document_artifact_service import parse_log_workbook_inputs_from_workbook
from app.services.historical_afg_import import (
    HistoricalAfgUpload,
    apply_historical_afg_import,
    preview_historical_afg_import,
)
from app.services.product_service import create_product


PHASES = ("afg", "inventory", "log")
TERMINAL_FILE_STATES = {"ready", "blocked", "already_imported", "skipped", "applied", "failed"}


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _plain(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime,)):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _plain(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_plain(item) for item in value]
    return value


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", _clean_text(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text.lower()).strip()


def _decimal(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = _clean_text(value).lower().replace("dkk", "").replace("kr.", "").replace("kr", "").replace(" ", "")
    text = re.sub(r"[^0-9,.-]", "", text)
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text or "0")
    except InvalidOperation:
        return Decimal("0")


def _section(label: str) -> str | None:
    if "guldbarrer" in label:
        return "gold_bar"
    if "guldmonter" in label:
        return "gold_coin"
    if "guldsmykker" in label:
        return "gold_jewelry"
    if "solvmonter" in label:
        return "silver_coin"
    if "solv" in label and ("oversigt" in label or "barrer" in label or "smykker" in label):
        return "silver"
    if "platin" in label or "palladium" in label:
        return "platinum_palladium"
    return None


def _purity(values: list[object], section: str, name: str) -> tuple[str | None, Decimal | None]:
    text = " ".join(_normalized(value) for value in values[:4])
    karat_match = re.search(r"(?<!\d)(8|9|14|18|21[.,]?6|22|24)\s*k(?:t|arat)?\b", text)
    if karat_match:
        karat = karat_match.group(1).replace(",", ".")
        mapping = {"8": "33.30", "9": "37.50", "14": "58.50", "18": "75.00", "21.6": "90.00", "22": "91.70", "24": "99.90"}
        return f"{karat}K", Decimal(mapping[karat])
    for token in re.findall(r"(?<!\d)(333|375|585|750|800|830|875|900|916|917|925|999)(?!\d)", text):
        numeric = Decimal(token) / Decimal("10")
        karat = {"333": "8K", "375": "9K", "585": "14K", "750": "18K", "875": "21K", "900": "21.6K", "916": "22K", "917": "22K", "999": "24K"}.get(token)
        return karat or token, numeric
    if section == "gold_bar":
        return "24K", Decimal("99.90")
    if section in {"silver", "silver_coin"} and "sterling" in _normalized(name):
        return "925", Decimal("92.50")
    if section == "platinum_palladium":
        return "999", Decimal("99.90")
    return None, None


def parse_legacy_inventory(content: bytes, source_hash: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    if "Lager" not in workbook.sheetnames:
        raise ValueError("Depolama dosyasında Lager sayfası bulunamadı.")
    sheet = workbook["Lager"]
    market_snapshot = {
        "gold": _plain(sheet["K4"].value),
        "silver": _plain(sheet["K5"].value),
        "platinum": _plain(sheet["K6"].value),
        "palladium": _plain(sheet["K7"].value),
        "policy": "historical_snapshot_only",
    }
    current_section: str | None = None
    records: list[dict[str, Any]] = []
    for row_idx in range(1, sheet.max_row + 1):
        values = [sheet.cell(row_idx, col).value for col in range(1, 12)]
        detected = _section(" ".join(_normalized(value) for value in values[:4]))
        if detected:
            current_section = detected
            continue
        if current_section is None:
            continue
        unit_weight = _decimal(values[3])
        unit_count_raw = _decimal(values[4])
        total_weight = _decimal(values[5])
        if total_weight <= 0 and unit_weight > 0 and unit_count_raw > 0:
            total_weight = unit_weight * unit_count_raw
        if unit_count_raw <= 0 and total_weight <= 0:
            continue
        unit_count = max(1, int(unit_count_raw or 1))
        if unit_weight <= 0 and total_weight > 0:
            unit_weight = total_weight / Decimal(unit_count)
        purchase_price = _decimal(values[7])
        shop_price = _decimal(values[10])
        reference = _clean_text(values[0])
        if reference and not re.fullmatch(r"[A-Za-z0-9._/-]+", reference):
            reference = ""
        name_candidates = [_clean_text(values[index]) for index in (2, 1, 0)]
        display_name = next((value for value in name_candidates if value and not re.fullmatch(r"\d+(?:[.,]\d+)?", value)), "")
        display_name = display_name or {
            "gold_bar": "Altın külçe",
            "gold_coin": "Altın sikke",
            "gold_jewelry": "Altın takı",
            "silver": "Gümüş",
            "silver_coin": "Gümüş sikke",
            "platinum_palladium": "Platin / Palladyum",
        }[current_section]
        purity_karat, purity_percentage = _purity(values, current_section, display_name)
        errors: list[str] = []
        warnings: list[str] = []
        if unit_weight <= 0:
            errors.append("Pozitif ürün ağırlığı bulunamadı.")
        if purchase_price <= 0:
            errors.append("Pozitif alış fiyatı bulunamadı.")
        if purity_percentage is None:
            errors.append("Saflık/ayar bilgisi çözümlenemedi.")
        if len(reference) > 10:
            errors.append("Eski stok numarası 10 karakter sınırını aşıyor.")
        if any(_clean_text(value).upper() == "#REF!" for value in values):
            warnings.append("Kaynak satırda #REF! hücresi var; hesaplanan alanlar kontrol edilmeli.")
        if not reference and current_section == "gold_jewelry":
            warnings.append("Takı satırında eski stok numarası yok.")
        if current_section == "platinum_palladium":
            metal_type = "palladium" if "palladium" in _normalized(display_name) else "platinum"
        elif current_section in {"silver", "silver_coin"}:
            metal_type = "silver"
        else:
            metal_type = "yellow_gold"
        category = {
            "gold_bar": "kulce",
            "gold_coin": "sikke",
            "gold_jewelry": "taki",
            "silver": "gumus",
            "silver_coin": "gumus",
            "platinum_palladium": "platin_pd",
        }[current_section]
        subcategory = None
        if current_section == "silver":
            subcategory = "barrer" if "bar" in _normalized(display_name) else "smykker"
        if current_section == "silver_coin":
            subcategory = "monter"
        if current_section == "platinum_palladium":
            subcategory = metal_type
        source_key = f"inventory:{source_hash}:Lager:{current_section}:{row_idx}"
        records.append(
            {
                "source_key": source_key,
                "status": "blocked" if errors else "ready",
                "errors": errors,
                "warnings": warnings,
                "product": {
                    "reference_number": reference or None,
                    "display_name": display_name,
                    "product_type": "bar" if current_section == "gold_bar" else "jewelry",
                    "metal_type": metal_type,
                    "weight_grams": str(unit_weight.quantize(Decimal("0.01"))),
                    "total_weight_grams": str(total_weight.quantize(Decimal("0.01"))),
                    "unit_count": unit_count,
                    "purity_karat": purity_karat,
                    "purity_percentage": str(purity_percentage) if purity_percentage is not None else None,
                    "purchase_date": _now().isoformat(),
                    "purchase_price_dkk": str(purchase_price.quantize(Decimal("0.01"))),
                    "shop_price_dkk": str(shop_price.quantize(Decimal("0.01"))) if shop_price > 0 else None,
                    "storage_location": "Legacy Lager",
                    "inventory_category": category,
                    "inventory_subcategory": subcategory,
                    "operation_destination": "inventory",
                    "notes": json.dumps({"kind": "legacy_inventory_import_v1", "source_row": row_idx, "section": current_section}, ensure_ascii=False),
                    "photos": [],
                },
                "source_row": row_idx,
                "section": current_section,
            }
        )
    if not records:
        raise ValueError("Lager sayfasında pozitif adet veya gram içeren gerçek stok satırı bulunamadı.")
    return records, market_snapshot


async def get_run(db: AsyncSession, run_id: uuid.UUID) -> LegacyMigrationRun:
    run = await db.get(LegacyMigrationRun, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Taşıma çalışması bulunamadı.")
    return run


async def create_run(db: AsyncSession, actor: User, *, log_year: int) -> LegacyMigrationRun:
    run = LegacyMigrationRun(
        status="in_progress",
        current_phase="afg",
        created_by_user_id=actor.id,
        settings_json={"log_year": log_year, "external_effects": "disabled", "market_rate_policy": "snapshot_only"},
    )
    db.add(run)
    await db.flush()
    return run


async def store_file(
    db: AsyncSession,
    *,
    run: LegacyMigrationRun,
    phase: str,
    file_name: str,
    content: bytes,
) -> LegacyMigrationFile:
    if phase not in PHASES:
        raise HTTPException(status_code=422, detail="Geçersiz taşıma adımı.")
    extension = Path(file_name).suffix.lower()
    if extension not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=422, detail=f"Desteklenmeyen dosya: {file_name}")
    source_hash = hashlib.sha256(content).hexdigest()
    existing = await db.scalar(
        select(LegacyMigrationFile).where(
            LegacyMigrationFile.run_id == run.id,
            LegacyMigrationFile.phase == phase,
            LegacyMigrationFile.sha256 == source_hash,
        )
    )
    if existing is not None:
        return existing
    root = get_settings().media_root_path() / "legacy-migrations" / str(run.id) / phase
    root.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(file_name).name)[:180] or f"upload{extension}"
    path = root / f"{source_hash[:16]}-{safe_name}"
    path.write_bytes(content)
    record = LegacyMigrationFile(
        run_id=run.id,
        phase=phase,
        file_name=Path(file_name).name,
        sha256=source_hash,
        stored_path=str(path),
        size_bytes=len(content),
        status="uploaded",
        summary_json={},
    )
    db.add(record)
    await db.flush()
    return record


async def _linked(db: AsyncSession, source_key: str) -> bool:
    return await db.scalar(select(LegacyMigrationLink.id).where(LegacyMigrationLink.source_key == source_key)) is not None


async def _replace_records(db: AsyncSession, file: LegacyMigrationFile, rows: list[dict[str, Any]], entity_type: str) -> None:
    await db.execute(delete(LegacyMigrationRecord).where(LegacyMigrationRecord.file_id == file.id))
    await db.flush()
    for row in rows:
        status = row["status"]
        if await _linked(db, row["source_key"]):
            status = "already_imported"
        db.add(
            LegacyMigrationRecord(
                file_id=file.id,
                source_key=row["source_key"],
                entity_type=entity_type,
                status=status,
                payload_json=_plain(row),
                warnings_json=_plain(row.get("warnings", [])),
                errors_json=_plain(row.get("errors", [])),
                resolution_json={},
            )
        )


async def analyze_phase(db: AsyncSession, run_id: uuid.UUID, phase: str) -> None:
    run = await get_run(db, run_id)
    files = list(
        (await db.scalars(select(LegacyMigrationFile).where(LegacyMigrationFile.run_id == run.id, LegacyMigrationFile.phase == phase))).all()
    )
    if not files:
        raise ValueError("Analiz edilecek dosya yüklenmedi.")
    run.status = "analyzing"
    await db.commit()
    for file in files:
        file.status = "analyzing"
        file.error_text = None
        await db.commit()
        try:
            content = Path(file.stored_path).read_bytes()
            if phase == "afg":
                upload = HistoricalAfgUpload.from_content(filename=file.file_name, content=content)
                preview = await preview_historical_afg_import(db, uploads=[upload])
                item = preview.items[0]
                row = item.model_dump(mode="json")
                row.update({"source_key": f"afg:{file.sha256}", "status": item.status, "warnings": item.warnings, "errors": item.errors})
                await _replace_records(db, file, [row], "afg_document")
                file.summary_json = preview.model_dump(mode="json")
                file.status = item.status
            elif phase == "inventory":
                rows, market_snapshot = parse_legacy_inventory(content, file.sha256)
                for row in rows:
                    reference = row["product"].get("reference_number")
                    if reference:
                        existing = await db.scalar(select(Product.id).where(Product.reference_number == reference, Product.deleted_at.is_(None)))
                        if existing is not None and not await _linked(db, row["source_key"]):
                            row["status"] = "blocked"
                            row["errors"].append(f"Stok numarası mevcut bir üründe kullanılıyor: {reference}")
                await _replace_records(db, file, rows, "inventory_product")
                file.summary_json = {
                    "market_snapshot": market_snapshot,
                    "record_count": len(rows),
                    "ready_count": sum(row["status"] == "ready" for row in rows),
                    "blocked_count": sum(row["status"] == "blocked" for row in rows),
                    "skipped_zero_stock": max(0, 420 - len(rows)),
                }
                file.status = "blocked" if any(row["status"] == "blocked" for row in rows) else "ready"
            else:
                year = int((run.settings_json or {}).get("log_year") or _now().year)
                workspace = await build_log_workspace(db, q=None, limit=10000, year=year)
                parsed = parse_log_workbook_inputs_from_workbook(content, year=year, current_workspace=workspace)
                source_key = f"log:{file.sha256}:{year}"
                status = "already_imported" if await _linked(db, source_key) else "ready"
                row = {
                    "source_key": source_key,
                    "status": status,
                    "warnings": ["Tarihsel Log importunda dış entegrasyonlar kapalıdır."],
                    "errors": [],
                    "year": year,
                    "route_count": len(parsed.route_updates),
                    "lot_create_count": len(parsed.lot_creates),
                    "lot_update_count": len(parsed.lot_updates),
                }
                await _replace_records(db, file, [row], "log_workbook")
                file.summary_json = _plain(row)
                file.status = status
            file.analyzed_at = _now()
        except Exception as exc:
            file.status = "blocked"
            file.error_text = str(getattr(exc, "detail", exc))
            await _replace_records(
                db,
                file,
                [{"source_key": f"{phase}:{file.sha256}:error", "status": "blocked", "warnings": [], "errors": [file.error_text]}],
                f"{phase}_file",
            )
        await db.commit()
    run = await get_run(db, run_id)
    run.status = "in_progress"
    await db.commit()


async def run_analysis_job(run_id: uuid.UUID, phase: str) -> None:
    async with AsyncSessionLocal() as db:
        try:
            await analyze_phase(db, run_id, phase)
        except Exception:
            await db.rollback()
            run = await db.get(LegacyMigrationRun, run_id)
            if run is not None:
                run.status = "failed"
                await db.commit()


async def _phase_records(db: AsyncSession, run_id: uuid.UUID, phase: str) -> list[LegacyMigrationRecord]:
    return list(
        (
            await db.scalars(
                select(LegacyMigrationRecord)
                .join(LegacyMigrationFile, LegacyMigrationFile.id == LegacyMigrationRecord.file_id)
                .where(LegacyMigrationFile.run_id == run_id, LegacyMigrationFile.phase == phase)
                .order_by(LegacyMigrationRecord.created_at.asc())
            )
        ).all()
    )


async def apply_phase(db: AsyncSession, run: LegacyMigrationRun, phase: str, actor: User) -> None:
    if phase != run.current_phase:
        raise HTTPException(status_code=409, detail=f"Önce {run.current_phase} adımı tamamlanmalı.")
    files = list(
        (await db.scalars(select(LegacyMigrationFile).where(LegacyMigrationFile.run_id == run.id, LegacyMigrationFile.phase == phase))).all()
    )
    records = await _phase_records(db, run.id, phase)
    if not files or not records:
        raise HTTPException(status_code=422, detail="Önce dosyaları yükleyip analiz edin.")
    blocked = [record for record in records if record.status in {"blocked", "failed"}]
    if blocked:
        raise HTTPException(status_code=409, detail=f"{len(blocked)} engelli kayıt çözülmeden bu adım uygulanamaz.")
    ready = [record for record in records if record.status == "ready"]
    if phase == "afg" and ready:
        ready_hashes = {record.source_key.split(":", 1)[1] for record in ready}
        uploads = [
            HistoricalAfgUpload.from_content(filename=file.file_name, content=Path(file.stored_path).read_bytes())
            for file in files
            if file.sha256 in ready_hashes
        ]
        result = await apply_historical_afg_import(db, uploads=uploads, selected_hashes=sorted(ready_hashes), actor=actor)
        if result.failed_count or result.skipped_count:
            raise HTTPException(status_code=409, detail="AFG adımı atomik uygulanamadı; hiçbir kayıt kaydedilmedi.")
        result_by_hash = {item.source_hash: item for item in result.items}
        for record in ready:
            item = result_by_hash[record.source_key.split(":", 1)[1]]
            record.status = "applied"
            db.add(LegacyMigrationLink(run_id=run.id, record_id=record.id, source_key=record.source_key, entity_type="afg_document", entity_id=str(item.sequence_no), before_json={}, after_json=item.model_dump(mode="json")))
    elif phase == "inventory":
        for record in ready:
            payload = ProductCreate.model_validate(record.payload_json["product"])
            created = await create_product(db, payload, actor.id, commit=False)
            record.status = "applied"
            db.add(LegacyMigrationLink(run_id=run.id, record_id=record.id, source_key=record.source_key, entity_type="inventory_product", entity_id=str(created.id), before_json={}, after_json=created.model_dump(mode="json")))
    elif phase == "log" and ready:
        from app.api.v2 import _apply_log_workbook_artifact_inputs

        year = int((run.settings_json or {}).get("log_year") or _now().year)
        for file in files:
            record = next((item for item in ready if item.source_key == f"log:{file.sha256}:{year}"), None)
            if record is None:
                continue
            await _apply_log_workbook_artifact_inputs(db, year=year, workbook_bytes=Path(file.stored_path).read_bytes(), create_snapshot=True)
            record.status = "applied"
            db.add(LegacyMigrationLink(run_id=run.id, record_id=record.id, source_key=record.source_key, entity_type="log_workbook", entity_id=str(year), before_json={}, after_json=record.payload_json))
    for file in files:
        if file.status != "already_imported":
            file.status = "applied"
    next_index = PHASES.index(phase) + 1
    if next_index < len(PHASES):
        run.current_phase = PHASES[next_index]
        run.status = "in_progress"
    else:
        run.status = "completed"
        run.completed_at = _now()


async def serialize_run(db: AsyncSession, run: LegacyMigrationRun) -> dict[str, Any]:
    files = list((await db.scalars(select(LegacyMigrationFile).where(LegacyMigrationFile.run_id == run.id).order_by(LegacyMigrationFile.created_at))).all())
    records = await _phase_records(db, run.id, "afg") + await _phase_records(db, run.id, "inventory") + await _phase_records(db, run.id, "log")
    phases: dict[str, Any] = {}
    for phase in PHASES:
        phase_files = [file for file in files if file.phase == phase]
        phase_records = [record for record in records if any(file.id == record.file_id and file.phase == phase for file in phase_files)]
        statuses = {status: sum(record.status == status for record in phase_records) for status in ("ready", "blocked", "already_imported", "applied", "skipped")}
        if phase_files and all(file.status in {"applied", "already_imported"} for file in phase_files):
            phase_status = "applied"
        elif any(file.status in {"blocked", "failed"} for file in phase_files):
            phase_status = "blocked"
        elif any(file.status == "analyzing" for file in phase_files):
            phase_status = "analyzing"
        elif phase_files and phase_records:
            phase_status = "ready"
        elif phase_files:
            phase_status = "uploaded"
        else:
            phase_status = "empty"
        phases[phase] = {"status": phase_status, "file_count": len(phase_files), "record_count": len(phase_records), **statuses}
    return {
        "id": str(run.id),
        "status": run.status,
        "current_phase": run.current_phase,
        "settings": run.settings_json,
        "phases": phases,
        "files": [
            {"id": str(file.id), "phase": file.phase, "file_name": file.file_name, "status": file.status, "summary": file.summary_json, "error": file.error_text}
            for file in files
        ],
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "completed_at": run.completed_at.isoformat() if run.completed_at else None,
    }


async def list_records(db: AsyncSession, run_id: uuid.UUID, phase: str, *, offset: int, limit: int) -> dict[str, Any]:
    base = (
        select(LegacyMigrationRecord)
        .join(LegacyMigrationFile, LegacyMigrationFile.id == LegacyMigrationRecord.file_id)
        .where(LegacyMigrationFile.run_id == run_id, LegacyMigrationFile.phase == phase)
    )
    total = int((await db.scalar(select(func.count()).select_from(base.subquery()))) or 0)
    records = list((await db.scalars(base.order_by(LegacyMigrationRecord.created_at).offset(offset).limit(limit))).all())
    return {
        "items": [
            {"id": str(record.id), "source_key": record.source_key, "entity_type": record.entity_type, "status": record.status, "payload": record.payload_json, "warnings": record.warnings_json, "errors": record.errors_json, "resolution": record.resolution_json}
            for record in records
        ],
        "total_count": total,
        "offset": offset,
        "limit": limit,
        "has_more": offset + len(records) < total,
    }
