from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable
from uuid import UUID

from openpyxl.utils.datetime import from_excel as excel_datetime_from_serial
from openpyxl.worksheet.datavalidation import DataValidation


def _core():
    from app.services import document_artifact_service as core

    return core


def _log_default_classification(metal_type: str | None, classification: str | None):
    if classification in {"jewelry_cleaning", "white_gold", "separate_storage"}:
        return classification
    if metal_type == "white_gold":
        return "white_gold"
    return "standard"


def _log_route_code(destination: str | None, classification: str | None, *, metal_type: str | None = None) -> str:
    if destination == "melt":
        return "M"
    if destination == "inventory":
        effective_classification = _log_default_classification(metal_type, classification)
        if effective_classification == "white_gold":
            return "H"
        if effective_classification == "separate_storage":
            return "D"
        return "S"
    return "-"


def _normalize_log_route_code(value: object) -> str:
    core = _core()
    text = (core._clean_text(value) or "-").upper()
    if text in core.LOG_ROUTE_CODE_OPTIONS:
        return text
    raise ValueError(f"Geçersiz log route code: {text}")


def _log_route_request_from_code(
    *,
    route_code: str,
    current_classification: str | None,
    metal_type: str | None,
    note: str | None,
    line_id: UUID,
):
    core = _core()
    preserved_classification = _log_default_classification(metal_type, current_classification)
    if route_code == "S":
        next_classification = preserved_classification if preserved_classification in {"standard", "jewelry_cleaning"} else "jewelry_cleaning"
        return core.AfgRouteRequest(
            line_ids=[line_id],
            destination="inventory",
            classification=next_classification,
            note=note,
        )
    if route_code == "H":
        return core.AfgRouteRequest(line_ids=[line_id], destination="inventory", classification="white_gold", note=note)
    if route_code == "D":
        return core.AfgRouteRequest(line_ids=[line_id], destination="inventory", classification="separate_storage", note=note)
    if route_code == "M":
        return core.AfgRouteRequest(
            line_ids=[line_id],
            destination="melt",
            classification=preserved_classification,
            note=note,
        )
    return core.AfgRouteRequest(
        line_ids=[line_id],
        destination="undecided",
        classification=preserved_classification,
        note=note,
    )


def _log_sheet_lines(workspace, metal_bucket: str) -> list[dict[str, Any]]:
    bucket = workspace.gold if metal_bucket == "gold" else workspace.silver
    rows: list[dict[str, Any]] = []
    for document in bucket.documents:
        for line in document.lines:
            classification = _log_default_classification(line.metal_type, line.operation_classification)
            rows.append(
                {
                    "metal_bucket": metal_bucket,
                    "document_number": document.document_number,
                    "document_sequence_no": document.sequence_no,
                    "issued_at": document.issued_at,
                    "line_id": str(line.id),
                    "line_no": line.line_no,
                    "customer_name": line.customer_name or document.customer_name or "—",
                    "weight_grams": _core().quantize_2(_core().to_decimal(line.weight_grams or 0)),
                    "line_total_dkk": _core().quantize_2(_core().to_decimal(line.line_total_dkk or 0)),
                    "pure_gold_grams": _core().quantize_2(_core().to_decimal(line.pure_gold_grams or 0)),
                    "destination": (line.operation_destination or "undecided"),
                    "classification": classification,
                    "metal_type": line.metal_type,
                    "note": line.product_notes or "",
                    "route_code": _log_route_code(line.operation_destination, classification, metal_type=line.metal_type),
                    "reference_number": line.reference_number or "—",
                }
            )
    rows.sort(key=lambda item: (item["document_sequence_no"], item["line_no"]))
    return rows


def _log_route_line_rows(workspace) -> list[dict[str, Any]]:
    return [*_log_sheet_lines(workspace, "gold"), *_log_sheet_lines(workspace, "silver")]


def _log_bucket_group_key(row: dict[str, Any]) -> str | None:
    if row["route_code"] == "H":
        return "white_gold"
    if row["route_code"] == "D":
        return "separate_storage"
    if row["route_code"] == "S":
        return "jewelry_cleaning"
    return None


def _log_active_lot_slot(workspace, metal_bucket: str) -> dict[str, Any]:
    core = _core()
    bucket = workspace.gold if metal_bucket == "gold" else workspace.silver
    if bucket.melt_lots:
        lot = bucket.melt_lots[0]
        return {
            "mode": "update",
            "metal_bucket": metal_bucket,
            "lot_id": str(lot.id),
            "sent_date": lot.sent_date.isoformat() if lot.sent_date else "",
            "purchased_from_date": lot.purchased_from_date.isoformat() if lot.purchased_from_date else "",
            "before_weight_grams": core.quantize_2(core.to_decimal(lot.before_weight_grams)),
            "before_amount_dkk": core.quantize_2(core.to_decimal(lot.before_amount_dkk)),
            "before_pure_gold_grams": core.quantize_2(core.to_decimal(lot.before_pure_gold_grams)),
            "after_pure_gold_grams": core.quantize_2(core.to_decimal(lot.after_pure_gold_grams)),
            "insurance_dkk": core.quantize_2(core.to_decimal(lot.insurance_dkk)),
            "shipping_dkk": core.quantize_2(core.to_decimal(lot.shipping_dkk)),
            "refining_dkk": core.quantize_2(core.to_decimal(lot.refining_dkk)),
            "sale_date": lot.sale_date.isoformat() if lot.sale_date else "",
            "quote_eur": core._fmt_optional_decimal(lot.quote_eur),
            "exchange_rate_dkk": core._fmt_decimal(lot.exchange_rate_dkk),
            "payout_total_dkk": core._fmt_optional_decimal(lot.payout_total_dkk),
        }
    return {
        "mode": "create",
        "metal_bucket": metal_bucket,
        "lot_id": f"new:{metal_bucket}",
        "sent_date": "",
        "purchased_from_date": "",
        "before_weight_grams": core.quantize_2(core.to_decimal(bucket.melt_queue.total_weight_grams)),
        "before_amount_dkk": core.quantize_2(core.to_decimal(bucket.melt_queue.total_amount_dkk)),
        "before_pure_gold_grams": core.quantize_2(core.to_decimal(bucket.melt_queue.total_pure_gold_grams)),
        "after_pure_gold_grams": "",
        "insurance_dkk": "",
        "shipping_dkk": "",
        "refining_dkk": "",
        "sale_date": "",
        "quote_eur": "",
        "exchange_rate_dkk": "",
        "payout_total_dkk": "",
    }


def _log_lot_rows(workspace) -> list[dict[str, Any]]:
    return [
        _log_active_lot_slot(workspace, "gold"),
        _log_active_lot_slot(workspace, "silver"),
    ]


def _set_log_lot_section(sheet, *, metal_bucket: str, row: dict[str, Any], editable_refs: list[str], mappings: list) -> None:
    core = _core()
    if metal_bucket == "gold":
        cells = {
            "sent_date": "B37",
            "purchased_from_date": "C37",
            "before_weight": "D37",
            "before_amount": "E37",
            "before_pure": "F37",
            "after_pure": "F38",
            "insurance": "B41",
            "shipping": "B42",
            "refining": "B43",
            "sale_date": "E44",
            "quote_eur": "E45",
            "exchange_rate": "E46",
            "payout_total": "B47",
        }
        section_row = 37
        core._set_sheet_cell(sheet, "D38", "=F38")
        core._set_sheet_cell(sheet, "D39", "=D37-D38")
        core._set_sheet_cell(sheet, "F39", "=F37-F38")
    else:
        cells = {
            "sent_date": "B88",
            "purchased_from_date": "C88",
            "before_weight": "D88",
            "before_amount": "E88",
            "before_pure": "F88",
            "after_pure": "F90",
            "insurance": "B93",
            "refining": "B94",
            "shipping": "B95",
            "sale_date": "E96",
            "quote_eur": "E97",
            "exchange_rate": "E98",
            "payout_total": "B99",
        }
        section_row = 88
        core._set_sheet_cell(sheet, "D90", "=F90")
        core._set_sheet_cell(sheet, "D91", "=D88-D90")
        core._set_sheet_cell(sheet, "F91", "=F88-F90")

    core._set_sheet_cell(sheet, cells["sent_date"], row["sent_date"] or None)
    core._set_sheet_cell(sheet, cells["purchased_from_date"], row["purchased_from_date"] or None)
    core._set_sheet_cell(sheet, cells["before_weight"], row["before_weight_grams"])
    core._set_sheet_cell(sheet, cells["before_amount"], row["before_amount_dkk"])
    core._set_sheet_cell(sheet, cells["before_pure"], row["before_pure_gold_grams"])
    core._set_sheet_cell(sheet, cells["after_pure"], row["after_pure_gold_grams"] or None)
    core._set_sheet_cell(sheet, cells["insurance"], row["insurance_dkk"] or None)
    core._set_sheet_cell(sheet, cells["shipping"], row["shipping_dkk"] or None)
    core._set_sheet_cell(sheet, cells["refining"], row["refining_dkk"] or None)
    core._set_sheet_cell(sheet, cells["sale_date"], row["sale_date"] or None)
    core._set_sheet_cell(sheet, cells["quote_eur"], row["quote_eur"] or None)
    core._set_sheet_cell(sheet, cells["exchange_rate"], row["exchange_rate_dkk"] or None)
    core._set_sheet_cell(sheet, cells["payout_total"], row["payout_total_dkk"] or None)

    editable_refs.extend(
        [
            cells["sent_date"],
            cells["purchased_from_date"],
            cells["after_pure"],
            cells["insurance"],
            cells["shipping"],
            cells["refining"],
            cells["sale_date"],
            cells["quote_eur"],
            cells["exchange_rate"],
            cells["payout_total"],
        ]
    )
    mappings.append(
        core.SyncSheetRowMapping(
            mapping_type="log_lot",
            sheet_name=core.LOG_SHEET,
            row_index=section_row,
            entity_id=row["lot_id"],
            entity_key=row["metal_bucket"],
            extra={
                "mode": row["mode"],
                "sent_date": row["sent_date"],
                "purchased_from_date": row["purchased_from_date"],
                "after_pure_gold_grams": str(row["after_pure_gold_grams"]),
                "insurance_dkk": str(row["insurance_dkk"]),
                "shipping_dkk": str(row["shipping_dkk"]),
                "refining_dkk": str(row["refining_dkk"]),
                "sale_date": row["sale_date"],
                "quote_eur": str(row["quote_eur"]),
                "exchange_rate_dkk": str(row["exchange_rate_dkk"]),
                "payout_total_dkk": str(row["payout_total_dkk"]),
            },
        )
    )


def _add_log_route_validation(sheet, refs: Iterable[str]) -> None:
    validation = DataValidation(type="list", formula1='"-,S,H,D,M"', allow_blank=False)
    validation.promptTitle = "Route Code"
    validation.prompt = "Kullanilabilir kodlar: -, S, H, D, M"
    validation.errorTitle = "Geçersiz rota"
    validation.error = "Route code yalniz -, S, H, D veya M olabilir."
    sheet.add_data_validation(validation)
    for cell_ref in refs:
        validation.add(sheet[cell_ref])


def _datetime_from_excel(value: object) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, (int, float)):
        parsed = excel_datetime_from_serial(value)
        if isinstance(parsed, datetime):
            return parsed.replace(tzinfo=None)
        if isinstance(parsed, date):
            return datetime.combine(parsed, datetime.min.time())
    text = str(value).strip()
    if not text or text in {"—", "–", "-"}:
        return None
    for candidate in (text, text.replace(".", "-")):
        try:
            return datetime.fromisoformat(candidate)
        except ValueError:
            continue
    return None


def _parse_log_lot_section(sheet, *, metal_bucket: str, row_idx: int, extra: dict[str, Any]):
    core = _core()
    if metal_bucket == "gold":
        refs = {
            "sent_date": "B37",
            "purchased_from_date": "C37",
            "after_pure_gold_grams": "F38",
            "insurance_dkk": "B41",
            "shipping_dkk": "B42",
            "refining_dkk": "B43",
            "sale_date": "E44",
            "quote_eur": "E45",
            "exchange_rate_dkk": "E46",
            "payout_total_dkk": "B47",
        }
    else:
        refs = {
            "sent_date": "B88",
            "purchased_from_date": "C88",
            "after_pure_gold_grams": "F90",
            "insurance_dkk": "B93",
            "shipping_dkk": "B95",
            "refining_dkk": "B94",
            "sale_date": "E96",
            "quote_eur": "E97",
            "exchange_rate_dkk": "E98",
            "payout_total_dkk": "B99",
        }

    sent_date = _datetime_from_excel(sheet[refs["sent_date"]].value)
    purchased_from_date = _datetime_from_excel(sheet[refs["purchased_from_date"]].value)
    sale_datetime = _datetime_from_excel(sheet[refs["sale_date"]].value)
    update_payload = core.AfgMeltLotUpdateRequest(
        sent_date=sent_date.date() if sent_date else None,
        purchased_from_date=purchased_from_date.date() if purchased_from_date else None,
        after_pure_gold_grams=core._decimal_from_excel(sheet[refs["after_pure_gold_grams"]].value)
        if core._clean_text(sheet[refs["after_pure_gold_grams"]].value) is not None
        else None,
        insurance_dkk=core._decimal_from_excel(sheet[refs["insurance_dkk"]].value)
        if core._clean_text(sheet[refs["insurance_dkk"]].value) is not None
        else None,
        shipping_dkk=core._decimal_from_excel(sheet[refs["shipping_dkk"]].value)
        if core._clean_text(sheet[refs["shipping_dkk"]].value) is not None
        else None,
        refining_dkk=core._decimal_from_excel(sheet[refs["refining_dkk"]].value)
        if core._clean_text(sheet[refs["refining_dkk"]].value) is not None
        else None,
        sale_date=sale_datetime.date() if sale_datetime else None,
        quote_eur=core._decimal_from_excel(sheet[refs["quote_eur"]].value)
        if core._clean_text(sheet[refs["quote_eur"]].value) is not None
        else None,
        exchange_rate_dkk=core._decimal_from_excel(sheet[refs["exchange_rate_dkk"]].value)
        if core._clean_text(sheet[refs["exchange_rate_dkk"]].value) is not None
        else None,
        payout_total_dkk=core._decimal_from_excel(sheet[refs["payout_total_dkk"]].value)
        if core._clean_text(sheet[refs["payout_total_dkk"]].value) is not None
        else None,
    )
    mode = str(extra.get("mode") or "update")
    if mode == "create":
        has_input = any(
            value is not None
            for value in (
                update_payload.sent_date,
                update_payload.purchased_from_date,
                update_payload.after_pure_gold_grams,
                update_payload.insurance_dkk,
                update_payload.shipping_dkk,
                update_payload.refining_dkk,
                update_payload.sale_date,
                update_payload.quote_eur,
                update_payload.exchange_rate_dkk,
                update_payload.payout_total_dkk,
            )
        )
        if not has_input:
            return None, None
        return (
            core.LogWorkbookLotCreate(
                metal_bucket=metal_bucket,
                create_payload=core.AfgMeltLotCreateRequest(
                    metal_bucket=metal_bucket,
                    sent_date=update_payload.sent_date,
                    purchased_from_date=update_payload.purchased_from_date,
                ),
                update_payload=update_payload,
            ),
            None,
        )

    normalized_current = {
        "sent_date": str(extra.get("sent_date") or ""),
        "purchased_from_date": str(extra.get("purchased_from_date") or ""),
        "after_pure_gold_grams": str(extra.get("after_pure_gold_grams") or ""),
        "insurance_dkk": str(extra.get("insurance_dkk") or ""),
        "shipping_dkk": str(extra.get("shipping_dkk") or ""),
        "refining_dkk": str(extra.get("refining_dkk") or ""),
        "sale_date": str(extra.get("sale_date") or ""),
        "quote_eur": str(extra.get("quote_eur") or ""),
        "exchange_rate_dkk": str(extra.get("exchange_rate_dkk") or ""),
        "payout_total_dkk": str(extra.get("payout_total_dkk") or ""),
    }
    normalized_next = {
        "sent_date": update_payload.sent_date.isoformat() if update_payload.sent_date else "",
        "purchased_from_date": update_payload.purchased_from_date.isoformat() if update_payload.purchased_from_date else "",
        "after_pure_gold_grams": core._fmt_optional_decimal(update_payload.after_pure_gold_grams),
        "insurance_dkk": core._fmt_optional_decimal(update_payload.insurance_dkk),
        "shipping_dkk": core._fmt_optional_decimal(update_payload.shipping_dkk),
        "refining_dkk": core._fmt_optional_decimal(update_payload.refining_dkk),
        "sale_date": update_payload.sale_date.isoformat() if update_payload.sale_date else "",
        "quote_eur": core._fmt_optional_decimal(update_payload.quote_eur),
        "exchange_rate_dkk": core._fmt_optional_decimal(update_payload.exchange_rate_dkk),
        "payout_total_dkk": core._fmt_optional_decimal(update_payload.payout_total_dkk),
    }
    if normalized_current == normalized_next:
        return None, None
    return None, core.LogWorkbookLotUpdate(lot_id=UUID(extra.get("lot_id") or ""), payload=update_payload)


def _log_raw_reference_route_code(
    *,
    document,
    metal_bucket: str,
    inventory_marked: bool,
    separate_storage_refs: set[int],
) -> str:
    if document.sequence_no in separate_storage_refs:
        if not inventory_marked:
            raise ValueError(
                f"Log raw import: ayrı depolama işaretli belge solda inventory olarak işaretlenmemiş: {document.sequence_no}"
            )
        return "D"
    if inventory_marked:
        if metal_bucket == "silver":
            return "S"
        all_white_gold = bool(document.lines) and all(line.metal_type == "white_gold" for line in document.lines)
        return "H" if all_white_gold else "S"
    return "M"


def _raw_log_route_request_for_document(
    *,
    document,
    route_code: str,
    metal_bucket: str,
):
    core = _core()
    if route_code == "D":
        return core.AfgRouteRequest(
            line_ids=[line.id for line in document.lines],
            destination="inventory",
            classification="separate_storage",
            note=None,
        )
    if route_code == "H":
        return core.AfgRouteRequest(
            line_ids=[line.id for line in document.lines],
            destination="inventory",
            classification="white_gold",
            note=None,
        )
    if route_code == "S":
        return core.AfgRouteRequest(
            line_ids=[line.id for line in document.lines],
            destination="inventory",
            classification=("standard" if metal_bucket == "silver" else "jewelry_cleaning"),
            note=None,
        )
    default_classification = "standard"
    if metal_bucket == "gold" and document.lines and all(line.metal_type == "white_gold" for line in document.lines):
        default_classification = "white_gold"
    return core.AfgRouteRequest(
        line_ids=[line.id for line in document.lines],
        destination="melt",
        classification=default_classification,
        note=None,
    )


def _parse_raw_log_reference_route_updates(
    sheet,
    *,
    current_workspace,
) -> list:
    core = _core()
    route_updates: list = []
    explicit_separate_storage_refs: set[int] = set()
    for row_idx in range(6, 18):
        seq_value = core._clean_text(sheet[f"R{row_idx}"].value)
        if seq_value is None:
            continue
        try:
            sequence_no = int(core.to_decimal(seq_value))
        except Exception as exc:  # pragma: no cover - defensive branch
            raise ValueError(f"Log raw import: geçersiz ayrı depolama AFG ref değeri: {seq_value}") from exc
        if sequence_no in explicit_separate_storage_refs:
            raise ValueError(f"Log raw import: ayrı depolama AFG ref tekrarı: {sequence_no}")
        explicit_separate_storage_refs.add(sequence_no)

    bucket_specs = (
        ("gold", current_workspace.gold, core.LOG_GOLD_LEDGER_START, core.LOG_GOLD_LEDGER_END),
        ("silver", current_workspace.silver, core.LOG_SILVER_LEDGER_START, core.LOG_SILVER_LEDGER_END),
    )
    for metal_bucket, bucket, start_row, end_row in bucket_specs:
        documents_by_sequence = {document.sequence_no: document for document in bucket.documents}
        seen_sequences: set[int] = set()
        for row_idx in range(start_row, end_row + 1):
            raw_sequence = sheet[f"A{row_idx}"].value
            if raw_sequence in (None, ""):
                continue
            try:
                sequence_no = int(core.to_decimal(raw_sequence))
            except Exception:
                continue
            if sequence_no in seen_sequences:
                raise ValueError(f"Log raw import: aynı belge birden fazla kez listelenmiş: {sequence_no}")
            seen_sequences.add(sequence_no)
            document = documents_by_sequence.get(sequence_no)
            if document is None:
                raise ValueError(f"Log raw import: sistemde bulunmayan belge numarası: {sequence_no}")
            marker = core._clean_text(sheet[f"G{row_idx}"].value)
            if marker is not None and marker.lower() != "x":
                raise ValueError(f"Log raw import: yalnız boş veya x rota işareti bekleniyor, belge {sequence_no}")
            route_code = _log_raw_reference_route_code(
                document=document,
                metal_bucket=metal_bucket,
                inventory_marked=(marker is not None),
                separate_storage_refs=explicit_separate_storage_refs,
            )
            route_updates.append(
                core.LogWorkbookRouteEdit(
                    line_id=document.lines[0].id,
                    payload=_raw_log_route_request_for_document(
                        document=document,
                        route_code=route_code,
                        metal_bucket=metal_bucket,
                    ),
                )
            )

    dangling_refs = explicit_separate_storage_refs.difference(
        {document.sequence_no for document in current_workspace.gold.documents}
    )
    if dangling_refs:
        raise ValueError(
            f"Log raw import: ayrı depolama alanında sistemde bulunmayan belge referansı var: {sorted(dangling_refs)}"
        )
    return route_updates


def _parse_raw_log_lot_inputs(
    sheet,
    *,
    current_workspace,
) -> tuple[list, list]:
    lot_creates: list = []
    lot_updates: list = []
    for metal_bucket, bucket in (("gold", current_workspace.gold), ("silver", current_workspace.silver)):
        slot = _log_active_lot_slot(current_workspace, metal_bucket)
        create_row, update_row = _parse_log_lot_section(sheet, metal_bucket=metal_bucket, row_idx=0, extra=slot)
        if slot["mode"] == "create" and create_row is not None:
            should_create = any(
                (
                    create_row.update_payload.sent_date is not None,
                    create_row.update_payload.purchased_from_date is not None,
                    create_row.update_payload.sale_date is not None,
                    (create_row.update_payload.after_pure_gold_grams or Decimal("0")) > 0,
                    (create_row.update_payload.payout_total_dkk or Decimal("0")) > 0,
                )
            )
            if not should_create:
                continue
        if create_row is not None:
            lot_creates.append(create_row)
        if update_row is not None:
            lot_updates.append(update_row)
    return lot_creates, lot_updates


def _parse_log_workbook_inputs_raw(
    workbook,
    *,
    current_workspace,
):
    core = _core()
    if core.LOG_SHEET not in workbook.sheetnames:
        raise ValueError("Log Ark1 sheet bulunamadı")
    sheet = workbook[core.LOG_SHEET]
    route_updates = _parse_raw_log_reference_route_updates(sheet, current_workspace=current_workspace)
    lot_creates, lot_updates = _parse_raw_log_lot_inputs(sheet, current_workspace=current_workspace)
    return core.LogWorkbookArtifactInputs(
        route_updates=route_updates,
        lot_creates=lot_creates,
        lot_updates=lot_updates,
        base_version=None,
    )


def _build_log_workbook_bytes(workspace, *, year: int, sync_context) -> bytes:
    core = _core()
    workbook = core._open_template(core.LOG_TEMPLATE_NAME)
    core._touch_calc_flags(workbook)
    sheet = workbook[core.LOG_SHEET]
    sheet["C5"] = f"År {year}"
    sheet["G6"] = "Rota"
    sheet["G57"] = "Rota"

    for row_idx in range(core.LOG_GOLD_LEDGER_START, core.LOG_GOLD_LEDGER_END + 1):
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            core._set_sheet_cell(sheet, f"{col}{row_idx}", None)
    for row_idx in range(core.LOG_SILVER_LEDGER_START, core.LOG_SILVER_LEDGER_END + 1):
        for col in ("A", "B", "C", "D", "E", "F", "G"):
            core._set_sheet_cell(sheet, f"{col}{row_idx}", None)
    for row_idx in range(7, 19):
        for col in ("H", "I", "J", "K", "N", "O", "P", "R", "S", "T", "U", "V"):
            core._set_sheet_cell(sheet, f"{col}{row_idx}", None)

    route_rows = _log_route_line_rows(workspace)
    gold_rows = [row for row in route_rows if row["metal_bucket"] == "gold"]
    silver_rows = [row for row in route_rows if row["metal_bucket"] == "silver"]
    editable_refs: list[str] = []
    route_editable_refs: list[str] = []
    mappings: list = []

    for offset, row in enumerate(gold_rows[: core.LOG_GOLD_LEDGER_END - core.LOG_GOLD_LEDGER_START + 1], start=core.LOG_GOLD_LEDGER_START):
        core._set_sheet_cell(sheet, f"A{offset}", row["document_sequence_no"])
        core._set_sheet_cell(sheet, f"B{offset}", core._excel_datetime(row["issued_at"]))
        core._set_sheet_cell(sheet, f"C{offset}", row["customer_name"])
        core._set_sheet_cell(sheet, f"D{offset}", row["weight_grams"])
        core._set_sheet_cell(sheet, f"E{offset}", row["line_total_dkk"])
        core._set_sheet_cell(sheet, f"F{offset}", row["pure_gold_grams"])
        core._set_sheet_cell(sheet, f"G{offset}", row["route_code"])
        route_editable_refs.append(f"G{offset}")
        mappings.append(
            core.SyncSheetRowMapping(
                mapping_type="log_route",
                sheet_name=core.LOG_SHEET,
                row_index=offset,
                entity_id=row["line_id"],
                entity_key=row["document_number"],
                extra={
                    "route_code": row["route_code"],
                    "classification": row["classification"],
                    "note": row["note"],
                    "metal_type": row["metal_type"],
                },
            )
        )

    for offset, row in enumerate(silver_rows[: core.LOG_SILVER_LEDGER_END - core.LOG_SILVER_LEDGER_START + 1], start=core.LOG_SILVER_LEDGER_START):
        core._set_sheet_cell(sheet, f"A{offset}", row["document_sequence_no"])
        core._set_sheet_cell(sheet, f"B{offset}", core._excel_datetime(row["issued_at"]))
        core._set_sheet_cell(sheet, f"C{offset}", row["customer_name"])
        core._set_sheet_cell(sheet, f"D{offset}", row["weight_grams"])
        core._set_sheet_cell(sheet, f"E{offset}", row["line_total_dkk"])
        core._set_sheet_cell(sheet, f"F{offset}", row["pure_gold_grams"])
        core._set_sheet_cell(sheet, f"G{offset}", row["route_code"])
        route_editable_refs.append(f"G{offset}")
        mappings.append(
            core.SyncSheetRowMapping(
                mapping_type="log_route",
                sheet_name=core.LOG_SHEET,
                row_index=offset,
                entity_id=row["line_id"],
                entity_key=row["document_number"],
                extra={
                    "route_code": row["route_code"],
                    "classification": row["classification"],
                    "note": row["note"],
                    "metal_type": row["metal_type"],
                },
            )
        )

    group_specs = {
        "jewelry_cleaning": ("H", "I", "J", "K"),
        "white_gold": ("N", "O", "P", None),
        "separate_storage": ("R", "S", "T", "U"),
    }
    grouped_rows = {
        "jewelry_cleaning": [],
        "white_gold": [],
        "separate_storage": [],
    }
    for row in gold_rows:
        key = _log_bucket_group_key(row)
        if key is not None:
            grouped_rows[key].append(row)

    for key, rows in grouped_rows.items():
        col_a, col_b, col_c, col_d = group_specs[key]
        for row_index, row in enumerate(rows[:11], start=7):
            if key == "jewelry_cleaning":
                core._set_sheet_cell(sheet, f"{col_a}{row_index}", row["reference_number"])
                core._set_sheet_cell(sheet, f"{col_b}{row_index}", row["weight_grams"])
                core._set_sheet_cell(sheet, f"{col_c}{row_index}", row["line_total_dkk"])
                core._set_sheet_cell(sheet, f"{col_d}{row_index}", row["pure_gold_grams"])
            elif key == "white_gold":
                core._set_sheet_cell(sheet, f"{col_a}{row_index}", row["weight_grams"])
                core._set_sheet_cell(sheet, f"{col_b}{row_index}", row["line_total_dkk"])
                core._set_sheet_cell(sheet, f"{col_c}{row_index}", row["pure_gold_grams"])
            else:
                core._set_sheet_cell(sheet, f"{col_a}{row_index}", row["document_sequence_no"])
                core._set_sheet_cell(sheet, f"{col_b}{row_index}", row["weight_grams"])
                core._set_sheet_cell(sheet, f"{col_c}{row_index}", row["line_total_dkk"])
                core._set_sheet_cell(sheet, f"{col_d}{row_index}", row["pure_gold_grams"])
        if key == "jewelry_cleaning":
            core._set_sheet_cell(sheet, "I18", "=SUM(I7:I17)")
            core._set_sheet_cell(sheet, "J18", "=SUM(J7:J17)")
            core._set_sheet_cell(sheet, "K18", "=SUM(K7:K17)")
        elif key == "white_gold":
            core._set_sheet_cell(sheet, "N18", "=SUM(N7:N17)")
            core._set_sheet_cell(sheet, "O18", "=SUM(O7:O17)")
            core._set_sheet_cell(sheet, "P18", "=SUM(P7:P17)")
        else:
            core._set_sheet_cell(sheet, "S18", "=SUM(S7:S17)")
            core._set_sheet_cell(sheet, "T18", "=SUM(T7:T17)")
            core._set_sheet_cell(sheet, "U18", "=SUM(U7:U17)")

    core._set_sheet_cell(sheet, "K21", "Gram")
    core._set_sheet_cell(sheet, "K22", "=I18+N18+S18")
    core._set_sheet_cell(sheet, "L22", "=J18+O18+T18")
    core._set_sheet_cell(sheet, "M22", "=K18+P18+U18")
    core._set_sheet_cell(sheet, "D34", "=-K22")
    core._set_sheet_cell(sheet, "E34", "=-L22")
    core._set_sheet_cell(sheet, "F34", "=-M22")

    gold_lot_row = _log_active_lot_slot(workspace, "gold")
    silver_lot_row = _log_active_lot_slot(workspace, "silver")
    _set_log_lot_section(sheet, metal_bucket="gold", row=gold_lot_row, editable_refs=editable_refs, mappings=mappings)
    _set_log_lot_section(sheet, metal_bucket="silver", row=silver_lot_row, editable_refs=editable_refs, mappings=mappings)

    core._write_sync_sheet(
        workbook,
        context=core.ArtifactSyncContext(
            kind=sync_context.kind,
            key=sync_context.key,
            artifact_key=sync_context.artifact_key,
            base_version=sync_context.base_version,
            contract_version=sync_context.contract_version,
            mappings=mappings,
        ),
    )
    editable_refs.extend(route_editable_refs)
    core._protect_sheet(sheet, editable_refs=editable_refs)
    _add_log_route_validation(sheet, route_editable_refs)
    workbook.active = workbook.sheetnames.index(core.LOG_SHEET)
    return core._save_workbook_bytes(workbook)


def _parse_log_workbook_inputs_v1(workbook, metadata):
    core = _core()
    if core.LOG_CONTROL_SHEET not in workbook.sheetnames:
        raise ValueError("Log control sheet bulunamadı")
    sheet = workbook[core.LOG_CONTROL_SHEET]
    route_updates = []
    lot_creates = []
    lot_updates = []
    for mapping in metadata.mappings:
        row_idx = mapping.row_index
        extra = mapping.extra or {}
        if mapping.mapping_type == "log_route":
            destination = (core._clean_text(sheet[f"H{row_idx}"].value) or "undecided").lower()
            classification = (core._clean_text(sheet[f"I{row_idx}"].value) or "standard").lower()
            note = core._clean_text(sheet[f"J{row_idx}"].value)
            if (
                destination == str(extra.get("destination") or "").lower()
                and classification == str(extra.get("classification") or "").lower()
                and (note or "") == str(extra.get("note") or "")
            ):
                continue
            route_updates.append(
                core.LogWorkbookRouteEdit(
                    line_id=UUID(mapping.entity_id),
                    payload=core.AfgRouteRequest(
                        line_ids=[UUID(mapping.entity_id)],
                        destination=destination,
                        classification=classification,
                        note=note,
                    ),
                )
            )
        elif mapping.mapping_type == "log_lot":
            sent_date = _datetime_from_excel(sheet[f"D{row_idx}"].value)
            purchased_from_date = _datetime_from_excel(sheet[f"E{row_idx}"].value)
            sale_datetime = _datetime_from_excel(sheet[f"M{row_idx}"].value)
            update_payload = core.AfgMeltLotUpdateRequest(
                sent_date=sent_date.date() if sent_date else None,
                purchased_from_date=purchased_from_date.date() if purchased_from_date else None,
                after_pure_gold_grams=core._decimal_from_excel(sheet[f"I{row_idx}"].value)
                if core._clean_text(sheet[f"I{row_idx}"].value) is not None
                else None,
                insurance_dkk=core._decimal_from_excel(sheet[f"J{row_idx}"].value)
                if core._clean_text(sheet[f"J{row_idx}"].value) is not None
                else None,
                shipping_dkk=core._decimal_from_excel(sheet[f"K{row_idx}"].value)
                if core._clean_text(sheet[f"K{row_idx}"].value) is not None
                else None,
                refining_dkk=core._decimal_from_excel(sheet[f"L{row_idx}"].value)
                if core._clean_text(sheet[f"L{row_idx}"].value) is not None
                else None,
                sale_date=sale_datetime.date() if sale_datetime else None,
                quote_eur=core._decimal_from_excel(sheet[f"N{row_idx}"].value)
                if core._clean_text(sheet[f"N{row_idx}"].value) is not None
                else None,
                exchange_rate_dkk=core._decimal_from_excel(sheet[f"O{row_idx}"].value)
                if core._clean_text(sheet[f"O{row_idx}"].value) is not None
                else None,
                payout_total_dkk=core._decimal_from_excel(sheet[f"P{row_idx}"].value)
                if core._clean_text(sheet[f"P{row_idx}"].value) is not None
                else None,
                notes=core._clean_text(sheet[f"Q{row_idx}"].value),
            )
            mode = str(extra.get("mode") or "update")
            if mode == "create":
                has_input = any(
                    value is not None
                    for value in (
                        update_payload.sent_date,
                        update_payload.purchased_from_date,
                        update_payload.after_pure_gold_grams,
                        update_payload.insurance_dkk,
                        update_payload.shipping_dkk,
                        update_payload.refining_dkk,
                        update_payload.sale_date,
                        update_payload.quote_eur,
                        update_payload.payout_total_dkk,
                        update_payload.notes,
                    )
                )
                if not has_input:
                    continue
                lot_creates.append(
                    core.LogWorkbookLotCreate(
                        metal_bucket=mapping.entity_key or "gold",
                        create_payload=core.AfgMeltLotCreateRequest(
                            metal_bucket=mapping.entity_key or "gold",
                            sent_date=update_payload.sent_date,
                            purchased_from_date=update_payload.purchased_from_date,
                            notes=update_payload.notes,
                        ),
                        update_payload=update_payload,
                    )
                )
                continue
            normalized_current = {
                "sent_date": str(extra.get("sent_date") or ""),
                "purchased_from_date": str(extra.get("purchased_from_date") or ""),
                "after_pure_gold_grams": str(extra.get("after_pure_gold_grams") or ""),
                "insurance_dkk": str(extra.get("insurance_dkk") or ""),
                "shipping_dkk": str(extra.get("shipping_dkk") or ""),
                "refining_dkk": str(extra.get("refining_dkk") or ""),
                "sale_date": str(extra.get("sale_date") or ""),
                "quote_eur": str(extra.get("quote_eur") or ""),
                "exchange_rate_dkk": str(extra.get("exchange_rate_dkk") or ""),
                "payout_total_dkk": str(extra.get("payout_total_dkk") or ""),
                "notes": str(extra.get("notes") or ""),
            }
            normalized_next = {
                "sent_date": update_payload.sent_date.isoformat() if update_payload.sent_date else "",
                "purchased_from_date": update_payload.purchased_from_date.isoformat() if update_payload.purchased_from_date else "",
                "after_pure_gold_grams": core._fmt_optional_decimal(update_payload.after_pure_gold_grams),
                "insurance_dkk": core._fmt_optional_decimal(update_payload.insurance_dkk),
                "shipping_dkk": core._fmt_optional_decimal(update_payload.shipping_dkk),
                "refining_dkk": core._fmt_optional_decimal(update_payload.refining_dkk),
                "sale_date": update_payload.sale_date.isoformat() if update_payload.sale_date else "",
                "quote_eur": core._fmt_optional_decimal(update_payload.quote_eur),
                "exchange_rate_dkk": core._fmt_optional_decimal(update_payload.exchange_rate_dkk),
                "payout_total_dkk": core._fmt_optional_decimal(update_payload.payout_total_dkk),
                "notes": update_payload.notes or "",
            }
            if normalized_current == normalized_next:
                continue
            lot_updates.append(core.LogWorkbookLotUpdate(lot_id=UUID(mapping.entity_id), payload=update_payload))

    return core.LogWorkbookArtifactInputs(
        route_updates=route_updates,
        lot_creates=lot_creates,
        lot_updates=lot_updates,
        base_version=metadata.base_version,
    )


def _parse_log_workbook_inputs_v2(workbook, metadata):
    core = _core()
    if core.LOG_SHEET not in workbook.sheetnames:
        raise ValueError("Log Ark1 sheet bulunamadı")
    sheet = workbook[core.LOG_SHEET]
    route_updates = []
    lot_creates = []
    lot_updates = []

    for mapping in metadata.mappings:
        row_idx = mapping.row_index
        extra = mapping.extra or {}
        if mapping.mapping_type == "log_route":
            route_code = _normalize_log_route_code(sheet[f"G{row_idx}"].value)
            current_code = str(extra.get("route_code") or "-").upper()
            if route_code == current_code:
                continue
            route_updates.append(
                core.LogWorkbookRouteEdit(
                    line_id=UUID(mapping.entity_id),
                    payload=_log_route_request_from_code(
                        route_code=route_code,
                        current_classification=core._clean_text(extra.get("classification")),
                        metal_type=core._clean_text(extra.get("metal_type")),
                        note=core._clean_text(extra.get("note")),
                        line_id=UUID(mapping.entity_id),
                    ),
                )
            )
        elif mapping.mapping_type == "log_lot":
            create_row, update_row = _parse_log_lot_section(
                sheet,
                metal_bucket=(mapping.entity_key or "gold"),
                row_idx=row_idx,
                extra={**extra, "lot_id": mapping.entity_id},
            )
            if create_row is not None:
                lot_creates.append(create_row)
            if update_row is not None:
                lot_updates.append(update_row)

    return core.LogWorkbookArtifactInputs(
        route_updates=route_updates,
        lot_creates=lot_creates,
        lot_updates=lot_updates,
        base_version=metadata.base_version,
    )
