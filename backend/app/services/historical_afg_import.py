from __future__ import annotations

import hashlib
import json
import re
import secrets
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import UUID

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.enums import (
    MetalTypeEnum,
    PosDocumentTypeEnum,
    PosRateSourceEnum,
    PosSessionStatusEnum,
    PosTradeSideEnum,
    ProductTypeEnum,
    RoleEnum,
)
from app.models.pos_document import PosDocument
from app.models.pos_session import PosSession
from app.models.pos_session_line import PosSessionLine
from app.models.transaction import Transaction
from app.models.transaction_line import TransactionLine
from app.models.user import User
from app.schemas.customer import CustomerCreate
from app.schemas.historical_afg_import import (
    HistoricalAfgImportApplyItemOut,
    HistoricalAfgImportApplyOut,
    HistoricalAfgImportPreviewItemOut,
    HistoricalAfgImportPreviewOut,
)
from app.services import pos_service as pos_core
from app.services.customer_service import create_customer, customer_identity_match
from app.services.document_artifact_service import parse_afg_workspace_inputs_from_workbook
from app.utils.helpers import quantize_2, to_decimal, utc_now


@dataclass(frozen=True)
class HistoricalAfgUpload:
    filename: str
    content: bytes
    source_hash: str

    @classmethod
    def from_content(cls, *, filename: str, content: bytes) -> "HistoricalAfgUpload":
        return cls(filename=filename, content=content, source_hash=hashlib.sha256(content).hexdigest())


@dataclass(frozen=True)
class HistoricalAfgLine:
    line_no: int
    row_key: str
    metal_type: MetalTypeEnum
    purity_karat: str | None
    purity_percentage: Decimal
    weight_grams: Decimal
    rate_dkk: Decimal
    margin_percent: Decimal
    line_total_dkk: Decimal
    type_label: str


@dataclass
class HistoricalAfgParsed:
    upload: HistoricalAfgUpload
    legacy_document_number: str | None
    issued_at: datetime | None
    customer: object | None
    sections: object | None
    lines: list[HistoricalAfgLine]
    total_weight_grams: Decimal
    total_amount_dkk: Decimal
    errors: list[str]
    # Legacy şablon: belgedeki tamamlanmış tutarlar birebir korunur, yeniden hesaplanmaz.
    template_profile: str = "current"
    net_amount_dkk: Decimal | None = None
    vat_amount_dkk: Decimal | None = None
    warnings: list[str] = field(default_factory=list)
    birth_date_text: str | None = None
    is_company: bool = False


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip() or None


def _normalise_name(value: str | None) -> str:
    return " ".join((value or "").casefold().split())


def _normalise_phone(value: str | None) -> str:
    return re.sub(r"[^0-9+]", "", value or "")


def _as_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=value.tzinfo or timezone.utc)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=timezone.utc)
    if isinstance(value, (int, float)):
        return _as_datetime(from_excel(value))
    if isinstance(value, str):
        for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), pattern).replace(tzinfo=timezone.utc)
            except ValueError:
                pass
    return None


def _metadata_from_workbook(content: bytes) -> tuple[str | None, datetime | None]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        if "Afregningsbilag" not in workbook.sheetnames:
            return None, None
        sheet = workbook["Afregningsbilag"]
        legacy_number = _clean_text(sheet["H6"].value)
        issued_at = _as_datetime(sheet["H7"].value) or _as_datetime(sheet["D2"].value)
        if not legacy_number and "Variable værdier" in workbook.sheetnames:
            legacy_number = _clean_text(workbook["Variable værdier"]["C14"].value)
        return legacy_number, issued_at
    finally:
        workbook.close()


def _make_lines(sections: object) -> tuple[list[HistoricalAfgLine], list[str]]:
    market_rates = getattr(sections, "market_rates", None)
    if market_rates is None:
        return [], ["Piyasa oranları bulunamadı; tutar güvenli biçimde hesaplanamaz."]

    lines: list[HistoricalAfgLine] = []
    errors: list[str] = []
    line_no = 1
    for row in getattr(sections, "gold_rows", []):
        gram = quantize_2(to_decimal(row.gram))
        if gram <= 0:
            continue
        definition = next(
            (
                item
                for item in pos_core.GOLD_WORKSPACE_ROWS
                if to_decimal(item["karat"]) == quantize_2(to_decimal(row.karat))
            ),
            None,
        )
        if definition is None:
            errors.append(f"Bilinmeyen altın saflığı: {row.karat}.")
            continue
        margin = quantize_2(to_decimal(row.avance_percent))
        rate = pos_core._workspace_market_rate_dkk(market_rates, str(definition["row_key"]))
        if rate <= 0:
            errors.append(f"{definition['label']} için oran bulunamadı.")
            continue
        unit_price = pos_core._workspace_row_unit_price_from_matrix(rate_dkk=rate, avance_percent=margin)
        lines.append(
            HistoricalAfgLine(
                line_no=line_no,
                row_key=str(definition["row_key"]),
                metal_type=MetalTypeEnum.YELLOW_GOLD,
                purity_karat=str(definition["label"]).upper(),
                purity_percentage=quantize_2(to_decimal(definition["purity_percentage"])),
                weight_grams=gram,
                rate_dkk=rate,
                margin_percent=margin,
                line_total_dkk=pos_core._workspace_row_line_total(unit_price_dkk=unit_price, gram=gram),
                type_label=f"Guld {str(definition['label']).upper()}",
            )
        )
        line_no += 1
    for row in getattr(sections, "silver_rows", []):
        gram = quantize_2(to_decimal(row.gram))
        if gram <= 0:
            continue
        definition = next(
            (item for item in pos_core.SILVER_WORKSPACE_ROWS if str(item["type_code"]) == row.type_code),
            None,
        )
        if definition is None:
            errors.append(f"Bilinmeyen gümüş türü: {row.type_code}.")
            continue
        margin = quantize_2(to_decimal(row.avance_percent))
        rate = pos_core._workspace_market_rate_dkk(market_rates, str(definition["row_key"]))
        if rate <= 0:
            errors.append(f"{definition['label']} için oran bulunamadı.")
            continue
        unit_price = pos_core._workspace_row_unit_price_from_matrix(rate_dkk=rate, avance_percent=margin)
        lines.append(
            HistoricalAfgLine(
                line_no=line_no,
                row_key=str(definition["row_key"]),
                metal_type=MetalTypeEnum.SILVER,
                purity_karat=None,
                purity_percentage=quantize_2(to_decimal(definition["purity_percentage"])),
                weight_grams=gram,
                rate_dkk=rate,
                margin_percent=margin,
                line_total_dkk=pos_core._workspace_row_line_total(unit_price_dkk=unit_price, gram=gram),
                type_label=str(definition["label"]),
            )
        )
        line_no += 1
    if not lines and not errors:
        errors.append("Pozitif gram içeren altın veya gümüş satırı bulunamadı.")
    return lines, errors


_LEGACY_SCAN_MAX_ROW = 60
_LEGACY_CUSTOMER_LABELS = (
    ("navn", "name"),
    ("adresse", "address"),
    ("postnr", "postal_code"),
    ("post nr", "postal_code"),
    ("cpr", "cpr_number"),
    ("kørekort", "identity_doc_number"),
    ("pas", "identity_doc_number"),
    ("tlf", "phone"),
    ("telefon", "phone"),
    ("e-mail", "email"),
    ("email", "email"),
    ("mail", "email"),
)
_LEGACY_NET_LABELS = ("subtotal", "netto")
_LEGACY_VAT_LABELS = ("moms",)
_LEGACY_GROSS_LABELS = ("i alt", "total")
_LEGACY_NUMBER_LABELS = ("afregningsnr", "afregnings nr", "bilag nr", "bilagsnr")
_LEGACY_DATE_LABELS = ("dato",)
_LEGACY_SILVER_LABELS = (
    ("finsølv", "silver:2"),
    ("sterling", "silver:3"),
    ("tårnet", "silver:4"),
    ("plet", "silver:5"),
    ("sølv", "silver:5"),
)
# Etiket/değer çiftleri: müşteri bloğu (C,D)+(F,G); belge no/tarih ve
# toplamlar (G,H) çiftinde yaşar.
_LEGACY_LABEL_VALUE_PAIRS = (("C", "D"), ("F", "G"), ("G", "H"))
_LEGACY_CVR_RE = re.compile(r"^(cvr|se)\b", re.IGNORECASE)


def _norm_label(value: object) -> str:
    text = _clean_text(value) or ""
    # Not: '-' korunur ki 'e-mail' etiketi ayırt edilebilsin.
    return " ".join(text.casefold().replace(":", " ").replace(".", " ").split())


def _legacy_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace("kr", "").replace("DKK", "").strip()
        cleaned = cleaned.replace(".", "").replace(",", ".") if "," in cleaned else cleaned
        try:
            return quantize_2(Decimal(cleaned))
        except (InvalidOperation, ValueError):
            return None
    try:
        return quantize_2(to_decimal(value))
    except (InvalidOperation, ValueError):
        return None


def _legacy_grid(content: bytes) -> tuple[dict[str, object], list[str]]:
    """Tek geçişte (read_only) tüm ilgili hücreleri sözlüğe alır."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
        sheet = (
            workbook["Afregningsbilag"]
            if "Afregningsbilag" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        grid: dict[str, object] = {}
        for row in sheet.iter_rows(min_row=1, max_row=_LEGACY_SCAN_MAX_ROW, min_col=1, max_col=8):
            for cell in row:
                if cell.value is not None:
                    grid[f"{cell.coordinate}"] = cell.value
        return grid, sheet_names
    finally:
        workbook.close()


def looks_like_legacy_afg_template(content: bytes) -> bool:
    """Eski şablon: etiketler C/F sütununda kalır, müşteri değerleri D/G'de.

    CRM üretimi dosyalar gizli __SERO_SYNC sayfası taşır — etiket sezgisine
    bakılmaksızın kesin olarak "legacy değil" sayılır (afg-v2 düzeni de
    değerleri D/G'ye yazdığı için bu ayrım şarttır).
    """
    from app.services.document_artifact_service import SYNC_SHEET_NAME

    grid, sheet_names = _legacy_grid(content)
    if SYNC_SHEET_NAME in sheet_names:
        return False
    label_hits = 0
    value_hits = 0
    for row in range(1, _LEGACY_SCAN_MAX_ROW + 1):
        for label_col, value_col in (("C", "D"), ("F", "G")):
            label = _norm_label(grid.get(f"{label_col}{row}"))
            if not label:
                continue
            if any(label.startswith(prefix) for prefix, _ in _LEGACY_CUSTOMER_LABELS):
                label_hits += 1
                if _clean_text(grid.get(f"{value_col}{row}")) is not None:
                    value_hits += 1
    return label_hits >= 2 and value_hits >= 1


def _legacy_gold_definition_from_row(grid: dict[str, object], row: int) -> dict | None:
    """Metal satırı imzası: C='Guld', karat D'de, lodighed E'de."""
    label = _norm_label(grid.get(f"C{row}"))
    if "guld" not in label or "barre" in label:
        return None
    karat_raw = _clean_text(grid.get(f"D{row}"))
    lodighed_raw = _clean_text(grid.get(f"E{row}"))
    if karat_raw is not None:
        try:
            karat = to_decimal(str(karat_raw).replace(",", "."))
        except Exception:  # noqa: BLE001
            karat = None
        if karat is not None:
            match = next(
                (item for item in pos_core.GOLD_WORKSPACE_ROWS if to_decimal(item["karat"]) == quantize_2(karat)),
                None,
            )
            if match is not None:
                return match
    if lodighed_raw is not None:
        # 22K eski dosyalarda 916 VEYA 917 yazabilir.
        normalized = "916" if lodighed_raw == "917" else lodighed_raw
        return next(
            (item for item in pos_core.GOLD_WORKSPACE_ROWS if str(item["lodighed"]) == normalized),
            None,
        )
    return None


def _legacy_silver_definition_from_row(grid: dict[str, object], row: int) -> dict | None:
    label = _norm_label(grid.get(f"C{row}"))
    if not label or "guld" in label or "barre" in label:
        return None
    for needle, row_key in _LEGACY_SILVER_LABELS:
        if needle in label:
            return next(
                (item for item in pos_core.SILVER_WORKSPACE_ROWS if str(item["row_key"]) == row_key),
                None,
            )
    return None


def _split_legacy_postal(value: str | None) -> tuple[str | None, str | None]:
    """"2200 København N" → ("2200", "København N")."""
    text = (value or "").strip()
    if not text:
        return None, None
    match = re.match(r"^(\d{3,4})\s+(.+)$", text)
    if match:
        return match.group(1), match.group(2).strip() or None
    if text.isdigit():
        return text, None
    return None, text


def _parse_legacy_upload(upload: HistoricalAfgUpload) -> HistoricalAfgParsed:
    try:
        grid, _sheet_names = _legacy_grid(upload.content)
    except Exception as exc:  # noqa: BLE001
        return HistoricalAfgParsed(
            upload=upload,
            legacy_document_number=None,
            issued_at=None,
            customer=None,
            sections=None,
            lines=[],
            total_weight_grams=Decimal("0.00"),
            total_amount_dkk=Decimal("0.00"),
            errors=[f"AFG dosyası okunamadı: {exc}"],
            template_profile="legacy",
        )

    customer_fields: dict[str, str | None] = {}
    lines: list[HistoricalAfgLine] = []
    errors: list[str] = []
    warnings: list[str] = []
    seen_row_keys: set[str] = set()
    legacy_number: str | None = None
    issued_at: datetime | None = None
    net_amount: Decimal | None = None
    vat_amount: Decimal | None = None
    gross_amount: Decimal | None = None
    birth_date_text: str | None = None
    line_no = 1

    for row in range(1, _LEGACY_SCAN_MAX_ROW + 1):
        # 1) Etiket/değer çiftleri: müşteri, belge no, tarih, toplamlar.
        for label_col, value_col in _LEGACY_LABEL_VALUE_PAIRS:
            label = _norm_label(grid.get(f"{label_col}{row}"))
            if not label:
                continue
            value = grid.get(f"{value_col}{row}")

            for prefix, field in _LEGACY_CUSTOMER_LABELS:
                if label.startswith(prefix) and field not in customer_fields:
                    text = _clean_text(value)
                    if text is not None:
                        customer_fields[field] = text
                    break

            if legacy_number is None and any(label.startswith(p) for p in _LEGACY_NUMBER_LABELS):
                legacy_number = _clean_text(value)
            if issued_at is None and any(label.startswith(p) for p in _LEGACY_DATE_LABELS):
                issued_at = _as_datetime(value)

            amount = _legacy_decimal(value)
            if amount is not None:
                if net_amount is None and any(label.startswith(p) for p in _LEGACY_NET_LABELS):
                    net_amount = amount
                elif vat_amount is None and any(label.startswith(p) for p in _LEGACY_VAT_LABELS):
                    vat_amount = amount
                elif gross_amount is None and any(label.startswith(p) for p in _LEGACY_GROSS_LABELS):
                    # C38 'I alt' satırı gram toplamıdır; genel toplam yalnız
                    # (G,H) çiftinden alınır.
                    if label_col == "G":
                        gross_amount = amount

        # 2) Metal satırları: C=tür etiketi, D=karat, E=lodighed,
        #    F=gram, G=birim fiyat, H=satır toplamı.
        gold_def = _legacy_gold_definition_from_row(grid, row)
        silver_def = None if gold_def is not None else _legacy_silver_definition_from_row(grid, row)
        definition = gold_def or silver_def
        if definition is None:
            continue
        gram = _legacy_decimal(grid.get(f"F{row}"))
        if gram is None or gram <= 0:
            continue
        line_total = _legacy_decimal(grid.get(f"H{row}"))
        unit_price = _legacy_decimal(grid.get(f"G{row}"))
        if (line_total is None or line_total <= 0) and unit_price is not None:
            line_total = quantize_2(gram * unit_price)
        if line_total is None or line_total <= 0:
            errors.append(f"{definition['label']} satırının tutarı (H{row}) okunamadı.")
            continue
        if unit_price is not None and unit_price > 0:
            expected_total = quantize_2(gram * unit_price)
            if abs(expected_total - line_total) > Decimal("0.05"):
                warnings.append(
                    f"{definition['label']} satırında gram × birim fiyat ({expected_total}) "
                    f"satır toplamından ({line_total}) sapıyor."
                )
        row_key = str(definition["row_key"])
        dedupe_key = f"{row_key}:{row}"
        if dedupe_key in seen_row_keys:
            continue
        seen_row_keys.add(dedupe_key)
        rate = unit_price if unit_price is not None and unit_price > 0 else quantize_2(line_total / gram)
        lines.append(
            HistoricalAfgLine(
                line_no=line_no,
                row_key=row_key,
                metal_type=MetalTypeEnum.YELLOW_GOLD if gold_def is not None else MetalTypeEnum.SILVER,
                purity_karat=str(definition["label"]).upper() if gold_def is not None else None,
                purity_percentage=quantize_2(to_decimal(definition["purity_percentage"])),
                weight_grams=quantize_2(gram),
                rate_dkk=rate,
                margin_percent=Decimal("0.00"),
                line_total_dkk=line_total,
                type_label=(
                    f"Guld {str(definition['label']).upper()}"
                    if gold_def is not None
                    else str(definition["label"])
                ),
            )
        )
        line_no += 1

    # CPR alanı: tam CPR (10 hane) doğrulanabilir; yalnız 6 hane doğum
    # tarihidir (CPR olarak KAYDEDİLMEZ); 'CVR ...' şirket kimliğidir.
    raw_cpr = customer_fields.get("cpr_number")
    is_company = False
    if raw_cpr:
        if _LEGACY_CVR_RE.match(raw_cpr):
            is_company = True
            customer_fields.pop("cpr_number", None)
            # CVR şirket kimliği olarak kimlik alanında izlenir.
            if not customer_fields.get("identity_doc_number"):
                customer_fields["identity_doc_number"] = raw_cpr
        else:
            digits = re.sub(r"\D", "", raw_cpr)
            if len(digits) == 10:
                customer_fields["cpr_number"] = digits
                birth_date_text = digits[:6]
            elif len(digits) >= 6:
                # Belgede yalnız doğum tarihi bölümü var; tam CPR uydurulmaz.
                customer_fields.pop("cpr_number", None)
                birth_date_text = digits[:6]
            else:
                customer_fields.pop("cpr_number", None)

    postal_code, city = _split_legacy_postal(customer_fields.get("postal_code"))
    if postal_code is not None or city is not None:
        customer_fields["postal_code"] = postal_code
        if city and not customer_fields.get("city"):
            customer_fields["city"] = city
    elif customer_fields.get("postal_code") is None:
        customer_fields.pop("postal_code", None)

    if not lines:
        errors.append("Tür/ayar imzasıyla eşleşen pozitif gramlı metal satırı bulunamadı.")
    if not legacy_number:
        errors.append("Eski AFG numarası bulunamadı.")
    elif len(legacy_number) > 80:
        errors.append("Eski AFG numarası 80 karakteri aşıyor.")
    if issued_at is None:
        errors.append("Belge tarihi bulunamadı.")

    line_sum = quantize_2(sum((line.line_total_dkk for line in lines), Decimal("0.00")))
    if gross_amount is None:
        gross_amount = line_sum if vat_amount is None else quantize_2(line_sum + vat_amount)
    if vat_amount is not None and net_amount is None:
        net_amount = quantize_2(gross_amount - vat_amount)
    if net_amount is not None and vat_amount is not None:
        expected = quantize_2(net_amount + vat_amount)
        if expected != gross_amount:
            errors.append(
                f"Belgedeki net+KDV ({expected}) genel toplamla ({gross_amount}) uyuşmuyor."
            )

    customer = (
        CustomerCreate.model_construct(
            **{key: value for key, value in customer_fields.items() if value is not None}
        )
        if customer_fields.get("name")
        else None
    )
    return HistoricalAfgParsed(
        upload=upload,
        legacy_document_number=legacy_number,
        issued_at=issued_at,
        customer=customer,
        sections=None,
        lines=lines,
        total_weight_grams=quantize_2(sum((line.weight_grams for line in lines), Decimal("0.00"))),
        total_amount_dkk=gross_amount,
        errors=errors,
        template_profile="legacy",
        net_amount_dkk=net_amount,
        vat_amount_dkk=vat_amount,
        warnings=warnings,
        birth_date_text=birth_date_text,
        is_company=is_company,
    )


def _parse_upload(upload: HistoricalAfgUpload) -> HistoricalAfgParsed:
    try:
        if looks_like_legacy_afg_template(upload.content):
            return _parse_legacy_upload(upload)
    except (OSError, KeyError, ValueError, InvalidOperation):
        pass
    try:
        legacy_number, issued_at = _metadata_from_workbook(upload.content)
        inputs = parse_afg_workspace_inputs_from_workbook(upload.content)
    except (OSError, KeyError, ValueError, InvalidOperation) as exc:
        return HistoricalAfgParsed(
            upload=upload,
            legacy_document_number=None,
            issued_at=None,
            customer=None,
            sections=None,
            lines=[],
            total_weight_grams=Decimal("0.00"),
            total_amount_dkk=Decimal("0.00"),
            errors=[f"AFG dosyası okunamadı: {exc}"],
        )

    numbering = getattr(inputs.sections, "numbering", None)
    if not legacy_number and numbering is not None:
        legacy_number = _clean_text(numbering.afregnings_number_next)
    lines, errors = _make_lines(inputs.sections)
    if not legacy_number:
        errors.append("Eski AFG numarası bulunamadı.")
    elif len(legacy_number) > 80:
        errors.append("Eski AFG numarası 80 karakteri aşıyor.")
    if issued_at is None:
        errors.append("Belge tarihi bulunamadı.")
    return HistoricalAfgParsed(
        upload=upload,
        legacy_document_number=legacy_number,
        issued_at=issued_at,
        customer=inputs.customer,
        sections=inputs.sections,
        lines=lines,
        total_weight_grams=quantize_2(sum((line.weight_grams for line in lines), Decimal("0.00"))),
        total_amount_dkk=quantize_2(sum((line.line_total_dkk for line in lines), Decimal("0.00"))),
        errors=errors,
    )


async def _resolve_customer(
    session: AsyncSession,
    customer: object | None,
    *,
    create: bool,
) -> tuple[User | None, str, list[str]]:
    if customer is None:
        return None, "blocked", ["Müşteri bilgisi okunamadı."]
    name = _clean_text(getattr(customer, "name", None))
    cpr = _clean_text(getattr(customer, "cpr_number", None))
    identity_doc = _clean_text(getattr(customer, "identity_doc_number", None))
    matched = await customer_identity_match(session, cpr_number=cpr, identity_doc_number=identity_doc)
    if matched.status == "conflict":
        return None, "blocked", ["CPR veya kimlik numarası birden fazla müşteriyle eşleşiyor."]
    if matched.status == "single":
        user = await session.get(User, UUID(matched.matches[0].id))
        if user is not None:
            return user, "matched_identity", []

    email = _clean_text(getattr(customer, "email", None))
    if email:
        email_matches = (
            await session.scalars(
                select(User).where(User.role == RoleEnum.CUSTOMER, func.lower(User.email) == email.casefold())
            )
        ).all()
        if email_matches:
            same_name = [item for item in email_matches if _normalise_name(item.name) == _normalise_name(name)]
            if len(same_name) == 1:
                return same_name[0], "matched_email", []
            return None, "blocked", ["E-posta başka bir müşteriyle eşleşiyor; otomatik birleştirme yapılmadı."]

    phone = _clean_text(getattr(customer, "phone", None))
    if phone:
        phone_matches = [
            item
            for item in (
                await session.scalars(
                    select(User).where(User.role == RoleEnum.CUSTOMER, User.phone.is_not(None))
                )
            ).all()
            if _normalise_phone(item.phone) == _normalise_phone(phone)
        ]
        if phone_matches:
            same_name = [item for item in phone_matches if _normalise_name(item.name) == _normalise_name(name)]
            if len(same_name) == 1:
                return same_name[0], "matched_phone", []
            return None, "blocked", ["Telefon başka bir müşteriyle eşleşiyor; otomatik birleştirme yapılmadı."]

    if not name:
        return None, "blocked", ["Müşteri adı bulunamadı."]
    if not any((phone, cpr, identity_doc)):
        return None, "blocked", ["Yeni müşteri için telefon, CPR veya kimlik numarasından biri gerekli."]
    try:
        payload = CustomerCreate(
            name=name,
            email=email,
            phone=phone,
            address=_clean_text(getattr(customer, "address", None)),
            postal_code=_clean_text(getattr(customer, "postal_code", None)),
            city=_clean_text(getattr(customer, "city", None)),
            cpr_number=cpr,
            identity_doc_number=identity_doc,
        )
    except ValidationError as exc:
        return None, "blocked", [f"Müşteri bilgisi geçersiz: {item['msg']}" for item in exc.errors()]
    if not create:
        # Preview/apply sapmasını öldür: apply'da create_customer'ın atacağı
        # doğrulamalar (posta kodu, telefon, CPR, kimlik) preview'da da koşar
        # ki dosya preview'da 'ready' görünüp apply'da patlamasın.
        from app.services.customer_service import _normalize_postal_code, _validate_customer_identity_inputs

        try:
            _normalize_postal_code(payload.postal_code)
            _validate_customer_identity_inputs(
                phone=payload.phone,
                cpr=payload.cpr_number,
                identity_doc_number=payload.identity_doc_number,
            )
        except HTTPException as exc:
            return None, "blocked", [str(exc.detail)]
        return None, "create_customer", []
    try:
        return await create_customer(session, payload), "created_customer", []
    except HTTPException as exc:
        return None, "blocked", [str(exc.detail)]


async def _duplicate_reason(session: AsyncSession, parsed: HistoricalAfgParsed) -> str | None:
    existing_hash = await session.scalar(
        select(PosDocument.sequence_no).where(PosDocument.historical_import_hash == parsed.upload.source_hash)
    )
    if existing_hash is not None:
        return "Bu dosya karması daha önce içe aktarıldı."
    if parsed.legacy_document_number:
        existing_number = await session.scalar(
            select(PosDocument.sequence_no).where(
                PosDocument.legacy_document_number == parsed.legacy_document_number
            )
        )
        if existing_number is not None:
            return "Bu eski AFG numarası zaten kayıtlı."
    return None


async def _preview_item(
    session: AsyncSession,
    parsed: HistoricalAfgParsed,
    *,
    duplicate_in_selection: bool,
) -> HistoricalAfgImportPreviewItemOut:
    errors = list(parsed.errors)
    if duplicate_in_selection:
        errors.append("Aynı dosya seçili listede birden fazla kez var.")
    existing = await _duplicate_reason(session, parsed)
    if existing:
        return HistoricalAfgImportPreviewItemOut(
            source_hash=parsed.upload.source_hash,
            file_name=parsed.upload.filename,
            status="already_imported",
            legacy_document_number=parsed.legacy_document_number,
            issued_at=parsed.issued_at,
            customer_name=_clean_text(getattr(parsed.customer, "name", None)),
            customer_action="already_imported",
            line_count=len(parsed.lines),
            total_weight_grams=parsed.total_weight_grams,
            total_amount_dkk=parsed.total_amount_dkk,
            template_profile=parsed.template_profile,
            source_net_amount_dkk=parsed.net_amount_dkk,
            source_vat_amount_dkk=parsed.vat_amount_dkk,
            source_gross_amount_dkk=parsed.total_amount_dkk,
            birth_date_text=parsed.birth_date_text,
            is_company=parsed.is_company,
            errors=[existing],
        )
    customer_action = "blocked"
    if not errors:
        _, customer_action, customer_errors = await _resolve_customer(session, parsed.customer, create=False)
        errors.extend(customer_errors)
    return HistoricalAfgImportPreviewItemOut(
        source_hash=parsed.upload.source_hash,
        file_name=parsed.upload.filename,
        status="ready" if not errors else "blocked",
        legacy_document_number=parsed.legacy_document_number,
        issued_at=parsed.issued_at,
        customer_name=_clean_text(getattr(parsed.customer, "name", None)),
        customer_action=customer_action,
        line_count=len(parsed.lines),
        total_weight_grams=parsed.total_weight_grams,
        total_amount_dkk=parsed.total_amount_dkk,
        template_profile=parsed.template_profile,
        source_net_amount_dkk=parsed.net_amount_dkk,
        source_vat_amount_dkk=parsed.vat_amount_dkk,
        source_gross_amount_dkk=parsed.total_amount_dkk,
        birth_date_text=parsed.birth_date_text,
        is_company=parsed.is_company,
        warnings=[
            *parsed.warnings,
            "Uniconta, WooCommerce, e-posta ve diğer dış entegrasyonlar bu işlemde kapalıdır.",
        ],
        errors=errors,
    )


async def preview_historical_afg_import(
    session: AsyncSession,
    *,
    uploads: list[HistoricalAfgUpload],
) -> HistoricalAfgImportPreviewOut:
    seen_hashes: set[str] = set()
    items: list[HistoricalAfgImportPreviewItemOut] = []
    for upload in uploads:
        parsed = _parse_upload(upload)
        items.append(
            await _preview_item(
                session,
                parsed,
                duplicate_in_selection=upload.source_hash in seen_hashes,
            )
        )
        seen_hashes.add(upload.source_hash)
    return HistoricalAfgImportPreviewOut(
        items=items,
        ready_count=sum(item.status == "ready" for item in items),
        blocked_count=sum(item.status == "blocked" for item in items),
        already_imported_count=sum(item.status == "already_imported" for item in items),
    )


def _historical_notes(parsed: HistoricalAfgParsed) -> str:
    market_rates = getattr(parsed.sections, "market_rates", None)
    return json.dumps(
        {
            "kind": "historical_afg_import_v1",
            "template_profile": parsed.template_profile,
            "source_file": parsed.upload.filename,
            "source_sha256": parsed.upload.source_hash,
            "legacy_document_number": parsed.legacy_document_number,
            "document_totals": {
                "gross_dkk": str(parsed.total_amount_dkk),
                "net_dkk": str(parsed.net_amount_dkk) if parsed.net_amount_dkk is not None else None,
                "vat_dkk": str(parsed.vat_amount_dkk) if parsed.vat_amount_dkk is not None else None,
            },
            "payment_method": getattr(parsed.sections, "payment_method", "bank") or "bank",
            "bank_info": getattr(parsed.sections, "bank_info", {}) or {},
            "market_rates": market_rates.model_dump(mode="json") if market_rates is not None else {},
            "external_effects": "disabled",
        },
        ensure_ascii=False,
        default=str,
    )


async def _persist(
    session: AsyncSession,
    *,
    parsed: HistoricalAfgParsed,
    actor: User,
) -> PosDocument:
    duplicate = await _duplicate_reason(session, parsed)
    if duplicate:
        raise ValueError(duplicate)
    customer, _, customer_errors = await _resolve_customer(session, parsed.customer, create=True)
    if customer is None:
        raise ValueError("; ".join(customer_errors) or "Müşteri çözümlenemedi.")
    if parsed.issued_at is None or not parsed.legacy_document_number or not parsed.lines:
        raise ValueError("İçe aktarma için zorunlu AFG bilgileri eksik.")

    rates = getattr(parsed.sections, "market_rates", None)
    notes = _historical_notes(parsed)
    # Tarihsel belgedeki tamamlanmış tutarlar aynen taşınır; KDV'li eski
    # belgelerde net/KDV yeniden hesaplanmaz.
    gross_amount = parsed.total_amount_dkk
    vat_amount = parsed.vat_amount_dkk if parsed.vat_amount_dkk is not None else Decimal("0.00")
    net_amount = parsed.net_amount_dkk if parsed.net_amount_dkk is not None else quantize_2(gross_amount - vat_amount)
    vat_rate = (
        quantize_2(vat_amount / net_amount * Decimal("100"))
        if vat_amount > 0 and net_amount > 0
        else Decimal("0.00")
    )
    pos_session = PosSession(
        session_code="IMP-" + parsed.upload.source_hash[:12].upper(),
        display_token=secrets.token_urlsafe(30),
        clerk_user_id=actor.id,
        customer_id=customer.id,
        trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
        product_type=ProductTypeEnum.JEWELRY,
        metal_type=parsed.lines[0].metal_type,
        weight_grams=parsed.total_weight_grams,
        purity_karat=parsed.lines[0].purity_karat,
        purity_percentage=parsed.lines[0].purity_percentage,
        live_rate_dkk=parsed.lines[0].rate_dkk,
        manual_rate_dkk=(quantize_2(to_decimal(rates.gold_24k_dkk)) if rates is not None else None),
        rate_source=PosRateSourceEnum.MANUAL,
        margin_percent_internal=parsed.lines[0].margin_percent,
        final_offer_dkk=parsed.total_amount_dkk,
        visible_snapshot={},
        notes=notes,
        status=PosSessionStatusEnum.CONFIRMED,
        created_at=parsed.issued_at,
        confirmed_at=parsed.issued_at,
    )
    session.add(pos_session)
    await session.flush()
    for line in parsed.lines:
        session.add(
            PosSessionLine(
                pos_session_id=pos_session.id,
                line_no=line.line_no,
                product_type=ProductTypeEnum.JEWELRY,
                metal_type=line.metal_type,
                weight_grams=line.weight_grams,
                purity_karat=line.purity_karat,
                purity_percentage=line.purity_percentage,
                rate_dkk=line.rate_dkk,
                margin_percent_internal=line.margin_percent,
                line_offer_dkk=line.line_total_dkk,
                notes=json.dumps(
                    {
                        "source": "historical_afg_import",
                        "row_key": line.row_key,
                        "type_label": line.type_label,
                    }
                ),
            )
        )
    await session.flush()
    await pos_core._sync_buy_session_summary_from_lines(session, pos_session=pos_session)

    document = PosDocument(
        pos_session_id=pos_session.id,
        document_type=PosDocumentTypeEnum.PURCHASE_RECEIPT,
        issued_at=parsed.issued_at,
        supply_at=parsed.issued_at,
        currency_code="DKK",
        gross_amount_dkk=gross_amount,
        net_amount_dkk=net_amount,
        vat_rate_percent=vat_rate,
        vat_amount_dkk=vat_amount,
        customer_name=customer.name,
        customer_phone=customer.phone,
        customer_email=customer.email,
        customer_address=getattr(parsed.customer, "address", None),
        customer_postal_code=getattr(parsed.customer, "postal_code", None),
        customer_city=getattr(parsed.customer, "city", None),
        notes=notes,
        legacy_document_number=parsed.legacy_document_number,
        historical_import_hash=parsed.upload.source_hash,
        historical_imported_at=utc_now(),
        historical_imported_by=actor.id,
        uniconta_sync_status="historical",
    )
    session.add(document)
    await session.flush()
    transaction = Transaction(
        pos_session_id=pos_session.id,
        pos_document_sequence_no=document.sequence_no,
        trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER.value,
        status="confirmed",
        customer_id=customer.id,
        clerk_user_id=actor.id,
        currency_code="DKK",
        gross_amount_dkk=gross_amount,
        net_amount_dkk=net_amount,
        vat_rate_percent=vat_rate,
        vat_amount_dkk=vat_amount,
        notes="Tarihsel AFG içe aktarımı; dış entegrasyonlar kapalı.",
        created_at=parsed.issued_at,
        confirmed_at=parsed.issued_at,
    )
    session.add(transaction)
    await session.flush()
    for line in parsed.lines:
        session.add(
            TransactionLine(
                transaction_id=transaction.id,
                line_no=line.line_no,
                product_type=ProductTypeEnum.JEWELRY.value,
                metal_type=line.metal_type.value,
                weight_grams=line.weight_grams,
                purity_karat=line.purity_karat,
                purity_percentage=line.purity_percentage,
                pure_gold_grams=quantize_2(
                    line.weight_grams * line.purity_percentage / Decimal("100")
                ),
                rate_dkk=line.rate_dkk,
                margin_percent=line.margin_percent,
                line_total_dkk=line.line_total_dkk,
            )
        )
    await session.flush()
    return document


def _archive_original_upload(upload: HistoricalAfgUpload) -> str | None:
    """Orijinal dosya bayt bayt ProgramData altına arşivlenir (değişmez kaynak).

    Yol: {document_root}/historical-imports/{sha256}.{ext} — hash zaten
    benzersiz olduğundan idempotenttir; yazılamazsa import engellenmez.
    """
    from pathlib import Path

    from app.config import get_settings

    try:
        settings = get_settings()
        root = Path(settings.document_root_path()) / "historical-imports"
        root.mkdir(parents=True, exist_ok=True)
        suffix = Path(upload.filename).suffix.lower() or ".xlsx"
        target = root / f"{upload.source_hash}{suffix}"
        if not target.exists():
            target.write_bytes(upload.content)
        return str(target)
    except OSError:
        return None


async def apply_historical_afg_import(
    session: AsyncSession,
    *,
    uploads: list[HistoricalAfgUpload],
    selected_hashes: list[str],
    actor: User,
) -> HistoricalAfgImportApplyOut:
    selected = set(selected_hashes)
    if not selected:
        raise HTTPException(status_code=422, detail="İçe aktarılacak en az bir hazır dosya seçin.")
    if not selected.issubset({upload.source_hash for upload in uploads}):
        raise HTTPException(
            status_code=422,
            detail="Seçilen dosyalar yüklenen listeyle eşleşmiyor; önizlemeyi tekrar çalıştırın.",
        )

    seen_hashes: set[str] = set()
    results: list[HistoricalAfgImportApplyItemOut] = []
    for upload in uploads:
        if upload.source_hash not in selected:
            continue
        parsed = _parse_upload(upload)
        preview = await _preview_item(
            session,
            parsed,
            duplicate_in_selection=upload.source_hash in seen_hashes,
        )
        seen_hashes.add(upload.source_hash)
        if preview.status != "ready":
            results.append(
                HistoricalAfgImportApplyItemOut(
                    source_hash=upload.source_hash,
                    file_name=upload.filename,
                    status="skipped",
                    legacy_document_number=parsed.legacy_document_number,
                    message="; ".join(preview.errors),
                    errors=list(preview.errors),
                )
            )
            continue
        try:
            async with session.begin_nested():
                document = await _persist(session, parsed=parsed, actor=actor)
        except (HTTPException, ValidationError, ValueError) as exc:
            results.append(
                HistoricalAfgImportApplyItemOut(
                    source_hash=upload.source_hash,
                    file_name=upload.filename,
                    status="failed",
                    legacy_document_number=parsed.legacy_document_number,
                    message=str(getattr(exc, "detail", exc)),
                    errors=[str(getattr(exc, "detail", exc))],
                )
            )
            continue
        archive_path = _archive_original_upload(upload)
        results.append(
            HistoricalAfgImportApplyItemOut(
                source_hash=upload.source_hash,
                file_name=upload.filename,
                status="imported",
                legacy_document_number=parsed.legacy_document_number,
                sequence_no=document.sequence_no,
                message="Belge, işlem ve satırlar oluşturuldu. Dış entegrasyon çalıştırılmadı.",
                archive_path=archive_path,
            )
        )
    return HistoricalAfgImportApplyOut(
        items=results,
        imported_count=sum(item.status == "imported" for item in results),
        skipped_count=sum(item.status == "skipped" for item in results),
        failed_count=sum(item.status == "failed" for item in results),
    )
