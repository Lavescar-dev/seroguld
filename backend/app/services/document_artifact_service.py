from __future__ import annotations

import hashlib
import io
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable
from uuid import UUID, uuid4

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Protection
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import ROOT_DIR, get_settings
from app.utils.cpr import cpr_birth_part
from app.models.document_artifact import DocumentArtifact
from app.schemas.afg import (
    AfgClassification,
    AfgLogWorkspaceOut,
    AfgMeltLotCreateRequest,
    AfgMeltLotUpdateRequest,
    AfgRouteRequest,
    AfgWorkspaceDocumentOut,
    AfgWorkspaceLineOut,
)
from app.schemas.document_artifact import (
    DocumentArtifactCellChangeOut,
    DocumentArtifactEditableCellOut,
    DocumentArtifactPreviewOut,
    DocumentArtifactReconcilePreviewOut,
    DocumentArtifactRecordOut,
    DocumentArtifactSheetPreviewOut,
)
from app.schemas.inventory import InventoryMarketPricesUpdate, InventoryWorkspaceOut
from app.schemas.pos import (
    PosWorkspaceCalculatorsUpdate,
    PosDocumentDetailOut,
    PosWorkspaceCustomerUpdate,
    PosWorkspaceRateMatrixEntry,
    PosWorkspaceGoldRowOut,
    PosWorkspaceGoldRowInput,
    PosWorkspaceInvoiceGoldSheetUpdate,
    PosWorkspaceInvoiceMiscSheetUpdate,
    PosWorkspaceMarketRates,
    PosWorkspaceNumberingUpdate,
    PosWorkspaceOut,
    PosWorkspaceSectionsUpdate,
    PosWorkspaceSilverRowOut,
    PosWorkspaceSilverRowInput,
)
from app.schemas.product import ProductUpdate
from app.services import document_artifact_inventory, document_artifact_log
from app.utils.helpers import quantize_2, to_decimal, utc_now

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSM_MIME = "application/vnd.ms-excel.sheet.macroEnabled.12"

SYNC_SHEET_NAME = "__SERO_SYNC"
SYNC_MAGIC = "SERO_SYNC_V1"
SYNC_CONTRACT_VERSION = "1"

AFG_TEMPLATE_NAME = "Afregningsbilag ( alis frontumuz).xlsm"
DEPOLAMA_TEMPLATE_NAME = "Depolama.xlsx"
LOG_TEMPLATE_NAME = "Log sistemi- afg verileri buraya yazdiriyorum..xlsx"

AFG_CONTRACT_VERSION = "afg-v2"
INVENTORY_CONTRACT_VERSION = "inventory-v1"
LOG_CONTRACT_VERSION = "log-v2"

AFG_GOLD_ROW_START = 22
AFG_SILVER_ROW_START = 30
AFG_FACTURA_GOLD_ROW_START = 26
AFG_FACTURA_GOLD_ROW_END = 37
AFG_FACTURA_GOLD_FOOTER_START = 38
AFG_FACTURA_MISC_ROW_START = 25
AFG_FACTURA_MISC_ROW_END = 39
AFG_GOLD_ROW_KEYS = ("gold:8", "gold:14", "gold:18", "gold:21", "gold:21.6", "gold:22", "gold:24")
AFG_SILVER_ROW_KEYS = ("silver:2", "silver:3", "silver:4", "silver:5")
AFG_PRIMARY_SHEET = "Afregningsbilag"
AFG_FACTURA_GOLD_SHEET = "Faktura guld og sølv"
AFG_FACTURA_MISC_SHEET = "Faktura diverse"
AFG_VARIABLES_SHEET = "Variable værdier"
AFG_GUIDE_SHEET = "Brugsanvisning"
AFG_SYSTEM_INPUT_SHEETS = {AFG_PRIMARY_SHEET, AFG_VARIABLES_SHEET, AFG_FACTURA_GOLD_SHEET, AFG_FACTURA_MISC_SHEET}
AFG_DERIVED_SHEETS: tuple[str, ...] = ()
INVENTORY_SHEET = "Lager"
LOG_SHEET = "Ark1"
LOG_CONTROL_SHEET = "Sero Control"

LOG_GOLD_LEDGER_START = 10
LOG_GOLD_LEDGER_END = 33
LOG_SILVER_LEDGER_START = 58
LOG_SILVER_LEDGER_END = 87
LOG_ROUTE_CODE_OPTIONS = ("-", "S", "H", "D", "M")

AFG_CUSTOMER_EDITABLE_CELLS = (
    # Değer hücreleri etiketlerin sağındadır (D/G); etiket hücreleri kilitli
    # kalır. CPR workbook'tan DÜZENLENEMEZ: G16 salt görüntü (yalnız doğum
    # tarihi bölümü), tam CPR CRM içinde kalır.
    {"cell_ref": "D16", "label": "Müşteri Adı", "input_kind": "text", "field": "name"},
    {"cell_ref": "D17", "label": "Adres (yalnız sokak)", "input_kind": "text", "field": "address"},
    {"cell_ref": "G17", "label": "Kimlik / Pas", "input_kind": "text", "field": "identity_doc_number"},
    {"cell_ref": "D18", "label": "Posta kodu + Şehir", "input_kind": "text", "field": "postal_line"},
    {"cell_ref": "G18", "label": "Telefon", "input_kind": "text", "field": "phone"},
    {"cell_ref": "G19", "label": "E-mail", "input_kind": "text", "field": "email"},
)

AFG_DOCUMENT_EDITABLE_CELLS = (
    {"cell_ref": "H7", "label": "Belge tarihi", "input_kind": "date", "field": "document_date"},
)

AFG_SUMMARY_EDITABLE_CELLS = (
    {"cell_ref": "C40", "label": "Ödeme Yöntemi", "input_kind": "payment_method", "field": "payment_method"},
    {"cell_ref": "D41", "label": "Reg. Nr.", "input_kind": "text", "field": "reg_number"},
    {"cell_ref": "D42", "label": "Kontonr.", "input_kind": "text", "field": "account_number"},
    {"cell_ref": "A42", "label": "%25 alış KDV'si", "input_kind": "boolean", "field": "purchase_vat_enabled"},
    {"cell_ref": "D44", "label": "AFG notu", "input_kind": "text", "field": "afg_note"},
)

AFG_VARIABLE_EDITABLE_CELLS = (
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "C10", "label": "EUR / DKK FX", "input_kind": "decimal", "field": "market_rates:eur_dkk_fx"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "C4", "label": "Au 24K DKK", "input_kind": "decimal", "field": "market_rates:gold_24k_dkk"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "C5", "label": "Ag DKK", "input_kind": "decimal", "field": "market_rates:silver_dkk"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J4", "label": "Gold 8K DKK/g", "input_kind": "decimal", "field": "market_rates:gold_rates_dkk:8"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J5", "label": "Gold 14K DKK/g", "input_kind": "decimal", "field": "market_rates:gold_rates_dkk:14"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J6", "label": "Gold 18K DKK/g", "input_kind": "decimal", "field": "market_rates:gold_rates_dkk:18"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J7", "label": "Gold 21K DKK/g", "input_kind": "decimal", "field": "market_rates:gold_rates_dkk:21"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J8", "label": "Gold 21.6K DKK/g", "input_kind": "decimal", "field": "market_rates:gold_rates_dkk:21.6"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J9", "label": "Gold 22K DKK/g", "input_kind": "decimal", "field": "market_rates:gold_rates_dkk:22"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J10", "label": "Gold 24K DKK/g", "input_kind": "decimal", "field": "market_rates:gold_rates_dkk:24"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J11", "label": "Silver 999 DKK/g", "input_kind": "decimal", "field": "market_rates:silver_rates_dkk:999"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J12", "label": "Silver 925 DKK/g", "input_kind": "decimal", "field": "market_rates:silver_rates_dkk:925"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J13", "label": "Silver 830 DKK/g", "input_kind": "decimal", "field": "market_rates:silver_rates_dkk:830"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J14", "label": "Plet DKK/g", "input_kind": "decimal", "field": "market_rates:plet_dkk"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J15", "label": "Guldbarre DKK/g", "input_kind": "decimal", "field": "market_rates:gold_bar_dkk"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J16", "label": "Sølvbarre DKK/g", "input_kind": "decimal", "field": "market_rates:silver_bar_dkk"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J17", "label": "Platin DKK/g", "input_kind": "decimal", "field": "market_rates:platinum_dkk"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "J18", "label": "Palladium DKK/g", "input_kind": "decimal", "field": "market_rates:palladium_dkk"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "C14", "label": "Afregningsnr.", "input_kind": "text", "field": "numbering:afregnings_number_next"},
    {"sheet": AFG_VARIABLES_SHEET, "cell_ref": "D14", "label": "Fakturanr.", "input_kind": "text", "field": "numbering:invoice_number_next"},
)

AFG_VARIABLE_MATRIX_ROWS = (
    ("I4", "J4", "K4", "Gold 8K", "333"),
    ("I5", "J5", "K5", "Gold 14K", "585"),
    ("I6", "J6", "K6", "Gold 18K", "750"),
    ("I7", "J7", "K7", "Gold 21K", "875"),
    ("I8", "J8", "K8", "Gold 21.6K", "900"),
    ("I9", "J9", "K9", "Gold 22K", "916"),
    ("I10", "J10", "K10", "Gold 24K", "999"),
    ("I11", "J11", "K11", "Finsølv", "999"),
    ("I12", "J12", "K12", "Sterling sølv", "925"),
    ("I13", "J13", "K13", "3 tårnet sølv", "830"),
    ("I14", "J14", "K14", "Plet", ""),
    ("I15", "J15", "K15", "Guldbarre", "999.9"),
    ("I16", "J16", "K16", "Sølvbarre", "999"),
    ("I17", "J17", "K17", "Platin", "950"),
    ("I18", "J18", "K18", "Palladium", "500"),
)

# Taslak workbook'ta bar giriş satırları: 29 (şablonda boş) ve 34 (ilk dolgu
# satırı). Gümüş satırları 30-33'te kalır; SUM(F22:F37)/SUM(H22:H37) kapsar.
AFG_BAR_GOLD_ROW = 29
AFG_BAR_SILVER_ROW = 34

# Pt/Pd giriş satırları: 35-36 (A-H kolonları şablonda boş; J/K/L kolonları
# gümüş hesaplayıcıya ait — çakışma yok). SUM(F22:F37)/SUM(H22:H37) kapsar.
AFG_PTPD_PLATINUM_ROW = 35
AFG_PTPD_PALLADIUM_ROW = 36

AFG_GOLD_CALCULATOR_ROWS = (
    ("calc_gold:1", "J22", "K22", "L22"),
    ("calc_gold:2", "J23", "K23", "L23"),
    ("calc_gold:3", "J24", "K24", "L24"),
    ("calc_gold:4", "J25", "K25", "L25"),
    ("calc_gold:5", "J26", "K26", "L26"),
)

AFG_SILVER_CALCULATOR_ROWS = (
    ("calc_silver:1", "J32", "K32", "L32"),
    ("calc_silver:2", "J33", "K33", "L33"),
    ("calc_silver:3", "J34", "K34", "L34"),
    ("calc_silver:4", "J35", "K35", "L35"),
    ("calc_silver:5", "J36", "K36", "L36"),
    ("calc_silver:6", "J37", "K37", "L37"),
)

# Dinamik satır görünürlüğü: boş grid satırları hidden bayrağıyla gizlenir
# (değerler dosyada kalır; F38/H38 SUM'ları ve round-trip parse etkilenmez).
AFG_GOLD_FILL_ARGB = "FFFFC000"  # şablondaki karat satırı sarısı (22-28)


def _afg_row_editable_cells() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for offset, row_key in enumerate(AFG_GOLD_ROW_KEYS, start=AFG_GOLD_ROW_START):
        karat = row_key.split(":", 1)[1]
        rows.append({"cell_ref": f"B{offset}", "label": f"Guld {karat}k Mer pris kr/g", "input_kind": "decimal", "field": f"{row_key}:avance"})
        rows.append({"cell_ref": f"F{offset}", "label": f"Guld {karat}k Gram", "input_kind": "decimal", "field": f"{row_key}:gram"})
    for offset, row_key in enumerate(AFG_SILVER_ROW_KEYS, start=AFG_SILVER_ROW_START):
        type_code = row_key.split(":", 1)[1]
        rows.append({"cell_ref": f"B{offset}", "label": f"Gümüş {type_code} Mer pris kr/g", "input_kind": "decimal", "field": f"{row_key}:avance"})
        rows.append({"cell_ref": f"F{offset}", "label": f"Gümüş {type_code} Gram", "input_kind": "decimal", "field": f"{row_key}:gram"})
    rows.append({"cell_ref": f"B{AFG_BAR_GOLD_ROW}", "label": "Guldbarre Mer pris kr/g", "input_kind": "decimal", "field": "bar:gold:avance"})
    rows.append({"cell_ref": f"F{AFG_BAR_GOLD_ROW}", "label": "Guldbarre Gram", "input_kind": "decimal", "field": "bar:gold:gram"})
    rows.append({"cell_ref": f"B{AFG_BAR_SILVER_ROW}", "label": "Sølvbarre Mer pris kr/g", "input_kind": "decimal", "field": "bar:silver:avance"})
    rows.append({"cell_ref": f"F{AFG_BAR_SILVER_ROW}", "label": "Sølvbarre Gram", "input_kind": "decimal", "field": "bar:silver:gram"})
    rows.append({"cell_ref": f"B{AFG_PTPD_PLATINUM_ROW}", "label": "Platin Mer pris kr/g", "input_kind": "decimal", "field": "ptpd:platinum:avance"})
    rows.append({"cell_ref": f"F{AFG_PTPD_PLATINUM_ROW}", "label": "Platin Gram", "input_kind": "decimal", "field": "ptpd:platinum:gram"})
    rows.append({"cell_ref": f"B{AFG_PTPD_PALLADIUM_ROW}", "label": "Palladium Mer pris kr/g", "input_kind": "decimal", "field": "ptpd:palladium:avance"})
    rows.append({"cell_ref": f"F{AFG_PTPD_PALLADIUM_ROW}", "label": "Palladium Gram", "input_kind": "decimal", "field": "ptpd:palladium:gram"})
    return rows


def _afg_factura_gold_editable_cells() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for index, row_idx in enumerate(range(AFG_FACTURA_GOLD_ROW_START, AFG_FACTURA_GOLD_ROW_END + 1), start=1):
        row_key = f"invoice_gold:{index}"
        rows.extend(
            [
                {"sheet": AFG_FACTURA_GOLD_SHEET, "cell_ref": f"A{row_idx}", "label": f"Faktura G/S satir {index} kod", "input_kind": "text", "field": f"{row_key}:code"},
                {"sheet": AFG_FACTURA_GOLD_SHEET, "cell_ref": f"C{row_idx}", "label": f"Faktura G/S satir {index} finhed", "input_kind": "text", "field": f"{row_key}:fineness"},
                {"sheet": AFG_FACTURA_GOLD_SHEET, "cell_ref": f"E{row_idx}", "label": f"Faktura G/S satir {index} gram", "input_kind": "decimal", "field": f"{row_key}:gram"},
            ]
        )
    for index, row_idx in enumerate(range(AFG_FACTURA_GOLD_FOOTER_START, AFG_FACTURA_GOLD_FOOTER_START + 3), start=1):
        rows.append(
            {
                "sheet": AFG_FACTURA_GOLD_SHEET,
                "cell_ref": f"B{row_idx}",
                "label": f"Faktura G/S serbest metin {index}",
                "input_kind": "text",
                "field": f"invoice_gold_footer:{index}",
            }
        )
    return rows


def _afg_factura_misc_editable_cells() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for index, row_idx in enumerate(range(AFG_FACTURA_MISC_ROW_START, AFG_FACTURA_MISC_ROW_END + 1), start=1):
        row_key = f"invoice_misc:{index}"
        rows.extend(
            [
                {"sheet": AFG_FACTURA_MISC_SHEET, "cell_ref": f"C{row_idx}", "label": f"Faktura diverse satir {index} metin", "input_kind": "text", "field": f"{row_key}:text"},
                {"sheet": AFG_FACTURA_MISC_SHEET, "cell_ref": f"E{row_idx}", "label": f"Faktura diverse satir {index} adet", "input_kind": "decimal", "field": f"{row_key}:quantity"},
                {"sheet": AFG_FACTURA_MISC_SHEET, "cell_ref": f"F{row_idx}", "label": f"Faktura diverse satir {index} birim fiyat", "input_kind": "decimal", "field": f"{row_key}:unit_price_dkk"},
            ]
        )
    return rows


def _afg_calculator_editable_cells() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row_key, unit_cell, count_cell, _total_cell in AFG_GOLD_CALCULATOR_ROWS:
        rows.append({"cell_ref": unit_cell, "label": f"{row_key} birim agirlik", "input_kind": "decimal", "field": f"{row_key}:unit_weight"})
        rows.append({"cell_ref": count_cell, "label": f"{row_key} adet", "input_kind": "decimal", "field": f"{row_key}:count"})
    for row_key, unit_cell, count_cell, _total_cell in AFG_SILVER_CALCULATOR_ROWS:
        rows.append({"cell_ref": unit_cell, "label": f"{row_key} birim agirlik", "input_kind": "decimal", "field": f"{row_key}:unit_weight"})
        rows.append({"cell_ref": count_cell, "label": f"{row_key} adet", "input_kind": "decimal", "field": f"{row_key}:count"})
    return rows


AFG_EDITABLE_CELLS = (
    *AFG_CUSTOMER_EDITABLE_CELLS,
    *AFG_DOCUMENT_EDITABLE_CELLS,
    *AFG_SUMMARY_EDITABLE_CELLS,
    *AFG_VARIABLE_EDITABLE_CELLS,
    *_afg_row_editable_cells(),
    *_afg_calculator_editable_cells(),
    *_afg_factura_gold_editable_cells(),
    *_afg_factura_misc_editable_cells(),
)
AFG_EDITABLE_CELL_MAP = {cell["cell_ref"]: cell for cell in AFG_EDITABLE_CELLS}


def office_contract_version_for_kind(kind: str) -> str:
    if kind in {"alis-workspace", "alis-document"}:
        return AFG_CONTRACT_VERSION
    if kind == "depolama":
        return INVENTORY_CONTRACT_VERSION
    if kind == "log":
        return LOG_CONTRACT_VERSION
    return SYNC_CONTRACT_VERSION


@dataclass(slots=True)
class AfgWorkspaceArtifactInputs:
    customer: PosWorkspaceCustomerUpdate
    sections: PosWorkspaceSectionsUpdate
    normalized_values: dict[str, str]
    base_version: str | None = None


@dataclass(slots=True)
class SyncSheetRowMapping:
    mapping_type: str
    sheet_name: str
    row_index: int
    entity_id: str
    entity_key: str | None = None
    extra: dict[str, Any] | None = None


@dataclass(slots=True)
class SyncSheetMetadata:
    kind: str
    key: str
    artifact_key: str
    base_version: str | None
    workspace_revision: str | None
    contract_version: str
    mappings: list[SyncSheetRowMapping]


@dataclass(slots=True)
class ArtifactSyncContext:
    kind: str
    key: str
    artifact_key: str
    base_version: str
    workspace_revision: str | None = None
    contract_version: str = SYNC_CONTRACT_VERSION
    mappings: list[SyncSheetRowMapping] | None = None


@dataclass(slots=True)
class InventoryWorkbookArtifactInputs:
    market_prices: InventoryMarketPricesUpdate
    product_updates: dict[UUID, ProductUpdate]
    base_version: str | None = None


@dataclass(slots=True)
class LogWorkbookRouteEdit:
    line_id: UUID
    payload: AfgRouteRequest


@dataclass(slots=True)
class LogWorkbookLotCreate:
    metal_bucket: str
    create_payload: AfgMeltLotCreateRequest
    update_payload: AfgMeltLotUpdateRequest


@dataclass(slots=True)
class LogWorkbookLotUpdate:
    lot_id: UUID
    payload: AfgMeltLotUpdateRequest


@dataclass(slots=True)
class LogWorkbookArtifactInputs:
    route_updates: list[LogWorkbookRouteEdit]
    lot_creates: list[LogWorkbookLotCreate]
    lot_updates: list[LogWorkbookLotUpdate]
    base_version: str | None = None


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in {"—", "–", "-"}:
        return None
    return text or None


def _decimal_from_excel(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    return quantize_2(to_decimal(value))


def _decimal4_from_excel(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0.0000")
    return _quantize_4(to_decimal(value))


def _percent_from_excel(value: object) -> Decimal:
    amount = _decimal_from_excel(value)
    if Decimal("0") <= amount <= Decimal("1"):
        return quantize_2(amount * Decimal("100"))
    return quantize_2(amount)


def _payment_method_from_excel(value: object) -> str:
    text = (str(value or "")).strip().lower()
    if "kontant" in text:
        return "cash"
    return "bank"


def _boolean_from_excel(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return bool(value)
    text = str(value or "").strip().casefold()
    return text in {"1", "true", "yes", "evet", "ja", "on", "aktif", "açık", "acik"}


def _normalized_cell_value(value: object, *, input_kind: str) -> str:
    if input_kind == "percent":
        return _fmt_decimal(_percent_from_excel(value))
    if input_kind == "decimal":
        return _fmt_decimal(_decimal_from_excel(value))
    if input_kind == "payment_method":
        return "Kontant" if _payment_method_from_excel(value) == "cash" else "Overførsel"
    if input_kind == "boolean":
        return "1" if _boolean_from_excel(value) else "0"
    if input_kind == "date":
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = _clean_text(value)
        if text:
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(text, fmt).date().isoformat()
                except ValueError:
                    continue
        return "—"
    return _clean_text(value) or "—"


def _editable_cell_out(cell: dict[str, str]) -> DocumentArtifactEditableCellOut:
    return DocumentArtifactEditableCellOut(
        sheet=cell.get("sheet", AFG_PRIMARY_SHEET),
        cell_ref=cell["cell_ref"],
        label=cell["label"],
        input_kind=cell["input_kind"],
    )


def _afg_editable_cell_key(cell: dict[str, str]) -> str:
    return f"{cell.get('sheet', AFG_PRIMARY_SHEET)}!{cell['cell_ref']}"


@dataclass(slots=True)
class WorkbookArtifactBundle:
    artifact: DocumentArtifact
    content: bytes


def _document_root() -> Path:
    root = get_settings().document_root_path()
    root.mkdir(parents=True, exist_ok=True)
    return root


def _reference_root() -> Path:
    return ROOT_DIR / "referans"


def _template_path(template_name: str) -> Path:
    path = _reference_root() / template_name
    if not path.exists():
        raise FileNotFoundError(f"Referans workbook bulunamadı: {template_name}")
    return path


def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _fmt_decimal(value: Decimal | str | int | float | None, places: str = "0.00") -> str:
    amount = quantize_2(to_decimal(value or 0))
    if places == "0":
        return f"{amount.quantize(Decimal('1'))}"
    return f"{amount.quantize(Decimal(places))}"


def _quantize_4(value: Decimal | str | int | float | None) -> Decimal:
    return to_decimal(value or 0).quantize(Decimal("0.0001"))


def _fmt_optional_decimal(value: Decimal | str | int | float | None, places: str = "0.00") -> str:
    if value is None or value == "":
        return ""
    return _fmt_decimal(value, places=places)


def _fmt_date(value: datetime | str | None) -> str:
    if value is None:
        return "—"
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return value
    return value.strftime("%d.%m.%Y")


def _excel_datetime(value: datetime | str | None):
    if value is None or isinstance(value, str):
        return value
    if value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _relative_storage_path(path: Path) -> str:
    try:
        return path.resolve().relative_to(_document_root()).as_posix()
    except ValueError:
        return path.name


def _record_out(record: DocumentArtifact) -> DocumentArtifactRecordOut:
    revision = int(getattr(record, "revision", 1) or 1)
    return DocumentArtifactRecordOut(
        id=record.id,
        artifact_key=record.artifact_key,
        module_name=record.module_name,
        document_type=record.document_type,
        business_key=record.business_key,
        version_kind=record.version_kind,
        is_live=record.is_live,
        file_name=record.file_name,
        mime_type=record.mime_type,
        template_name=record.template_name,
        size_bytes=record.size_bytes,
        checksum_sha256=record.checksum_sha256,
        revision=revision,
        workbook_revision=record.checksum_sha256,
        base_revision=None,
        crm_revision=str(revision),
        conflict_state=ARTIFACT_CONFLICT_CLEAN,
        updated_at=record.updated_at,
    )


def _artifact_version_token(value: datetime | None) -> str:
    stamp = value or utc_now()
    return str(int(stamp.timestamp() * 1000))


ARTIFACT_CONFLICT_CLEAN = "clean"
ARTIFACT_CONFLICT_STALE = "stale"
ARTIFACT_CONFLICT_CONFLICT = "conflict"
ARTIFACT_CONFLICT_INVALID = "invalid"


def resolve_artifact_conflict_state(*, current_revision: int, incoming_revision: object | None) -> str:
    """Return the deterministic state for an artifact revision comparison.

    A workbook is only safe to apply when its embedded base revision is exactly
    the current CRM revision. Field-level merging is intentionally unsupported.
    """
    try:
        current = int(current_revision)
        incoming = int(incoming_revision) if incoming_revision is not None else None
    except (TypeError, ValueError):
        return ARTIFACT_CONFLICT_INVALID
    if current < 0 or incoming is None or incoming < 0:
        return ARTIFACT_CONFLICT_INVALID
    if incoming == current:
        return ARTIFACT_CONFLICT_CLEAN
    if incoming < current:
        return ARTIFACT_CONFLICT_STALE
    return ARTIFACT_CONFLICT_CONFLICT


def _next_artifact_revision(record: DocumentArtifact | None) -> int:
    return int(getattr(record, "revision", 0) or 0) + 1


def _set_cell_unlocked(sheet, cell_ref: str) -> None:
    cell = sheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    cell.protection = Protection(locked=False)


def _sync_sheet(workbook):
    if SYNC_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[SYNC_SHEET_NAME]
        sheet.sheet_state = "hidden"
        return sheet
    sheet = workbook.create_sheet(SYNC_SHEET_NAME)
    sheet.sheet_state = "hidden"
    return sheet


def _write_sync_sheet(workbook, *, context: ArtifactSyncContext) -> None:
    sheet = _sync_sheet(workbook)
    for row_idx in range(1, max(sheet.max_row, 64) + 1):
        for col in ("A", "B", "C", "D", "E", "F"):
            _set_sheet_cell(sheet, f"{col}{row_idx}", None)
    _set_sheet_cell(sheet, "A1", "magic")
    _set_sheet_cell(sheet, "B1", SYNC_MAGIC)
    _set_sheet_cell(sheet, "A2", "kind")
    _set_sheet_cell(sheet, "B2", context.kind)
    _set_sheet_cell(sheet, "A3", "key")
    _set_sheet_cell(sheet, "B3", context.key)
    _set_sheet_cell(sheet, "A4", "artifact_key")
    _set_sheet_cell(sheet, "B4", context.artifact_key)
    _set_sheet_cell(sheet, "A5", "base_version")
    _set_sheet_cell(sheet, "B5", context.base_version)
    _set_sheet_cell(sheet, "A7", "workspace_revision")
    _set_sheet_cell(sheet, "B7", context.workspace_revision)
    _set_sheet_cell(sheet, "A6", "contract_version")
    _set_sheet_cell(sheet, "B6", context.contract_version)
    _set_sheet_cell(sheet, "A8", "mapping_type")
    _set_sheet_cell(sheet, "B8", "sheet")
    _set_sheet_cell(sheet, "C8", "row_index")
    _set_sheet_cell(sheet, "D8", "entity_id")
    _set_sheet_cell(sheet, "E8", "entity_key")
    _set_sheet_cell(sheet, "F8", "extra_json")
    for offset, mapping in enumerate(context.mappings or [], start=9):
        _set_sheet_cell(sheet, f"A{offset}", mapping.mapping_type)
        _set_sheet_cell(sheet, f"B{offset}", mapping.sheet_name)
        _set_sheet_cell(sheet, f"C{offset}", mapping.row_index)
        _set_sheet_cell(sheet, f"D{offset}", mapping.entity_id)
        _set_sheet_cell(sheet, f"E{offset}", mapping.entity_key)
        _set_sheet_cell(sheet, f"F{offset}", json.dumps(mapping.extra or {}, ensure_ascii=True, sort_keys=True))


def _read_sync_sheet(workbook) -> SyncSheetMetadata:
    if SYNC_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("Workbook sync metadata bulunamadı")
    sheet = workbook[SYNC_SHEET_NAME]
    magic = _clean_text(sheet["B1"].value)
    if magic != SYNC_MAGIC:
        raise ValueError("Workbook sync metadata geçersiz")
    mappings: list[SyncSheetRowMapping] = []
    row_idx = 9
    while True:
        mapping_type = _clean_text(sheet[f"A{row_idx}"].value)
        if mapping_type is None:
            break
        sheet_name = _clean_text(sheet[f"B{row_idx}"].value)
        row_value = sheet[f"C{row_idx}"].value
        entity_id = _clean_text(sheet[f"D{row_idx}"].value)
        entity_key = _clean_text(sheet[f"E{row_idx}"].value)
        extra_raw = _clean_text(sheet[f"F{row_idx}"].value)
        if sheet_name is None or row_value is None or entity_id is None:
            raise ValueError("Workbook sync satırı eksik")
        mappings.append(
            SyncSheetRowMapping(
                mapping_type=mapping_type,
                sheet_name=sheet_name,
                row_index=int(to_decimal(row_value)),
                entity_id=entity_id,
                entity_key=entity_key,
                extra=json.loads(extra_raw) if extra_raw else {},
            )
        )
        row_idx += 1
    return SyncSheetMetadata(
        kind=_clean_text(sheet["B2"].value) or "",
        key=_clean_text(sheet["B3"].value) or "",
        artifact_key=_clean_text(sheet["B4"].value) or "",
        base_version=_clean_text(sheet["B5"].value),
        workspace_revision=_clean_text(sheet["B7"].value),
        contract_version=_clean_text(sheet["B6"].value) or SYNC_CONTRACT_VERSION,
        mappings=mappings,
    )


def _require_sync_metadata(workbook, *, expected_kind: str, expected_key: str) -> SyncSheetMetadata:
    metadata = _read_sync_sheet(workbook)
    if metadata.kind != expected_kind or metadata.key != expected_key:
        raise ValueError("Workbook farklı bir belgeye ait")
    return metadata


def _sync_metadata_if_present(workbook, *, expected_kind: str, expected_key: str) -> SyncSheetMetadata | None:
    if SYNC_SHEET_NAME not in workbook.sheetnames:
        return None
    return _require_sync_metadata(workbook, expected_kind=expected_kind, expected_key=expected_key)


def read_artifact_sync_metadata(content: bytes, *, expected_kind: str, expected_key: str) -> SyncSheetMetadata:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    try:
        return _require_sync_metadata(workbook, expected_kind=expected_kind, expected_key=expected_key)
    finally:
        try:
            workbook.close()
        except Exception:
            pass


async def _upsert_record(
    session: AsyncSession,
    *,
    artifact_key: str,
    module_name: str,
    document_type: str,
    business_key: str,
    version_kind: str,
    is_live: bool,
    file_name: str,
    file_path: Path,
    mime_type: str,
    template_name: str | None,
    content: bytes,
    updated_at: datetime | None = None,
    revision: int | None = None,
) -> DocumentArtifact:
    checksum = _sha256(content)
    record = await session.scalar(select(DocumentArtifact).where(DocumentArtifact.artifact_key == artifact_key))
    if record is None:
        record = DocumentArtifact(
            artifact_key=artifact_key,
            module_name=module_name,
            document_type=document_type,
            business_key=business_key,
            version_kind=version_kind,
            is_live=is_live,
            file_name=file_name,
            file_path=_relative_storage_path(file_path),
            mime_type=mime_type,
            template_name=template_name,
            size_bytes=len(content),
            checksum_sha256=checksum,
            revision=max(int(revision or 1), 1),
        )
        if updated_at is not None:
            record.updated_at = updated_at
        session.add(record)
    else:
        next_file_path = _relative_storage_path(file_path)
        changed = any(
            (
                record.module_name != module_name,
                record.document_type != document_type,
                record.business_key != business_key,
                record.version_kind != version_kind,
                record.is_live != is_live,
                record.file_name != file_name,
                record.file_path != next_file_path,
                record.mime_type != mime_type,
                record.template_name != template_name,
                record.size_bytes != len(content),
                record.checksum_sha256 != checksum,
            )
        )
        current_revision = int(getattr(record, "revision", 0) or 0)
        requested_revision = int(revision) if revision is not None else current_revision + 1
        next_revision = max(current_revision + 1, requested_revision, 1)
        revision_changed = next_revision != current_revision
        record.module_name = module_name
        record.document_type = document_type
        record.business_key = business_key
        record.version_kind = version_kind
        record.is_live = is_live
        record.file_name = file_name
        record.file_path = next_file_path
        record.mime_type = mime_type
        record.template_name = template_name
        record.size_bytes = len(content)
        record.checksum_sha256 = checksum
        record.revision = next_revision
        if (changed or revision_changed) and updated_at is not None:
            record.updated_at = updated_at
    await session.flush()
    await session.refresh(record)
    return record


async def _store_artifact(
    session: AsyncSession,
    *,
    artifact_key: str,
    module_name: str,
    document_type: str,
    business_key: str,
    version_kind: str,
    is_live: bool,
    file_name: str,
    relative_path: Path,
    mime_type: str,
    template_name: str | None,
    content: bytes,
    updated_at: datetime | None = None,
    revision: int | None = None,
) -> WorkbookArtifactBundle:
    absolute_path = _document_root() / relative_path
    absolute_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = absolute_path.with_name(f".{absolute_path.name}.{uuid4().hex}.tmp")
    try:
        temporary.write_bytes(content)
        record = await _upsert_record(
            session,
            artifact_key=artifact_key,
            module_name=module_name,
            document_type=document_type,
            business_key=business_key,
            version_kind=version_kind,
            is_live=is_live,
            file_name=file_name,
            file_path=absolute_path,
            mime_type=mime_type,
            template_name=template_name,
            content=content,
            updated_at=updated_at,
            revision=revision,
        )
        os.replace(temporary, absolute_path)
    finally:
        temporary.unlink(missing_ok=True)
    return WorkbookArtifactBundle(artifact=record, content=content)


async def list_artifact_records(
    session: AsyncSession,
    *,
    module_name: str,
    document_type: str,
    business_key: str | None = None,
    version_kind: str | None = None,
    limit: int = 20,
) -> list[DocumentArtifactRecordOut]:
    stmt = (
        select(DocumentArtifact)
        .where(DocumentArtifact.module_name == module_name, DocumentArtifact.document_type == document_type)
        .order_by(DocumentArtifact.updated_at.desc())
        .limit(limit)
    )
    if business_key is not None:
        stmt = stmt.where(DocumentArtifact.business_key == business_key)
    if version_kind is not None:
        stmt = stmt.where(DocumentArtifact.version_kind == version_kind)
    rows = (await session.scalars(stmt)).all()
    return [_record_out(row) for row in rows]


async def get_artifact_record(session: AsyncSession, artifact_key: str) -> DocumentArtifact | None:
    return await session.scalar(select(DocumentArtifact).where(DocumentArtifact.artifact_key == artifact_key))


def artifact_absolute_path(record: DocumentArtifact) -> Path:
    return _document_root() / record.file_path


def _open_template(template_name: str, *, keep_vba: bool = False):
    return load_workbook(_template_path(template_name), keep_vba=keep_vba, data_only=False)


def _touch_calc_flags(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        setattr(calculation, "fullCalcOnLoad", True)
        setattr(calculation, "forceFullCalc", True)


def _save_workbook_bytes(workbook) -> bytes:
    payload = io.BytesIO()
    try:
        workbook.save(payload)
        return payload.getvalue()
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _sheet_row(label: str, value: str) -> list[str]:
    return [label, value]


def _set_sheet_cell(sheet, cell_ref: str, value) -> None:
    cell = sheet[cell_ref]
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _protect_sheet(sheet, *, editable_refs: Iterable[str] = ()) -> None:
    sheet.protection.sheet = True
    sheet.protection.enable()
    for cell_ref in editable_refs:
        _set_cell_unlocked(sheet, cell_ref)


def _stringify(value: object) -> str:
    if value is None:
        return "—"
    if isinstance(value, Decimal):
        return _fmt_decimal(value)
    return str(value)


def _normalized_customer_name(name: str | None) -> str:
    text = (name or "").strip()
    return text or "—"


def _afg_contract_sheet(
    name: str,
    *,
    mode: str,
    system_sync: bool,
    note: str,
    columns: list[str] | None = None,
    rows: list[list[str]] | None = None,
) -> DocumentArtifactSheetPreviewOut:
    return DocumentArtifactSheetPreviewOut(
        name=name,
        mode=mode,
        system_sync=system_sync,
        columns=columns or [],
        rows=rows or [],
        note=note,
    )


def _apply_afg_sheet_modes(workbook, *, can_write: bool) -> None:
    editable_by_sheet: dict[str, list[str]] = {}
    for cell in AFG_EDITABLE_CELLS:
        sheet_name = cell.get("sheet", AFG_PRIMARY_SHEET)
        editable_by_sheet.setdefault(sheet_name, []).append(cell["cell_ref"])
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name == SYNC_SHEET_NAME:
            sheet.sheet_state = "hidden"
            continue
        is_controlled = sheet_name in AFG_SYSTEM_INPUT_SHEETS
        if can_write and is_controlled:
            _protect_sheet(sheet, editable_refs=editable_by_sheet.get(sheet_name, ()))
        else:
            _protect_sheet(sheet)
    if AFG_PRIMARY_SHEET in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index(AFG_PRIMARY_SHEET)


def _afg_gold_rows_from_workspace(workspace: PosWorkspaceOut) -> list[PosWorkspaceGoldRowOut]:
    ordered = {row.row_key: row for row in workspace.gold_rows}
    return [ordered[key] for key in AFG_GOLD_ROW_KEYS if key in ordered]


def _afg_silver_rows_from_workspace(workspace: PosWorkspaceOut) -> list[PosWorkspaceSilverRowOut]:
    ordered = {row.row_key: row for row in workspace.silver_rows}
    return [ordered[key] for key in AFG_SILVER_ROW_KEYS if key in ordered]


def _aggregate_detail_gold_rows(detail: PosDocumentDetailOut) -> list[dict[str, Decimal | str]]:
    rows: dict[str, dict[str, Decimal | str]] = {
        key: {
            "row_key": key,
            "label": f"Guld {key.split(':', 1)[1]}k" if key != "gold:21.6" else "Guld 21.6k",
            "karat": Decimal(key.split(":", 1)[1]),
            "lodighed": "—",
            "gram": Decimal("0.00"),
            "avance_percent": Decimal("0.00"),
            "unit_price_dkk": Decimal("0.00"),
            "line_total_dkk": Decimal("0.00"),
        }
        for key in AFG_GOLD_ROW_KEYS
    }
    purity_to_key = {
        Decimal("8"): "gold:8",
        Decimal("14"): "gold:14",
        Decimal("18"): "gold:18",
        Decimal("21"): "gold:21",
        Decimal("21.6"): "gold:21.6",
        Decimal("22"): "gold:22",
        Decimal("24"): "gold:24",
    }
    for line in detail.lines:
        if str(line.metal_type or "").lower() != "yellow_gold":
            continue
        if str(line.product_type or "").lower() == "bar":
            continue
        karat = to_decimal(line.purity_karat.replace("K", "").replace("k", "")) if line.purity_karat else Decimal("0")
        row_key = purity_to_key.get(karat)
        if row_key is None:
            continue
        row = rows[row_key]
        gram = to_decimal(line.weight_grams or 0)
        line_total = to_decimal(line.line_total_dkk or 0)
        row["gram"] = quantize_2(to_decimal(row["gram"]) + gram)
        row["line_total_dkk"] = quantize_2(to_decimal(row["line_total_dkk"]) + line_total)
        row["unit_price_dkk"] = quantize_2(line_total / gram) if gram > 0 else to_decimal(row["unit_price_dkk"])
        row["lodighed"] = str(int(to_decimal(line.purity_percentage or 0) * Decimal("10")))
        row["avance_percent"] = quantize_2(to_decimal(line.margin_percent or 0))
    return [rows[key] for key in AFG_GOLD_ROW_KEYS]


def _aggregate_detail_silver_rows(detail: PosDocumentDetailOut) -> list[dict[str, Decimal | str]]:
    silver_defs = {
        "silver:2": ("2", "Finsølv999‰", "999", Decimal("99.90")),
        "silver:3": ("3", "Sterling sølv925‰", "925", Decimal("92.50")),
        "silver:4": ("4", "3 tårnet sølv830‰", "830", Decimal("83.00")),
        "silver:5": ("5", "Plet", "", Decimal("0.00")),
    }
    rows: dict[str, dict[str, Decimal | str]] = {
        key: {
            "row_key": key,
            "type_code": type_code,
            "label": label,
            "lodighed": lodighed,
            "gram": Decimal("0.00"),
            "avance_percent": Decimal("0.00"),
            "unit_price_dkk": Decimal("0.00"),
            "line_total_dkk": Decimal("0.00"),
        }
        for key, (type_code, label, lodighed, _purity) in silver_defs.items()
    }
    purity_to_key = {
        Decimal("99.90"): "silver:2",
        Decimal("92.50"): "silver:3",
        Decimal("83.00"): "silver:4",
        # Plet saflıksızdır (0.00); eski kayıtlardaki 80.00 de Plet'e eşlenir.
        Decimal("0.00"): "silver:5",
        Decimal("80.00"): "silver:5",
    }
    for line in detail.lines:
        if str(line.metal_type or "").lower() != "silver":
            continue
        if str(line.product_type or "").lower() == "bar":
            continue
        key = purity_to_key.get(quantize_2(to_decimal(line.purity_percentage or 0)))
        if key is None:
            continue
        row = rows[key]
        gram = to_decimal(line.weight_grams or 0)
        line_total = to_decimal(line.line_total_dkk or 0)
        row["gram"] = quantize_2(to_decimal(row["gram"]) + gram)
        row["line_total_dkk"] = quantize_2(to_decimal(row["line_total_dkk"]) + line_total)
        row["unit_price_dkk"] = quantize_2(line_total / gram) if gram > 0 else to_decimal(row["unit_price_dkk"])
        row["avance_percent"] = quantize_2(to_decimal(line.margin_percent or 0))
    return [rows[key] for key in AFG_SILVER_ROW_KEYS]


def _aggregate_detail_bar_rows(detail: PosDocumentDetailOut) -> list[dict[str, Decimal | str]]:
    rows: dict[str, dict[str, Decimal | str]] = {
        bar_type: {
            "bar_type": bar_type,
            "gram": Decimal("0.00"),
            "avance_percent": Decimal("0.00"),
            "unit_price_dkk": Decimal("0.00"),
            "line_total_dkk": Decimal("0.00"),
        }
        for bar_type in ("gold", "silver")
    }
    for line in detail.lines:
        if str(line.product_type or "").lower() != "bar":
            continue
        bar_type = "gold" if str(line.metal_type or "").lower() == "yellow_gold" else "silver"
        row = rows[bar_type]
        gram = to_decimal(line.weight_grams or 0)
        line_total = to_decimal(line.line_total_dkk or 0)
        row["gram"] = quantize_2(to_decimal(row["gram"]) + gram)
        row["line_total_dkk"] = quantize_2(to_decimal(row["line_total_dkk"]) + line_total)
        row["unit_price_dkk"] = quantize_2(line_total / gram) if gram > 0 else to_decimal(row["unit_price_dkk"])
        row["avance_percent"] = quantize_2(to_decimal(line.margin_percent or 0))
    return [rows["gold"], rows["silver"]]


def _aggregate_detail_ptpd_rows(detail: PosDocumentDetailOut) -> list[dict[str, Decimal | str]]:
    rows: dict[str, dict[str, Decimal | str]] = {
        metal: {
            "metal": metal,
            "gram": Decimal("0.00"),
            "avance_percent": Decimal("0.00"),
            "unit_price_dkk": Decimal("0.00"),
            "line_total_dkk": Decimal("0.00"),
        }
        for metal in ("platinum", "palladium")
    }
    for line in detail.lines:
        metal = str(line.metal_type or "").lower()
        if metal not in rows:
            continue
        row = rows[metal]
        gram = to_decimal(line.weight_grams or 0)
        line_total = to_decimal(line.line_total_dkk or 0)
        row["gram"] = quantize_2(to_decimal(row["gram"]) + gram)
        row["line_total_dkk"] = quantize_2(to_decimal(row["line_total_dkk"]) + line_total)
        row["unit_price_dkk"] = quantize_2(line_total / gram) if gram > 0 else to_decimal(row["unit_price_dkk"])
        row["avance_percent"] = quantize_2(to_decimal(line.margin_percent or 0))
    return [rows["platinum"], rows["palladium"]]


def _apply_afg_customer_cells(
    sheet,
    *,
    name: str | None,
    cpr_number: str | None,
    address: str | None,
    city: str | None,
    postal_code: str | None,
    phone: str | None,
    email: str | None,
    identity_doc: str | None = None,
) -> None:
    # Değerler etiketlerin SAĞINDAKİ hücrelere yazılır; C16-C18/F16-F19
    # etiketleri (Navn:/Adresse:/Postnr.:/CPR nr./...) asla ezilmez.
    # D17 yalnız sokak adresi, D18 "posta_kodu şehir" birlikte.
    sheet["D16"] = name or "—"
    # Veri minimizasyonu: belgeye yalnız CPR'nin doğum tarihi bölümü yazılır.
    sheet["G16"] = cpr_birth_part(cpr_number) or "—"
    sheet["D17"] = str(address or "").strip() or "—"
    sheet["G17"] = identity_doc or "—"
    sheet["D18"] = _compose_afg_postal_line(postal_code, city) or "—"
    sheet["G18"] = phone or "—"
    sheet["G19"] = email or "—"


def _apply_afg_footer_cells(sheet) -> None:
    """C53/C54 firma alt bilgisi — ayarlardan, 'Sero Guld' bölümü kalın.

    Şablon C53'ü zaten rich-text taşır ama openpyxl rich_text=False ile
    yüklerken düzleştirir; her üretimde açıkça yeniden yazarak bold korunur.
    Global rich_text=True bilinçli olarak KULLANILMAZ (üç parser'da hücre
    tipi denetimi gerektirir).
    """
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    settings = get_settings()
    display_name = str(settings.invoice_seller_name or "Sero Guld").strip()
    if display_name.lower().endswith(" aps"):
        display_name = display_name[: -len(" aps")].strip()
    address = str(settings.invoice_seller_address_line1 or "").strip()
    postal_city = " ".join(
        part for part in (str(settings.invoice_seller_postal_code or "").strip(), str(settings.invoice_seller_city or "").strip()) if part
    )
    country = str(settings.invoice_seller_country or "").strip()
    cvr = str(settings.invoice_seller_cvr or "").strip()
    line1_rest = " -  " + " - ".join(part for part in (address, postal_city, country, f"CVR-nr: {cvr}" if cvr else "") if part)

    bold_font = InlineFont(b=True, sz=10, rFont="Calibri", family=2, scheme="minor")
    plain_font = InlineFont(sz=10, rFont="Calibri", family=2, scheme="minor")
    sheet["C53"] = CellRichText(TextBlock(bold_font, display_name), TextBlock(plain_font, line1_rest))

    phone = str(settings.invoice_seller_phone or "").strip()
    email = str(settings.invoice_seller_email or "").strip()
    website = str(getattr(settings, "invoice_seller_website", "") or "").strip()
    line2_parts = [f"Tlf.: {phone}" if phone else "", f"E-mail: {email}" if email else "", website]
    sheet["C54"] = "         " + " - ".join(part for part in line2_parts if part)


# R2-09 — üç maddelik beyan (AML/PEP dahil). ÇEVİRİ KATMANI DIŞINDA: metin
# sabit Danca'dır, arayüz dili değişse de değişmez. Şablonda C47 başlık +
# C48/C49 ilk iki madde zaten var; C50 (aynı stille boş hücre) 3. PEP maddesi
# için kod tarafından doldurulur. Idempotent: her üretimde tam metin yazılır.
from app.services.pos_value_helpers import AFG_DECLARATION_HEADER, AFG_DECLARATION_ITEMS  # noqa: E402


def _apply_afg_declaration_cells(sheet) -> None:
    sheet["C47"] = AFG_DECLARATION_HEADER
    sheet["C48"] = AFG_DECLARATION_ITEMS[0]
    sheet["C49"] = AFG_DECLARATION_ITEMS[1]
    sheet["C50"] = AFG_DECLARATION_ITEMS[2]


def _compose_afg_postal_line(postal_code: str | None, city: str | None) -> str | None:
    postal_text = str(postal_code or "").strip()
    city_text = str(city or "").strip()
    if postal_text and city_text:
        return f"{postal_text} {city_text}"
    return postal_text or city_text or None


def _split_afg_postal_line(value: object) -> tuple[str | None, str | None]:
    """"2650 Hvidovre" → ("2650", "Hvidovre"); baştaki rakam bloğu posta kodudur."""
    text = _clean_text(value)
    if text is None:
        return None, None
    match = re.match(r"^(\d{3,4})\s+(.+)$", text)
    if match:
        return match.group(1), match.group(2).strip() or None
    if text.isdigit():
        return text, None
    return None, text


def _compose_afg_address_line(address: str | None, city: str | None) -> str | None:
    # Eski (afg-v1) belgelerin adres satırı: "sokak · şehir".
    address_text = str(address or "").strip()
    city_text = str(city or "").strip()
    if address_text and city_text:
        return f"{address_text} · {city_text}"
    return address_text or city_text or None


def _split_afg_address_line(value: object) -> tuple[str | None, str | None]:
    text = _clean_text(value)
    if text is None:
        return None, None
    if " · " in text:
        address, city = text.rsplit(" · ", 1)
        return str(address).strip() or None, str(city).strip() or None
    return text, None


def _build_afg_market_rates_from_workspace(market_rates: PosWorkspaceMarketRates) -> PosWorkspaceMarketRates:
    fx = quantize_2(to_decimal(market_rates.eur_dkk_fx))
    gold_rates_dkk = {str(key): quantize_2(to_decimal(value)) for key, value in (market_rates.gold_rates_dkk or {}).items()}
    silver_rates_dkk = {str(key): quantize_2(to_decimal(value)) for key, value in (market_rates.silver_rates_dkk or {}).items()}
    gold_24k_dkk = quantize_2(
        to_decimal(market_rates.gold_24k_dkk or gold_rates_dkk.get("24", Decimal("0.00")))
    )
    silver_dkk = quantize_2(
        to_decimal(market_rates.silver_dkk or silver_rates_dkk.get("999", Decimal("0.00")))
    )
    gold_matrix = [
        PosWorkspaceRateMatrixEntry(
            row_key=f"gold:{key}",
            label=f"{key}K",
            lodighed=lodighed,
            dkk_per_gram=gold_rates_dkk.get(key, Decimal("0.00")),
            karat=to_decimal(key),
            type_code="1",
        )
        for key, lodighed in (("8", "333"), ("14", "585"), ("18", "750"), ("21", "875"), ("21.6", "900"), ("22", "916"), ("24", "999"))
    ]
    plet_dkk = quantize_2(to_decimal(market_rates.plet_dkk))
    silver_matrix = [
        PosWorkspaceRateMatrixEntry(
            row_key=f"silver:{type_code}",
            label=label,
            lodighed=lodighed or "—",
            dkk_per_gram=silver_rates_dkk.get(lodighed, Decimal("0.00")) if lodighed else plet_dkk,
            karat=None,
            type_code=type_code,
        )
        for type_code, label, lodighed in (("2", "Finsølv", "999"), ("3", "Sterling sølv", "925"), ("4", "3 tårnet sølv", "830"), ("5", "Plet", ""))
    ]
    return PosWorkspaceMarketRates(
        eur_dkk_fx=fx,
        gold_rates_dkk=gold_rates_dkk,
        silver_rates_dkk=silver_rates_dkk,
        gold_24k_dkk=gold_24k_dkk,
        silver_dkk=silver_dkk,
        plet_dkk=plet_dkk,
        gold_bar_dkk=quantize_2(to_decimal(market_rates.gold_bar_dkk)),
        silver_bar_dkk=quantize_2(to_decimal(market_rates.silver_bar_dkk)),
        gold_matrix=gold_matrix,
        silver_matrix=silver_matrix,
    )


def _afg_calculator_rows_payload(calculators) -> dict[str, list[dict[str, Decimal | str | None]]]:
    raw = calculators.model_dump() if hasattr(calculators, "model_dump") else (calculators or {})
    return {
        "gold_rows": raw.get("gold_rows") if isinstance(raw.get("gold_rows"), list) else [],
        "silver_rows": raw.get("silver_rows") if isinstance(raw.get("silver_rows"), list) else [],
    }


def _apply_afg_market_rate_cells(vars_sheet, market_rates: PosWorkspaceMarketRates) -> None:
    fx = quantize_2(to_decimal(market_rates.eur_dkk_fx))
    vars_sheet["C10"] = fx
    vars_sheet["C4"] = quantize_2(to_decimal(market_rates.gold_24k_dkk))
    vars_sheet["C5"] = quantize_2(to_decimal(market_rates.silver_dkk))
    vars_sheet["C6"] = quantize_2(to_decimal(market_rates.silver_rates_dkk.get("925", 0)))
    vars_sheet["C7"] = quantize_2(to_decimal(market_rates.silver_rates_dkk.get("830", 0)))
    # 'Variable værdier' VLOOKUP tablosu ($A$4:$B$12): Plet adı ve bar satırları.
    vars_sheet["B8"] = "Plet"
    vars_sheet["A9"] = 6
    vars_sheet["B9"] = "Guldbarre"
    vars_sheet["A10"] = 7
    vars_sheet["B10"] = "Sølvbarre"
    vars_sheet["A11"] = 8
    vars_sheet["B11"] = "Platin"
    vars_sheet["A12"] = 9
    vars_sheet["B12"] = "Palladium"
    vars_sheet["G2"] = "Kategori"
    vars_sheet["H2"] = "Beskrivelse"
    # I sütunundaki eski EUR aynası artık yazılmaz; kanonik birim DKK/g (J).
    vars_sheet["I2"] = None
    vars_sheet["J2"] = "DKK/g"
    vars_sheet["K2"] = "Lødighed / Karat"
    values = (
        ("gold", "Gold 8K", market_rates.gold_rates_dkk.get("8"), "333"),
        ("gold", "Gold 14K", market_rates.gold_rates_dkk.get("14"), "585"),
        ("gold", "Gold 18K", market_rates.gold_rates_dkk.get("18"), "750"),
        ("gold", "Gold 21K", market_rates.gold_rates_dkk.get("21"), "875"),
        ("gold", "Gold 21.6K", market_rates.gold_rates_dkk.get("21.6"), "900"),
        ("gold", "Gold 22K", market_rates.gold_rates_dkk.get("22"), "916"),
        ("gold", "Gold 24K", market_rates.gold_rates_dkk.get("24"), "999"),
        ("silver", "Finsølv", market_rates.silver_rates_dkk.get("999"), "999"),
        ("silver", "Sterling sølv", market_rates.silver_rates_dkk.get("925"), "925"),
        ("silver", "3 tårnet sølv", market_rates.silver_rates_dkk.get("830"), "830"),
        ("silver", "Plet", market_rates.plet_dkk, ""),
        ("bar", "Guldbarre", market_rates.gold_bar_dkk, "999.9"),
        ("bar", "Sølvbarre", market_rates.silver_bar_dkk, "999"),
        ("ptpd", "Platin", market_rates.platinum_dkk, "950"),
        ("ptpd", "Palladium", market_rates.palladium_dkk, "500"),
    )
    for row, (kind, label, dkk_value, lodighed) in zip(AFG_VARIABLE_MATRIX_ROWS, values, strict=False):
        eur_cell, dkk_cell, meta_cell, _legacy_label, _legacy_lodighed = row
        row_idx = int(eur_cell[1:])
        vars_sheet[f"G{row_idx}"] = kind
        vars_sheet[f"H{row_idx}"] = label
        vars_sheet[eur_cell] = None
        vars_sheet[dkk_cell] = quantize_2(to_decimal(dkk_value))
        vars_sheet[meta_cell] = lodighed


def _apply_afg_calculator_cells(sheet, calculators) -> None:
    payload = _afg_calculator_rows_payload(calculators)
    by_key = {
        str(row.get("row_key") or "").strip(): row
        for row in payload["gold_rows"] + payload["silver_rows"]
        if isinstance(row, dict) and str(row.get("row_key") or "").strip()
    }
    for row_key, unit_cell, count_cell, _total_cell in (*AFG_GOLD_CALCULATOR_ROWS, *AFG_SILVER_CALCULATOR_ROWS):
        row = by_key.get(row_key, {})
        sheet[unit_cell] = quantize_2(to_decimal(row.get("unit_weight") or 0))
        sheet[count_cell] = quantize_2(to_decimal(row.get("count") or 0))


def _parse_afg_calculators_from_sheet(sheet) -> PosWorkspaceCalculatorsUpdate:
    gold_rows = []
    for row_key, unit_cell, count_cell, _total_cell in AFG_GOLD_CALCULATOR_ROWS:
        gold_rows.append(
            {
                "row_key": row_key,
                "unit_weight": _decimal_from_excel(sheet[unit_cell].value),
                "count": _decimal_from_excel(sheet[count_cell].value),
                "target_row_key": None,
            }
        )
    silver_rows = []
    for row_key, unit_cell, count_cell, _total_cell in AFG_SILVER_CALCULATOR_ROWS:
        silver_rows.append(
            {
                "row_key": row_key,
                "unit_weight": _decimal_from_excel(sheet[unit_cell].value),
                "count": _decimal_from_excel(sheet[count_cell].value),
                "target_row_key": None,
            }
        )
    target_defaults = {
        "calc_gold:1": "gold:8",
        "calc_gold:2": "gold:14",
        "calc_gold:3": "gold:18",
        "calc_gold:4": "gold:21",
        "calc_gold:5": "gold:21.6",
        "calc_silver:1": "silver:2",
        "calc_silver:2": "silver:3",
        "calc_silver:3": "silver:4",
        "calc_silver:4": "silver:5",
    }
    for row in gold_rows + silver_rows:
        row["target_row_key"] = target_defaults.get(str(row["row_key"]))
    return PosWorkspaceCalculatorsUpdate(gold_rows=gold_rows, silver_rows=silver_rows)


def _apply_afg_summary_cells(
    sheet,
    *,
    net_amount_dkk: Decimal,
    vat_amount_dkk: Decimal,
    gross_amount_dkk: Decimal,
    payment_method: str,
    reg_number: str | None,
    account_number: str | None,
    note: str | None,
) -> None:
    payment_label = "Kontant" if payment_method == "cash" else "Overførsel"
    sheet["C40"] = payment_label
    sheet["D40"] = gross_amount_dkk
    sheet["D41"] = "—" if payment_method == "cash" else (reg_number or "—")
    sheet["D42"] = "—" if payment_method == "cash" else (account_number or "—")
    sheet["H38"] = net_amount_dkk
    sheet["H40"] = net_amount_dkk
    # R2-15: KDV yokken (varsayılan alış akışı) belgede "Moms 0,00" satırı
    # GÖRÜNMEZ — G42 etiketi ve H42 tutarı boşaltılır. A42 sözleşme hücresi
    # (round-trip parse) her durumda yazılmaya devam eder. H43 kodda doğrudan
    # gross ile yazıldığı için SUM kırılmaz.
    sheet["A42"] = 1 if vat_amount_dkk > 0 else 0
    if vat_amount_dkk > 0:
        sheet["G42"] = "Moms"
        sheet["H42"] = vat_amount_dkk
    else:
        sheet["G42"] = None
        sheet["H42"] = None
    sheet["H43"] = gross_amount_dkk
    sheet["C44"] = "Not:"
    sheet["D44"] = note or None


def _write_afg_mer_pris(sheet, idx: int, avance_percent) -> None:
    """R2-07: B kolonu artık 'Mer pris' (kr/g) taşır — yüzde-fraksiyon DEĞİL.

    Değer HAM kr/g olarak yazılır (eski /100 bölmesi kaldırıldı) ve hücrenin
    yüzde numFmt'i kalıntısı '1500%' göstermesin diye sayı formatı sabitlenir.
    Negatif serbest (−15 → birim fiyattan düşer)."""
    cell = sheet[f"B{idx}"]
    cell.value = quantize_2(to_decimal(avance_percent))
    cell.number_format = "#,##0.00"


def _apply_afg_mer_pris_header(sheet) -> None:
    # Şablondaki B20 başlığı '↓' — belge kolonunu adlandır (R2-07/R2-15).
    sheet["B20"] = "Mer pris"


AFG_EXTRA_ROW = 37  # şablondaki tek boş grid satırı — kniv/çeyrek "i alt" agregatı


def _apply_afg_workspace_rows(
    sheet,
    gold_rows: Iterable[PosWorkspaceGoldRowOut],
    silver_rows: Iterable[PosWorkspaceSilverRowOut],
    bar_rows: Iterable["PosWorkspaceBarRowOut"] = (),
    ptpd_rows: Iterable["PosWorkspacePtPdRowOut"] = (),
    extra_rows: Iterable = (),
) -> None:
    _apply_afg_mer_pris_header(sheet)
    total_gold_grams = Decimal("0.00")
    total_gold_amount = Decimal("0.00")
    total_gold_pure = Decimal("0.00")
    total_silver_grams = Decimal("0.00")
    total_silver_amount = Decimal("0.00")
    total_silver_pure = Decimal("0.00")

    for idx, row in enumerate(gold_rows, start=AFG_GOLD_ROW_START):
        gram = quantize_2(to_decimal(row.gram))
        unit_price = quantize_2(to_decimal(row.unit_price_dkk))
        line_total = quantize_2(to_decimal(row.line_total_dkk))
        pure = quantize_2(gram * (to_decimal(row.purity_percentage) / Decimal("100")))
        total_gold_grams += gram
        total_gold_amount += line_total
        total_gold_pure += pure
        sheet[f"A{idx}"] = 1
        _write_afg_mer_pris(sheet, idx, row.avance_percent)
        sheet[f"C{idx}"] = "Guld"
        sheet[f"D{idx}"] = to_decimal(row.karat)
        sheet[f"E{idx}"] = row.lodighed
        sheet[f"F{idx}"] = gram
        sheet[f"G{idx}"] = unit_price
        sheet[f"H{idx}"] = line_total

    for idx, row in enumerate(silver_rows, start=AFG_SILVER_ROW_START):
        gram = quantize_2(to_decimal(row.gram))
        unit_price = quantize_2(to_decimal(row.unit_price_dkk))
        line_total = quantize_2(to_decimal(row.line_total_dkk))
        pure = quantize_2(gram * (to_decimal(row.purity_percentage) / Decimal("100")))
        total_silver_grams += gram
        total_silver_amount += line_total
        total_silver_pure += pure
        sheet[f"A{idx}"] = row.type_code
        _write_afg_mer_pris(sheet, idx, row.avance_percent)
        sheet[f"C{idx}"] = row.label
        sheet[f"E{idx}"] = row.lodighed if row.lodighed != "—" else None
        sheet[f"F{idx}"] = gram
        sheet[f"G{idx}"] = unit_price
        sheet[f"H{idx}"] = line_total

    for row in bar_rows:
        idx = AFG_BAR_GOLD_ROW if row.bar_type == "gold" else AFG_BAR_SILVER_ROW
        gram = quantize_2(to_decimal(row.gram))
        unit_price = quantize_2(to_decimal(row.unit_price_dkk))
        line_total = quantize_2(to_decimal(row.line_total_dkk))
        pure = quantize_2(gram * (to_decimal(row.purity_percentage) / Decimal("100")))
        if row.bar_type == "gold":
            total_gold_grams += gram
            total_gold_amount += line_total
            total_gold_pure += pure
        else:
            total_silver_grams += gram
            total_silver_amount += line_total
            total_silver_pure += pure
        sheet[f"A{idx}"] = 6 if row.bar_type == "gold" else 7
        _write_afg_mer_pris(sheet, idx, row.avance_percent)
        sheet[f"C{idx}"] = row.label
        sheet[f"D{idx}"] = Decimal("24") if row.bar_type == "gold" else None
        sheet[f"E{idx}"] = row.lodighed
        sheet[f"F{idx}"] = gram
        sheet[f"G{idx}"] = unit_price
        sheet[f"H{idx}"] = line_total

    for row in ptpd_rows:
        idx = AFG_PTPD_PLATINUM_ROW if row.metal == "platinum" else AFG_PTPD_PALLADIUM_ROW
        gram = quantize_2(to_decimal(row.gram))
        unit_price = quantize_2(to_decimal(row.unit_price_dkk))
        line_total = quantize_2(to_decimal(row.line_total_dkk))
        # Pt/Pd altın/gümüş toplam bloklarına (D75-I75) girmez; belge içi
        # SUM(F22:F37)/SUM(H22:H37) satırları zaten kapsar.
        sheet[f"A{idx}"] = 8 if row.metal == "platinum" else 9
        _write_afg_mer_pris(sheet, idx, row.avance_percent)
        sheet[f"C{idx}"] = row.label
        sheet[f"D{idx}"] = None
        sheet[f"E{idx}"] = row.lodighed
        sheet[f"F{idx}"] = gram
        sheet[f"G{idx}"] = unit_price
        sheet[f"H{idx}"] = line_total

    # R2-01 — dinamik kniv/çeyrek satırları: şablon gridinde tek boş satır (37)
    # olduğundan hepsi tek "i alt" agregatı olarak yazılır. Böylece belgedeki
    # satır toplamı, uygulamanın (extra dahil) özet toplamıyla UYUŞUR. Satır 37
    # round-trip parse'ında okunmaz — Excel'de bu satıra yapılan düzenleme
    # bilinçli olarak yok sayılır (kaynak: uygulamadaki hesaplayıcı).
    extra_list = list(extra_rows or [])
    if extra_list:
        extra_gram = Decimal("0.00")
        extra_total = Decimal("0.00")
        for row in extra_list:
            gram = quantize_2(to_decimal(row.gram))
            line_total = quantize_2(to_decimal(row.line_total_dkk))
            extra_gram += gram
            extra_total += line_total
            pure = quantize_2(gram * (to_decimal(row.purity_percentage) / Decimal("100")))
            if str(getattr(row, "metal", "")) == "gold":
                total_gold_grams += gram
                total_gold_amount += line_total
                total_gold_pure += pure
            else:
                total_silver_grams += gram
                total_silver_amount += line_total
                total_silver_pure += pure
        idx = AFG_EXTRA_ROW
        sheet[f"A{idx}"] = None
        sheet[f"B{idx}"] = None
        # Etiket gerçek bileşimi yansıtır: yalnız kniv → "Kniv", yalnız çeyrek →
        # "Møntguld" (22K-2 çeyrekler dahil), karışık → ikisi birden. R2-10 çeyrek
        # satırları Møntguld'dur; agregata bıçak etiketi yapışmasın.
        kinds = {str(getattr(row, "kind", "")) for row in extra_list}
        if kinds == {"kniv"}:
            extra_label = "Kniv (i alt)"
        elif kinds == {"quarter"}:
            extra_label = "Møntguld (i alt)"
        else:
            extra_label = "Kniv / Møntguld (i alt)"
        sheet[f"C{idx}"] = extra_label
        sheet[f"D{idx}"] = None
        sheet[f"E{idx}"] = None
        sheet[f"F{idx}"] = quantize_2(extra_gram)
        sheet[f"G{idx}"] = None
        sheet[f"H{idx}"] = quantize_2(extra_total)

    sheet["D75"] = total_gold_grams
    sheet["E75"] = total_gold_amount
    sheet["F75"] = total_gold_pure
    sheet["G75"] = total_silver_grams
    sheet["H75"] = total_silver_amount
    sheet["I75"] = total_silver_pure


def _apply_afg_factura_gold_sheet(
    sheet,
    *,
    customer_name: str | None,
    issued_at: datetime | None,
    invoice_number: str | None,
    rows,
    footer_lines: list[str],
    vat_enabled: bool,
    note: str | None,
) -> None:
    sheet["C2"] = _excel_datetime(issued_at)
    sheet["B11"] = customer_name or "—"
    sheet["F15"] = invoice_number or "—"
    sheet["F16"] = _excel_datetime(issued_at)
    for row_idx in range(AFG_FACTURA_GOLD_ROW_START, AFG_FACTURA_GOLD_ROW_END + 1):
        sheet[f"A{row_idx}"] = None
        sheet[f"C{row_idx}"] = None
        sheet[f"E{row_idx}"] = None
        sheet[f"F{row_idx}"] = None
        sheet[f"G{row_idx}"] = None
    for row in rows:
        row_index = int(str(row.row_key).split(":", 1)[1]) + AFG_FACTURA_GOLD_ROW_START - 1
        if row_index < AFG_FACTURA_GOLD_ROW_START or row_index > AFG_FACTURA_GOLD_ROW_END:
            continue
        sheet[f"A{row_index}"] = row.code or None
        sheet[f"C{row_index}"] = row.fineness or None
        sheet[f"E{row_index}"] = quantize_2(to_decimal(row.gram)) if to_decimal(row.gram) > 0 else None
        sheet[f"F{row_index}"] = quantize_2(to_decimal(row.unit_price_dkk)) if to_decimal(row.unit_price_dkk) > 0 else None
        sheet[f"G{row_index}"] = quantize_2(to_decimal(row.line_total_dkk)) if to_decimal(row.line_total_dkk) > 0 else None
    for index, row_idx in enumerate(range(AFG_FACTURA_GOLD_FOOTER_START, AFG_FACTURA_GOLD_FOOTER_START + 3)):
        sheet[f"B{row_idx}"] = footer_lines[index] if index < len(footer_lines) and footer_lines[index] else None
    sheet["A49"] = "Not:" if note else None
    sheet["B49"] = note or None
    sheet["A46"] = 1 if vat_enabled else 0


def _apply_afg_factura_misc_sheet(
    sheet,
    *,
    issued_at: datetime | None,
    invoice_number: str | None,
    rows,
    vat_enabled: bool,
    note: str | None,
) -> None:
    sheet["C2"] = _excel_datetime(issued_at)
    sheet["F14"] = invoice_number or "—"
    sheet["F15"] = _excel_datetime(issued_at)
    for row_idx in range(AFG_FACTURA_MISC_ROW_START, AFG_FACTURA_MISC_ROW_END + 1):
        sheet[f"C{row_idx}"] = None
        sheet[f"E{row_idx}"] = None
        sheet[f"F{row_idx}"] = None
    for row in rows:
        row_index = int(str(row.row_key).split(":", 1)[1]) + AFG_FACTURA_MISC_ROW_START - 1
        if row_index < AFG_FACTURA_MISC_ROW_START or row_index > AFG_FACTURA_MISC_ROW_END:
            continue
        sheet[f"C{row_index}"] = row.text or None
        sheet[f"E{row_index}"] = quantize_2(to_decimal(row.quantity)) if row.quantity is not None else None
        sheet[f"F{row_index}"] = quantize_2(to_decimal(row.unit_price_dkk)) if to_decimal(row.unit_price_dkk) > 0 else None
    sheet["A45"] = 1 if vat_enabled else 0
    sheet["C48"] = f"Not: {note}" if note else None


AFG_DETAIL_LAST_METAL_ROW = 37
# Gizli hesap bloğu (74-88) metal kimliğiyle sabitlenmiştir; kompakt düzende
# görünür satır pozisyonuna güvenen formüller (E79=+F22 vb.) backend'den
# literal yazılır.
_AFG_HIDDEN_GOLD_KARATS = ("8", "9", "10", "14", "18", "21", "21.6", "22", "24")
_AFG_HIDDEN_GOLD_FACTORS = {
    "8": Decimal("0.333"),
    "9": Decimal("0.375"),
    "10": Decimal("0.416"),
    "14": Decimal("0.585"),
    "18": Decimal("0.75"),
    "21": Decimal("0.875"),
    "21.6": Decimal("0.9"),
    "22": Decimal("0.916"),
    "24": Decimal("0.999"),
}


def _apply_afg_detail_rows(
    sheet,
    gold_rows: Iterable[dict[str, Decimal | str]],
    silver_rows: Iterable[dict[str, Decimal | str]],
    bar_rows: Iterable[dict[str, Decimal | str]] = (),
    ptpd_rows: Iterable[dict[str, Decimal | str]] = (),
) -> None:
    """Nihai AFG kompakt yazılır: yalnız dolu satırlar, 22'den ardışık.

    Tek 14K alış → tek metal satırı; kullanılmayan satırlar (şablonun
    G30/G31/G32 hayalet literalleri dahil) 37'ye kadar temizlenir. Sıra:
    altın 8→24, Guldbarre, Sølvbarre, gümüş Finsølv→Plet, Platin, Palladium.
    """
    _apply_afg_mer_pris_header(sheet)
    total_gold_grams = Decimal("0.00")
    total_gold_amount = Decimal("0.00")
    total_gold_pure = Decimal("0.00")
    total_silver_grams = Decimal("0.00")
    total_silver_amount = Decimal("0.00")
    total_silver_pure = Decimal("0.00")
    hidden_gold_grams: dict[str, Decimal] = {key: Decimal("0.00") for key in _AFG_HIDDEN_GOLD_KARATS}
    hidden_silver_grams: dict[str, Decimal] = {"999": Decimal("0.00"), "925": Decimal("0.00"), "830": Decimal("0.00")}

    filled: list[dict[str, Decimal | str | int]] = []
    for row in gold_rows:
        gram = quantize_2(to_decimal(row["gram"]))
        if gram <= 0:
            continue
        karat = to_decimal(row["karat"])
        karat_key = str(row["karat"]).rstrip("0").rstrip(".") if "." in str(row["karat"]) else str(row["karat"])
        line_total = quantize_2(to_decimal(row["line_total_dkk"]))
        pure = quantize_2(gram * (karat / Decimal("24")))
        total_gold_grams += gram
        total_gold_amount += line_total
        total_gold_pure += pure
        if karat_key in hidden_gold_grams:
            hidden_gold_grams[karat_key] += gram
        filled.append(
            {
                "type_code": 1,
                "avance": quantize_2(to_decimal(row["avance_percent"])),
                "label": "Guld",
                "karat": karat,
                "lodighed": row["lodighed"],
                "gram": gram,
                "unit_price": quantize_2(to_decimal(row["unit_price_dkk"])),
                "line_total": line_total,
            }
        )
    for row in bar_rows:
        gram = quantize_2(to_decimal(row["gram"]))
        if gram <= 0:
            continue
        line_total = quantize_2(to_decimal(row["line_total_dkk"]))
        is_gold = str(row.get("bar_type")) == "gold"
        pure_factor = Decimal("0.9999") if is_gold else Decimal("0.999")
        if is_gold:
            total_gold_grams += gram
            total_gold_amount += line_total
            total_gold_pure += quantize_2(gram * pure_factor)
        else:
            total_silver_grams += gram
            total_silver_amount += line_total
            total_silver_pure += quantize_2(gram * pure_factor)
        filled.append(
            {
                "type_code": 6 if is_gold else 7,
                "avance": quantize_2(to_decimal(row.get("avance_percent") or 0)),
                "label": "Guldbarre" if is_gold else "Sølvbarre",
                "karat": Decimal("24") if is_gold else None,
                "lodighed": "999.9" if is_gold else "999",
                "gram": gram,
                "unit_price": quantize_2(to_decimal(row["unit_price_dkk"])),
                "line_total": line_total,
            }
        )
    for row in silver_rows:
        gram = quantize_2(to_decimal(row["gram"]))
        if gram <= 0:
            continue
        line_total = quantize_2(to_decimal(row["line_total_dkk"]))
        lodighed = str(row["lodighed"])
        pure = quantize_2(gram * (Decimal(lodighed) / Decimal("1000"))) if lodighed.isdigit() else Decimal("0.00")
        total_silver_grams += gram
        total_silver_amount += line_total
        total_silver_pure += pure
        if lodighed in hidden_silver_grams:
            hidden_silver_grams[lodighed] += gram
        filled.append(
            {
                "type_code": row["type_code"],
                "avance": quantize_2(to_decimal(row["avance_percent"])),
                "label": row["label"],
                "karat": None,
                "lodighed": lodighed if lodighed.isdigit() else None,
                "gram": gram,
                "unit_price": quantize_2(to_decimal(row["unit_price_dkk"])),
                "line_total": line_total,
            }
        )

    for row in ptpd_rows:
        gram = quantize_2(to_decimal(row["gram"]))
        if gram <= 0:
            continue
        line_total = quantize_2(to_decimal(row["line_total_dkk"]))
        is_platinum = str(row.get("metal")) == "platinum"
        # Pt/Pd altın/gümüş toplam ve gizli saf-metal bloklarına girmez;
        # belge toplamı SUM satırlarından gelir.
        filled.append(
            {
                "type_code": 8 if is_platinum else 9,
                "avance": quantize_2(to_decimal(row.get("avance_percent") or 0)),
                "label": "Platin" if is_platinum else "Palladium",
                "karat": None,
                "lodighed": "950" if is_platinum else "500",
                "gram": gram,
                "unit_price": quantize_2(to_decimal(row["unit_price_dkk"])),
                "line_total": line_total,
            }
        )

    idx = AFG_GOLD_ROW_START
    for entry in filled:
        sheet[f"A{idx}"] = entry["type_code"]
        _write_afg_mer_pris(sheet, idx, entry["avance"])
        sheet[f"C{idx}"] = entry["label"]
        sheet[f"D{idx}"] = entry["karat"]
        sheet[f"E{idx}"] = entry["lodighed"]
        sheet[f"F{idx}"] = entry["gram"]
        sheet[f"G{idx}"] = entry["unit_price"]
        sheet[f"H{idx}"] = entry["line_total"]
        idx += 1
    for blank_idx in range(idx, AFG_DETAIL_LAST_METAL_ROW + 1):
        for column in "ABCDEFGH":
            sheet[f"{column}{blank_idx}"] = None
    # Şablon hesaplayıcı hayaletleri (J22=10, J32=1.754 vb.) final belgeye
    # taşınmaz: detail yolu hesaplayıcı verisi yazmaz.
    for calc_idx in range(AFG_GOLD_ROW_START, AFG_DETAIL_LAST_METAL_ROW + 1):
        for column in "JKL":
            sheet[f"{column}{calc_idx}"] = None

    sheet["D75"] = total_gold_grams
    sheet["E75"] = total_gold_amount
    sheet["F75"] = total_gold_pure
    sheet["G75"] = total_silver_grams
    sheet["H75"] = total_silver_amount
    sheet["I75"] = total_silver_pure
    fine_gold_total = Decimal("0.00")
    for offset, karat_key in enumerate(_AFG_HIDDEN_GOLD_KARATS):
        row_idx = 79 + offset
        gram = hidden_gold_grams[karat_key]
        fine = quantize_2(gram * _AFG_HIDDEN_GOLD_FACTORS[karat_key])
        fine_gold_total += fine
        sheet[f"E{row_idx}"] = gram
        sheet[f"F{row_idx}"] = fine
    fine_silver_total = Decimal("0.00")
    for offset, lodighed in enumerate(("999", "925", "830")):
        row_idx = 79 + offset
        gram = hidden_silver_grams[lodighed]
        fine = quantize_2(gram * (Decimal(lodighed) / Decimal("1000")))
        fine_silver_total += fine
        sheet[f"I{row_idx}"] = gram
        sheet[f"J{row_idx}"] = fine
    sheet["J82"] = fine_silver_total
    sheet["F88"] = fine_gold_total


def _afg_cell_amount(sheet, cell_ref: str) -> Decimal:
    cell = sheet[cell_ref]
    if cell.data_type == "f":  # formül hücresi görünürlük sinyali taşımaz
        return Decimal("0.00")
    return quantize_2(to_decimal(cell.value or 0))


def _afg_reset_row_geometry(sheet, idx: int) -> None:
    """İnce ayraç satırının (29: ht=7.5, thickBot) ölçüsünü normal grid
    satırı ölçüsüne döndürür — Guldbarre doluyken karat satırları gibi dursun.
    customHeight height'tan türetilir (salt-okunur), ayrıca set edilmez."""
    sibling = sheet.row_dimensions.get(idx - 1)
    dim = sheet.row_dimensions[idx]
    dim.height = sibling.height if sibling is not None else None
    dim.thickTop = getattr(sibling, "thickTop", None) if sibling is not None else None
    dim.thickBot = getattr(sibling, "thickBot", None) if sibling is not None else None


def _afg_calc_cells_filled(sheet, idx: int) -> bool:
    return any(_afg_cell_amount(sheet, f"{column}{idx}") > 0 for column in "JKL")


def _apply_afg_row_visibility(sheet, *, blank_splitter_visible: bool = True) -> None:
    """AFG grid satırlarını (21-37) veriye göre gizler/açar; her build'in
    sonunda (satır + hesaplayıcı yazımından sonra) çağrılır.

    Görünürlük yalnız gramdan (F kolonu) türetilir — hesaplayıcı (J/K)
    varsayılanları (kniv ağırlıkları) görünürlüğü zorlamaz, aksi halde tüm
    karat satırları her taslakta açık kalırdı. Satır 29 çift rol oynar:
    doluyken Guldbarre sarı karat satırı, boşken altın bloğunu gümüş/pt/pd
    bloğundan ayıran ince mavi şerit (yalnız iki taraf da doluyken görünür).
    Final yolu (blank_splitter_visible=False) satırları 22'den paketler;
    boş kalan şerit verinin altında başıboş görünmemesi için gizlenir.
    """
    gold_filled = {
        idx: _afg_cell_amount(sheet, f"F{idx}") > 0
        for idx in range(AFG_GOLD_ROW_START, AFG_BAR_GOLD_ROW)
    }
    silver_filled = {
        idx: _afg_cell_amount(sheet, f"F{idx}") > 0
        for idx in range(AFG_SILVER_ROW_START, AFG_EXTRA_ROW + 1)
    }
    if not silver_filled[AFG_EXTRA_ROW]:  # agregat: gram 0 ama toplam yazılmış olabilir
        silver_filled[AFG_EXTRA_ROW] = _afg_cell_amount(sheet, f"H{AFG_EXTRA_ROW}") > 0

    row_dims = sheet.row_dimensions
    # 21: kniv hesaplayıcı tablo başlığı — tablosu (22-26) tamamen boşsa gizli:
    # satır görünür olsa bile J/K/L içeriği yoksa başlık başıboş kalır.
    row_dims[21].hidden = not any(
        gold_filled[idx] and _afg_calc_cells_filled(sheet, idx)
        for idx in range(AFG_GOLD_ROW_START, AFG_GOLD_ROW_START + 5)
    )
    for idx in range(AFG_GOLD_ROW_START, AFG_BAR_GOLD_ROW):  # 22-28 karat satırları
        row_dims[idx].hidden = not gold_filled[idx]

    bar_gold_filled = (
        to_decimal(sheet[f"A{AFG_BAR_GOLD_ROW}"].value or 0) == 6
        and _afg_cell_amount(sheet, f"F{AFG_BAR_GOLD_ROW}") > 0
    )
    bar_row_has_data = _afg_cell_amount(sheet, f"F{AFG_BAR_GOLD_ROW}") > 0
    gold_side = any(gold_filled.values()) or bar_gold_filled
    silver_side = any(silver_filled.values())
    if bar_row_has_data:
        # Guldbarre (veya final yolunda paketlenmiş satır) → normal satır görünümü.
        row_dims[AFG_BAR_GOLD_ROW].hidden = False
        _afg_reset_row_geometry(sheet, AFG_BAR_GOLD_ROW)
        if bar_gold_filled:
            gold_fill = PatternFill(fill_type="solid", fgColor=AFG_GOLD_FILL_ARGB)
            for column in "CDEFGH":
                sheet[f"{column}{AFG_BAR_GOLD_ROW}"].fill = gold_fill
    else:
        # Boş şerit: yazı artıklarını temizle ("Guldbarre 24 0,00" kalmasın),
        # yalnız iki taraf da doluyken ayraç olarak göster.
        for column in "ACDEFGH":
            sheet[f"{column}{AFG_BAR_GOLD_ROW}"] = None
        row_dims[AFG_BAR_GOLD_ROW].hidden = not (
            blank_splitter_visible and gold_side and silver_side
        )

    # 31: Sterling satırı VE gümüş hesaplayıcı tablo başlığı — tablosu (32-37)
    # hesaplayıcı içeriğiyle görünürken başlık olarak da açık kalır.
    silver_calc_block = any(
        silver_filled[idx] and _afg_calc_cells_filled(sheet, idx)
        for idx in range(32, AFG_EXTRA_ROW + 1)
    )
    for idx in range(AFG_SILVER_ROW_START, AFG_EXTRA_ROW + 1):  # 30-37
        if idx == AFG_SILVER_ROW_START + 1:
            row_dims[idx].hidden = not (silver_filled[idx] or silver_calc_block)
        else:
            row_dims[idx].hidden = not silver_filled[idx]


def _build_afg_workbook_bytes_from_workspace(workspace: PosWorkspaceOut, *, sync_context: ArtifactSyncContext) -> bytes:
    from app.services import document_artifact_afg

    return document_artifact_afg.build_afg_workbook_bytes_from_workspace(workspace, sync_context=sync_context)


def _build_afg_workbook_bytes_from_detail(detail: PosDocumentDetailOut, *, sync_context: ArtifactSyncContext) -> bytes:
    from app.services import document_artifact_afg

    return document_artifact_afg.build_afg_workbook_bytes_from_detail(detail, sync_context=sync_context)


def _workspace_normalized_afg_values(workspace: PosWorkspaceOut) -> dict[str, str]:
    from app.services import document_artifact_afg

    return document_artifact_afg.workspace_normalized_afg_values(workspace)


def parse_afg_workspace_inputs_from_workbook(
    content: bytes, *, legacy_percent_avance: bool = False
) -> AfgWorkspaceArtifactInputs:
    from app.services import document_artifact_afg

    return document_artifact_afg.parse_afg_workspace_inputs_from_workbook(
        content, legacy_percent_avance=legacy_percent_avance
    )


def build_afg_workspace_reconcile_preview(
    workspace: PosWorkspaceOut,
    parsed: AfgWorkspaceArtifactInputs,
) -> DocumentArtifactReconcilePreviewOut:
    from app.services import document_artifact_afg

    return document_artifact_afg.build_afg_workspace_reconcile_preview(workspace, parsed)


def build_inventory_reconcile_preview(
    workspace: InventoryWorkspaceOut,
    parsed: InventoryWorkbookArtifactInputs,
) -> DocumentArtifactReconcilePreviewOut:
    return document_artifact_inventory.build_inventory_reconcile_preview(workspace, parsed)


def _inventory_sync_mappings(workspace: InventoryWorkspaceOut) -> list[SyncSheetRowMapping]:
    return document_artifact_inventory._inventory_sync_mappings(workspace)


def _inventory_editable_refs(mappings: Iterable[SyncSheetRowMapping]) -> list[str]:
    return document_artifact_inventory._inventory_editable_refs(mappings)


def _inventory_raw_row_payload(sheet, row_idx: int) -> tuple[datetime | None, dict[str, str]]:
    return document_artifact_inventory._inventory_raw_row_payload(sheet, row_idx)


def _inventory_raw_product_updates(
    sheet,
    *,
    current_workspace: InventoryWorkspaceOut,
) -> dict[UUID, ProductUpdate]:
    return document_artifact_inventory._inventory_raw_product_updates(
        sheet,
        current_workspace=current_workspace,
    )


def _build_inventory_workbook_bytes(
    workspace: InventoryWorkspaceOut,
    *,
    sync_context: ArtifactSyncContext,
    display_updated_at: datetime,
) -> bytes:
    return document_artifact_inventory._build_inventory_workbook_bytes(
        workspace,
        sync_context=sync_context,
        display_updated_at=display_updated_at,
    )


def _log_default_classification(metal_type: str | None, classification: str | None) -> AfgClassification:
    return document_artifact_log._log_default_classification(metal_type, classification)


def _log_route_code(destination: str | None, classification: str | None, *, metal_type: str | None = None) -> str:
    return document_artifact_log._log_route_code(destination, classification, metal_type=metal_type)


def _normalize_log_route_code(value: object) -> str:
    return document_artifact_log._normalize_log_route_code(value)


def _log_route_request_from_code(
    *,
    route_code: str,
    current_classification: str | None,
    metal_type: str | None,
    note: str | None,
    line_id: UUID,
) -> AfgRouteRequest:
    return document_artifact_log._log_route_request_from_code(
        route_code=route_code,
        current_classification=current_classification,
        metal_type=metal_type,
        note=note,
        line_id=line_id,
    )


def _log_sheet_lines(workspace: AfgLogWorkspaceOut, metal_bucket: str) -> list[dict[str, Any]]:
    return document_artifact_log._log_sheet_lines(workspace, metal_bucket)


def _log_route_line_rows(workspace: AfgLogWorkspaceOut) -> list[dict[str, Any]]:
    return document_artifact_log._log_route_line_rows(workspace)


def _log_bucket_group_key(row: dict[str, Any]) -> str | None:
    return document_artifact_log._log_bucket_group_key(row)


def _log_active_lot_slot(workspace: AfgLogWorkspaceOut, metal_bucket: str) -> dict[str, Any]:
    return document_artifact_log._log_active_lot_slot(workspace, metal_bucket)


def _log_lot_rows(workspace: AfgLogWorkspaceOut) -> list[dict[str, Any]]:
    return document_artifact_log._log_lot_rows(workspace)


def _set_log_lot_section(sheet, *, metal_bucket: str, row: dict[str, Any], editable_refs: list[str], mappings: list[SyncSheetRowMapping]) -> None:
    document_artifact_log._set_log_lot_section(
        sheet,
        metal_bucket=metal_bucket,
        row=row,
        editable_refs=editable_refs,
        mappings=mappings,
    )


def _add_log_route_validation(sheet, refs: Iterable[str]) -> None:
    document_artifact_log._add_log_route_validation(sheet, refs)


def _build_log_workbook_bytes(workspace: AfgLogWorkspaceOut, *, year: int, sync_context: ArtifactSyncContext) -> bytes:
    return document_artifact_log._build_log_workbook_bytes(workspace, year=year, sync_context=sync_context)


def _datetime_from_excel(value: object) -> datetime | None:
    return document_artifact_log._datetime_from_excel(value)


def parse_inventory_workbook_inputs_from_workbook(
    content: bytes,
    *,
    current_workspace: InventoryWorkspaceOut | None = None,
) -> InventoryWorkbookArtifactInputs:
    return document_artifact_inventory.parse_inventory_workbook_inputs_from_workbook(
        content,
        current_workspace=current_workspace,
    )


def _parse_log_workbook_inputs_v1(workbook, metadata: SyncSheetMetadata) -> LogWorkbookArtifactInputs:
    return document_artifact_log._parse_log_workbook_inputs_v1(workbook, metadata)


def _parse_log_lot_section(sheet, *, metal_bucket: str, row_idx: int, extra: dict[str, Any]) -> tuple[LogWorkbookLotCreate | None, LogWorkbookLotUpdate | None]:
    return document_artifact_log._parse_log_lot_section(
        sheet,
        metal_bucket=metal_bucket,
        row_idx=row_idx,
        extra=extra,
    )


def _parse_log_workbook_inputs_v2(workbook, metadata: SyncSheetMetadata) -> LogWorkbookArtifactInputs:
    return document_artifact_log._parse_log_workbook_inputs_v2(workbook, metadata)


def _log_raw_reference_route_code(
    *,
    document: AfgWorkspaceDocumentOut,
    metal_bucket: str,
    inventory_marked: bool,
    separate_storage_refs: set[int],
) -> str:
    return document_artifact_log._log_raw_reference_route_code(
        document=document,
        metal_bucket=metal_bucket,
        inventory_marked=inventory_marked,
        separate_storage_refs=separate_storage_refs,
    )


def _raw_log_route_request_for_document(
    *,
    document: AfgWorkspaceDocumentOut,
    route_code: str,
    metal_bucket: str,
) -> AfgRouteRequest:
    return document_artifact_log._raw_log_route_request_for_document(
        document=document,
        route_code=route_code,
        metal_bucket=metal_bucket,
    )


def _parse_raw_log_reference_route_updates(
    sheet,
    *,
    current_workspace: AfgLogWorkspaceOut,
) -> list[LogWorkbookRouteEdit]:
    return document_artifact_log._parse_raw_log_reference_route_updates(
        sheet,
        current_workspace=current_workspace,
    )


def _parse_raw_log_lot_inputs(
    sheet,
    *,
    current_workspace: AfgLogWorkspaceOut,
) -> tuple[list[LogWorkbookLotCreate], list[LogWorkbookLotUpdate]]:
    return document_artifact_log._parse_raw_log_lot_inputs(
        sheet,
        current_workspace=current_workspace,
    )


def _parse_log_workbook_inputs_raw(
    workbook,
    *,
    current_workspace: AfgLogWorkspaceOut,
) -> LogWorkbookArtifactInputs:
    return document_artifact_log._parse_log_workbook_inputs_raw(
        workbook,
        current_workspace=current_workspace,
    )


def parse_log_workbook_inputs_from_workbook(
    content: bytes,
    *,
    year: int,
    current_workspace: AfgLogWorkspaceOut | None = None,
) -> LogWorkbookArtifactInputs:
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    metadata = _sync_metadata_if_present(workbook, expected_kind="log", expected_key=str(year))
    if metadata is not None:
        if metadata.contract_version == "log-v1":
            return _parse_log_workbook_inputs_v1(workbook, metadata)
        return _parse_log_workbook_inputs_v2(workbook, metadata)
    if current_workspace is None:
        raise ValueError("Log raw import için mevcut workspace gerekir")
    values_workbook = load_workbook(io.BytesIO(content), data_only=True)
    return _parse_log_workbook_inputs_raw(values_workbook, current_workspace=current_workspace)


def _afg_preview_sheet_from_workspace(workspace: PosWorkspaceOut) -> DocumentArtifactSheetPreviewOut:
    rows: list[list[str]] = []
    for row in _afg_gold_rows_from_workspace(workspace):
        rows.append(
            [
                row.row_key,
                "Guld",
                row.label,
                _stringify(row.avance_percent),
                _stringify(row.karat),
                _stringify(row.lodighed),
                _stringify(row.gram),
                _stringify(row.unit_price_dkk),
                _stringify(row.line_total_dkk),
            ]
        )
    rows.append(["", "SØLV — GÜMÜŞ", "999 · 925 · 830 · 800", "", "", "", "", "", ""])
    for row in _afg_silver_rows_from_workspace(workspace):
        rows.append(
            [
                row.row_key,
                row.type_code,
                row.label,
                _stringify(row.avance_percent),
                "—",
                _stringify(row.lodighed),
                _stringify(row.gram),
                _stringify(row.unit_price_dkk),
                _stringify(row.line_total_dkk),
            ]
        )
    return DocumentArtifactSheetPreviewOut(
        name="Afregningsbilag",
        mode="editable",
        system_sync=True,
        columns=["#", "Type", "Açıklama", "Avance %", "Karat", "Lødighed", "Vægt i g", "Enhedspris/g", "I alt (DKK)"],
        rows=rows,
        note=f"Toplam: {_fmt_decimal(workspace.summary.total_amount_dkk)} DKK",
    )


def _afg_variables_preview_sheet_from_workspace(workspace: PosWorkspaceOut) -> DocumentArtifactSheetPreviewOut:
    return _afg_contract_sheet(
        AFG_VARIABLES_SHEET,
        mode="editable",
        system_sync=True,
        note="Kontrollü piyasa girdileri. Save/apply sırasında workspace market rate alanlarına yazılır.",
        columns=["Hücre", "Açıklama", "Değer"],
        rows=[
            ["C4", "Au 24K DKK", _fmt_decimal(workspace.market_rates.gold_24k_dkk)],
            ["C5", "Ag DKK", _fmt_decimal(workspace.market_rates.silver_dkk)],
            ["C6", "Ag 925", _fmt_decimal(to_decimal(workspace.market_rates.silver_dkk) * Decimal("0.925"))],
            ["C7", "Ag 830", _fmt_decimal(to_decimal(workspace.market_rates.silver_dkk) * Decimal("0.83"))],
            ["C14", "Afregningsnr.", workspace.numbering_preview.afregnings_number_next],
            ["D14", "Fakturanr.", workspace.numbering_preview.invoice_number_next],
        ],
    )


def _afg_factura_gold_preview_sheet_from_workspace(workspace: PosWorkspaceOut) -> DocumentArtifactSheetPreviewOut:
    rows = [
        [
            row.row_key,
            row.code or "—",
            row.label or "—",
            row.fineness or "—",
            row.lodighed or "—",
            _fmt_decimal(row.gram),
            _fmt_decimal(row.unit_price_dkk),
            _fmt_decimal(row.line_total_dkk),
        ]
        for row in workspace.invoice_gold.rows
    ]
    footer_rows = [[f"Metin {index}", line or "—", "", "", "", "", "", ""] for index, line in enumerate(workspace.invoice_gold.footer_lines, start=1)]
    return _afg_contract_sheet(
        AFG_FACTURA_GOLD_SHEET,
        mode="editable",
        system_sync=True,
        note=f"Companion invoice sheet. Fakturanr: {workspace.numbering_preview.invoice_number_next or '—'}",
        columns=["#", "Kod", "Type", "Finhed", "Lødighed", "Gram", "Enhedspris / g", "I alt"],
        rows=[*rows, *footer_rows],
    )


def _afg_factura_misc_preview_sheet_from_workspace(workspace: PosWorkspaceOut) -> DocumentArtifactSheetPreviewOut:
    return _afg_contract_sheet(
        AFG_FACTURA_MISC_SHEET,
        mode="editable",
        system_sync=True,
        note=f"Companion misc invoice sheet. Fakturanr: {workspace.numbering_preview.invoice_number_next or '—'}",
        columns=["#", "Tekst", "Antal", "Pris", "I alt"],
        rows=[
            [
                row.row_key,
                row.text or "—",
                (_fmt_decimal(row.quantity) if row.quantity is not None else "—"),
                _fmt_decimal(row.unit_price_dkk),
                _fmt_decimal(row.line_total_dkk),
            ]
            for row in workspace.invoice_misc.rows
        ],
    )


def _afg_preview_sheet_from_detail(detail: PosDocumentDetailOut) -> DocumentArtifactSheetPreviewOut:
    rows: list[list[str]] = []
    for row in _aggregate_detail_gold_rows(detail):
        rows.append(
            [
                str(row["row_key"]),
                "Guld",
                "Guld",
                _stringify(row["avance_percent"]),
                _stringify(row["karat"]),
                _stringify(row["lodighed"]),
                _stringify(row["gram"]),
                _stringify(row["unit_price_dkk"]),
                _stringify(row["line_total_dkk"]),
            ]
        )
    rows.append(["", "SØLV — GÜMÜŞ", "999 · 925 · 830 · 800", "", "", "", "", "", ""])
    for row in _aggregate_detail_silver_rows(detail):
        rows.append(
            [
                str(row["row_key"]),
                _stringify(row["type_code"]),
                _stringify(row["label"]),
                _stringify(row["avance_percent"]),
                "—",
                _stringify(row["lodighed"]),
                _stringify(row["gram"]),
                _stringify(row["unit_price_dkk"]),
                _stringify(row["line_total_dkk"]),
            ]
        )
    return DocumentArtifactSheetPreviewOut(
        name="Afregningsbilag",
        mode="readonly",
        system_sync=False,
        columns=["#", "Type", "Açıklama", "Avance %", "Karat", "Lødighed", "Vægt i g", "Enhedspris/g", "I alt (DKK)"],
        rows=rows,
        note=f"Toplam: {_fmt_decimal(detail.net_amount_dkk)} DKK",
    )


def _afg_variables_preview_sheet_from_detail(detail: PosDocumentDetailOut) -> DocumentArtifactSheetPreviewOut:
    gold_rate = max(
        (to_decimal(line.rate_dkk or 0) for line in detail.lines if str(line.metal_type or "").lower() != "silver"),
        default=Decimal("0"),
    )
    silver_rate = max(
        (to_decimal(line.rate_dkk or 0) for line in detail.lines if str(line.metal_type or "").lower() == "silver"),
        default=Decimal("0"),
    )
    return _afg_contract_sheet(
        AFG_VARIABLES_SHEET,
        mode="readonly",
        system_sync=False,
        note="Final belgede Variable værdier korunur ama sistem tarafında tekrar yazılamaz.",
        columns=["Hücre", "Açıklama", "Değer"],
        rows=[
            ["C4", "Au 24K DKK", _fmt_decimal(gold_rate)],
            ["C5", "Ag DKK", _fmt_decimal(silver_rate)],
            ["C6", "Ag 925", _fmt_decimal(silver_rate * Decimal("0.925"))],
            ["C7", "Ag 830", _fmt_decimal(silver_rate * Decimal("0.83"))],
            ["C14", "Afregningsnr.", detail.numbering_preview.afregnings_number_next or detail.document_number],
            ["D14", "Fakturanr.", detail.numbering_preview.invoice_number_next or "—"],
        ],
    )


def _afg_factura_gold_preview_sheet_from_detail(detail: PosDocumentDetailOut) -> DocumentArtifactSheetPreviewOut:
    rows = [
        [
            row.row_key,
            row.code or "—",
            row.label or "—",
            row.fineness or "—",
            row.lodighed or "—",
            _fmt_decimal(row.gram),
            _fmt_decimal(row.unit_price_dkk),
            _fmt_decimal(row.line_total_dkk),
        ]
        for row in detail.invoice_gold.rows
    ]
    footer_rows = [[f"Metin {index}", line or "—", "", "", "", "", "", ""] for index, line in enumerate(detail.invoice_gold.footer_lines, start=1)]
    return _afg_contract_sheet(
        AFG_FACTURA_GOLD_SHEET,
        mode="readonly",
        system_sync=False,
        note=f"Final companion invoice sheet. Fakturanr: {detail.numbering_preview.invoice_number_next or '—'}",
        columns=["#", "Kod", "Type", "Finhed", "Lødighed", "Gram", "Enhedspris / g", "I alt"],
        rows=[*rows, *footer_rows],
    )


def _afg_factura_misc_preview_sheet_from_detail(detail: PosDocumentDetailOut) -> DocumentArtifactSheetPreviewOut:
    return _afg_contract_sheet(
        AFG_FACTURA_MISC_SHEET,
        mode="readonly",
        system_sync=False,
        note=f"Final companion misc invoice sheet. Fakturanr: {detail.numbering_preview.invoice_number_next or '—'}",
        columns=["#", "Tekst", "Antal", "Pris", "I alt"],
        rows=[
            [
                row.row_key,
                row.text or "—",
                (_fmt_decimal(row.quantity) if row.quantity is not None else "—"),
                _fmt_decimal(row.unit_price_dkk),
                _fmt_decimal(row.line_total_dkk),
            ]
            for row in detail.invoice_misc.rows
        ],
    )


def build_afg_workspace_preview(workspace: PosWorkspaceOut, artifact: DocumentArtifact | None = None) -> DocumentArtifactPreviewOut:
    from app.services.document_artifact_preview import build_afg_workspace_preview as _impl

    return _impl(workspace, artifact)


def build_afg_document_preview(detail: PosDocumentDetailOut, artifact: DocumentArtifact | None = None) -> DocumentArtifactPreviewOut:
    from app.services.document_artifact_preview import build_afg_document_preview as _impl

    return _impl(detail, artifact)


def build_inventory_preview(workspace: InventoryWorkspaceOut, artifact: DocumentArtifact | None = None) -> DocumentArtifactPreviewOut:
    from app.services.document_artifact_preview import build_inventory_preview as _impl

    return _impl(workspace, artifact)


def build_log_preview(workspace: AfgLogWorkspaceOut, *, year: int, artifact: DocumentArtifact | None = None) -> DocumentArtifactPreviewOut:
    from app.services.document_artifact_preview import build_log_preview as _impl

    return _impl(workspace, year=year, artifact=artifact)

async def sync_afg_workspace_artifact(session: AsyncSession, workspace: PosWorkspaceOut) -> WorkbookArtifactBundle:
    session_code = workspace.session.session_code
    document_number = workspace.numbering_preview.afregnings_number_next or session_code
    artifact_key = f"alis.workspace.{workspace.session.id}"
    existing_record = await get_artifact_record(session, artifact_key)
    revision = _next_artifact_revision(existing_record)
    stamp = utc_now()
    content = _build_afg_workbook_bytes_from_workspace(
        workspace,
        sync_context=ArtifactSyncContext(
            kind="alis-workspace",
            key=str(workspace.session.id),
            artifact_key=artifact_key,
            base_version=str(revision),
            workspace_revision=str(workspace.workspace_revision),
        ),
    )
    return await _store_artifact(
        session,
        artifact_key=artifact_key,
        module_name="alis",
        document_type="afg_workdraft",
        business_key=session_code,
        version_kind="draft",
        is_live=True,
        file_name=f"{document_number}.xlsm",
        relative_path=Path("alis") / "afg" / "drafts" / f"{document_number}.xlsm",
        mime_type=XLSM_MIME,
        template_name=AFG_TEMPLATE_NAME,
        content=content,
        updated_at=stamp,
        revision=revision,
    )


async def sync_afg_document_artifact(session: AsyncSession, detail: PosDocumentDetailOut) -> WorkbookArtifactBundle:
    artifact_key = f"alis.document.{detail.sequence_no}"
    existing_record = await get_artifact_record(session, artifact_key)
    revision = _next_artifact_revision(existing_record)
    stamp = utc_now()
    content = _build_afg_workbook_bytes_from_detail(
        detail,
        sync_context=ArtifactSyncContext(
            kind="alis-document",
            key=str(detail.sequence_no),
            artifact_key=artifact_key,
            base_version=str(revision),
        ),
    )
    year = str(detail.issued_at.year)
    file_name = f"{detail.sequence_no}.xlsm"
    return await _store_artifact(
        session,
        artifact_key=artifact_key,
        module_name="alis",
        document_type="afg_document",
        business_key=str(detail.sequence_no),
        version_kind="final",
        is_live=True,
        file_name=file_name,
        relative_path=Path("alis") / "afg" / "final" / year / file_name,
        mime_type=XLSM_MIME,
        template_name=AFG_TEMPLATE_NAME,
        content=content,
        updated_at=stamp,
        revision=revision,
    )


async def sync_inventory_workbook_artifact(
    session: AsyncSession,
    workspace: InventoryWorkspaceOut,
    *,
    create_snapshot: bool,
) -> WorkbookArtifactBundle:
    stamp = utc_now()
    existing_record = await get_artifact_record(session, "depolama.live")
    revision = _next_artifact_revision(existing_record)
    content = _build_inventory_workbook_bytes(
        workspace,
        sync_context=ArtifactSyncContext(
            kind="depolama",
            key="live",
            artifact_key="depolama.live",
            base_version=str(revision),
            mappings=_inventory_sync_mappings(workspace),
        ),
        display_updated_at=stamp if create_snapshot or existing_record is None else existing_record.updated_at,
    )
    live_bundle = await _store_artifact(
        session,
        artifact_key="depolama.live",
        module_name="depolama",
        document_type="inventory_workbook",
        business_key="live",
        version_kind="live",
        is_live=True,
        file_name="Depolama.xlsx",
        relative_path=Path("depolama") / "live" / "Depolama.xlsx",
        mime_type=XLSX_MIME,
        template_name=DEPOLAMA_TEMPLATE_NAME,
        content=content,
        updated_at=stamp,
        revision=revision,
    )
    if create_snapshot:
        now = utc_now()
        snapshot_stamp = now.strftime("%Y%m%d-%H%M%S")
        await _store_artifact(
            session,
            artifact_key=f"depolama.snapshot.{snapshot_stamp}",
            module_name="depolama",
            document_type="inventory_workbook",
            business_key="live",
            version_kind="snapshot",
            is_live=False,
            file_name=f"{snapshot_stamp}-Depolama.xlsx",
            relative_path=Path("depolama") / "snapshots" / now.strftime("%Y") / now.strftime("%m") / f"{snapshot_stamp}-Depolama.xlsx",
            mime_type=XLSX_MIME,
            template_name=DEPOLAMA_TEMPLATE_NAME,
            content=content,
            updated_at=now,
        )
    return live_bundle


async def sync_log_workbook_artifact(
    session: AsyncSession,
    workspace: AfgLogWorkspaceOut,
    *,
    year: int,
    create_snapshot: bool,
) -> WorkbookArtifactBundle:
    stamp = utc_now()
    existing_record = await get_artifact_record(session, f"log.live.{year}")
    revision = _next_artifact_revision(existing_record)
    content = _build_log_workbook_bytes(
        workspace,
        year=year,
        sync_context=ArtifactSyncContext(
            kind="log",
            key=str(year),
            artifact_key=f"log.live.{year}",
            base_version=str(revision),
        ),
    )
    live_bundle = await _store_artifact(
        session,
        artifact_key=f"log.live.{year}",
        module_name="log",
        document_type="log_workbook",
        business_key=str(year),
        version_kind="live",
        is_live=True,
        file_name=f"Log-{year}.xlsx",
        relative_path=Path("log") / "live" / f"Log-{year}.xlsx",
        mime_type=XLSX_MIME,
        template_name=LOG_TEMPLATE_NAME,
        content=content,
        updated_at=stamp,
        revision=revision,
    )
    if create_snapshot:
        now = utc_now()
        snapshot_stamp = now.strftime("%Y%m%d-%H%M%S")
        await _store_artifact(
            session,
            artifact_key=f"log.snapshot.{year}.{snapshot_stamp}",
            module_name="log",
            document_type="log_workbook",
            business_key=str(year),
            version_kind="snapshot",
            is_live=False,
            file_name=f"{snapshot_stamp}-Log-{year}.xlsx",
            relative_path=Path("log") / "snapshots" / now.strftime("%Y") / now.strftime("%m") / f"{snapshot_stamp}-Log-{year}.xlsx",
            mime_type=XLSX_MIME,
            template_name=LOG_TEMPLATE_NAME,
            content=content,
            updated_at=now,
        )
    return live_bundle
