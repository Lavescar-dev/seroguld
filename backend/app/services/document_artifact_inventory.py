from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Iterable
from uuid import UUID

from openpyxl import load_workbook

from app.schemas.document_artifact import DocumentArtifactCellChangeOut, DocumentArtifactReconcilePreviewOut
from app.schemas.inventory import InventoryMarketPricesUpdate, InventoryWorkspaceOut
from app.schemas.product import ProductUpdate


def _core():
    from app.services import document_artifact_service as core

    return core


def _inventory_sync_mappings(workspace: InventoryWorkspaceOut):
    core = _core()
    return [
        core.SyncSheetRowMapping(
            mapping_type="inventory_row",
            sheet_name=core.INVENTORY_SHEET,
            row_index=offset,
            entity_id=str(item.id),
            entity_key=item.product_number,
            extra={
                "lager_dato": item.lager_dato,
                "urun": item.urun or "",
                "birim_gram": core._fmt_decimal(item.birim_gram),
                "adet": str(item.adet),
                "alis_fiyati_dkk": core._fmt_decimal(item.alis_fiyati_dkk),
                "length_cm": item.length_cm or "",
                "width_mm": core._fmt_decimal(item.width_mm),
                "thickness_mm": core._fmt_decimal(item.thickness_mm),
                "producer": item.producer or "",
            },
        )
        for offset, item in enumerate(workspace.rows, start=11)
    ]


def _inventory_editable_refs(mappings: Iterable) -> list[str]:
    refs = ["K4", "K5", "K6", "K7"]
    for mapping in mappings:
        row_idx = mapping.row_index
        refs.extend(
            [
                f"B{row_idx}",
                f"C{row_idx}",
                f"D{row_idx}",
                f"E{row_idx}",
                f"H{row_idx}",
                f"N{row_idx}",
                f"O{row_idx}",
                f"P{row_idx}",
                f"Q{row_idx}",
            ]
        )
    return refs


def _inventory_raw_row_payload(sheet, row_idx: int) -> tuple[datetime | None, dict[str, str]]:
    core = _core()
    purchase_date = core._datetime_from_excel(sheet[f"B{row_idx}"].value)
    next_values = {
        "lager_dato": purchase_date.date().isoformat() if purchase_date else "",
        "urun": core._clean_text(sheet[f"C{row_idx}"].value) or "",
        "birim_gram": core._fmt_decimal(core._decimal_from_excel(sheet[f"D{row_idx}"].value)),
        "adet": str(int(core.to_decimal(sheet[f"E{row_idx}"].value or 1))),
        "alis_fiyati_dkk": core._fmt_decimal(core._decimal_from_excel(sheet[f"H{row_idx}"].value)),
        "length_cm": core._clean_text(sheet[f"N{row_idx}"].value) or "",
        "width_mm": core._fmt_decimal(core._decimal_from_excel(sheet[f"O{row_idx}"].value))
        if core._clean_text(sheet[f"O{row_idx}"].value) is not None
        else "",
        "thickness_mm": core._fmt_decimal(core._decimal_from_excel(sheet[f"P{row_idx}"].value))
        if core._clean_text(sheet[f"P{row_idx}"].value) is not None
        else "",
        "producer": core._clean_text(sheet[f"Q{row_idx}"].value) or "",
    }
    return purchase_date, next_values


def _inventory_raw_product_updates(
    sheet,
    *,
    current_workspace: InventoryWorkspaceOut,
) -> dict[UUID, ProductUpdate]:
    core = _core()
    row_map = {item.product_number: item for item in current_workspace.rows}
    seen_product_numbers: set[str] = set()
    product_updates: dict[UUID, ProductUpdate] = {}

    for row_idx in range(11, max(sheet.max_row, 420) + 1):
        product_number = core._clean_text(sheet[f"K{row_idx}"].value)
        if product_number is None:
            continue
        if product_number in seen_product_numbers:
            raise ValueError(f"Aynı ürün numarası workbook içinde tekrar ediyor: {product_number}")
        seen_product_numbers.add(product_number)
        current_row = row_map.get(product_number)
        if current_row is None:
            raise ValueError(f"Workbook satırı sistemde bulunmayan ürün numarası taşıyor: {product_number}")

        purchase_date, next_values = _inventory_raw_row_payload(sheet, row_idx)
        current_values = {
            "lager_dato": current_row.lager_dato or "",
            "urun": current_row.urun or "",
            "birim_gram": core._fmt_decimal(current_row.birim_gram),
            "adet": str(current_row.adet),
            "alis_fiyati_dkk": core._fmt_decimal(current_row.alis_fiyati_dkk),
            "length_cm": current_row.length_cm or "",
            "width_mm": core._fmt_decimal(current_row.width_mm),
            "thickness_mm": core._fmt_decimal(current_row.thickness_mm),
            "producer": current_row.producer or "",
        }
        if current_values == next_values:
            continue

        product_updates[current_row.id] = ProductUpdate(
            purchase_date=purchase_date,
            display_name=next_values["urun"] or None,
            weight_grams=core._decimal_from_excel(sheet[f"D{row_idx}"].value),
            unit_count=int(core.to_decimal(sheet[f"E{row_idx}"].value or 1)),
            purchase_price_dkk=core._decimal_from_excel(sheet[f"H{row_idx}"].value),
            length_cm=next_values["length_cm"] or None,
            width_mm=core._decimal_from_excel(sheet[f"O{row_idx}"].value)
            if core._clean_text(sheet[f"O{row_idx}"].value) is not None
            else None,
            thickness_mm=core._decimal_from_excel(sheet[f"P{row_idx}"].value)
            if core._clean_text(sheet[f"P{row_idx}"].value) is not None
            else None,
            producer=next_values["producer"] or None,
        )
    return product_updates


def build_inventory_reconcile_preview(
    workspace: InventoryWorkspaceOut,
    parsed,
) -> DocumentArtifactReconcilePreviewOut:
    core = _core()
    changes: list[DocumentArtifactCellChangeOut] = []
    market_specs = (
        ("K4", "Au", core._fmt_decimal(workspace.market_prices.gold), core._fmt_decimal(parsed.market_prices.gold)),
        ("K5", "Ag", core._fmt_decimal(workspace.market_prices.silver), core._fmt_decimal(parsed.market_prices.silver)),
        ("K6", "Pt", core._fmt_decimal(workspace.market_prices.platinum), core._fmt_decimal(parsed.market_prices.platinum)),
        ("K7", "Pd", core._fmt_decimal(workspace.market_prices.palladium), core._fmt_decimal(parsed.market_prices.palladium)),
    )
    for cell_ref, label, old_value, new_value in market_specs:
        if old_value == new_value:
            continue
        changes.append(
            DocumentArtifactCellChangeOut(
                sheet=core.INVENTORY_SHEET,
                cell_ref=cell_ref,
                label=f"Piyasa Değeri · {label}",
                old_value=old_value,
                new_value=new_value,
            )
        )

    row_by_id = {row.id: row for row in workspace.rows}
    row_index_by_id = {row.id: index for index, row in enumerate(workspace.rows, start=11)}
    field_specs = (
        ("B", "Lager Tarihi", lambda row, update: (core._fmt_date(row.lager_dato), core._fmt_date(update.purchase_date))),
        ("C", "Ürün", lambda row, update: (row.urun or "—", update.display_name or "—")),
        ("D", "Birim Gram", lambda row, update: (core._fmt_decimal(row.birim_gram), core._fmt_decimal(update.weight_grams))),
        ("E", "Adet", lambda row, update: (str(row.adet), str(update.unit_count))),
        ("H", "Alış Fiyatı", lambda row, update: (core._fmt_decimal(row.alis_fiyati_dkk), core._fmt_decimal(update.purchase_price_dkk))),
        ("N", "Uzunluk", lambda row, update: (row.length_cm or "—", update.length_cm or "—")),
        ("O", "Genişlik", lambda row, update: (core._fmt_optional_decimal(row.width_mm), core._fmt_optional_decimal(update.width_mm))),
        ("P", "Kalınlık", lambda row, update: (core._fmt_optional_decimal(row.thickness_mm), core._fmt_optional_decimal(update.thickness_mm))),
        ("Q", "Üretici", lambda row, update: (row.producer or "—", update.producer or "—")),
    )
    for product_id, update in sorted(parsed.product_updates.items(), key=lambda item: row_index_by_id.get(item[0], 99999)):
        current_row = row_by_id.get(product_id)
        if current_row is None:
            continue
        row_index = row_index_by_id.get(product_id, 11)
        for column, label, formatter in field_specs:
            old_value, new_value = formatter(current_row, update)
            normalized_old = old_value if old_value not in {None, ""} else "—"
            normalized_new = new_value if new_value not in {None, ""} else "—"
            if normalized_old == normalized_new:
                continue
            changes.append(
                DocumentArtifactCellChangeOut(
                    sheet=core.INVENTORY_SHEET,
                    cell_ref=f"{column}{row_index}",
                    label=f"{current_row.product_number} · {label}",
                    old_value=normalized_old,
                    new_value=normalized_new,
                )
            )

    warnings = [
        "Yalnız kontrollü market prices ve satır alanları sisteme uygulanır; toplam, has metal ve spot formülleri write-back kaynağı değildir.",
    ]
    if not changes:
        warnings.append("Workbook mevcut depolama state’iyle aynı; uygulanacak değişiklik yok.")
    return DocumentArtifactReconcilePreviewOut(
        editable=True,
        changes=changes,
        warnings=warnings,
        blocking_errors=[],
    )


def _build_inventory_workbook_bytes(
    workspace: InventoryWorkspaceOut,
    *,
    sync_context,
    display_updated_at: datetime,
) -> bytes:
    core = _core()
    workbook = core._open_template(core.DEPOLAMA_TEMPLATE_NAME)
    core._touch_calc_flags(workbook)
    sheet = workbook[core.INVENTORY_SHEET]
    sheet["E2"] = core._excel_datetime(display_updated_at)
    sheet["K4"] = core.quantize_2(core.to_decimal(workspace.market_prices.gold))
    sheet["K5"] = core.quantize_2(core.to_decimal(workspace.market_prices.silver))
    sheet["K6"] = core.quantize_2(core.to_decimal(workspace.market_prices.platinum))
    sheet["K7"] = core.quantize_2(core.to_decimal(workspace.market_prices.palladium))

    for row_idx in range(11, max(sheet.max_row, 420) + 1):
        for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "N", "O", "P", "Q"):
            core._set_sheet_cell(sheet, f"{col}{row_idx}", None)

    for offset, item in enumerate(workspace.rows, start=11):
        core._set_sheet_cell(sheet, f"A{offset}", "Foto" if item.primary_photo else None)
        core._set_sheet_cell(sheet, f"B{offset}", item.lager_dato)
        core._set_sheet_cell(sheet, f"C{offset}", item.urun)
        core._set_sheet_cell(sheet, f"D{offset}", core.quantize_2(core.to_decimal(item.birim_gram)))
        core._set_sheet_cell(sheet, f"E{offset}", item.adet)
        core._set_sheet_cell(sheet, f"F{offset}", core.quantize_2(core.to_decimal(item.toplam_gram)))
        core._set_sheet_cell(sheet, f"G{offset}", core.quantize_2(core.to_decimal(item.has_metal_grams or 0)))
        core._set_sheet_cell(sheet, f"H{offset}", core.quantize_2(core.to_decimal(item.alis_fiyati_dkk)))
        core._set_sheet_cell(sheet, f"I{offset}", core.quantize_2(core.to_decimal(item.spot_degeri_dkk)))
        core._set_sheet_cell(sheet, f"J{offset}", item.shop_sync_status or item.status)
        core._set_sheet_cell(sheet, f"K{offset}", item.product_number)
        core._set_sheet_cell(sheet, f"N{offset}", item.length_cm)
        core._set_sheet_cell(sheet, f"O{offset}", item.width_mm)
        core._set_sheet_cell(sheet, f"P{offset}", item.thickness_mm)
        core._set_sheet_cell(sheet, f"Q{offset}", item.producer)

    mappings = list(sync_context.mappings or _inventory_sync_mappings(workspace))
    core._write_sync_sheet(
        workbook,
        context=core.ArtifactSyncContext(
            kind=sync_context.kind,
            key=sync_context.key,
            artifact_key=sync_context.artifact_key,
            base_version=sync_context.base_version,
            contract_version=sync_context.contract_version,
            mappings=mappings,
        ),
    )
    core._protect_sheet(sheet, editable_refs=_inventory_editable_refs(mappings))
    workbook.active = workbook.sheetnames.index(core.INVENTORY_SHEET)
    return core._save_workbook_bytes(workbook)


def parse_inventory_workbook_inputs_from_workbook(
    content: bytes,
    *,
    current_workspace: InventoryWorkspaceOut | None = None,
):
    core = _core()
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    metadata = core._sync_metadata_if_present(workbook, expected_kind="depolama", expected_key="live")
    if core.INVENTORY_SHEET not in workbook.sheetnames:
        raise ValueError("Depolama workbook sheet bulunamadı")

    if metadata is not None:
        sheet = workbook[core.INVENTORY_SHEET]
        mappings = [mapping for mapping in metadata.mappings if mapping.mapping_type == "inventory_row"]
        product_updates: dict[UUID, ProductUpdate] = {}
        for mapping in mappings:
            row_idx = mapping.row_index
            purchase_date, next_values = _inventory_raw_row_payload(sheet, row_idx)
            current_values = {key: str((mapping.extra or {}).get(key) or "") for key in next_values}
            if current_values == next_values:
                continue
            payload = ProductUpdate(
                purchase_date=purchase_date,
                display_name=next_values["urun"] or None,
                weight_grams=core._decimal_from_excel(sheet[f"D{row_idx}"].value),
                unit_count=int(core.to_decimal(sheet[f"E{row_idx}"].value or 1)),
                purchase_price_dkk=core._decimal_from_excel(sheet[f"H{row_idx}"].value),
                length_cm=next_values["length_cm"] or None,
                width_mm=core._decimal_from_excel(sheet[f"O{row_idx}"].value)
                if core._clean_text(sheet[f"O{row_idx}"].value) is not None
                else None,
                thickness_mm=core._decimal_from_excel(sheet[f"P{row_idx}"].value)
                if core._clean_text(sheet[f"P{row_idx}"].value) is not None
                else None,
                producer=next_values["producer"] or None,
            )
            product_updates[UUID(mapping.entity_id)] = payload
        return core.InventoryWorkbookArtifactInputs(
            market_prices=InventoryMarketPricesUpdate(
                gold=core._decimal_from_excel(sheet["K4"].value),
                silver=core._decimal_from_excel(sheet["K5"].value),
                platinum=core._decimal_from_excel(sheet["K6"].value),
                palladium=core._decimal_from_excel(sheet["K7"].value),
            ),
            product_updates=product_updates,
            base_version=metadata.base_version,
        )

    if current_workspace is None:
        raise ValueError("Depolama raw import için mevcut workspace gerekir")

    values_workbook = load_workbook(io.BytesIO(content), data_only=True)
    if core.INVENTORY_SHEET not in values_workbook.sheetnames:
        raise ValueError("Depolama workbook sheet bulunamadı")
    values_sheet = values_workbook[core.INVENTORY_SHEET]
    return core.InventoryWorkbookArtifactInputs(
        market_prices=InventoryMarketPricesUpdate(
            gold=core._decimal_from_excel(values_sheet["K4"].value),
            silver=core._decimal_from_excel(values_sheet["K5"].value),
            platinum=core._decimal_from_excel(values_sheet["K6"].value),
            palladium=core._decimal_from_excel(values_sheet["K7"].value),
        ),
        product_updates=_inventory_raw_product_updates(values_sheet, current_workspace=current_workspace),
        base_version=None,
    )
