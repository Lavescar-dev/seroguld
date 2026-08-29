from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from html import escape as html_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas.pos import PosWorkspaceBankInfo, PosWorkspaceCustomerOut, PosWorkspaceOut
from app.services.pos_value_helpers import AFG_DECLARATION_HEADER, AFG_DECLARATION_ITEMS, fmt_decimal as _fmt_decimal
from app.utils.helpers import quantize_2, to_decimal


def workspace_payment_label(payment_method: str) -> str:
    return "Kontant" if payment_method == "cash" else "Bankoverførsel"


def workspace_preview_document_number(workspace: PosWorkspaceOut) -> str:
    return workspace.numbering_preview.afregnings_number_next or workspace.session.session_code


def workspace_preview_lines(workspace: PosWorkspaceOut) -> list[dict[str, str]]:
    lines: list[dict[str, str]] = []
    line_no = 1
    for row in workspace.gold_rows:
        gram = quantize_2(to_decimal(row.gram))
        if gram <= 0:
            continue
        purity = quantize_2(to_decimal(row.purity_percentage))
        lines.append(
            {
                "line_no": str(line_no),
                "type": str(row.label),
                "fineness": f"{_fmt_decimal(row.karat)}K / {_fmt_decimal(purity)}%",
                "lodighed": str(row.lodighed),
                "gram": _fmt_decimal(gram),
                "avance": _fmt_decimal(row.avance_percent),
                "unit_price": _fmt_decimal(row.unit_price_dkk),
                "line_total": _fmt_decimal(row.line_total_dkk),
                "pure_metal_grams": _fmt_decimal(quantize_2(gram * (purity / Decimal('100')))),
                "metal_label": "Guld",
            }
        )
        line_no += 1

    for row in workspace.silver_rows:
        gram = quantize_2(to_decimal(row.gram))
        if gram <= 0:
            continue
        purity = quantize_2(to_decimal(row.purity_percentage))
        lines.append(
            {
                "line_no": str(line_no),
                "type": str(row.label),
                "fineness": f"{_fmt_decimal(purity)}%",
                "lodighed": str(row.lodighed),
                "gram": _fmt_decimal(gram),
                "avance": _fmt_decimal(row.avance_percent),
                "unit_price": _fmt_decimal(row.unit_price_dkk),
                "line_total": _fmt_decimal(row.line_total_dkk),
                "pure_metal_grams": _fmt_decimal(quantize_2(gram * (purity / Decimal('100')))),
                "metal_label": "Sølv",
            }
        )
        line_no += 1

    # Bar / Pt-Pd / dinamik (kniv-çeyrek) satırları da önizleme/CSV'ye girer —
    # eksik kalmaları taslak çıktısında tutar-satır uyuşmazlığı yaratıyordu.
    for row in getattr(workspace, "bar_rows", None) or []:
        gram = quantize_2(to_decimal(row.gram))
        if gram <= 0:
            continue
        purity = quantize_2(to_decimal(row.purity_percentage))
        lines.append(
            {
                "line_no": str(line_no),
                "type": str(row.label),
                "fineness": f"{_fmt_decimal(purity)}%",
                "lodighed": str(row.lodighed),
                "gram": _fmt_decimal(gram),
                "avance": _fmt_decimal(row.avance_percent),
                "unit_price": _fmt_decimal(row.unit_price_dkk),
                "line_total": _fmt_decimal(row.line_total_dkk),
                "pure_metal_grams": _fmt_decimal(quantize_2(gram * (purity / Decimal('100')))),
                "metal_label": "Guld" if str(row.bar_type) == "gold" else "Sølv",
            }
        )
        line_no += 1

    for row in getattr(workspace, "ptpd_rows", None) or []:
        gram = quantize_2(to_decimal(row.gram))
        if gram <= 0:
            continue
        purity = quantize_2(to_decimal(row.purity_percentage))
        lines.append(
            {
                "line_no": str(line_no),
                "type": str(row.label),
                "fineness": f"{_fmt_decimal(purity)}%",
                "lodighed": str(row.lodighed),
                "gram": _fmt_decimal(gram),
                "avance": _fmt_decimal(row.avance_percent),
                "unit_price": _fmt_decimal(row.unit_price_dkk),
                "line_total": _fmt_decimal(row.line_total_dkk),
                "pure_metal_grams": _fmt_decimal(quantize_2(gram * (purity / Decimal('100')))),
                "metal_label": str(row.label),
            }
        )
        line_no += 1

    for row in getattr(workspace, "extra_rows", None) or []:
        gram = quantize_2(to_decimal(row.gram))
        if gram <= 0:
            continue
        purity = quantize_2(to_decimal(row.purity_percentage))
        lines.append(
            {
                "line_no": str(line_no),
                "type": str(row.label),
                "fineness": f"{_fmt_decimal(purity)}%",
                "lodighed": str(row.karat),
                "gram": _fmt_decimal(gram),
                "avance": _fmt_decimal(row.avance_percent),
                "unit_price": _fmt_decimal(row.unit_price_dkk),
                "line_total": _fmt_decimal(row.line_total_dkk),
                "pure_metal_grams": _fmt_decimal(quantize_2(gram * (purity / Decimal('100')))),
                "metal_label": "Guld" if str(row.metal) == "gold" else "Sølv",
            }
        )
        line_no += 1
    return lines


def workspace_csv_escape(value: str | Decimal | int | None) -> str:
    text = "" if value is None else str(value)
    return '"' + text.replace('"', '""') + '"'


def build_purchase_workbook_bytes(
    *,
    document_number: str,
    issued_at: datetime,
    customer: PosWorkspaceCustomerOut,
    payment_method: str,
    bank_info: PosWorkspaceBankInfo,
    lines: list[dict[str, str]],
    net_amount_dkk: Decimal,
    vat_rate_percent: Decimal,
    vat_amount_dkk: Decimal,
    gross_amount_dkk: Decimal,
    note: str | None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Afregningsbilag"

    header_fill = PatternFill(fill_type="solid", fgColor="6F5B45")
    sub_fill = PatternFill(fill_type="solid", fgColor="F4E9DA")
    total_fill = PatternFill(fill_type="solid", fgColor="1F6B4F")

    sheet["A1"] = "SERO GULD"
    sheet["A1"].font = Font(bold=True, size=18)
    sheet["A2"] = "Køb og salg af guld, sølv og smykker"
    sheet["F1"] = "Afregningsnr."
    sheet["F1"].font = Font(bold=True)
    sheet["G1"] = document_number
    sheet["G1"].font = Font(bold=True, size=15)
    sheet["F2"] = "Dato"
    sheet["G2"] = issued_at.strftime("%d.%m.%Y %H:%M")

    detail_rows = [
        ("Navn", customer.name or "—"),
        ("CPR", customer.cpr_number or "—"),
        ("Kørekort / Pas", customer.identity_doc_number or "—"),
        ("Telefon", customer.phone or "—"),
        ("E-mail", customer.email or "—"),
        ("Adresse", customer.address or "—"),
        ("Postnr.", customer.postal_code or "—"),
        (
            "Betaling",
            workspace_payment_label(payment_method)
            if payment_method == "cash"
            else f"Bank {bank_info.reg_number or '-'} / {bank_info.account_number or '-'}",
        ),
    ]
    start_row = 4
    for idx, (label, value) in enumerate(detail_rows, start=start_row):
        sheet[f"A{idx}"] = label
        sheet[f"A{idx}"].font = Font(bold=True)
        sheet[f"B{idx}"] = value
        sheet.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)

    header_row = start_row + len(detail_rows) + 2
    headers = ["Type", "Saflık", "Lødighed", "Gram", "Avance %", "Birim", "Toplam"]
    for col_idx, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    row_cursor = header_row + 1
    for line in lines:
        values = [
            line["type"],
            line["fineness"],
            line["lodighed"],
            line["gram"],
            line["avance"],
            line["unit_price"],
            line["line_total"],
        ]
        for col_idx, value in enumerate(values, start=1):
            sheet.cell(row=row_cursor, column=col_idx, value=value)
        row_cursor += 1

    total_row = row_cursor + 1
    for offset, (label, value) in enumerate(
        (
            ("Netto", net_amount_dkk),
            (f"Moms %{_fmt_decimal(vat_rate_percent)}", vat_amount_dkk),
            ("I alt", gross_amount_dkk),
        )
    ):
        row_idx = total_row + offset
        sheet.cell(row=row_idx, column=6, value=label).font = Font(bold=True, color="FFFFFF")
        sheet.cell(row=row_idx, column=6).fill = total_fill
        sheet.cell(row=row_idx, column=7, value=_fmt_decimal(value)).font = Font(
            bold=True,
            color="FFFFFF",
            size=14 if offset == 2 else 11,
        )
        sheet.cell(row=row_idx, column=7).fill = total_fill
    if note:
        note_row = total_row + 4
        sheet.cell(row=note_row, column=1, value="Not").font = Font(bold=True)
        sheet.cell(row=note_row, column=2, value=note)
        sheet.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=7)

    for row_idx in range(header_row + 1, row_cursor):
        for col_idx in range(1, 8):
            sheet.cell(row=row_idx, column=col_idx).fill = (
                sub_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
            )

    widths = [22, 18, 14, 12, 12, 14, 16]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + col_idx)].width = width
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["F"].width = 16
    sheet.column_dimensions["G"].width = 18

    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return payload.getvalue()


def render_purchase_workspace_print_html(
    *,
    workspace: PosWorkspaceOut,
    payment_method: str,
    lines: list[dict[str, str]],
    auto_print: bool,
) -> str:
    line_rows = "".join(
        (
            "<tr>"
            f"<td>{html_escape(line['type'])}</td>"
            f"<td>{html_escape(line['fineness'])}</td>"
            f"<td>{html_escape(line['lodighed'])}</td>"
            f"<td style=\"text-align:right\">{html_escape(line['gram'])} g</td>"
            f"<td style=\"text-align:right\">{html_escape(line['avance'])}%</td>"
            f"<td style=\"text-align:right\">{html_escape(line['unit_price'])} DKK</td>"
            f"<td style=\"text-align:right;font-weight:800\">{html_escape(line['line_total'])} DKK</td>"
            "</tr>"
        )
        for line in lines
    )
    print_script = "<script>window.addEventListener('load', () => window.print());</script>" if auto_print else ""
    return f"""<!doctype html>
<html lang="da">
  <head>
    <meta charset="UTF-8" />
    <title>AFG {html_escape(workspace_preview_document_number(workspace))}</title>
    <style>
      body {{ font-family: 'IBM Plex Sans', system-ui, sans-serif; padding: 40px; color: #23170f; }}
      .mono {{ font-family: 'IBM Plex Mono', monospace; }}
      .header {{ display:flex; justify-content:space-between; align-items:flex-end; border-bottom:3px solid #6f5233; padding-bottom:18px; margin-bottom:24px; }}
      .title {{ font-size:32px; font-weight:900; letter-spacing:.08em; color:#6f5233; }}
      .meta {{ text-align:right; }}
      .meta p {{ margin:0; }}
      .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:12px 24px; margin-bottom:24px; }}
      .field strong {{ display:block; font-size:11px; text-transform:uppercase; color:#7c6a59; margin-bottom:4px; }}
      .field span {{ font-size:14px; font-weight:700; }}
      table {{ width:100%; border-collapse:collapse; margin-bottom:24px; }}
      th, td {{ border:1px solid #dac8b0; padding:10px 12px; font-size:13px; }}
      th {{ background:#f4e9da; text-align:left; text-transform:uppercase; font-size:11px; letter-spacing:.08em; }}
      .total {{ display:flex; justify-content:space-between; align-items:center; background:#221710; color:#fff; padding:16px 18px; }}
      .disclaimer {{ margin-top:28px; font-size:11px; line-height:1.65; color:#655341; }}
      @media print {{
        body {{ padding: 16px; }}
      }}
    </style>
  </head>
  <body>
    <div class="header">
      <div>
        <div class="title mono">SERO GULD</div>
        <p>Køb og salg af guld, sølv og smykker</p>
      </div>
      <div class="meta mono">
        <p><strong>Afregningsnr.</strong></p>
        <p style="font-size:24px;font-weight:900">{html_escape(workspace_preview_document_number(workspace))}</p>
        <p>{html_escape(workspace.session.updated_at.strftime("%d.%m.%Y %H:%M"))}</p>
      </div>
    </div>

    <div class="grid">
      <div class="field"><strong>Navn</strong><span>{html_escape(workspace.customer.name or '—')}</span></div>
      <div class="field"><strong>CPR</strong><span>{html_escape(workspace.customer.cpr_number or '—')}</span></div>
      <div class="field"><strong>Telefon</strong><span>{html_escape(workspace.customer.phone or '—')}</span></div>
      <div class="field"><strong>E-mail</strong><span>{html_escape(workspace.customer.email or '—')}</span></div>
      <div class="field"><strong>Adresse</strong><span>{html_escape(workspace.customer.address or '—')}</span></div>
      <div class="field"><strong>Postnr.</strong><span>{html_escape(workspace.customer.postal_code or '—')}</span></div>
      <div class="field"><strong>Kørekort / Pas</strong><span>{html_escape(workspace.customer.identity_doc_number or '—')}</span></div>
      <div class="field"><strong>Betaling</strong><span>{html_escape(workspace_payment_label(payment_method)) if payment_method == 'cash' else f"Bank {html_escape(workspace.bank_info.reg_number or '-')}/{html_escape(workspace.bank_info.account_number or '-')}"}</span></div>
    </div>

    <table>
      <thead>
        <tr>
          <th>Type</th>
          <th>Saflık</th>
          <th>Lødighed</th>
          <th>Gram</th>
          <th>Avance</th>
          <th>Birim</th>
          <th>Toplam</th>
        </tr>
      </thead>
      <tbody>{line_rows}</tbody>
    </table>

    <div class="total">
      <span>Netto {_fmt_decimal(workspace.summary.net_amount_dkk)} DKK · Moms %{_fmt_decimal(workspace.summary.vat_rate_percent)}: {_fmt_decimal(workspace.summary.vat_amount_dkk)} DKK<br><strong>Genel Toplam / I alt</strong></span>
      <span class="mono" style="font-size:28px;font-weight:900">{html_escape(_fmt_decimal(workspace.summary.gross_amount_dkk))} DKK</span>
    </div>

    {f'<div class="disclaimer"><strong>Not:</strong> {html_escape(workspace.afg_note)}</div>' if workspace.afg_note else ''}

    <div class="disclaimer">
      <strong>{html_escape(AFG_DECLARATION_HEADER)}</strong><br>
      1. {html_escape(AFG_DECLARATION_ITEMS[0])}<br>
      2. {html_escape(AFG_DECLARATION_ITEMS[1])}<br>
      3. {html_escape(AFG_DECLARATION_ITEMS[2])}
    </div>
    {print_script}
  </body>
</html>"""
