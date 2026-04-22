from __future__ import annotations

import csv
import io
from datetime import datetime, timedelta
from decimal import Decimal
from typing import Any, Iterable

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import require_admin
from app.database import get_db
from app.models.enums import ProductStatusEnum
from app.models.product import Product
from app.schemas.report import ReportSummaryOut
from app.utils.helpers import utc_now

router = APIRouter()

EXPORT_HEADERS = [
    ("product_number", "Ürün No"),
    ("product_type", "Ürün Tipi"),
    ("metal_type", "Metal"),
    ("status", "Durum"),
    ("purchase_price_dkk", "Alım (DKK)"),
    ("sale_price_dkk", "Satış (DKK)"),
    ("profit_dkk", "Kar (DKK)"),
    ("purchase_date", "Alım Tarihi"),
]


async def _summary_for_range(db: AsyncSession, start, end) -> ReportSummaryOut:
    purchased_count = await db.scalar(
        select(func.count(Product.id)).where(Product.purchase_date >= start, Product.purchase_date < end)
    )
    sold_count = await db.scalar(
        select(func.count(Product.id)).where(
            Product.status == ProductStatusEnum.SOLD,
            Product.sale_date >= start,
            Product.sale_date < end,
        )
    )
    melted_count = await db.scalar(
        select(func.count(Product.id)).where(
            Product.status == ProductStatusEnum.MELTED,
            Product.melt_date >= start,
            Product.melt_date < end,
        )
    )

    purchase_total = await db.scalar(
        select(func.coalesce(func.sum(Product.purchase_price_dkk), Decimal("0"))).where(
            Product.purchase_date >= start,
            Product.purchase_date < end,
        )
    )
    sale_total = await db.scalar(
        select(func.coalesce(func.sum(Product.sale_price_dkk), Decimal("0"))).where(
            Product.sale_date >= start,
            Product.sale_date < end,
        )
    )
    profit_total = await db.scalar(
        select(func.coalesce(func.sum(Product.profit_dkk), Decimal("0"))).where(
            Product.sale_date >= start,
            Product.sale_date < end,
        )
    )

    return ReportSummaryOut(
        period_start=start,
        period_end=end,
        purchased_count=int(purchased_count or 0),
        sold_count=int(sold_count or 0),
        melted_count=int(melted_count or 0),
        total_purchase_value_dkk=str(purchase_total or Decimal("0")),
        total_sale_value_dkk=str(sale_total or Decimal("0")),
        total_profit_dkk=str(profit_total or Decimal("0")),
    )


@router.get("/daily", response_model=ReportSummaryOut)
async def daily(
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
) -> ReportSummaryOut:
    now = utc_now()
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end = start + timedelta(days=1)
    return await _summary_for_range(db, start, end)


@router.get("/weekly", response_model=ReportSummaryOut)
async def weekly(
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
) -> ReportSummaryOut:
    now = utc_now()
    start = now - timedelta(days=7)
    return await _summary_for_range(db, start, now)


@router.get("/monthly", response_model=ReportSummaryOut)
async def monthly(
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
) -> ReportSummaryOut:
    now = utc_now()
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return await _summary_for_range(db, start, now)


def _resolve_period_bounds(period: str, now: datetime) -> tuple[datetime | None, datetime | None]:
    if period == "all":
        return None, None
    if period == "daily":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=1)
        return start, end
    if period == "weekly":
        return now - timedelta(days=7), now
    if period == "monthly":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, now
    raise ValueError(f"Bilinmeyen period: {period}")


def _enum_to_value(value: Any) -> Any:
    return getattr(value, "value", value)


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    value = _enum_to_value(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def _serialize_export_rows(records: Iterable[Any]) -> list[list[str]]:
    rows: list[list[str]] = []
    for record in records:
        rows.append(
            [
                _format_cell(record.product_number),
                _format_cell(record.product_type),
                _format_cell(record.metal_type),
                _format_cell(record.status),
                _format_cell(record.purchase_price_dkk),
                _format_cell(record.sale_price_dkk),
                _format_cell(record.profit_dkk),
                _format_cell(record.purchase_date),
            ]
        )
    return rows


def _to_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _compute_export_totals(records: Iterable[Any]) -> dict[str, Any]:
    purchased_count = 0
    sold_count = 0
    melted_count = 0
    total_purchase = Decimal("0")
    total_sale = Decimal("0")
    total_profit = Decimal("0")

    for record in records:
        purchased_count += 1
        status_value = str(_enum_to_value(getattr(record, "status", "")))
        if status_value == ProductStatusEnum.SOLD.value:
            sold_count += 1
        if status_value == ProductStatusEnum.MELTED.value:
            melted_count += 1

        total_purchase += _to_decimal(getattr(record, "purchase_price_dkk", None))
        total_sale += _to_decimal(getattr(record, "sale_price_dkk", None))
        total_profit += _to_decimal(getattr(record, "profit_dkk", None))

    return {
        "purchased_count": purchased_count,
        "sold_count": sold_count,
        "melted_count": melted_count,
        "total_purchase_value_dkk": format(total_purchase, "f"),
        "total_sale_value_dkk": format(total_sale, "f"),
        "total_profit_dkk": format(total_profit, "f"),
    }


def _build_csv_content(records: list[Any]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([item[0] for item in EXPORT_HEADERS])
    for row in _serialize_export_rows(records):
        writer.writerow(row)
    output.seek(0)
    return output.getvalue()


def _build_xlsx_content(records: list[Any], period: str, generated_at: datetime) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapor"

    totals = _compute_export_totals(records)
    generated_label = generated_at.strftime("%Y-%m-%d %H:%M")

    sheet["A1"] = "SERO GULD CRM RAPORU"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.merge_cells("A1:H1")
    sheet["A2"] = f"Periyot: {period}"
    sheet["A3"] = f"Oluşturulma: {generated_label}"

    summary_row = 5
    sheet[f"A{summary_row}"] = "Alım"
    sheet[f"B{summary_row}"] = totals["purchased_count"]
    sheet[f"C{summary_row}"] = "Satış"
    sheet[f"D{summary_row}"] = totals["sold_count"]
    sheet[f"E{summary_row}"] = "Eritilen"
    sheet[f"F{summary_row}"] = totals["melted_count"]
    sheet[f"G{summary_row}"] = "Toplam Kâr"
    sheet[f"H{summary_row}"] = totals["total_profit_dkk"]

    header_row = 7
    for idx, (_, title) in enumerate(EXPORT_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=idx, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(fill_type="solid", fgColor="6F5B45")

    for row_idx, row in enumerate(_serialize_export_rows(records), start=header_row + 1):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    widths = [12, 14, 12, 12, 14, 14, 12, 24]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + col_idx)].width = width

    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return payload.getvalue()


def _build_pdf_content(records: list[Any], period: str, generated_at: datetime) -> bytes:
    payload = io.BytesIO()
    document = SimpleDocTemplate(
        payload,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()

    totals = _compute_export_totals(records)
    generated_label = generated_at.strftime("%Y-%m-%d %H:%M")

    elements: list[Any] = []
    elements.append(Paragraph("SERO GULD CRM RAPORU", styles["Heading1"]))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"Periyot: {period}", styles["Normal"]))
    elements.append(Paragraph(f"Oluşturulma: {generated_label}", styles["Normal"]))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(
            " / ".join(
                [
                    f"Alım: {totals['purchased_count']}",
                    f"Satış: {totals['sold_count']}",
                    f"Eritilen: {totals['melted_count']}",
                    f"Toplam Alım: {totals['total_purchase_value_dkk']} DKK",
                    f"Toplam Satış: {totals['total_sale_value_dkk']} DKK",
                    f"Toplam Kâr: {totals['total_profit_dkk']} DKK",
                ]
            ),
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))

    data = [[title for _, title in EXPORT_HEADERS]]
    max_pdf_rows = 500
    serialized = _serialize_export_rows(records)
    for row in serialized[:max_pdf_rows]:
        data.append(row)

    table = Table(
        data,
        repeatRows=1,
        colWidths=[60, 80, 75, 75, 80, 80, 70, 120],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6F5B45")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B9AB96")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f6f1e7")]),
            ]
        )
    )
    elements.append(table)

    if len(serialized) > max_pdf_rows:
        elements.append(Spacer(1, 10))
        elements.append(
            Paragraph(
                f"Not: PDF performansı için ilk {max_pdf_rows} satır gösterildi. "
                f"Tam veri için XLSX/CSV export kullanın.",
                styles["Italic"],
            )
        )

    document.build(elements)
    payload.seek(0)
    return payload.getvalue()


@router.get("/export")
async def export_report(
    format: str = Query(default="csv", pattern="^(csv|xlsx|pdf)$"),
    period: str = Query(default="all", pattern="^(all|daily|weekly|monthly)$"),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_admin),
):
    now = utc_now()
    start, end = _resolve_period_bounds(period, now)

    stmt = (
        select(
            Product.product_number,
            Product.product_type,
            Product.metal_type,
            Product.status,
            Product.purchase_price_dkk,
            Product.sale_price_dkk,
            Product.profit_dkk,
            Product.purchase_date,
        )
        .order_by(Product.created_at.desc())
    )
    if start and end:
        stmt = stmt.where(Product.purchase_date >= start, Product.purchase_date < end)

    records = (await db.execute(stmt)).all()

    timestamp = now.strftime("%Y%m%d-%H%M%S")
    filename_base = f"seroguld-report-{period}-{timestamp}"

    if format == "csv":
        content = _build_csv_content(records)
        return StreamingResponse(
            iter([content]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    if format == "xlsx":
        content = _build_xlsx_content(records, period=period, generated_at=now)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )

    content = _build_pdf_content(records, period=period, generated_at=now)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
    )
