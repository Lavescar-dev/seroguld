from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal
import httpx
import io
import json
import logging
import os
from types import SimpleNamespace
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile, WebSocket
from jose import JWTError, jwt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.afg import (
    apply_afg_route_requests,
    build_log_workspace,
    create_afg_melt_lot,
    post_afg_lines_route as legacy_post_afg_lines_route,
    update_afg_melt_lot,
)
from app.api.bootstrap import get_bootstrap as get_legacy_bootstrap
from app.api.customers import (
    delete_customer as legacy_delete_customer,
    get_customer as legacy_get_customer,
    get_customer_history as legacy_get_customer_history,
    get_customers as legacy_get_customers,
    post_customer as legacy_post_customer,
    put_customer as legacy_put_customer,
    search_customers as legacy_search_customers,
)
from app.api.deps import require_admin, require_password_change_complete
from app.api.inventory import (
    get_inventory_workspace as get_legacy_inventory_workspace,
)
from app.api.v2_inventory import router as inventory_router
from app.api.v2_document_artifacts import router as document_artifacts_router
from app.api.v2_excel_sessions import router as excel_sessions_router
from app.api.v2_support import apply_inventory_workbook_artifact_inputs, artifact_file_response
from app.api.pos import get_pos_documents as get_legacy_pos_documents
from app.api.pos import get_pos_document_detail as get_legacy_pos_document_detail
from app.api.pos import get_pos_receipt as get_legacy_pos_receipt
from app.api.pos import display_socket as legacy_display_socket
from app.api.antifraud import (
    get_order_detail as get_legacy_antifraud_order_detail,
    get_recent_orders as get_legacy_antifraud_recent_orders,
)
from app.api.products import (
    ai_describe as legacy_ai_describe,
    approve_product_manual_review as legacy_approve_product_manual_review,
    delete_photo as legacy_delete_photo,
    get_product_history as legacy_get_product_history,
    get_product_sync_log as legacy_get_product_sync_log,
    get_product_woocommerce_raw as legacy_get_product_woocommerce_raw,
    publish as legacy_publish_product,
    sync_product_sale_status as legacy_sync_product_sale_status,
    unpublish as legacy_unpublish_product,
    update_ai_describe as legacy_update_ai_describe,
    upload_photos as legacy_upload_photos,
)
from app.config import ROOT_ENV_FILE, get_settings, resolve_desktop_onlyoffice_jwt_secret
from app.database import get_db
from app.models.customer_identity import CustomerIdentityDocument
from app.models.enums import PosDocumentTypeEnum, PosSessionStatusEnum, PosTradeSideEnum, RoleEnum
from app.models.pos_document import PosDocument
from app.models.pos_session import PosSession
from app.models.transaction import Transaction
from app.models.transaction_line import TransactionLine
from app.models.user import User
from app.schemas.bootstrap import DesktopBootstrapOut
from app.schemas.antifraud import AntiFraudOrderOut, AntiFraudOrdersResponse
from app.schemas.afg import (
    AfgRouteBatchApplyRequest,
    AfgLogWorkspaceOut,
    AfgMeltLotCreateRequest,
    AfgMeltLotOut,
    AfgMeltLotUpdateRequest,
    AfgRouteRequest,
    AfgRouteResponse,
)
from app.schemas.customer import (
    CustomerAlisSummaryOut,
    CustomerCreate,
    CustomerDetailOut,
    CustomerListResponse,
    CustomerOut,
    CustomerUpdate,
)
from app.schemas.desktop_views import (
    DashboardCategorySpotOut,
    DashboardMonthlyPurchasePointOut,
    DashboardRecentCustomerOut,
    DashboardRecentPurchaseOut,
    DashboardScreenOut,
    SettingsScreenOut,
    SettingsScreenUpdateIn,
    UnicontaBulkRetryOut,
    UnicontaConfigOut,
    UnicontaConnectIn,
    UnicontaConnectOut,
    UnicontaFailedSyncRowOut,
    UnicontaHealthOut,
    UnicontaInvoiceCustomerOut,
    UnicontaInvoiceLineOut,
    UnicontaInvoiceOut,
    UnicontaInvoicesOut,
    UnicontaSyncSummaryOut,
)
from app.schemas.document_artifact import (
    DocumentArtifactCellEditsIn,
    OfficeForceSaveOut,
    DocumentArtifactPreviewOut,
    DocumentArtifactSheetPreviewOut,
    DocumentArtifactRecordOut,
    DocumentArtifactReconcilePreviewOut,
    OfficeDocumentLaunchOut,
    OfficeDocumentStatusOut,
    OfficeRuntimeStatusOut,
)
from app.schemas.inventory import InventoryWorkspaceOut
from app.schemas.pos import (
    PosPostalLookupOut,
    PosDisplayPreviewOut,
    PosDocumentDetailOut,
    PosDocumentListItemOut,
    PosSavedPurchaseListItemOut,
    PosSessionDisplayOut,
    PosSessionOutClerk,
    PosWorkspaceCustomerSelectRequest,
    PosWorkspaceCustomerUpdate,
    PosWorkspaceFinalizeRequest,
    PosWorkspaceFinalizeResponse,
    PosWorkspaceOpenRequest,
    PosWorkspaceOut,
    PosWorkspaceSectionsUpdate,
)
from app.schemas.product import (
    ProductAIDescriptionUpdate,
    ProductHistoryOut,
    ProductOut,
    ProductPublishRequest,
    ProductPublishResponse,
    WooSyncLogOut,
)
from app.schemas.runtime import DesktopDevSessionOut, RuntimeReadinessOut, RuntimeStatusOut
from app.schemas.woocommerce import WooWorkspaceOut, WooWorkspaceSummaryOut
from app.services.pos_service import (
    build_purchase_workspace,
    build_purchase_workspace_csv_export,
    build_purchase_workspace_xlsx_export,
    build_purchase_workspace_print_html,
    cancel_session,
    delete_purchase_document,
    display_snapshot,
    extract_purchase_payment_method,
    find_latest_draft_pos_session,
    finalize_purchase_workspace,
    get_next_reference_number_preview,
    get_pos_session_by_display_token_or_404,
    get_pos_session_or_404,
    open_purchase_document_for_edit,
    replace_purchase_workspace_sections,
    select_purchase_workspace_customer,
    store_purchase_workspace_preferences,
    update_purchase_workspace_draft_customer,
    update_purchase_workspace_customer,
)
from app.services.document_artifact_service import (
    XLSM_MIME,
    XLSX_MIME,
    artifact_absolute_path,
    build_afg_workspace_reconcile_preview,
    build_inventory_reconcile_preview,
    build_afg_document_preview,
    build_afg_workspace_preview,
    build_inventory_preview,
    build_log_preview,
    get_artifact_record,
    list_artifact_records,
    office_contract_version_for_kind,
    parse_afg_workspace_inputs_from_workbook,
    parse_inventory_workbook_inputs_from_workbook,
    parse_log_workbook_inputs_from_workbook,
    read_artifact_sync_metadata,
    resolve_artifact_conflict_state,
    sync_afg_document_artifact,
    sync_afg_workspace_artifact,
    sync_inventory_workbook_artifact,
    sync_log_workbook_artifact,
)
from app.services.office_host_service import office_host_service
from app.services.runtime_readiness import collect_runtime_readiness
from app.services.sequence_service import preview_afregnings_number, preview_invoice_number, preview_product_number
from app.services.uniconta_service import (
    UNICONTA_WEB_API_BASE,
    UnicontaError,
    UnicontaClient,
    get_uniconta_client,
    map_uniconta_invoice_to_dto,
    reset_uniconta_client,
)
from app.services.pos_document_service import format_document_number
from app.services.product_service import get_product_or_404, to_product_out, update_product, update_status
from app.services.pos_service import create_pos_session
from app.schemas.pos import PosSessionCreate
from app.utils.env_file import upsert_env_values
from app.utils.helpers import utc_now
from app.utils.security import decrypt_field, mask_cpr

router = APIRouter()
router.include_router(inventory_router)
logger = logging.getLogger(__name__)

MONTH_NAMES_TR = ["Oca", "Sub", "Mar", "Nis", "May", "Haz", "Tem", "Agu", "Eyl", "Eki", "Kas", "Ara"]
CAT_COLORS = {
    "kulce": "#b8860b",
    "sikke": "#d4a017",
    "taki": "#c09a3e",
    "gumus": "#8c8c8c",
    "platin_pd": "#6b7280",
}
CAT_LABELS = {
    "kulce": "Kulce (Au)",
    "sikke": "Sikke (Au)",
    "taki": "Taki (Au)",
    "gumus": "Gumus",
    "platin_pd": "Platin/Pd",
}


def _as_utc_datetime(value: datetime) -> datetime:
    """Normalize SQLite-naive and PostgreSQL-aware timestamps to UTC."""

    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _normalize_postal_lookup_code(value: str) -> str:
    return "".join(ch for ch in (value or "") if ch.isdigit())[:4]


def _first_named_item(value) -> str | None:
    if isinstance(value, list) and value:
        first = value[0]
        if isinstance(first, dict):
            name = str(first.get("navn") or "").strip()
            return name or None
        name = str(first or "").strip()
        return name or None
    if isinstance(value, dict):
        name = str(value.get("navn") or "").strip()
        return name or None
    text = str(value or "").strip()
    return text or None


async def _lookup_danish_postal_code(postal_code: str) -> PosPostalLookupOut:
    normalized = _normalize_postal_lookup_code(postal_code)
    if len(normalized) != 4:
        raise HTTPException(status_code=422, detail="Postnr. 4 rakam olmalı.")

    url = f"https://api.dataforsyningen.dk/postnumre/{normalized}"
    try:
        async with httpx.AsyncClient(timeout=4.0, follow_redirects=True) as client:
            response = await client.get(url, headers={"Accept": "application/json"})
    except httpx.HTTPError:
        return PosPostalLookupOut(postal_code=normalized, found=False, available=False)

    if response.status_code == 404:
        return PosPostalLookupOut(postal_code=normalized, found=False, available=True)

    response.raise_for_status()
    payload = response.json() if response.content else {}
    postal_district = str(payload.get("navn") or "").strip() or None
    municipality_name = _first_named_item(payload.get("kommuner"))
    region_name = _first_named_item(payload.get("regioner"))

    return PosPostalLookupOut(
        postal_code=normalized,
        found=bool(postal_district),
        available=True,
        postal_district=postal_district,
        municipality_name=municipality_name,
        region_name=region_name,
    )
BACKEND_STARTED_AT = utc_now()
DESKTOP_SESSION_FILE = ROOT_ENV_FILE.parent / ".run" / "desktop-dev-session.json"


def _to_float(value: Decimal | str | float | int | None) -> float:
    return float(Decimal(str(value or 0)))


def _to_bool_status(value: str | None) -> bool:
    return bool((value or "").strip())


def _format_purchase_line_fineness(*, purity_karat: str | None, purity_percentage: Decimal | str | None) -> str:
    purity = Decimal(str(purity_percentage or 0))
    if purity_karat:
        return f"{purity_karat} / {purity:.2f}%"
    return f"{purity:.2f}%"


def _humanize_pos_label(value: str | None) -> str:
    if not value:
        return "Kalem"
    return value.replace("_", " ").title()


def _read_desktop_dev_session() -> DesktopDevSessionOut | None:
    if not DESKTOP_SESSION_FILE.exists():
        return None
    try:
        payload = json.loads(DESKTOP_SESSION_FILE.read_text(encoding="utf-8"))
    except Exception:
        return None

    session = DesktopDevSessionOut(
        mode=str(payload.get("mode") or "desktop-dev"),
        started_at=str(payload.get("started_at") or ""),
        backend_url=str(payload.get("backend_url") or "http://127.0.0.1:8100"),
        frontend_url=str(payload.get("frontend_url") or "http://127.0.0.1:3300"),
        frontend_mode=str(payload.get("frontend_mode") or "vite-dev"),
        tauri_mode=str(payload.get("tauri_mode") or "tauri-dev-url"),
        backend_pid=int(payload["backend_pid"]) if payload.get("backend_pid") else None,
        frontend_pid=int(payload["frontend_pid"]) if payload.get("frontend_pid") else None,
        tauri_pid=int(payload["tauri_pid"]) if payload.get("tauri_pid") else None,
    )
    tracked_pids = [pid for pid in (session.backend_pid, session.frontend_pid, session.tauri_pid) if pid]
    if tracked_pids and not any(_pid_is_alive(pid) for pid in tracked_pids):
        return None
    return session


def _pid_is_alive(pid: int) -> bool:
    try:
        os.kill(pid, 0)
    except OSError:
        return False
    return True


async def _ensure_alis_workspace_artifact(
    db: AsyncSession,
    workspace: PosWorkspaceOut,
    *,
    force_sync: bool,
) -> None:
    if not force_sync and await get_artifact_record(db, f"alis.workspace.{workspace.session.id}"):
        return
    await sync_afg_workspace_artifact(db, workspace)
    await db.commit()


async def _ensure_log_artifact(
    db: AsyncSession,
    workspace: AfgLogWorkspaceOut,
    *,
    year: int,
    create_snapshot: bool,
    force_sync: bool,
) -> None:
    if not force_sync and await get_artifact_record(db, f"log.live.{year}"):
        return
    await sync_log_workbook_artifact(db, workspace, year=year, create_snapshot=create_snapshot)
    await db.commit()


def _workspace_has_business_inputs(workspace: PosWorkspaceOut) -> bool:
    if any(row.gram > 0 for row in workspace.gold_rows):
        return True
    if any(row.gram > 0 for row in workspace.silver_rows):
        return True
    customer = workspace.customer
    if any(
        [
            str(customer.name or "").strip(),
            str(customer.email or "").strip(),
            str(customer.phone or "").strip(),
            str(customer.address or "").strip(),
            str(customer.postal_code or "").strip(),
            str(customer.cpr_number or "").strip(),
            str(customer.identity_doc_number or "").strip(),
        ]
    ):
        return True
    if any([str(workspace.bank_info.reg_number or "").strip(), str(workspace.bank_info.account_number or "").strip()]):
        return True
    if any(
        str(row.code or "").strip() or str(row.fineness or "").strip() or row.gram > 0
        for row in workspace.invoice_gold.rows
    ):
        return True
    if any(str(value or "").strip() for value in workspace.invoice_gold.footer_lines):
        return True
    if any(
        str(row.text or "").strip() or (row.quantity is not None and row.quantity > 0) or row.unit_price_dkk > 0
        for row in workspace.invoice_misc.rows
    ):
        return True
    return False


def _parsed_afg_inputs_look_blank(parsed) -> bool:
    if any(row.gram > 0 for row in parsed.sections.gold_rows):
        return False
    if any(row.gram > 0 for row in parsed.sections.silver_rows):
        return False
    if any(
        [
            str(parsed.customer.name or "").strip(),
            str(parsed.customer.email or "").strip(),
            str(parsed.customer.phone or "").strip(),
            str(parsed.customer.address or "").strip(),
            str(parsed.customer.postal_code or "").strip(),
            str(parsed.customer.cpr_number or "").strip(),
            str(parsed.customer.identity_doc_number or "").strip(),
        ]
    ):
        return False
    bank_info = parsed.sections.bank_info
    if bank_info and any([str(bank_info.reg_number or "").strip(), str(bank_info.account_number or "").strip()]):
        return False
    invoice_gold = parsed.sections.invoice_gold
    if invoice_gold and any(
        str(row.code or "").strip() or str(row.fineness or "").strip() or row.gram > 0
        for row in invoice_gold.rows
    ):
        return False
    if invoice_gold and any(str(value or "").strip() for value in invoice_gold.footer_lines):
        return False
    invoice_misc = parsed.sections.invoice_misc
    if invoice_misc and any(
        str(row.text or "").strip() or (row.quantity is not None and row.quantity > 0) or row.unit_price_dkk > 0
        for row in invoice_misc.rows
    ):
        return False
    return True


def _customer_update_has_business_inputs(customer) -> bool:
    return any(
        [
            str(customer.name or "").strip(),
            str(customer.email or "").strip(),
            str(customer.phone or "").strip(),
            str(customer.address or "").strip(),
            str(customer.postal_code or "").strip(),
            str(customer.city or "").strip(),
            str(customer.cpr_number or "").strip(),
            str(customer.identity_doc_number or "").strip(),
            str(customer.identity_doc_country or "").strip(),
            str(customer.identity_doc_type or "").strip(),
        ]
    )


async def _apply_afg_workspace_artifact_inputs(
    db: AsyncSession,
    *,
    pos_session: PosSession,
    workbook_bytes: bytes,
    office_lineage: bool = False,
    allow_full_clear: bool = False,
    office_workspace_revision: int | None = None,
    enforce_base_version: bool = True,
) -> PosWorkspaceOut:
    try:
        parsed = parse_afg_workspace_inputs_from_workbook(workbook_bytes)
    except (ValidationError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if enforce_base_version and parsed.base_version and not office_lineage:
        record = await get_artifact_record(db, f"alis.workspace.{pos_session.id}")
        if record is not None:
            conflict_state = resolve_artifact_conflict_state(
                current_revision=getattr(record, "revision", 1),
                incoming_revision=parsed.base_version,
            )
            if conflict_state != "clean":
                raise HTTPException(
                    status_code=409,
                    detail=f"AFG artifact conflict_state={conflict_state}; önce yenileyin.",
                )
    current_workspace = await build_purchase_workspace(db, pos_session=pos_session)
    if not office_lineage and not allow_full_clear and _workspace_has_business_inputs(current_workspace) and _parsed_afg_inputs_look_blank(parsed):
        logger.warning(
            "Rejected blank AFG callback for workspace %s because parsed workbook had no business inputs",
            pos_session.id,
        )
        raise HTTPException(
            status_code=409,
            detail="ONLYOFFICE callback boş veya eski workbook döndürdü; mevcut taslak korunuyor.",
        )
    workspace = current_workspace
    customer_mutation_started = False
    customer_payload = parsed.customer.model_copy(
        update={"base_revision": office_workspace_revision}
    ) if office_workspace_revision is not None else parsed.customer
    if pos_session.customer_id is not None:
        workspace = await update_purchase_workspace_customer(
            db,
            pos_session=pos_session,
            payload=customer_payload,
            commit=False,
            emit=False,
            claim_revision=office_workspace_revision is not None,
        )
        customer_mutation_started = True
    elif _customer_update_has_business_inputs(parsed.customer):
        workspace = await update_purchase_workspace_draft_customer(
            db,
            pos_session=pos_session,
            payload=customer_payload,
            commit=False,
            emit=False,
            claim_revision=office_workspace_revision is not None,
        )
        customer_mutation_started = True
    workspace = await replace_purchase_workspace_sections(
        db,
        pos_session=pos_session,
        payload=parsed.sections.model_copy(
            update={
                "market_rates": parsed.sections.market_rates or current_workspace.market_rates,
                **(
                    {"base_revision": office_workspace_revision}
                    if office_workspace_revision is not None and not customer_mutation_started
                    else {}
                ),
            }
        ),
        commit=True,
        emit=True,
        lock=not customer_mutation_started,
        claim_revision=not customer_mutation_started,
    )
    await _ensure_alis_workspace_artifact(db, workspace, force_sync=True)
    return workspace


async def _apply_log_workbook_artifact_inputs(
    db: AsyncSession,
    *,
    year: int,
    workbook_bytes: bytes,
    create_snapshot: bool,
) -> AfgLogWorkspaceOut:
    current_workspace = await build_log_workspace(db, q=None, limit=200)
    try:
        parsed = parse_log_workbook_inputs_from_workbook(
            workbook_bytes,
            year=year,
            current_workspace=current_workspace,
        )
    except (ValidationError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if parsed.base_version:
        record = await get_artifact_record(db, f"log.live.{year}")
        if record is not None:
            conflict_state = resolve_artifact_conflict_state(
                current_revision=getattr(record, "revision", 1),
                incoming_revision=parsed.base_version,
            )
            if conflict_state != "clean":
                raise HTTPException(
                    status_code=409,
                    detail=f"Log artifact conflict_state={conflict_state}; önce yenileyin.",
                )
    system_actor = SimpleNamespace(id=None)
    if parsed.route_updates:
        await apply_afg_route_requests(
            db=db,
            route_requests=[edit.payload for edit in parsed.route_updates],
            actor_id=system_actor.id,
        )
    for create in parsed.lot_creates:
        lot = await create_afg_melt_lot(db, payload=create.create_payload)
        await update_afg_melt_lot(db, lot_id=UUID(str(lot.id)), payload=create.update_payload)
    for update in parsed.lot_updates:
        await update_afg_melt_lot(db, lot_id=update.lot_id, payload=update.payload)

    workspace = await build_log_workspace(db, q=None, limit=200)
    await _ensure_log_artifact(
        db,
        workspace,
        year=year,
        create_snapshot=create_snapshot,
        force_sync=True,
    )
    return workspace


async def _apply_office_session_content(
    db: AsyncSession,
    *,
    entry,
    workbook_bytes: bytes,
) -> PosSession | None:
    office_pos_session: PosSession | None = None
    applied_workspace_revision: int | None = None
    if entry.artifact_key:
        record = await get_artifact_record(db, entry.artifact_key)
        if record is None:
            raise HTTPException(status_code=409, detail="Office artifact conflict_state=invalid; artifact bulunamadı.")
        expected_key = "live" if entry.kind == "depolama" else entry.key
        try:
            metadata = read_artifact_sync_metadata(
                workbook_bytes,
                expected_kind=entry.kind,
                expected_key=expected_key,
            )
        except (OSError, ValueError, KeyError, TypeError) as exc:
            raise HTTPException(
                status_code=409,
                detail="Office artifact conflict_state=invalid; revision metadata bulunamadı.",
            ) from exc
        current_revision = int(getattr(record, "revision", 1) or 1)
        if entry.kind == "alis-workspace":
            office_pos_session = await get_pos_session_or_404(db, UUID(entry.key))
            current_workspace = await build_purchase_workspace(db, pos_session=office_pos_session)
            workbook_workspace_revision = getattr(metadata, "workspace_revision", None)
            if workbook_workspace_revision is None:
                raise HTTPException(
                    status_code=409,
                    detail="Office artifact conflict_state=stale_lineage; workspace revision metadata bulunamadı.",
                )
            try:
                incoming_workspace_revision = int(workbook_workspace_revision)
            except (TypeError, ValueError) as exc:
                raise HTTPException(
                    status_code=409,
                    detail="Office artifact conflict_state=invalid; workspace revision metadata geçersiz.",
                ) from exc
            launch_workspace_revision = getattr(entry, "workspace_launch_revision", None)
            applied_workspace_revision = getattr(entry, "workspace_applied_revision", None)
            if launch_workspace_revision is None:
                launch_workspace_revision = incoming_workspace_revision
            if applied_workspace_revision is None:
                applied_workspace_revision = launch_workspace_revision
            if incoming_workspace_revision not in {int(launch_workspace_revision), int(applied_workspace_revision)}:
                raise HTTPException(
                    status_code=409,
                    detail="Office artifact conflict_state=stale_lineage; workbook workspace revision ileride.",
                )
            if int(current_workspace.workspace_revision) != int(applied_workspace_revision):
                raise HTTPException(
                    status_code=409,
                    detail="Office artifact conflict_state=external_write; workspace önce güncellendi.",
                )
            launch_revision = int(getattr(entry, "launch_revision", entry.artifact_revision) or entry.artifact_revision or 1)
            applied_revision = int(getattr(entry, "artifact_revision", launch_revision) or launch_revision)
            try:
                incoming_base_revision = int(metadata.base_version) if metadata.base_version is not None else None
            except (TypeError, ValueError):
                incoming_base_revision = None
            if (
                incoming_base_revision is None
                or incoming_base_revision < launch_revision
                or incoming_base_revision > applied_revision
            ):
                raise HTTPException(
                    status_code=409,
                    detail="Office artifact conflict_state=stale_lineage; Office belgesini yeniden açın.",
                )
            if current_revision != applied_revision:
                raise HTTPException(
                    status_code=409,
                    detail="Office artifact conflict_state=external_write; önce güncel workspace alın.",
                )
        else:
            conflict_state = resolve_artifact_conflict_state(
                current_revision=current_revision,
                incoming_revision=metadata.base_version,
            )
            if conflict_state != "clean":
                raise HTTPException(
                    status_code=409,
                    detail=f"Office artifact conflict_state={conflict_state}; önce yenileyin.",
                )
    if entry.kind == "alis-workspace":
        pos_session = office_pos_session or await get_pos_session_or_404(db, UUID(entry.key))
        return await _apply_afg_workspace_artifact_inputs(
            db,
            pos_session=pos_session,
            workbook_bytes=workbook_bytes,
            office_lineage=True,
            office_workspace_revision=int(applied_workspace_revision) if applied_workspace_revision is not None else None,
        )
    if entry.kind == "depolama":
        await apply_inventory_workbook_artifact_inputs(db, workbook_bytes=workbook_bytes, create_snapshot=False)
        return None
    if entry.kind == "log":
        raise HTTPException(status_code=409, detail="Log workbook salt okunurdur")
    raise HTTPException(status_code=409, detail="Bu office oturumu yazma desteklemiyor")


def _verify_onlyoffice_callback_token(request: Request, payload: dict) -> None:
    settings = get_settings()
    header_value = request.headers.get("authorization", "").strip()
    raw_token = ""
    if header_value.lower().startswith("bearer "):
        raw_token = header_value[7:].strip()
    elif isinstance(payload.get("token"), str):
        raw_token = str(payload["token"]).strip()
    if not raw_token:
        return
    try:
        jwt.decode(
            raw_token,
            resolve_desktop_onlyoffice_jwt_secret(settings.onlyoffice_jwt_secret),
            algorithms=["HS256"],
        )
    except JWTError as exc:
        raise HTTPException(status_code=401, detail="ONLYOFFICE callback token geçersiz") from exc


async def _office_preview_for_kind(
    db: AsyncSession,
    *,
    kind: str,
    key: str,
    admin: User,
) -> tuple[DocumentArtifactPreviewOut, bool]:
    try:
        if kind == "alis-workspace":
            session_id = UUID(key)
            pos_session = await get_pos_session_or_404(db, session_id)
            workspace = await build_purchase_workspace(db, pos_session=pos_session)
            bundle = await sync_afg_workspace_artifact(db, workspace)
            await db.commit()
            preview = build_afg_workspace_preview(workspace, artifact=bundle.artifact)
            can_write = pos_session.status == PosSessionStatusEnum.DRAFT
            if not can_write:
                preview = preview.model_copy(
                    update={
                        "import_supported": False,
                        "external_edit_supported": False,
                        "editable_cells": [],
                    }
                )
            return preview, can_write
        if kind == "alis-document":
            sequence_no = int(key)
            detail = await get_legacy_pos_document_detail(sequence_no=sequence_no, db=db, _=admin)
            bundle = await sync_afg_document_artifact(db, detail)
            await db.commit()
            return build_afg_document_preview(detail, artifact=bundle.artifact), False
        if kind == "depolama":
            workspace = await get_legacy_inventory_workspace(q=None, db=db, _=admin)
            bundle = await sync_inventory_workbook_artifact(db, workspace, create_snapshot=False)
            await db.commit()
            return build_inventory_preview(workspace, artifact=bundle.artifact), True
        if kind == "log":
            year = _default_artifact_year(int(key))
            workspace = await build_log_workspace(db, q=None, limit=200)
            bundle = await sync_log_workbook_artifact(db, workspace, year=year, create_snapshot=False)
            await db.commit()
            return build_log_preview(workspace, year=year, artifact=bundle.artifact), False
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Geçersiz office belge anahtarı") from exc
    raise HTTPException(status_code=404, detail="Office belge bulunamadı")


def _office_launch_sheet_meta(sheet: DocumentArtifactSheetPreviewOut) -> DocumentArtifactSheetPreviewOut:
    return DocumentArtifactSheetPreviewOut(
        name=sheet.name,
        mode=sheet.mode,
        system_sync=sheet.system_sync,
        note=sheet.note,
    )


async def _office_status_for_kind(
    db: AsyncSession,
    *,
    kind: str,
    key: str,
    admin: User,
    access_token: str | None = None,
) -> OfficeDocumentStatusOut:
    try:
        can_write = False
        import_supported = False
        artifact = None
        if kind == "alis-workspace":
            session_id = UUID(key)
            artifact = await get_artifact_record(db, f"alis.workspace.{session_id}")
            if artifact is None:
                pos_session = await get_pos_session_or_404(db, session_id)
                workspace = await build_purchase_workspace(db, pos_session=pos_session)
                artifact = (await sync_afg_workspace_artifact(db, workspace)).artifact
                await db.commit()
            pos_session = await get_pos_session_or_404(db, session_id)
            can_write = pos_session.status == PosSessionStatusEnum.DRAFT
            import_supported = can_write
        elif kind == "alis-document":
            sequence_no = int(key)
            artifact = await get_artifact_record(db, f"alis.document.{sequence_no}")
            if artifact is None:
                detail = await get_legacy_pos_document_detail(sequence_no=sequence_no, db=db, _=admin)
                artifact = (await sync_afg_document_artifact(db, detail)).artifact
                await db.commit()
        elif kind == "depolama":
            artifact = await get_artifact_record(db, "depolama.live")
            if artifact is None:
                workspace = await get_legacy_inventory_workspace(q=None, db=db, _=admin)
                artifact = (await sync_inventory_workbook_artifact(db, workspace, create_snapshot=False)).artifact
                await db.commit()
            can_write = True
            import_supported = True
        elif kind == "log":
            year = _default_artifact_year(int(key))
            artifact = await get_artifact_record(db, f"log.live.{year}")
            if artifact is None:
                workspace = await build_log_workspace(db, q=None, limit=200)
                artifact = (await sync_log_workbook_artifact(db, workspace, year=year, create_snapshot=False)).artifact
                await db.commit()
            # Log/history workbooks are projections and intentionally cannot
            # write through the controlled grid or Excel bridge.
            can_write = False
            import_supported = False
        else:
            raise HTTPException(status_code=404, detail="Office belge bulunamadı")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Geçersiz office belge anahtarı") from exc

    live_sync_state, live_sync_message, last_callback_at = office_host_service.live_sync_status(
        access_token,
        kind=kind,
        key=key,
    )
    return OfficeDocumentStatusOut(
        kind=kind,
        key=key,
        provider=office_host_service.provider_for_kind(kind),
        provider_label=office_host_service.provider_label_for_kind(kind),
        provider_branding_level=office_host_service.provider_branding_level_for_kind(kind),
        contract_version=office_contract_version_for_kind(kind),
        artifact=DocumentArtifactRecordOut.model_validate(artifact) if artifact is not None else None,
        can_write=can_write,
        import_supported=import_supported,
        office_available=await office_host_service.is_available(kind),
        live_sync_state=live_sync_state,
        live_sync_message=live_sync_message,
        last_callback_at=last_callback_at,
    )


def _resolve_office_access_token(request: Request) -> str | None:
    token = request.query_params.get("access_token")
    if token:
        return token
    auth_header = request.headers.get("authorization", "")
    if auth_header.lower().startswith("bearer "):
        return auth_header[7:].strip()
    return None


async def _office_artifact_record_or_404(db: AsyncSession, access_token: str):
    entry = office_host_service.get_session(access_token)
    if entry is None:
        raise HTTPException(status_code=404, detail="Office oturumu bulunamadı")
    if not entry.artifact_key:
        raise HTTPException(status_code=404, detail="Office artifact anahtarı bulunamadı")
    record = await get_artifact_record(db, entry.artifact_key)
    if record is None:
        raise HTTPException(status_code=404, detail="Office artifact bulunamadı")
    return entry, record


def _default_artifact_year(year: int | None) -> int:
    return year or utc_now().year


async def _build_alis_saved_purchase_items(
    db: AsyncSession,
    *,
    admin: User,
    q: str | None,
    date: str | None,
    limit: int,
) -> list[PosSavedPurchaseListItemOut]:
    raw_items = await get_legacy_pos_documents(
        q=q,
        kind="afregningsbilag",
        limit=max(limit, 200),
        db=db,
        _=admin,
    )
    items = [
        item
        for item in raw_items
        if str(item.status or "").strip().lower() != "cancelled"
        if not date or item.issued_at.date().isoformat() == date
    ][:limit]
    if not items:
        return []

    sequence_numbers = [item.sequence_no for item in items]
    extra_rows = (
        await db.execute(
            select(
                PosDocument.sequence_no,
                PosDocument.notes,
                PosSession.customer_id,
                User.address_encrypted,
                User.postal_code,
                User.cpr_number_encrypted,
                CustomerIdentityDocument.identity_doc_number_encrypted,
                PosDocument.uniconta_sync_status,
                PosDocument.uniconta_invoice_number,
                PosDocument.uniconta_sync_error,
            )
            .join(PosSession, PosSession.id == PosDocument.pos_session_id)
            .outerjoin(User, User.id == PosSession.customer_id)
            .outerjoin(CustomerIdentityDocument, CustomerIdentityDocument.user_id == PosSession.customer_id)
            .where(PosDocument.sequence_no.in_(sequence_numbers))
        )
    ).all()
    extra_map: dict[int, dict[str, str | None]] = {}
    for (
        sequence_no,
        notes,
        customer_id,
        address_encrypted,
        postal_code,
        cpr_encrypted,
        identity_encrypted,
        uc_status,
        uc_invoice_no,
        uc_error,
    ) in extra_rows:
        cpr_masked = None
        cpr_plain = None
        identity_plain = None
        address = None
        if address_encrypted:
            try:
                address = decrypt_field(address_encrypted)
            except Exception:
                address = None
        if cpr_encrypted:
            try:
                cpr_plain = decrypt_field(cpr_encrypted)
            except Exception:
                cpr_plain = None
            try:
                cpr_masked = mask_cpr(decrypt_field(cpr_encrypted))
            except Exception:
                cpr_masked = None
        if identity_encrypted:
            try:
                identity_plain = decrypt_field(identity_encrypted)
            except Exception:
                identity_plain = None
        extra_map[int(sequence_no)] = {
            "customer_id": str(customer_id) if customer_id else None,
            "address": address,
            "payment_method": extract_purchase_payment_method(notes),
            "postal_code": postal_code,
            "cpr_masked": cpr_masked,
            "cpr": cpr_plain,
            "identity_doc_number": identity_plain,
            "uniconta_sync_status": uc_status,
            "uniconta_invoice_number": uc_invoice_no,
            "uniconta_sync_error": uc_error,
        }

    preview_rows = (
        await db.execute(
            select(
                Transaction.pos_document_sequence_no,
                TransactionLine.line_no,
                TransactionLine.product_type,
                TransactionLine.metal_type,
                TransactionLine.weight_grams,
                TransactionLine.purity_karat,
                TransactionLine.purity_percentage,
                TransactionLine.line_total_dkk,
            )
            .join(TransactionLine, TransactionLine.transaction_id == Transaction.id)
            .where(Transaction.pos_document_sequence_no.in_(sequence_numbers))
            .order_by(Transaction.pos_document_sequence_no.asc(), TransactionLine.line_no.asc())
        )
    ).all()
    preview_map: dict[int, dict[str, list[dict[str, object]]]] = defaultdict(lambda: {"gold": [], "silver": []})
    for sequence_no, line_no, product_type, metal_type, weight_grams, purity_karat, purity_percentage, line_total_dkk in preview_rows:
        if sequence_no is None:
            continue
        bucket = "silver" if str(metal_type or "").lower() == "silver" else "gold"
        preview_map[int(sequence_no)][bucket].append(
            {
                "line_no": int(line_no),
                "type_label": f"{purity_karat or f'{Decimal(str(purity_percentage or 0)):.0f}%'} {_humanize_pos_label(product_type)}",
                "weight_grams": Decimal(str(weight_grams or 0)),
                "purity_label": _format_purchase_line_fineness(
                    purity_karat=purity_karat,
                    purity_percentage=purity_percentage,
                ),
                "line_total_dkk": Decimal(str(line_total_dkk or 0)),
            }
        )

    return [
        PosSavedPurchaseListItemOut(
            sequence_no=item.sequence_no,
            session_id=item.session_id,
            session_code=item.session_code,
            document_number=item.document_number,
            issued_at=item.issued_at,
            customer_id=extra_map.get(item.sequence_no, {}).get("customer_id"),
            customer_name=item.customer_name,
            customer_phone=item.customer_phone,
            customer_email=item.customer_email,
            customer_address=extra_map.get(item.sequence_no, {}).get("address"),
            customer_postal_code=extra_map.get(item.sequence_no, {}).get("postal_code"),
            customer_cpr=extra_map.get(item.sequence_no, {}).get("cpr"),
            customer_cpr_masked=extra_map.get(item.sequence_no, {}).get("cpr_masked"),
            customer_identity_doc_number=extra_map.get(item.sequence_no, {}).get("identity_doc_number"),
            gross_amount_dkk=item.gross_amount_dkk,
            total_weight_grams=item.total_weight_grams,
            line_count=item.line_count,
            payment_method=extra_map.get(item.sequence_no, {}).get("payment_method"),
            gold_preview_items=preview_map.get(item.sequence_no, {}).get("gold", []),
            silver_preview_items=preview_map.get(item.sequence_no, {}).get("silver", []),
            can_edit=True,
            can_delete=True,
            uniconta_sync_status=extra_map.get(item.sequence_no, {}).get("uniconta_sync_status"),
            uniconta_invoice_number=extra_map.get(item.sequence_no, {}).get("uniconta_invoice_number"),
            uniconta_sync_error=extra_map.get(item.sequence_no, {}).get("uniconta_sync_error"),
        )
        for item in items
    ]


def _build_alis_list_xlsx(items: list[PosSavedPurchaseListItemOut]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kayıtlı Alışlar"
    headers = ["Afg. Nr.", "Dato", "Müşteri", "Tlf / E-mail", "Ödeme", "Toplam Gram", "I alt (DKK)", "Kalem"]
    fill = PatternFill(fill_type="solid", fgColor="6F5B45")
    for col_idx, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill

    for row_idx, item in enumerate(items, start=2):
        values = [
            item.document_number,
            item.issued_at.strftime("%d.%m.%Y"),
            item.customer_name or "—",
            item.customer_phone or item.customer_email or "—",
            "Kontant" if item.payment_method == "cash" else "Bankoverførsel",
            float(Decimal(str(item.total_weight_grams or 0))),
            float(Decimal(str(item.gross_amount_dkk or 0))),
            item.line_count,
        ]
        for col_idx, value in enumerate(values, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    widths = [20, 14, 26, 24, 18, 14, 16, 10]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + col_idx)].width = width

    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return payload.getvalue()


def _build_alis_document_xlsx(detail: PosDocumentDetailOut) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Afregningsbilag"

    sheet["A1"] = "SERO GULD"
    sheet["A1"].font = Font(bold=True, size=18)
    sheet["A2"] = "Køb og salg af guld, sølv og smykker"
    sheet["F1"] = "Afregningsnr."
    sheet["F1"].font = Font(bold=True)
    sheet["G1"] = detail.document_number
    sheet["G1"].font = Font(bold=True, size=15)
    sheet["F2"] = "Dato"
    sheet["G2"] = detail.issued_at.strftime("%d.%m.%Y %H:%M")

    payment_method = extract_purchase_payment_method(detail.notes)
    detail_rows = [
        ("Navn", detail.customer_name or "—"),
        ("CPR", "—"),
        ("Kørekort / Pas", "—"),
        ("Telefon", detail.customer_phone or "—"),
        ("E-mail", detail.customer_email or "—"),
        ("Adresse", detail.customer_address or "—"),
        ("Betaling", "Kontant" if payment_method == "cash" else "Bankoverførsel"),
    ]
    start_row = 4
    for idx, (label, value) in enumerate(detail_rows, start=start_row):
        sheet[f"A{idx}"] = label
        sheet[f"A{idx}"].font = Font(bold=True)
        sheet[f"B{idx}"] = value
        sheet.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=4)

    header_row = start_row + len(detail_rows) + 2
    headers = ["Type", "Saflık", "Lødighed", "Gram", "Birim", "Toplam"]
    fill = PatternFill(fill_type="solid", fgColor="6F5B45")
    for col_idx, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill

    row_idx = header_row + 1
    for line in detail.lines:
        sheet.cell(row=row_idx, column=1, value=line.product_type or "Kalem")
        sheet.cell(
            row=row_idx,
            column=2,
            value=_format_purchase_line_fineness(
                purity_karat=line.purity_karat,
                purity_percentage=line.purity_percentage,
            ),
        )
        sheet.cell(row=row_idx, column=3, value=f"{(Decimal(str(line.purity_percentage or 0)) * Decimal('10')):.0f}")
        sheet.cell(row=row_idx, column=4, value=float(Decimal(str(line.weight_grams or 0))))
        sheet.cell(row=row_idx, column=5, value=float(Decimal(str(line.rate_dkk or 0))))
        sheet.cell(row=row_idx, column=6, value=float(Decimal(str(line.line_total_dkk or 0))))
        row_idx += 1

    total_row = row_idx + 1
    total_fill = PatternFill(fill_type="solid", fgColor="1F6B4F")
    sheet.cell(row=total_row, column=5, value="I alt").font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=total_row, column=5).fill = total_fill
    sheet.cell(row=total_row, column=6, value=float(Decimal(str(detail.gross_amount_dkk or 0)))).font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=total_row, column=6).fill = total_fill

    widths = [22, 18, 14, 12, 14, 16]
    for col_idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + col_idx)].width = width

    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return payload.getvalue()


def _build_settings_screen_out() -> SettingsScreenOut:
    settings = get_settings()
    configured_secret_fields = [
        field_name
        for field_name, value in (
            ("openai_api_key", settings.openai_api_key),
            ("opmc_api_key", settings.opmc_api_key),
            ("opmc_webhook_secret", settings.opmc_webhook_secret),
            ("woo_consumer_key", settings.woocommerce_consumer_key),
            ("woo_consumer_secret", settings.woocommerce_consumer_secret),
            ("woo_webhook_secret", settings.woocommerce_webhook_secret),
            ("wp_app_password", settings.wp_app_password),
            ("uniconta_password", settings.uniconta_password),
            ("uniconta_api_key", settings.uniconta_api_key),
            ("metals_dev_api_key", settings.metals_dev_api_key),
        )
        if str(value or "").strip()
    ]
    return SettingsScreenOut(
        openai_api_key="",
        openai_model=settings.openai_model,
        openai_max_tokens=str(settings.openai_max_tokens),
        opmc_api_url=settings.opmc_api_url,
        opmc_api_key="",
        opmc_webhook_secret="",
        woo_store_url=settings.woocommerce_base_url,
        woo_consumer_key="",
        woo_consumer_secret="",
        woo_webhook_secret="",
        wp_site_url=settings.wordpress_base_url,
        wp_username=settings.wp_app_username,
        wp_app_password="",
        uniconta_api_url=UNICONTA_WEB_API_BASE,
        uniconta_username=settings.uniconta_username,
        uniconta_password="",
        uniconta_company_id=settings.uniconta_company_id,
        uniconta_api_key="",
        uniconta_purchase_vat_code_25=settings.uniconta_purchase_vat_code_25,
        uniconta_purchase_vat_code_0=settings.uniconta_purchase_vat_code_0,
        metals_dev_api_key="",
        market_gold=str(settings.inventory_market_gold_dkk),
        market_silver=str(settings.inventory_market_silver_dkk),
        market_platin=str(settings.inventory_market_platinum_dkk),
        market_palladyum=str(settings.inventory_market_palladium_dkk),
        market_rates_live_enabled=bool(settings.market_rates_live_enabled),
        market_rates_live_fx_enabled=bool(settings.market_rates_live_fx_enabled),
        market_rates_live_platinum_enabled=bool(settings.market_rates_live_platinum_enabled),
        market_rates_live_palladium_enabled=bool(settings.market_rates_live_palladium_enabled),
        firma_adi=settings.invoice_seller_name,
        firma_cvr=settings.invoice_seller_cvr,
        firma_telefon=settings.invoice_seller_phone,
        firma_email=settings.invoice_seller_email,
        firma_adres=settings.invoice_seller_address_line1,
        secret_fields_configured=configured_secret_fields,
    )


def _build_uniconta_config_out(
    message: str | None = None,
    *,
    connection_status: str | None = None,
    last_refreshed_at: str | None = None,
) -> UnicontaConfigOut:
    settings = get_settings()
    health = get_uniconta_client().get_health_snapshot()
    configured = bool(
        settings.uniconta_company_id.strip()
        and settings.uniconta_username.strip()
        and settings.uniconta_password.strip()
    )
    derived_status = connection_status
    if derived_status is None:
        if not configured:
            derived_status = "bagli_degil"
        elif health.get("last_call_ok") is True and health.get("has_token"):
            derived_status = "bagli"
        elif health.get("last_call_ok") is False:
            derived_status = "hata"
        else:
            derived_status = "bagli_degil"
    return UnicontaConfigOut(
        companyId=settings.uniconta_company_id,
        username=settings.uniconta_username,
        env="production",
        apiUrl=UNICONTA_WEB_API_BASE,
        connectionStatus=derived_status,
        configured=configured,
        passwordConfigured=bool(settings.uniconta_password.strip()),
        apiKeyConfigured=bool(settings.uniconta_api_key.strip()),
        lastRefreshedAt=last_refreshed_at,
        message=message or ("Uniconta proxy hazir." if configured else "Uniconta baglantisi henuz yapilandirilmadi."),
        sendEmailOnFinalize=bool(getattr(settings, "uniconta_send_email_on_finalize", False)),
        sendXmlOnFinalize=bool(getattr(settings, "uniconta_send_xml_on_finalize", False)),
    )


def _build_uniconta_invoice(
    document: PosDocument,
    pos_session: PosSession | None,
    transaction: Transaction | None,
    lines: list[TransactionLine],
) -> UnicontaInvoiceOut:
    invoice_number = format_document_number(document)
    signed_total_amount = _to_float(document.gross_amount_dkk)
    invoice_type = "Kreditnota" if signed_total_amount < 0 else "Salgsfaktura"
    amount_direction = (
        "income"
        if signed_total_amount > 0
        else "expense"
        if signed_total_amount < 0
        else "neutral"
    )
    line_items = [
        UnicontaInvoiceLineOut(
            id=str(line.id),
            beskrivelse=((line.product_type or "Kalem").replace("_", " ").title()),
            antal=1,
            enhedspris=_to_float(line.rate_dkk),
            rabat=_to_float(line.margin_percent),
            moms=_to_float(document.vat_rate_percent),
            liniepris=_to_float(line.line_total_dkk),
            dato=document.issued_at.date().isoformat(),
        )
        for line in lines
    ]
    account_label = document.customer_email or document.customer_phone or (pos_session.session_code if pos_session else f"UC-{invoice_number}")
    customer_id = str(transaction.customer_id) if transaction and transaction.customer_id else (str(pos_session.customer_id) if pos_session and pos_session.customer_id else f"customer-{document.sequence_no}")
    return UnicontaInvoiceOut(
        id=str(document.sequence_no),
        fakturanummer=invoice_number,
        ordrenummer=(pos_session.session_code if pos_session else None),
        type=invoice_type,
        fakturadato=document.issued_at.date().isoformat(),
        konto=account_label,
        mailSendt=None,
        eFakturaSendt=None,
        kunde=UnicontaInvoiceCustomerOut(
            id=customer_id,
            navn=document.customer_name or "Bilinmeyen Musteri",
            email=document.customer_email,
            telefon=document.customer_phone,
            adresse=document.customer_address,
            postnr=None,
            cvr=None,
        ),
        kalemler=line_items,
        subtotal=_to_float(document.net_amount_dkk),
        momsTotal=_to_float(document.vat_amount_dkk),
        total=signed_total_amount,
        signedTotalAmount=signed_total_amount,
        amountDirection=amount_direction,
        valuta=document.currency_code,
        note=document.notes,
        wooOrderId=None,
        unicontaRef=f"UC-{invoice_number}",
    )


async def _load_uniconta_invoices(
    db: AsyncSession,
    *,
    sequence_no: int | None = None,
    limit: int = 200,
    skip: int = 0,
) -> list[UnicontaInvoiceOut]:
    stmt = (
        select(PosDocument, PosSession, Transaction)
        .join(PosSession, PosSession.id == PosDocument.pos_session_id)
        .outerjoin(Transaction, Transaction.pos_document_sequence_no == PosDocument.sequence_no)
        .where(PosDocument.document_type == PosDocumentTypeEnum.SALE_INVOICE)
        .order_by(PosDocument.issued_at.desc(), PosDocument.sequence_no.desc())
    )
    if sequence_no is not None:
        stmt = stmt.where(PosDocument.sequence_no == sequence_no)
    else:
        stmt = stmt.offset(skip).limit(limit)

    rows = (await db.execute(stmt)).all()
    if not rows:
        return []

    transaction_ids = [transaction.id for _, _, transaction in rows if transaction is not None]
    line_rows = (
        await db.execute(
            select(TransactionLine)
            .where(TransactionLine.transaction_id.in_(transaction_ids))
            .order_by(TransactionLine.transaction_id.asc(), TransactionLine.line_no.asc())
        )
    ).scalars().all() if transaction_ids else []

    lines_by_transaction: dict[UUID, list[TransactionLine]] = defaultdict(list)
    for line in line_rows:
        lines_by_transaction[line.transaction_id].append(line)

    return [
        _build_uniconta_invoice(
            document,
            pos_session,
            transaction,
            lines_by_transaction.get(transaction.id, []) if transaction is not None else [],
        )
        for document, pos_session, transaction in rows
    ]


async def _build_dashboard_screen(db: AsyncSession, admin: User) -> DashboardScreenOut:
    settings = get_settings()
    inventory_workspace = await get_legacy_inventory_workspace(q=None, db=db, _=admin)
    log_workspace = await build_log_workspace(db, q=None, limit=200)

    purchase_rows = (
        await db.execute(
            select(PosDocument, PosSession)
            .join(PosSession, PosSession.id == PosDocument.pos_session_id)
            .where(PosDocument.document_type == PosDocumentTypeEnum.PURCHASE_RECEIPT)
            .order_by(PosDocument.issued_at.desc(), PosDocument.sequence_no.desc())
        )
    ).all()
    alis_sayisi = len(purchase_rows)
    alis_toplam_kr = sum(_to_float(document.gross_amount_dkk) for document, _ in purchase_rows)
    son_alislar = [
        DashboardRecentPurchaseOut(
            id=str(document.sequence_no),
            afregningsnr=format_document_number(document),
            dato=document.issued_at.isoformat(),
            musteri=document.customer_name or "—",
            total=_to_float(document.gross_amount_dkk),
            paymentMethod=extract_purchase_payment_method(document.notes or pos_session.notes),
        )
        for document, pos_session in purchase_rows[:6]
    ]

    monthly_map: dict[str, dict[str, float | int]] = defaultdict(lambda: {"adet": 0, "kr": 0.0})
    for document, _ in purchase_rows:
        month_key = document.issued_at.strftime("%Y-%m")
        monthly_map[month_key]["adet"] = int(monthly_map[month_key]["adet"]) + 1
        monthly_map[month_key]["kr"] = float(monthly_map[month_key]["kr"]) + _to_float(document.gross_amount_dkk)

    aylik_alis = []
    for month_key in sorted(monthly_map.keys())[-6:]:
        year, month = month_key.split("-")
        aylik_alis.append(
            DashboardMonthlyPurchasePointOut(
                ay=f"{MONTH_NAMES_TR[int(month) - 1]} {year[-2:]}",
                adet=int(monthly_map[month_key]["adet"]),
                kr=float(monthly_map[month_key]["kr"]),
            )
        )

    total_customers = int(
        await db.scalar(select(func.count(User.id)).where(User.role == RoleEnum.CUSTOMER, User.is_active.is_(True))) or 0
    )
    recent_customer_rows = (
        await db.execute(
            select(User)
            .where(User.role == RoleEnum.CUSTOMER, User.is_active.is_(True))
            .order_by(User.created_at.desc())
            .limit(5)
        )
    ).scalars().all()
    recent_customers = [
        DashboardRecentCustomerOut(
            id=str(customer.id),
            navn=customer.name,
            kayitTarihi=customer.created_at.isoformat(),
        )
        for customer in recent_customer_rows
    ]

    depo_by_cat_map: dict[str, dict[str, float | str]] = defaultdict(lambda: {"gram": 0.0, "spot": 0.0})
    woo_hazir = 0
    woo_foto = 0
    woo_listelendi = 0
    for row in inventory_workspace.rows:
        key = row.main_category
        depo_by_cat_map[key]["gram"] = float(depo_by_cat_map[key]["gram"]) + _to_float(row.has_metal_grams or row.toplam_gram)
        depo_by_cat_map[key]["spot"] = float(depo_by_cat_map[key]["spot"]) + _to_float(row.spot_degeri_dkk)
        if row.shop_sync_status == "hazir":
            woo_hazir += 1
        elif row.shop_sync_status == "mangler_foto":
            woo_foto += 1
        elif row.shop_sync_status == "listelendi":
            woo_listelendi += 1
    depo_by_cat = [
        DashboardCategorySpotOut(
            name=CAT_LABELS.get(key, key),
            gram=float(values["gram"]),
            spot=float(values["spot"]),
            color=CAT_COLORS.get(key, "#888888"),
        )
        for key, values in depo_by_cat_map.items()
    ]

    opmc_yuksek = 0
    opmc_orta = 0
    opmc_dusuk = 0
    opmc_belirsiz = 0
    opmc_manuel = 0
    try:
        opmc_orders = await get_legacy_antifraud_recent_orders(
            days=30,
            per_page=50,
            include_notes=False,
            notes_per_order=5,
            detail_mode=False,
            _=admin,
        )
        opmc_yuksek = opmc_orders.summary.high_risk_count
        opmc_orta = opmc_orders.summary.medium_risk_count
        opmc_dusuk = opmc_orders.summary.low_risk_count
        opmc_belirsiz = opmc_orders.summary.unknown_risk_count
        opmc_manuel = opmc_orders.summary.manual_review_count
    except Exception:
        pass

    invoice_rows = (
        await db.execute(
            select(PosDocument)
            .where(PosDocument.document_type == PosDocumentTypeEnum.SALE_INVOICE)
            .order_by(PosDocument.issued_at.desc(), PosDocument.sequence_no.desc())
        )
    ).scalars().all()

    return DashboardScreenOut(
        alisSayisi=alis_sayisi,
        alisToplamKr=alis_toplam_kr,
        sonAlislar=son_alislar,
        aylikAlis=aylik_alis,
        musteriSayisi=total_customers,
        sonMusteriler=recent_customers,
        depoToplamItem=inventory_workspace.summary.total_items,
        depoSpotDeger=_to_float(inventory_workspace.summary.total_spot_value_dkk),
        depoAlisDeger=_to_float(inventory_workspace.summary.total_purchase_value_dkk),
        depoByCat=depo_by_cat,
        wooHazir=woo_hazir,
        wooFoto=woo_foto,
        wooLisitlendi=woo_listelendi,
        logSayisi=log_workspace.summary.total_documents,
        ayirmaSayisi=log_workspace.gold.summary.split_line_count + log_workspace.silver.summary.split_line_count,
        eritmeSayisi=log_workspace.gold.summary.melt_lot_count + log_workspace.silver.summary.melt_lot_count,
        eritmeToplamHasAltin=sum(_to_float(item.after_pure_gold_grams) for item in log_workspace.gold.melt_lots) + sum(_to_float(item.after_pure_gold_grams) for item in log_workspace.silver.melt_lots),
        eritmeToplamPayout=sum(_to_float(item.payout_total_dkk) for item in log_workspace.gold.melt_lots) + sum(_to_float(item.payout_total_dkk) for item in log_workspace.silver.melt_lots),
        goldPrice=_to_float(settings.inventory_market_gold_dkk),
        silverPrice=_to_float(settings.inventory_market_silver_dkk),
        platinPrice=_to_float(settings.inventory_market_platinum_dkk),
        palladyumPrice=_to_float(settings.inventory_market_palladium_dkk),
        opmcYuksek=opmc_yuksek,
        opmcOrta=opmc_orta,
        opmcDusuk=opmc_dusuk,
        opmcBelirsiz=opmc_belirsiz,
        opmcManuel=opmc_manuel,
        faturaAdedi=len(invoice_rows),
        faturaToplamKr=sum(_to_float(document.gross_amount_dkk) for document in invoice_rows),
    )


def _build_woocommerce_workspace(inventory_workspace: InventoryWorkspaceOut) -> WooWorkspaceOut:
    rows = inventory_workspace.rows
    return WooWorkspaceOut(
        summary=WooWorkspaceSummaryOut(
            total_products=len(rows),
            published_products=sum(1 for row in rows if row.shop_sync_status == "listelendi"),
            draft_products=sum(1 for row in rows if row.shop_sync_status == "hazir"),
            unpublished_products=sum(1 for row in rows if not row.shop_sync_status or row.shop_sync_status == "mangler_foto"),
            photo_pending_products=sum(1 for row in rows if not row.primary_photo),
        ),
        rows=rows,
    )


@router.get("/bootstrap", response_model=DesktopBootstrapOut)
async def get_bootstrap_v2(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_password_change_complete),
) -> DesktopBootstrapOut:
    return await get_legacy_bootstrap(db=db, current_user=current_user)


@router.get("/dashboard", response_model=DashboardScreenOut)
async def get_dashboard_v2(
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> DashboardScreenOut:
    return await _build_dashboard_screen(db, admin)


@router.get("/settings", response_model=SettingsScreenOut)
async def get_settings_v2(
    _: User = Depends(require_admin),
) -> SettingsScreenOut:
    return _build_settings_screen_out()


@router.put("/settings", response_model=SettingsScreenOut)
async def put_settings_v2(
    payload: SettingsScreenUpdateIn,
    _: User = Depends(require_admin),
) -> SettingsScreenOut:
    updates = {
        "OPENAI_MODEL": payload.openai_model.strip() or "gpt-5.4",
        "OPENAI_MAX_TOKENS": payload.openai_max_tokens.strip() or "4096",
        "OPMC_API_URL": payload.opmc_api_url.strip(),
        "WOOCOMMERCE_BASE_URL": payload.woo_store_url.strip(),
        "WORDPRESS_BASE_URL": payload.wp_site_url.strip(),
        "WP_APP_USERNAME": payload.wp_username.strip(),
        "UNICONTA_API_URL": UNICONTA_WEB_API_BASE,
        "UNICONTA_USERNAME": payload.uniconta_username.strip(),
        "UNICONTA_COMPANY_ID": payload.uniconta_company_id.strip(),
        "UNICONTA_PURCHASE_VAT_CODE_25": payload.uniconta_purchase_vat_code_25.strip() or "Købsmoms",
        "UNICONTA_PURCHASE_VAT_CODE_0": payload.uniconta_purchase_vat_code_0.strip() or "KøbBrugtmoms",
        "INVENTORY_MARKET_GOLD_DKK": payload.market_gold.strip() or "2850",
        "INVENTORY_MARKET_SILVER_DKK": payload.market_silver.strip() or "8.5",
        "INVENTORY_MARKET_PLATINUM_DKK": payload.market_platin.strip() or "280",
        "INVENTORY_MARKET_PALLADIUM_DKK": payload.market_palladyum.strip() or "335",
        "MARKET_RATES_LIVE_ENABLED": "true" if payload.market_rates_live_enabled else "false",
        "MARKET_RATES_LIVE_FX_ENABLED": "true" if payload.market_rates_live_fx_enabled else "false",
        "MARKET_RATES_LIVE_PLATINUM_ENABLED": "true" if payload.market_rates_live_platinum_enabled else "false",
        "MARKET_RATES_LIVE_PALLADIUM_ENABLED": "true" if payload.market_rates_live_palladium_enabled else "false",
        "INVOICE_SELLER_NAME": payload.firma_adi.strip(),
        "INVOICE_SELLER_CVR": payload.firma_cvr.strip(),
        "INVOICE_SELLER_PHONE": payload.firma_telefon.strip(),
        "INVOICE_SELLER_EMAIL": payload.firma_email.strip(),
        "INVOICE_SELLER_ADDRESS_LINE1": payload.firma_adres.strip(),
    }
    secret_updates = {
        "OPENAI_API_KEY": payload.openai_api_key,
        "OPMC_API_KEY": payload.opmc_api_key,
        "OPMC_WEBHOOK_SECRET": payload.opmc_webhook_secret,
        "WOOCOMMERCE_CONSUMER_KEY": payload.woo_consumer_key,
        "WOOCOMMERCE_CONSUMER_SECRET": payload.woo_consumer_secret,
        "WOOCOMMERCE_WEBHOOK_SECRET": payload.woo_webhook_secret,
        "WP_APP_PASSWORD": payload.wp_app_password,
        "UNICONTA_PASSWORD": payload.uniconta_password,
        "UNICONTA_API_KEY": payload.uniconta_api_key,
        "METALS_DEV_API_KEY": payload.metals_dev_api_key,
    }
    updates.update(
        {
            key: value.strip()
            for key, value in secret_updates.items()
            if value is not None and value.strip()
        }
    )
    upsert_env_values(ROOT_ENV_FILE, updates)
    get_settings.cache_clear()
    reset_uniconta_client()
    return _build_settings_screen_out()


@router.get("/uniconta/config", response_model=UnicontaConfigOut)
async def get_uniconta_config_v2(
    _: User = Depends(require_admin),
) -> UnicontaConfigOut:
    return _build_uniconta_config_out()


@router.post("/uniconta/connect", response_model=UnicontaConnectOut)
async def post_uniconta_connect_v2(
    payload: UnicontaConnectIn,
    _: User = Depends(require_admin),
) -> UnicontaConnectOut:
    current = get_settings()
    company_id = payload.companyId.strip() or current.uniconta_company_id.strip()
    username = payload.username.strip() or current.uniconta_username.strip()
    password = (payload.password or "").strip() or current.uniconta_password.strip()
    configured = bool(company_id and username and password)
    if not configured:
        message = "Uniconta baglanti bilgileri eksik."
        config = _build_uniconta_config_out(message=message, connection_status="bagli_degil")
        return UnicontaConnectOut(
            connectionStatus="bagli_degil",
            configured=False,
            message=message,
            config=config,
        )
    # Aday değerleri önce ayrı bir istemciyle doğrula. Başarısız bir test mevcut
    # çalışan kimlik bilgilerini veya token cache'ini değiştirmemelidir.
    client = UnicontaClient(
        base_url=UNICONTA_WEB_API_BASE,
        company_id=company_id,
        username=username,
        password=password,
    )
    result = await client.test_connection()
    last_refreshed_at = utc_now().isoformat() if result.get("ok") else None
    if not result.get("ok"):
        message = result.get("message", "Uniconta baglantisi basarisiz.")
        config = _build_uniconta_config_out(message=message, connection_status="hata")
        return UnicontaConnectOut(
            connectionStatus="hata",
            configured=config.configured,
            message=message,
            config=config,
        )

    existing_api_key = current.uniconta_api_key.strip()
    supplied_api_key = (payload.apiKey or "").strip()
    upsert_env_values(
        ROOT_ENV_FILE,
        {
            "UNICONTA_API_URL": UNICONTA_WEB_API_BASE,
            "UNICONTA_USERNAME": username,
            "UNICONTA_PASSWORD": password,
            "UNICONTA_COMPANY_ID": company_id,
            "UNICONTA_API_KEY": supplied_api_key or existing_api_key,
            "UNICONTA_SEND_EMAIL_ON_FINALIZE": "true" if payload.sendEmailOnFinalize else "false",
            "UNICONTA_SEND_XML_ON_FINALIZE": "true" if payload.sendXmlOnFinalize else "false",
        },
    )
    get_settings.cache_clear()
    reset_uniconta_client()
    company_name = (result.get("company") or {}).get("CompanyName")
    message = f"Uniconta'ya baglandi: {company_name}" if company_name else result.get("message", "Uniconta baglantisi basarili.")
    config = _build_uniconta_config_out(
        message=message,
        connection_status="bagli",
        last_refreshed_at=last_refreshed_at,
    )
    return UnicontaConnectOut(
        connectionStatus="bagli",
        configured=True,
        message=message,
        config=config,
    )


@router.get("/uniconta/invoices", response_model=UnicontaInvoicesOut)
async def get_uniconta_invoices_v2(
    limit: int = Query(default=200, ge=1, le=500),
    skip: int = Query(default=0, ge=0),
    source: str = Query(default="local", pattern="^(local|remote)$"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> UnicontaInvoicesOut:
    if source == "remote":
        client = get_uniconta_client()
        try:
            rows = await client.get_sale_invoices(top=limit, skip=skip)
        except UnicontaError as exc:
            raise HTTPException(status_code=502, detail=f"Uniconta remote: {exc}") from exc
        invoices_remote = [UnicontaInvoiceOut(**map_uniconta_invoice_to_dto(r)) for r in rows]
        # Uniconta liste kaydı kalem taşımaz; CRM'in senkronladığı faturalar
        # için satırlar ve müşteri detayı yerel kayıttan hidre edilir.
        numbers = [inv.fakturanummer for inv in invoices_remote if inv.fakturanummer]
        if numbers:
            local_rows = (
                await db.execute(
                    select(PosDocument, PosSession, Transaction)
                    .join(PosSession, PosSession.id == PosDocument.pos_session_id)
                    .outerjoin(Transaction, Transaction.pos_document_sequence_no == PosDocument.sequence_no)
                    .where(PosDocument.uniconta_invoice_number.in_(numbers))
                )
            ).all()
            transaction_ids = [t.id for _, _, t in local_rows if t is not None]
            line_rows = (
                (
                    await db.execute(
                        select(TransactionLine)
                        .where(TransactionLine.transaction_id.in_(transaction_ids))
                        .order_by(TransactionLine.transaction_id.asc(), TransactionLine.line_no.asc())
                    )
                ).scalars().all()
                if transaction_ids
                else []
            )
            lines_by_transaction: dict[UUID, list[TransactionLine]] = defaultdict(list)
            for line in line_rows:
                lines_by_transaction[line.transaction_id].append(line)
            local_by_number: dict[str, UnicontaInvoiceOut] = {}
            for document, pos_session, transaction in local_rows:
                built = _build_uniconta_invoice(
                    document,
                    pos_session,
                    transaction,
                    lines_by_transaction.get(transaction.id, []) if transaction is not None else [],
                )
                if document.uniconta_invoice_number:
                    local_by_number[str(document.uniconta_invoice_number)] = built
            for invoice in invoices_remote:
                local = local_by_number.get(invoice.fakturanummer)
                if local is None:
                    continue
                invoice.kalemler = local.kalemler
                invoice.kunde = local.kunde
        return UnicontaInvoicesOut(
            source="uniconta_remote",
            generatedAt=datetime.utcnow().isoformat(),
            invoices=invoices_remote,
            skip=skip,
            limit=limit,
            hasMore=len(invoices_remote) == limit,
        )
    invoices = await _load_uniconta_invoices(db, limit=limit, skip=skip)
    return UnicontaInvoicesOut(
        source="crm_sale_invoices",
        generatedAt=datetime.utcnow().isoformat(),
        invoices=invoices,
        skip=skip,
        limit=limit,
        hasMore=len(invoices) == limit,
    )


@router.get("/uniconta/invoices/{sequence_no}", response_model=UnicontaInvoiceOut)
async def get_uniconta_invoice_v2(
    sequence_no: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> UnicontaInvoiceOut:
    invoices = await _load_uniconta_invoices(db, sequence_no=sequence_no, limit=1)
    if not invoices:
        raise HTTPException(status_code=404, detail="Fatura bulunamadi.")
    return invoices[0]


@router.get("/uniconta/invoice-pdf")
async def get_uniconta_invoice_pdf_v2(
    invoiceNumber: int = Query(..., description="Uniconta InvoiceNumber"),
    account: str = Query(..., description="DebtorClient Account kodu"),
    date: str = Query(..., description="Fatura tarihi (YYYY-MM-DD)"),
    _: User = Depends(require_admin),
) -> Response:
    """Var olan bir Uniconta DebtorInvoice'ın PDF'ini stream eder."""
    client = get_uniconta_client()
    try:
        pdf_bytes = await client.get_invoice_pdf(
            invoice_number=invoiceNumber,
            account=account,
            date=date,
        )
    except UnicontaError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="uniconta-{invoiceNumber}.pdf"',
            "Cache-Control": "no-store",
        },
    )


@router.get("/uniconta/invoice-pdf/from-pos/{sequence_no}")
async def get_uniconta_invoice_pdf_from_pos_v2(
    sequence_no: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> Response:
    """PosDocument referansından Uniconta DebtorInvoice PDF'i serve et.

    Önce data/documents/uniconta/{seq}.pdf cache'ine bak; yoksa Uniconta'dan
    canlı çek (PosDocument.uniconta_invoice_number + account + date varsa).
    """
    from pathlib import Path as _Path

    from app.models.pos_document import PosDocument

    stmt = select(PosDocument).where(PosDocument.sequence_no == sequence_no)
    row = (await db.execute(stmt)).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="PosDocument bulunamadi.")

    if row.uniconta_pdf_path:
        cached = _Path(row.uniconta_pdf_path)
        if cached.exists():
            return Response(
                content=cached.read_bytes(),
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'inline; filename="uniconta-{sequence_no}.pdf"',
                    "Cache-Control": "no-store",
                },
            )

    if not (row.uniconta_invoice_number and row.uniconta_account and row.uniconta_invoice_date):
        raise HTTPException(
            status_code=409,
            detail="Bu belge için henüz Uniconta sync tamamlanmadı veya başarısız.",
        )

    client = get_uniconta_client()
    try:
        pdf_bytes = await client.get_invoice_pdf(
            invoice_number=int(row.uniconta_invoice_number),
            account=row.uniconta_account,
            date=row.uniconta_invoice_date,
        )
    except UnicontaError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="uniconta-{sequence_no}.pdf"',
            "Cache-Control": "no-store",
        },
    )


def _categorize_sync_error(msg: str | None) -> str:
    """Hatayı kabaca kategorize et — Sync summary'de gruplandırma için."""
    if not msg:
        return "unknown"
    lower = msg.lower()
    if "timeout" in lower or "network" in lower or "connect" in lower:
        return "network"
    if "401" in lower or "auth" in lower or "credential" in lower:
        return "auth"
    if "400" in lower or "validation" in lower or "invalid" in lower:
        return "validation"
    if "5" in lower[:4]:
        return "server"
    if "skipped" in lower or "credentials missing" in lower:
        return "skipped"
    return "other"


@router.get("/uniconta/sync-summary", response_model=UnicontaSyncSummaryOut)
async def get_uniconta_sync_summary_v2(
    hours: int = Query(default=24, ge=1, le=720),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> UnicontaSyncSummaryOut:
    """Son N saatte (default 24) PosDocument sync istatistikleri."""
    from datetime import timedelta, timezone as _tz

    cutoff = datetime.now(_tz.utc) - timedelta(hours=hours)
    stmt = (
        select(PosDocument)
        .where(PosDocument.created_at >= cutoff)
        .order_by(PosDocument.created_at.desc())
    )
    docs = list((await db.execute(stmt)).scalars().all())

    synced = failed = skipped = pending = 0
    by_cat: dict[str, int] = {}
    last_synced: datetime | None = None
    last_failure: datetime | None = None
    for doc in docs:
        s = doc.uniconta_sync_status or "pending"
        if s == "synced":
            synced += 1
            if doc.uniconta_synced_at and (last_synced is None or doc.uniconta_synced_at > last_synced):
                last_synced = doc.uniconta_synced_at
        elif s == "failed":
            failed += 1
            cat = _categorize_sync_error(doc.uniconta_sync_error)
            by_cat[cat] = by_cat.get(cat, 0) + 1
            if doc.created_at and (last_failure is None or doc.created_at > last_failure):
                last_failure = doc.created_at
        elif s == "skipped":
            skipped += 1
            by_cat["skipped"] = by_cat.get("skipped", 0) + 1
        else:
            pending += 1

    return UnicontaSyncSummaryOut(
        period_hours=hours,
        total=len(docs),
        synced=synced,
        failed=failed,
        skipped=skipped,
        pending=pending,
        by_error_category=by_cat,
        last_synced_at=last_synced.isoformat() if last_synced else None,
        last_failure_at=last_failure.isoformat() if last_failure else None,
    )


@router.get("/uniconta/failed-syncs", response_model=list[UnicontaFailedSyncRowOut])
async def get_uniconta_failed_syncs_v2(
    status_filter: str = Query(default="failed", pattern="^(failed|skipped|all)$"),
    limit: int = Query(default=100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> list[UnicontaFailedSyncRowOut]:
    """Sync edilmemiş AFG'leri döner — UI'da accordion + bulk retry için."""
    stmt = select(PosDocument).order_by(PosDocument.sequence_no.desc()).limit(limit)
    if status_filter == "failed":
        stmt = stmt.where(PosDocument.uniconta_sync_status == "failed")
    elif status_filter == "skipped":
        stmt = stmt.where(PosDocument.uniconta_sync_status == "skipped")
    else:
        stmt = stmt.where(PosDocument.uniconta_sync_status.in_(("failed", "skipped")))
    docs = list((await db.execute(stmt)).scalars().all())
    out: list[UnicontaFailedSyncRowOut] = []
    for doc in docs:
        out.append(
            UnicontaFailedSyncRowOut(
                sequence_no=doc.sequence_no,
                document_number=format_document_number(doc) if hasattr(doc, "sequence_no") else None,
                issued_at=doc.issued_at.isoformat() if doc.issued_at else None,
                customer_name=doc.customer_name,
                gross_amount_dkk=str(doc.gross_amount_dkk) if doc.gross_amount_dkk is not None else None,
                uniconta_sync_status=doc.uniconta_sync_status,
                uniconta_sync_error=doc.uniconta_sync_error,
                uniconta_synced_at=doc.uniconta_synced_at.isoformat() if doc.uniconta_synced_at else None,
            )
        )
    return out


@router.post("/uniconta/sync-retry-all", response_model=UnicontaBulkRetryOut)
async def post_uniconta_sync_retry_all_v2(
    request: Request,
    limit: int = Query(default=50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> UnicontaBulkRetryOut:
    """Failed/skipped PosDocument'lar için toplu Uniconta sync retry.

    Max `limit` adet PosDocument'ı sırayla retry eder; her birinin sonucu
    `results` listesinde dönderir.
    """
    from pathlib import Path as _Path
    from app.models.pos_document_audit import PosDocumentAudit
    from app.models.pos_session import PosSession
    from app.models.pos_session_line import PosSessionLine
    from app.services.uniconta_service import sync_pos_document_to_uniconta

    stmt = (
        select(PosDocument)
        .where(PosDocument.uniconta_sync_status.in_(("failed", "skipped")))
        .order_by(PosDocument.sequence_no.desc())
        .limit(limit)
    )
    docs = list((await db.execute(stmt)).scalars().all())
    settings = get_settings()
    cache_dir = str(_Path(settings.document_root_path()) / "uniconta")

    attempted = 0
    succeeded = 0
    failed = 0
    results: list[dict] = []
    for doc in docs:
        attempted += 1
        sess = (
            await db.execute(select(PosSession).where(PosSession.id == doc.pos_session_id))
        ).scalar_one_or_none()
        if sess is None:
            failed += 1
            results.append({"sequence_no": doc.sequence_no, "ok": False, "message": "PosSession bulunamadı"})
            continue
        lines = list(
            (
                await db.execute(
                    select(PosSessionLine)
                    .where(PosSessionLine.pos_session_id == sess.id)
                    .order_by(PosSessionLine.line_no.asc())
                )
            ).scalars().all()
        )
        result = await sync_pos_document_to_uniconta(
            db, doc, pos_session=sess, pos_lines=lines, pdf_cache_dir=cache_dir
        )
        if result.get("ok"):
            succeeded += 1
        else:
            failed += 1
        # Audit
        db.add(
            PosDocumentAudit(
                sequence_no=doc.sequence_no,
                pos_session_id=sess.id,
                action="uniconta_bulk_retry",
                actor_user_id=getattr(admin, "id", None),
                actor_email=getattr(admin, "email", None),
                payload_json=json.dumps(
                    {
                        "ok": bool(result.get("ok")),
                        "uniconta_sync_status": doc.uniconta_sync_status,
                        "uniconta_invoice_number": doc.uniconta_invoice_number,
                    },
                    default=str,
                    ensure_ascii=False,
                ),
                note=str(result.get("message") or "") or None,
                request_ip=request.client.host if request.client else None,
            )
        )
        results.append(
            {
                "sequence_no": doc.sequence_no,
                "ok": bool(result.get("ok")),
                "message": result.get("message"),
                "uniconta_sync_status": doc.uniconta_sync_status,
                "uniconta_invoice_number": doc.uniconta_invoice_number,
            }
        )
    await db.commit()
    return UnicontaBulkRetryOut(
        attempted=attempted,
        succeeded=succeeded,
        failed=failed,
        results=results,
    )


@router.get("/uniconta/health", response_model=UnicontaHealthOut)
async def get_uniconta_health_v2(
    _: User = Depends(require_admin),
) -> UnicontaHealthOut:
    """Token + son çağrı sağlığı (memory snapshot, no remote ping)."""
    from app.services.uniconta_service import get_uniconta_client

    snap = get_uniconta_client().get_health_snapshot()
    return UnicontaHealthOut(**snap)


@router.post("/uniconta/invoice/from-pos/{sequence_no}")
async def post_uniconta_invoice_from_pos_v2(
    sequence_no: int,
    request: Request,
    force: bool = Query(default=False, description="True ise zaten 'synced' olan PosDocument bile yeniden Uniconta'ya gönderilir (duplicate riski operatöre)."),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, object]:
    """Manuel "Tekrar Dene" — başarısız Uniconta sync için yeniden gönderim.

    Idempotency: default `force=False` ise zaten synced kayıt için Uniconta'ya
    yeni istek atmadan mevcut bilgi döner (duplicate koruması).
    """
    from pathlib import Path as _Path

    from app.models.pos_document import PosDocument
    from app.models.pos_document_audit import PosDocumentAudit
    from app.models.pos_session import PosSession
    from app.models.pos_session_line import PosSessionLine
    from app.services.uniconta_service import sync_pos_document_to_uniconta

    stmt = select(PosDocument).where(PosDocument.sequence_no == sequence_no)
    pos_document = (await db.execute(stmt)).scalar_one_or_none()
    if pos_document is None:
        raise HTTPException(status_code=404, detail="PosDocument bulunamadi.")

    session_stmt = select(PosSession).where(PosSession.id == pos_document.pos_session_id)
    pos_session = (await db.execute(session_stmt)).scalar_one_or_none()
    if pos_session is None:
        raise HTTPException(status_code=404, detail="PosSession bulunamadi.")

    lines_stmt = (
        select(PosSessionLine)
        .where(PosSessionLine.pos_session_id == pos_session.id)
        .order_by(PosSessionLine.line_no.asc())
    )
    pos_lines = list((await db.execute(lines_stmt)).scalars().all())

    settings = get_settings()
    cache_dir = str(_Path(settings.document_root_path()) / "uniconta")
    result = await sync_pos_document_to_uniconta(
        db,
        pos_document,
        pos_session=pos_session,
        pos_lines=pos_lines,
        pdf_cache_dir=cache_dir,
        force=force,
    )

    db.add(
        PosDocumentAudit(
            sequence_no=sequence_no,
            pos_session_id=pos_session.id,
            action="uniconta_retry",
            actor_user_id=getattr(admin, "id", None),
            actor_email=getattr(admin, "email", None),
            payload_json=json.dumps(
                {
                    "ok": bool(result.get("ok")),
                    "uniconta_sync_status": pos_document.uniconta_sync_status,
                    "uniconta_invoice_number": pos_document.uniconta_invoice_number,
                },
                default=str,
                ensure_ascii=False,
            ),
            note=str(result.get("message") or "") or None,
            request_ip=request.client.host if request.client else None,
        )
    )
    await db.commit()
    await db.refresh(pos_document)
    return {
        "ok": bool(result.get("ok")),
        "message": result.get("message"),
        "idempotent": bool(result.get("idempotent")),
        "uniconta_sync_status": pos_document.uniconta_sync_status,
        "uniconta_invoice_number": pos_document.uniconta_invoice_number,
        "uniconta_sync_error": pos_document.uniconta_sync_error,
    }


@router.get("/musteriler", response_model=CustomerListResponse)
async def get_musteriler_v2(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200),
    sort_by: str = Query(default="created_at"),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> CustomerListResponse:
    return await legacy_get_customers(page=page, page_size=page_size, sort_by=sort_by, db=db, _=admin)


@router.get("/musteriler/search", response_model=list[CustomerOut])
async def search_musteriler_v2(
    q: str = Query(min_length=2),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> list[CustomerOut]:
    return await legacy_search_customers(q=q, db=db, _=admin)


@router.post("/musteriler", response_model=CustomerOut)
async def post_musteriler_v2(
    payload: CustomerCreate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> CustomerOut:
    return await legacy_post_customer(payload=payload, db=db, _=admin)


@router.get("/musteriler/{customer_id}", response_model=CustomerDetailOut)
async def get_musteri_detail_v2(
    customer_id: UUID,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> CustomerDetailOut:
    return await legacy_get_customer(customer_id=customer_id, db=db, _=admin)


@router.get("/musteriler/{customer_id}/history", response_model=list[PosDocumentListItemOut])
async def get_musteri_history_v2(
    customer_id: UUID,
    limit: int = Query(default=100, ge=1, le=300),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> list[PosDocumentListItemOut]:
    return await legacy_get_customer_history(customer_id=customer_id, limit=limit, db=db, _=admin)


@router.get("/musteriler/{customer_id}/alis-summary", response_model=CustomerAlisSummaryOut)
async def get_musteri_alis_summary_v2(
    customer_id: UUID,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
) -> CustomerAlisSummaryOut:
    from datetime import timedelta, timezone

    from app.models.pos_session_line import PosSessionLine

    base_stmt = (
        select(PosDocument)
        .join(PosSession, PosSession.id == PosDocument.pos_session_id)
        .where(PosSession.customer_id == customer_id)
    )
    documents = (await db.execute(base_stmt)).scalars().all()

    total_documents = len(documents)
    total_amount = sum((doc.gross_amount_dkk or Decimal("0")) for doc in documents)
    total_amount_dec = Decimal(total_amount) if not isinstance(total_amount, Decimal) else total_amount
    avg_amount = (total_amount_dec / Decimal(total_documents)) if total_documents else Decimal("0")

    issued_dates = [doc.issued_at for doc in documents if doc.issued_at is not None]
    last_purchase_at = max(issued_dates).isoformat() if issued_dates else None
    first_purchase_at = min(issued_dates).isoformat() if issued_dates else None

    weight_stmt = (
        select(func.coalesce(func.sum(PosSessionLine.weight_grams), 0))
        .select_from(PosSessionLine)
        .join(PosSession, PosSession.id == PosSessionLine.pos_session_id)
        .join(PosDocument, PosDocument.pos_session_id == PosSession.id)
        .where(PosSession.customer_id == customer_id)
    )
    total_weight_value = (await db.execute(weight_stmt)).scalar_one()
    total_weight = Decimal(total_weight_value or 0)

    now = datetime.now(timezone.utc)
    cutoff_30 = now - timedelta(days=30)
    cutoff_365 = now - timedelta(days=365)
    docs_30 = [d for d in documents if d.issued_at and d.issued_at >= cutoff_30]
    docs_365 = [d for d in documents if d.issued_at and d.issued_at >= cutoff_365]
    amount_30 = sum((d.gross_amount_dkk or Decimal("0")) for d in docs_30)
    amount_365 = sum((d.gross_amount_dkk or Decimal("0")) for d in docs_365)

    return CustomerAlisSummaryOut(
        customer_id=str(customer_id),
        total_documents=total_documents,
        total_amount_dkk=f"{total_amount_dec:.2f}",
        total_weight_grams=f"{total_weight:.2f}",
        last_purchase_at=last_purchase_at,
        first_purchase_at=first_purchase_at,
        avg_amount_dkk=f"{avg_amount:.2f}",
        last_30d_documents=len(docs_30),
        last_30d_amount_dkk=f"{Decimal(amount_30):.2f}",
        last_365d_documents=len(docs_365),
        last_365d_amount_dkk=f"{Decimal(amount_365):.2f}",
    )


@router.put("/musteriler/{customer_id}", response_model=CustomerOut)
async def put_musteriler_v2(
    customer_id: UUID,
    payload: CustomerUpdate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> CustomerOut:
    return await legacy_put_customer(customer_id=customer_id, payload=payload, db=db, _=admin)


@router.delete("/musteriler/{customer_id}", response_class=Response, status_code=204)
async def delete_musteriler_v2(
    customer_id: UUID,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> Response:
    return await legacy_delete_customer(customer_id=customer_id, db=db, _=admin)


@router.get("/display/preview", response_model=PosDisplayPreviewOut)
async def get_display_preview_v2(
    db: AsyncSession = Depends(get_db),
    clerk_user: User = Depends(require_admin),
) -> PosDisplayPreviewOut:
    draft = await find_latest_draft_pos_session(
        db,
        clerk_user_id=clerk_user.id,
        trade_side=PosTradeSideEnum.BUY_FROM_CUSTOMER,
    )
    if draft is None:
        return PosDisplayPreviewOut()
    return PosDisplayPreviewOut(
        display_token=draft.display_token,
        snapshot=await display_snapshot(db, draft),
    )


@router.get("/runtime/status", response_model=RuntimeStatusOut)
async def get_runtime_status_v2(
    request: Request,
    _: User = Depends(require_admin),
) -> RuntimeStatusOut:
    return RuntimeStatusOut(
        app_name=get_settings().app_name,
        env=get_settings().env,
        backend_pid=os.getpid(),
        backend_started_at=BACKEND_STARTED_AT.isoformat(),
        backend_url=str(request.base_url).rstrip("/"),
        office_runtime_url=get_settings().office_runtime_url,
        office_wopi_base_url=get_settings().office_wopi_base_url,
        desktop_session=_read_desktop_dev_session(),
    )


@router.get("/runtime/readiness", response_model=RuntimeReadinessOut)
async def get_runtime_readiness_v2(
    _: User = Depends(require_admin),
) -> RuntimeReadinessOut:
    return await collect_runtime_readiness()


@router.websocket("/display/{display_token}/ws")
async def display_socket_v2(websocket: WebSocket, display_token: str):
    await legacy_display_socket(websocket, display_token)


from app.api.v2_alis import router as alis_router
from app.api.v2_log import router as log_router
from app.api.v2_office_runtime import router as office_runtime_router
from app.api.v2_woocommerce import router as woocommerce_router

router.include_router(alis_router)
router.include_router(log_router)
router.include_router(office_runtime_router)
router.include_router(woocommerce_router)
router.include_router(document_artifacts_router)
router.include_router(excel_sessions_router)
